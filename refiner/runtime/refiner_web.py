"""Primary Flask web/API server for Refiner.

This module hosts:
- authentication and session management,
- job orchestration/work queues,
- LLM assistant endpoints,
- voice/STT ingestion and gesture planning,
- RAG and MCP integrations, and
- operational features (metrics, auditing, token ledger, notifications).

It is intentionally monolithic for deployment simplicity, while route groups
are split into `refiner_routes/*` where practical.
"""

import json
import base64
import hashlib
import csv
import io
import logging
import math
import os
import posixpath
import queue
import re
import shlex
import signal
import smtplib
import secrets as secrets_lib
import subprocess
import sys
import threading
import time
import uuid
import datetime as dt
from collections import deque
from concurrent.futures import Future, ThreadPoolExecutor, TimeoutError as FutureTimeoutError, as_completed
from dataclasses import dataclass, field
from email.message import EmailMessage
from typing import Any, Callable, Deque, Dict, List, Optional, Tuple
from urllib.parse import urlparse, urlencode, quote
from zoneinfo import ZoneInfo

import requests
import shutil
import statistics
import tempfile
from flask import (
    Flask,
    Response,
    jsonify,
    make_response,
    render_template,
    request,
    redirect,
    session,
    url_for,
    send_from_directory,
    g,
    has_request_context,
)
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

from refiner.llm_providers import get_provider, LLMError, LLMProvider, register_event_callback
from refiner.rag_engine import RagIndex, RagStore
from refiner.stt_learning import SttLearningStore
from refiner.stt_gesture_planner import plan_stt_avatar_motion, sanitize_avatar_mode, sanitize_gesture_mode
from refiner.nmstt_contracts import NmsttGesturePlanRequest, sanitize_nmstt_motion_response
from refiner.mcp_client import MCPClient, MCPServerConfig, MCPServerStore, PostgresMCPServerStore
from refiner.nmchain_client import NmChainClient, NmChainError
from refiner.customers_client import CustomersClient, CustomersServiceError
from refiner.billing_client import BillingClient, BillingServiceError
from refiner.capabilities import get_capabilities, capability_summary, select_skills, format_skill_brief
from refiner.runtime.llm_access import (
    accessible_configured_llm_provider as _accessible_configured_llm_provider,
    build_effective_llm_env_for_user as _effective_llm_env_for_user,
    build_job_runtime_env as _build_job_runtime_env_impl,
    gemini_credential_from_env as _gemini_credential_from_env,
    provider_billing_map as _provider_billing_map,
    provider_billing_metadata as _provider_billing_metadata,
    provider_has_accessible_credentials as _provider_has_accessible_credentials,
    provider_base_url as _provider_base_url,
    user_can_use_shared_llm_credentials as _user_can_use_shared_llm_credentials,
)
from refiner.solver_memory import SolverEpisode, SolverEpisodeStore
from refiner.thought_inbox import (
    build_fingerprint as inbox_build_fingerprint,
    build_route_suggestion,
    build_thought_item,
    extract_keywords as inbox_extract_keywords,
    infer_kind as inbox_infer_kind,
    infer_priority as inbox_infer_priority,
    merge_duplicate_capture,
    normalize_text as normalize_thought_text,
    score_query_match,
)
from refiner_routes.voice import register_voice_routes
from refiner_routes.assistant import register_assistant_routes
from refiner_routes.rag import register_rag_routes
from refiner_routes.admin import register_admin_routes
from refiner_routes.auth import register_auth_routes
from refiner_routes.jobs import register_jobs_routes
from assistant_api.admin_handlers import build_assistant_admin_handlers
from assistant_api.assistant_handlers import build_assistant_handlers
from assistant_api.rag_handlers import build_rag_handlers
from assistant_pipeline import service as assistant_service
from assistant_pipeline.contracts import ServiceError
from assistant_pipeline.dependencies import AssistantPipelineDependencies
from assistant_pipeline.ingestion import (
    build_rag_documents as _build_rag_documents_impl,
    coerce_rag_sources as _coerce_rag_sources_impl,
)
from assistant_pipeline.retrieval import (
    render_rag_context as _render_rag_context_impl,
    serialize_rag_match as _serialize_rag_match_impl,
)
from assistant_pipeline.runtime import OwnerAwareCapacityLimiter
from assistant_pipeline.memory.episodic_store import (
    assistant_memory_entry_line as _assistant_memory_entry_line_impl,
    assistant_memory_matches as _assistant_memory_matches_impl,
    assistant_memory_prompt_block as _assistant_memory_prompt_block_impl,
    assistant_memory_query_text as _assistant_memory_query_text_impl,
    assistant_memory_reference_payload as _assistant_memory_reference_payload_impl,
    assistant_memory_requirement_ids as _assistant_memory_requirement_ids_impl,
    assistant_memory_scope as _assistant_memory_scope_impl,
    assistant_memory_summary as _assistant_memory_summary_impl,
    record_assistant_memory as _record_assistant_memory_impl,
    should_use_assistant_ask_memory as _should_use_assistant_ask_memory_impl,
)
from refiner.security_utils import (
    AuditLogger,
    attach_redaction_filter,
    ensure_dir_permissions,
    ensure_file_permissions,
    hash_identifier,
    url_allowed,
)
from refiner.refiner_central_store import (
    PostgresScheduleDocumentStore,
    PostgresSessionRoomStore,
    PostgresTodoDocumentStore,
    create_central_store_from_env,
)
from refiner.refiner_settings import (
    SettingsValidationError,
    default_settings as default_user_settings,
    llm_defaults_from_settings,
    metadata_with_settings,
    settings_from_metadata,
    validate_settings_patch,
)
from refiner.versioning import get_public_version_info, get_version_info
from refiner.refiner_ai_orchestration import build_workflow_provider, orchestration_status
from refiner.refiner_ai_model_inventory import AIModelInventoryMonitor
from refiner.integrations.atlassian.actions import execute_atlassian_action as _execute_atlassian_action_impl
from refiner.integrations.platform.continuum_client import (
    ContinuumClientConfig,
    continuum_enabled as _continuum_enabled_impl,
    continuum_headers as _continuum_headers_impl,
    continuum_json_payload as _continuum_json_payload_impl,
    continuum_request as _continuum_request_impl,
    continuum_ready as _continuum_ready_impl,
    friendly_continuum_error as _friendly_continuum_error_impl,
    workspace_action_create as _workspace_action_create_impl,
    workspace_action_refresh as _workspace_action_refresh_impl,
    workspace_template_vars as _workspace_template_vars_impl,
)
from refiner.runtime.continuum_autoscaler import (
    ContinuumQueueAutoscaler as _ContinuumQueueAutoscalerImpl,
    continuum_cluster_snapshot as _continuum_cluster_snapshot_impl,
    workers_telemetry_payload as _workers_telemetry_payload_impl,
)

logger = logging.getLogger(__name__)
try:
    attach_redaction_filter(logging.getLogger())
except Exception:
    pass


def _env_flag(name: str, default: bool = False) -> bool:
    """Read a boolean env var using a standard truthy value set."""
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "on"}


def _env_first(*names: str, default: str = "") -> str:
    """Return the first non-empty environment variable from a list."""
    for name in names:
        value = os.getenv(name)
        if value is None:
            continue
        value = value.strip()
        if value:
            return value
    return default


def _normalize_opencode_server_url() -> Optional[str]:
    """Normalize and validate the optional OpenCode server URL."""
    raw = os.getenv("OPENCODE_SERVER_URL")
    if not raw:
        return None
    cleaned = raw.strip()
    if not cleaned:
        return None
    if "://" not in cleaned:
        cleaned = f"http://{cleaned}"
    parsed = urlparse(cleaned)
    if not parsed.scheme or not parsed.netloc:
        return None
    return cleaned


def _opencode_available_for_playground() -> bool:
    """Detect whether OpenCode tooling is reachable for playground features."""
    if _normalize_opencode_server_url():
        return True
    opencode_bin = os.getenv("OPENCODE_BIN", "opencode")
    return bool(shutil.which(opencode_bin))

try:
    import redis  # type: ignore
except Exception:
    redis = None

try:
    from cryptography.fernet import Fernet, InvalidToken  # type: ignore
except Exception:
    Fernet = None
    InvalidToken = Exception
try:
    from cryptography import x509  # type: ignore
    from cryptography.hazmat.primitives import hashes  # type: ignore
    from cryptography.hazmat.primitives.asymmetric import padding  # type: ignore
    from cryptography.hazmat.primitives.serialization import Encoding, PublicFormat  # type: ignore
except Exception:
    x509 = None
    hashes = None
    padding = None
    Encoding = None
    PublicFormat = None
try:
    import jwt  # type: ignore
except Exception:
    jwt = None

RUNTIME_DIR = os.path.dirname(os.path.abspath(__file__))
PACKAGE_DIR = os.path.dirname(RUNTIME_DIR)
BASE_DIR = os.path.dirname(PACKAGE_DIR)


def _resolve_python_script_path(script_name: str) -> str:
    """Resolve a runtime script path, preferring source but allowing legacy bytecode."""
    for base_dir in (PACKAGE_DIR, BASE_DIR):
        script_path = os.path.join(base_dir, script_name)
        if os.path.exists(script_path):
            return script_path
        bytecode_path = f"{script_path}c"
        if os.path.exists(bytecode_path):
            return bytecode_path
    return os.path.join(PACKAGE_DIR, script_name)


PUBLIC_DIR = os.path.join(BASE_DIR, "web", "public")
JOB_ROOT = os.getenv("REFINER_JOB_DIR", os.path.join(BASE_DIR, "job_data"))
PROJECTS_ROOT = os.path.join(JOB_ROOT, "projects")
SECRET_STORE_ROOT = os.path.join(JOB_ROOT, "secrets")
ASSISTANT_MEMORY_ROOT = os.path.join(JOB_ROOT, "assistant_memory")
USERS_PATH = os.path.join(JOB_ROOT, "users.json")
WORKSPACE_ROOT = os.path.join(JOB_ROOT, "workspaces")
AUDIT_LOG_PATH = os.getenv("REFINER_AUDIT_LOG_PATH", os.path.join(JOB_ROOT, "audit.log"))
ACCESS_STORE_PATH = os.path.join(JOB_ROOT, "access.json")
SESSIONS_ROOT = os.path.join(JOB_ROOT, "sessions")
TODO_ROOT = os.path.join(JOB_ROOT, "todos")
SCHEDULE_ROOT = os.path.join(JOB_ROOT, "schedules")
TODO_CLAIM_TTL_SEC = max(60, int(os.getenv("REFINER_TODO_CLAIM_TTL_SEC", "300")))
TODO_RETENTION_DAYS = max(0, int(os.getenv("REFINER_TODO_RETENTION_DAYS", "0")))
TODO_SCHEDULER_ENABLED = _env_flag("REFINER_TODO_SCHEDULER_ENABLED", True)
TODO_SCHEDULER_POLL_SEC = max(1.0, float(os.getenv("REFINER_TODO_SCHEDULER_POLL_SEC", "5")))
TODO_SCHEDULE_CLAIM_TTL_SEC = max(30, int(os.getenv("REFINER_TODO_SCHEDULE_CLAIM_TTL_SEC", "180")))
TODO_SCHEDULE_EXECUTION_TIMEOUT_SEC = max(
    5.0,
    float(os.getenv("REFINER_TODO_SCHEDULE_EXECUTION_TIMEOUT_SEC", "90")),
)
LLM_TELEMETRY_RETENTION_HOURS = max(0, int(os.getenv("REFINER_LLM_TELEMETRY_RETENTION_HOURS", "2160")))
LLM_TELEMETRY_PRUNE_POLL_SEC = max(300.0, float(os.getenv("REFINER_LLM_TELEMETRY_PRUNE_POLL_SEC", "3600")))
VOICE_TOKEN_PATH = os.path.join(JOB_ROOT, "voice_tokens.json")
VOICE_DEFAULT_USER = (os.getenv("REFINER_VOICE_DEFAULT_USER") or "").strip()
VOICE_ENV_TOKEN = (os.getenv("REFINER_VOICE_TOKEN") or "").strip()
VOICE_ENV_TOKENS = (os.getenv("REFINER_VOICE_TOKENS") or "").strip()
VOICE_USER_MAP_RAW = (os.getenv("REFINER_VOICE_USER_MAP") or "").strip()
VOICE_ALLOW_TOKENS = _env_flag("REFINER_VOICE_ALLOW_TOKENS", True)
VOICE_ALLOW_TOKENS_WITH_SIGNATURE = _env_flag("REFINER_VOICE_ALLOW_TOKENS_WITH_SIGNATURE", False)
VOICE_VERIFY_ALEXA = _env_flag("REFINER_VOICE_VERIFY_ALEXA", False)
VOICE_VERIFY_GOOGLE = _env_flag("REFINER_VOICE_VERIFY_GOOGLE", False)
VOICE_ALLOW_NETWORK = _env_flag("REFINER_VOICE_ALLOW_NETWORK", False)
STT_PUBLIC = _env_flag("REFINER_STT_PUBLIC", False)
STT_TOKEN = (os.getenv("REFINER_STT_TOKEN") or "").strip()
STT_COMMAND = (os.getenv("REFINER_STT_COMMAND") or "").strip()
STT_ARGS = (os.getenv("REFINER_STT_ARGS") or "{audio}").strip()
STT_OUTPUT_MODE = (os.getenv("REFINER_STT_OUTPUT") or "stdout").strip().lower()
STT_OUTPUT_PATH_TEMPLATE = (os.getenv("REFINER_STT_OUTPUT_PATH") or "{audio}.json").strip()
STT_TIMEOUT = float(os.getenv("REFINER_STT_TIMEOUT", "30"))
STT_MAX_BYTES = int(os.getenv("REFINER_STT_MAX_BYTES", "6000000"))
STT_PREPROCESS_COMMAND = (os.getenv("REFINER_STT_PREPROCESS_COMMAND") or "").strip()
STT_PREPROCESS_ARGS = (os.getenv("REFINER_STT_PREPROCESS_ARGS") or "{input} {output}").strip()
STT_PREPROCESS_EXT = (os.getenv("REFINER_STT_PREPROCESS_EXT") or ".wav").strip()
STT_LANG_DEFAULT = (os.getenv("REFINER_STT_LANG") or "en-GB").strip()
STT_SERVER_URL = (os.getenv("REFINER_STT_SERVER_URL") or "").strip()


def _resolve_stt_backend(raw_backend: Optional[str], stt_server_url: str) -> str:
    backend = (raw_backend or "").strip().lower()
    if backend:
        return backend
    return "server" if (stt_server_url or "").strip() else "command"


STT_BACKEND = _resolve_stt_backend(os.getenv("REFINER_STT_BACKEND"), STT_SERVER_URL)
STT_SERVER_TIMEOUT = float(os.getenv("REFINER_STT_SERVER_TIMEOUT", "25"))
STT_SERVER_PREPROCESS = _env_flag("REFINER_STT_SERVER_PREPROCESS", False)
STT_SERVER_SEND_PROMPT = _env_flag("REFINER_STT_SERVER_SEND_PROMPT", False)
STT_SERVER_RETRIES = max(0, int(os.getenv("REFINER_STT_SERVER_RETRIES", "2")))
STT_SERVER_BACKOFF_BASE = max(0.0, float(os.getenv("REFINER_STT_SERVER_BACKOFF_BASE", "0.2")))
STT_SERVER_BACKOFF_MAX = max(
    STT_SERVER_BACKOFF_BASE,
    float(os.getenv("REFINER_STT_SERVER_BACKOFF_MAX", "1.5")),
)
STT_SERVER_POOL_CONNECTIONS = max(1, int(os.getenv("REFINER_STT_SERVER_POOL_CONNECTIONS", "16")))
STT_SERVER_POOL_MAXSIZE = max(1, int(os.getenv("REFINER_STT_SERVER_POOL_MAXSIZE", "32")))
STT_MAX_CONCURRENT = max(1, int(os.getenv("REFINER_STT_MAX_CONCURRENT", "8")))
STT_MAX_CONCURRENT_PER_USER = max(
    1,
    min(
        STT_MAX_CONCURRENT,
        int(os.getenv("REFINER_STT_MAX_CONCURRENT_PER_USER", str(max(1, STT_MAX_CONCURRENT // 2)))),
    ),
)
STT_CAPACITY_WAIT_SEC = max(0.0, float(os.getenv("REFINER_STT_CAPACITY_WAIT_SEC", "0.35")))
CUSTOMERS_API_BASE = (_env_first("REFINER_CUSTOMERS_API_BASE", "CUSTOMERS_API_BASE") or "").strip().rstrip("/")
CUSTOMERS_TIMEOUT = max(1.0, float(_env_first("REFINER_CUSTOMERS_TIMEOUT", default="5")))
BILLING_API_BASE = (_env_first("REFINER_BILLING_API_BASE", "BILLING_API_BASE") or "").strip().rstrip("/")
BILLING_TIMEOUT = max(1.0, float(_env_first("REFINER_BILLING_TIMEOUT", default="10")))
STT_LEARNING_ENABLED = _env_flag("REFINER_STT_LEARNING_ENABLED", True)
STT_LEARNING_ROOT = os.path.join(JOB_ROOT, "stt_learning")
STT_LEARNING_ALLOW_NETWORK = _env_flag("REFINER_STT_LEARNING_ALLOW_NETWORK", True)
STT_LEARNING_PROMPT_TERMS = int(os.getenv("REFINER_STT_LEARNING_PROMPT_TERMS", "40"))
STT_LEARNING_MIN_COUNT = int(os.getenv("REFINER_STT_LEARNING_MIN_COUNT", "3"))
STT_LEARNING_MAX_MEMORY_DOCS = int(os.getenv("REFINER_STT_LEARNING_MAX_MEMORY_DOCS", "350"))
STT_GESTURE_ENABLED = _env_flag("REFINER_STT_GESTURE_ENABLED", True)
STT_BSL_ENABLED = _env_flag("REFINER_STT_BSL_ENABLED", True)
STT_GESTURE_PREFER_SERVER = _env_flag("REFINER_STT_GESTURE_PREFER_SERVER", True)
STT_GESTURE_NMSTT_FALLBACK = _env_flag(
    "REFINER_STT_GESTURE_NMSTT_FALLBACK",
    _env_flag("REFINER_STT_GESTURE_RUST_FALLBACK", True),
)
STT_GESTURE_NMSTT_TIMEOUT = max(
    0.1,
    float(
        _env_first(
            "REFINER_STT_GESTURE_NMSTT_TIMEOUT",
            "REFINER_STT_GESTURE_RUST_TIMEOUT",
            default="4.0",
        )
    ),
)
STT_GESTURE_DEFAULT_MODE = (os.getenv("REFINER_STT_GESTURE_DEFAULT_MODE") or "gesticulation").strip().lower()
STT_GESTURE_DEFAULT_AVATAR_MODE = (os.getenv("REFINER_STT_GESTURE_DEFAULT_AVATAR_MODE") or "chat").strip().lower()
_STT_SERVER_SESSION_LOCAL = threading.local()
ASSISTANT_MAX_CONCURRENT = max(1, int(os.getenv("REFINER_ASSISTANT_MAX_CONCURRENT", "6")))
ASSISTANT_MAX_CONCURRENT_PER_USER = max(
    1,
    min(
        ASSISTANT_MAX_CONCURRENT,
        int(
            os.getenv(
                "REFINER_ASSISTANT_MAX_CONCURRENT_PER_USER",
                str(max(1, ASSISTANT_MAX_CONCURRENT // 2)),
            )
        ),
    ),
)
ASSISTANT_CAPACITY_WAIT_SEC = max(0.0, float(os.getenv("REFINER_ASSISTANT_CAPACITY_WAIT_SEC", "0.5")))
ASSISTANT_SECURITY_POLICY_ENABLED = _env_flag("REFINER_ASSISTANT_SECURITY_POLICY_ENABLED", True)
ASSISTANT_SECURITY_STRICT_MESSAGE_ROLES = _env_flag("REFINER_ASSISTANT_SECURITY_STRICT_MESSAGE_ROLES", False)
ASSISTANT_SECURITY_BLOCK_PROMPT_LEAK = _env_flag("REFINER_ASSISTANT_SECURITY_BLOCK_PROMPT_LEAK", False)
ASSISTANT_MCP_ADMIN_ONLY = _env_flag("REFINER_ASSISTANT_MCP_ADMIN_ONLY", True)
ASSISTANT_BLOCK_UNSAFE_TOOL_REQUESTS = _env_flag("REFINER_ASSISTANT_BLOCK_UNSAFE_TOOL_REQUESTS", False)
ASSISTANT_SECURITY_VALIDATE_RAG_SOURCE_URLS = _env_flag(
    "REFINER_ASSISTANT_SECURITY_VALIDATE_RAG_SOURCE_URLS",
    False,
)
ASSISTANT_OUTPUT_REDACT_PII = _env_flag("REFINER_ASSISTANT_OUTPUT_REDACT_PII", False)
ASSISTANT_OUTPUT_VALIDATE_SHAPES = _env_flag("REFINER_ASSISTANT_OUTPUT_VALIDATE_SHAPES", True)
ASSISTANT_INTENT_ROUTING_ENABLED = _env_flag("REFINER_ASSISTANT_INTENT_ROUTING_ENABLED", False)
ASSISTANT_ROUTING_SKILL_HINT_LIMIT = max(1, int(os.getenv("REFINER_ASSISTANT_ROUTING_SKILL_HINT_LIMIT", "4")))
ASSISTANT_ROUTING_CAPABILITY_MAX_ITEMS = max(
    1,
    int(os.getenv("REFINER_ASSISTANT_ROUTING_CAPABILITY_MAX_ITEMS", "4")),
)
ASSISTANT_SEMANTIC_CACHE_ENABLED = _env_flag("REFINER_ASSISTANT_SEMANTIC_CACHE_ENABLED", False)
ASSISTANT_SEMANTIC_CACHE_TTL_HOURS = max(0.0, float(os.getenv("REFINER_ASSISTANT_SEMANTIC_CACHE_TTL_HOURS", "12")))
ASSISTANT_SEMANTIC_CACHE_MIN_SIMILARITY = min(
    1.0,
    max(0.0, float(os.getenv("REFINER_ASSISTANT_SEMANTIC_CACHE_MIN_SIMILARITY", "0.94"))),
)
ASSISTANT_SEMANTIC_CACHE_MAX_CANDIDATES = max(
    1,
    int(os.getenv("REFINER_ASSISTANT_SEMANTIC_CACHE_MAX_CANDIDATES", "20")),
)
ASSISTANT_HYBRID_RETRIEVAL_ENABLED = _env_flag("REFINER_ASSISTANT_HYBRID_RETRIEVAL_ENABLED", False)
ASSISTANT_HYBRID_RETRIEVAL_SPARSE_WEIGHT = min(
    1.0,
    max(0.0, float(os.getenv("REFINER_ASSISTANT_HYBRID_RETRIEVAL_SPARSE_WEIGHT", "0.65"))),
)
ASSISTANT_HYBRID_RETRIEVAL_DENSE_WEIGHT = min(
    1.0,
    max(0.0, float(os.getenv("REFINER_ASSISTANT_HYBRID_RETRIEVAL_DENSE_WEIGHT", "0.35"))),
)
ASSISTANT_HYBRID_RETRIEVAL_CANDIDATE_MULTIPLIER = max(
    1,
    int(os.getenv("REFINER_ASSISTANT_HYBRID_RETRIEVAL_CANDIDATE_MULTIPLIER", "4")),
)
ASSISTANT_HYBRID_RETRIEVAL_MIN_DENSE_SCORE = min(
    1.0,
    max(0.0, float(os.getenv("REFINER_ASSISTANT_HYBRID_RETRIEVAL_MIN_DENSE_SCORE", "0.18"))),
)
ASSISTANT_RETRIEVAL_COVERAGE_ENABLED = _env_flag("REFINER_ASSISTANT_RETRIEVAL_COVERAGE_ENABLED", False)
ASSISTANT_RETRIEVAL_MIN_QUERY_TERM_COVERAGE = min(
    1.0,
    max(0.0, float(os.getenv("REFINER_ASSISTANT_RETRIEVAL_MIN_QUERY_TERM_COVERAGE", "0.5"))),
)
ASSISTANT_RETRIEVAL_MIN_MATCH_COUNT = max(
    0,
    int(os.getenv("REFINER_ASSISTANT_RETRIEVAL_MIN_MATCH_COUNT", "1")),
)
ASSISTANT_RETRIEVAL_MIN_CONTEXT_CHARS = max(
    0,
    int(os.getenv("REFINER_ASSISTANT_RETRIEVAL_MIN_CONTEXT_CHARS", "24")),
)
ASSISTANT_RETRIEVAL_RETRY_ENABLED = _env_flag("REFINER_ASSISTANT_RETRIEVAL_RETRY_ENABLED", False)
ASSISTANT_RETRIEVAL_MAX_RETRY_QUERIES = max(
    0,
    int(os.getenv("REFINER_ASSISTANT_RETRIEVAL_MAX_RETRY_QUERIES", "3")),
)
ASSISTANT_RETRIEVAL_MIN_CLAUSE_TERMS = max(
    1,
    int(os.getenv("REFINER_ASSISTANT_RETRIEVAL_MIN_CLAUSE_TERMS", "2")),
)
ASSISTANT_RETRIEVAL_RERANK_ENABLED = _env_flag("REFINER_ASSISTANT_RETRIEVAL_RERANK_ENABLED", False)
ASSISTANT_RETRIEVAL_RERANK_MAX_PHRASE_TERMS = max(
    2,
    int(os.getenv("REFINER_ASSISTANT_RETRIEVAL_RERANK_MAX_PHRASE_TERMS", "6")),
)
ASSISTANT_RETRIEVAL_REFUSE_ON_INSUFFICIENT = _env_flag(
    "REFINER_ASSISTANT_RETRIEVAL_REFUSE_ON_INSUFFICIENT",
    True,
)
_STT_REQUEST_CAPACITY = OwnerAwareCapacityLimiter(
    STT_MAX_CONCURRENT,
    max_concurrent_per_owner=STT_MAX_CONCURRENT_PER_USER,
)
_ASSISTANT_REQUEST_CAPACITY = OwnerAwareCapacityLimiter(
    ASSISTANT_MAX_CONCURRENT,
    max_concurrent_per_owner=ASSISTANT_MAX_CONCURRENT_PER_USER,
)
JOB_ACTION_WORKERS = max(1, int(os.getenv("REFINER_JOB_ACTION_WORKERS", "2")))
JOB_ACTION_MAX_QUEUE = max(1, int(os.getenv("REFINER_JOB_ACTION_MAX_QUEUE", "64")))
JOB_ACTION_TASK_TTL_SEC = max(60, int(os.getenv("REFINER_JOB_ACTION_TASK_TTL_SEC", "1800")))
JOB_ACTION_TIMEOUT_SEC = max(5.0, float(os.getenv("REFINER_JOB_ACTION_TIMEOUT_SEC", "45")))
JOB_ACTION_MAX_OUTSTANDING_PER_OWNER = max(
    1,
    min(
        JOB_ACTION_MAX_QUEUE,
        int(os.getenv("REFINER_JOB_ACTION_MAX_OUTSTANDING_PER_OWNER", str(max(4, JOB_ACTION_WORKERS * 4)))),
    ),
)
JOB_ACTION_MAX_INFLIGHT_PER_OWNER = max(
    1,
    min(
        JOB_ACTION_WORKERS,
        int(os.getenv("REFINER_JOB_ACTION_MAX_INFLIGHT_PER_OWNER", str(max(1, JOB_ACTION_WORKERS // 2)))),
    ),
)
SUBTASK_WORKERS = max(1, int(os.getenv("REFINER_SUBTASK_WORKERS", "2")))
SUBTASK_MAX_QUEUE = max(1, int(os.getenv("REFINER_SUBTASK_MAX_QUEUE", "64")))
SUBTASK_TASK_TTL_SEC = max(60, int(os.getenv("REFINER_SUBTASK_TASK_TTL_SEC", "1800")))
SUBTASK_TIMEOUT_SEC = max(5.0, float(os.getenv("REFINER_SUBTASK_TIMEOUT_SEC", "90")))
SUBTASK_ORPHAN_TTL_SEC = max(60, int(os.getenv("REFINER_SUBTASK_ORPHAN_TTL_SEC", "900")))
SUBTASK_MAX_OUTSTANDING_PER_OWNER = max(
    1,
    min(
        SUBTASK_MAX_QUEUE,
        int(os.getenv("REFINER_SUBTASK_MAX_OUTSTANDING_PER_OWNER", str(max(4, SUBTASK_WORKERS * 4)))),
    ),
)
SUBTASK_MAX_INFLIGHT_PER_OWNER = max(
    1,
    min(
        SUBTASK_WORKERS,
        int(os.getenv("REFINER_SUBTASK_MAX_INFLIGHT_PER_OWNER", str(max(1, SUBTASK_WORKERS // 2)))),
    ),
)
EXTERNAL_HTTP_RETRIES = max(0, int(os.getenv("REFINER_EXTERNAL_HTTP_RETRIES", "2")))
EXTERNAL_HTTP_BACKOFF_BASE = max(0.0, float(os.getenv("REFINER_EXTERNAL_HTTP_BACKOFF_BASE", "0.25")))
EXTERNAL_HTTP_BACKOFF_MAX = max(
    EXTERNAL_HTTP_BACKOFF_BASE,
    float(os.getenv("REFINER_EXTERNAL_HTTP_BACKOFF_MAX", "2.0")),
)
EXTERNAL_HTTP_WORKERS = max(4, int(os.getenv("REFINER_EXTERNAL_HTTP_WORKERS", "24")))
EXTERNAL_HTTP_FUTURE_GRACE_SEC = max(0.05, float(os.getenv("REFINER_EXTERNAL_HTTP_FUTURE_GRACE_SEC", "0.5")))
CUSTOMERS_PROFILE_PREFETCH_WAIT_SEC = max(
    0.05,
    float(os.getenv("REFINER_CUSTOMERS_PROFILE_PREFETCH_WAIT_SEC", "0.75")),
)
CUSTOMERS_SESSION_LOOKUP_WAIT_SEC = max(
    0.05,
    float(os.getenv("REFINER_CUSTOMERS_SESSION_LOOKUP_WAIT_SEC", "1.25")),
)
CUSTOMERS_TEAM_DETAIL_MAX_WORKERS = max(
    1,
    int(os.getenv("REFINER_CUSTOMERS_TEAM_DETAIL_MAX_WORKERS", "12")),
)
STT_KB_LOCAL_PATHS = [
    p.strip()
    for p in (os.getenv("REFINER_STT_KB_LOCAL_PATHS") or "/home/pbisaacs/Developer/neuralmimicry.ai-website").split(",")
    if p.strip()
]
STT_KB_SEED_URLS = [
    u.strip()
    for u in (os.getenv("REFINER_STT_KB_SEED_URLS") or "https://neuralmimicry.ai").split(",")
    if u.strip()
]
ALEXA_CERT_TTL_SEC = int(os.getenv("REFINER_ALEXA_CERT_TTL_SEC", "3600"))
ALEXA_REQUEST_TTL_SEC = int(os.getenv("REFINER_ALEXA_REQUEST_TTL_SEC", "150"))
ALEXA_CERT_CACHE_PATH = (os.getenv("REFINER_ALEXA_CERT_CACHE_PATH") or "").strip()
GOOGLE_CERTS_URL = (os.getenv("REFINER_GOOGLE_CERTS_URL") or "https://www.googleapis.com/oauth2/v3/certs").strip()
GOOGLE_CERTS_PATH = (os.getenv("REFINER_GOOGLE_CERTS_PATH") or "").strip()
GOOGLE_AUDIENCE_RAW = (os.getenv("REFINER_GOOGLE_ASSISTANT_AUDIENCE") or os.getenv("REFINER_GOOGLE_PROJECT_ID") or "").strip()
GOOGLE_ISSUERS_RAW = (os.getenv("REFINER_GOOGLE_ASSISTANT_ISSUERS") or "").strip()


def _parse_voice_env_tokens(raw: str, default_user: str) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    if not raw:
        return mapping
    for item in raw.replace(";", ",").split(","):
        item = item.strip()
        if not item:
            continue
        token = ""
        user = ""
        if ":" in item:
            token, user = item.split(":", 1)
        elif "=" in item:
            token, user = item.split("=", 1)
        else:
            token = item
            user = default_user
        token = token.strip()
        user = user.strip() if user else default_user
        if token:
            mapping[token] = user
    return mapping


VOICE_ENV_TOKEN_MAP = _parse_voice_env_tokens(VOICE_ENV_TOKENS, VOICE_DEFAULT_USER)
if VOICE_ENV_TOKEN:
    if VOICE_ENV_TOKEN not in VOICE_ENV_TOKEN_MAP:
        VOICE_ENV_TOKEN_MAP[VOICE_ENV_TOKEN] = VOICE_DEFAULT_USER


def _split_csv(value: str) -> List[str]:
    return [item.strip() for item in value.split(",") if item and item.strip()]


def _parse_voice_user_map(raw: str) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    if not raw:
        return mapping
    for entry in raw.replace(";", ",").split(","):
        entry = entry.strip()
        if not entry:
            continue
        if "=" not in entry:
            continue
        key, user = entry.split("=", 1)
        key = key.strip()
        user = user.strip()
        if not key or not user:
            continue
        mapping[key] = user
    return mapping


VOICE_USER_MAP = _parse_voice_user_map(VOICE_USER_MAP_RAW)
GOOGLE_AUDIENCES = _split_csv(GOOGLE_AUDIENCE_RAW)
GOOGLE_ISSUERS = _split_csv(GOOGLE_ISSUERS_RAW) or [
    "https://accounts.google.com",
    "accounts.google.com",
]

# SHA-256 SPKI hashes from Amazon Trust Services for Alexa request validation.
ALEXA_TRUSTED_ROOT_SPKI = {
    "fbe3018031f9586bcbf41727e417b7d1c45c2f47f93be372a17b96b50757d5a2",
    "7f4296fc5b6a4e3b35d3c369623e364ab1af381d8fa7121533c9d6c633ea2461",
    "36abc32656acfc645c61b71613c4bf21c787f5cabbee48348d58597803d7abc9",
    "f7ecded5c66047d28ed6466b543c40e0743abe81d109254dcf845d4c2c7853c5",
    "2b071c59a0a0ae76b0eadb2bad23bad4580b69c3601b630c2eaf0613afa83f92",
}
_alexa_env_spki = _split_csv(os.getenv("REFINER_ALEXA_ROOT_SPKI") or "")
if _alexa_env_spki:
    ALEXA_TRUSTED_ROOT_SPKI = {item.replace(":", "").lower() for item in _alexa_env_spki if item}
SECRET_STORE_KEY = (os.getenv("REFINER_SECRET_STORE_KEY") or "").strip()
SECRET_STORE_REQUIRE_ENCRYPTION = _env_flag("REFINER_SECRET_STORE_REQUIRE_ENCRYPTION", False)
SECRET_STORE_ALLOW_PLAINTEXT = _env_flag("REFINER_SECRET_STORE_ALLOW_PLAINTEXT", False)
if SECRET_STORE_REQUIRE_ENCRYPTION and not SECRET_STORE_KEY:
    raise RuntimeError("REFINER_SECRET_STORE_KEY is required when encryption is enforced.")
DEFAULT_FORK_ORG = os.getenv("REFINER_FORK_ORG", "neuralmimicry")
DEFAULT_WORKERS = int(os.getenv("REFINER_WORKERS", "2"))
DEFAULT_TAIL = int(os.getenv("REFINER_LOG_TAIL", "200"))
DEFAULT_REPO_PRIVATE = str(os.getenv("REFINER_REPO_PRIVATE", "1")).lower() in {"1", "true", "yes"}
REQUIREMENTS_MAX_BYTES = int(os.getenv("REFINER_REQUIREMENTS_SCAN_BYTES", "20000"))
REQUIREMENTS_IMPORT_MAX_BYTES = int(os.getenv("REFINER_REQUIREMENTS_IMPORT_MAX_BYTES", "4000000"))
REFUND_SCREENSHOT_MAX_BYTES = int(os.getenv("REFINER_REFUND_SCREENSHOT_MAX_BYTES", "5000000"))
REFUND_MAX_FILES = int(os.getenv("REFINER_REFUND_MAX_FILES", "6"))
SITE_BASE = (os.getenv("NEURALMIMICRY_SITE_BASE") or "https://neuralmimicry.ai").rstrip("/")
JOB_META_FILENAME = "job.json"
JOB_META_VERSION = 1
ACTIVE_WINDOW_SEC = int(os.getenv("REFINER_ACTIVE_WINDOW_SEC", "120"))
LEDGER_ROOT = os.path.join(JOB_ROOT, "ledger")
TEAM_LEDGER_ROOT = os.path.join(JOB_ROOT, "team_ledger")
REFINER_CHAIN_SYSTEM = (_env_first("REFINER_CHAIN_SYSTEM", default="refiner") or "refiner").strip() or "refiner"
ESTIMATE_REPO_TTL_SEC = int(os.getenv("REFINER_ESTIMATE_REPO_TTL", "30"))
ESTIMATE_REPO_MAX_FILES = int(os.getenv("REFINER_ESTIMATE_REPO_MAX_FILES", "900"))
ESTIMATE_REPO_MAX_SEC = float(os.getenv("REFINER_ESTIMATE_REPO_MAX_SEC", "0.35"))
ESTIMATE_REPO_MAX_FILE_BYTES = int(os.getenv("REFINER_ESTIMATE_REPO_MAX_FILE_BYTES", "300000"))
ESTIMATE_REPO_SAMPLE_MULTIPLIER = float(os.getenv("REFINER_ESTIMATE_REPO_SAMPLE_MULTIPLIER", "1.6"))
ESTIMATE_CALIBRATION_TTL_SEC = int(os.getenv("REFINER_ESTIMATE_CALIBRATION_TTL", "90"))
DEFAULT_LLM_MAX_TOKENS = int(os.getenv("REFINER_DEFAULT_LLM_MAX_TOKENS", "48000"))
PLAYGROUND_LLM_MAX_TOKENS = max(
    1000,
    int(os.getenv("REFINER_PLAYGROUND_LLM_MAX_TOKENS", str(min(DEFAULT_LLM_MAX_TOKENS, 12000)))),
)
PLAYGROUND_PROJECT_MAX_STEPS = max(25, int(os.getenv("REFINER_PLAYGROUND_PROJECT_MAX_STEPS", "120")))
PLAYGROUND_PROJECT_MIN_ITERATIONS = max(1, int(os.getenv("REFINER_PLAYGROUND_PROJECT_MIN_ITERATIONS", "6")))
PLAYGROUND_PROJECT_MAX_ITERATIONS = max(
    PLAYGROUND_PROJECT_MIN_ITERATIONS,
    int(os.getenv("REFINER_PLAYGROUND_PROJECT_MAX_ITERATIONS", "12")),
)
RESUME_LLM_MAX_TOKENS_CAP = int(os.getenv("REFINER_RESUME_LLM_MAX_TOKENS_CAP", "96000"))
JOB_RETENTION_DAYS = int(os.getenv("REFINER_JOB_RETENTION_DAYS", "0"))
SESSION_TTL_SEC = int(os.getenv("REFINER_SESSION_TTL_SEC", "14400"))
SESSION_HISTORY_MAX = int(os.getenv("REFINER_SESSION_HISTORY_MAX", "200"))
UK_TZ = ZoneInfo("Europe/London")
UK_DATETIME_FORMAT = "%d/%m/%Y %H:%M:%S"
RAG_STORE_ROOT = os.path.join(JOB_ROOT, "rag")
MCP_STORE_ROOT = os.path.join(JOB_ROOT, "mcp")
RAG_MAX_DOCS = int(os.getenv("REFINER_RAG_MAX_DOCS", "60"))
RAG_MAX_DOC_BYTES = int(os.getenv("REFINER_RAG_MAX_DOC_BYTES", "600000"))
RAG_DEFAULT_CHUNK_SIZE = int(os.getenv("REFINER_RAG_CHUNK_SIZE", "1200"))
RAG_DEFAULT_CHUNK_OVERLAP = int(os.getenv("REFINER_RAG_CHUNK_OVERLAP", "200"))
RAG_DEFAULT_MAX_CHUNKS = int(os.getenv("REFINER_RAG_MAX_CHUNKS", "2000"))
RAG_ASYNC_INDEX_BUILDS = _env_flag("REFINER_RAG_ASYNC_INDEX_BUILDS", False)
RAG_BUILD_TIMEOUT_SEC = max(5.0, float(os.getenv("REFINER_RAG_BUILD_TIMEOUT_SEC", str(SUBTASK_TIMEOUT_SEC))))
NMCHAIN = NmChainClient.from_env()
CUSTOMERS_SERVICE = CustomersClient.from_env()
BILLING_SERVICE = BillingClient.from_env()
if NMCHAIN:
    logger.info("nmchain integration enabled for Refiner via %s", NMCHAIN.base_url)
if CUSTOMERS_SERVICE:
    logger.info("customers integration enabled for Refiner via %s", CUSTOMERS_SERVICE.base_url)
if BILLING_SERVICE:
    logger.info("billing integration enabled for Refiner via %s", BILLING_SERVICE.base_url)
RAG_ALLOWED_ROOTS = [
    p for p in (os.getenv("REFINER_RAG_ALLOWED_ROOTS") or "").split(",") if p.strip()
]
if BASE_DIR not in RAG_ALLOWED_ROOTS:
    RAG_ALLOWED_ROOTS.append(BASE_DIR)
if JOB_ROOT not in RAG_ALLOWED_ROOTS:
    RAG_ALLOWED_ROOTS.append(JOB_ROOT)
try:
    TOKEN_BTC_RATE = float(os.getenv("REFINER_TOKEN_BTC_RATE", "0.000016"))
except Exception:
    TOKEN_BTC_RATE = 0.000016
PORTAL_WEBHOOK_URL = (os.getenv("REFINER_PORTAL_WEBHOOK") or "").strip()
PORTAL_WEBHOOK_TOKEN = (os.getenv("REFINER_PORTAL_WEBHOOK_TOKEN") or "").strip()
PORTAL_WEBHOOK_TIMEOUT = int(os.getenv("REFINER_PORTAL_WEBHOOK_TIMEOUT", "12"))
CONTINUUM_API_BASE = _env_first("CONTINUUM_API_BASE", "NMC_API_BASE", "NMC_SERVER_URL", default="").rstrip("/")
CONTINUUM_AUTH_TOKEN = _env_first(
    "CONTINUUM_BEARER_TOKEN",
    "CONTINUUM_AUTH_TOKEN",
    "NMC_OIDC_ACCESS_TOKEN",
    "NMC_BEARER_TOKEN",
    "NMC_AUTH_TOKEN",
    default="",
)
CONTINUUM_TIMEOUT = float(os.getenv("CONTINUUM_TIMEOUT", "20"))
CONTINUUM_VM_REGION = _env_first("CONTINUUM_VM_REGION", "NMC_VM_REGION", default="gb-mids")
CONTINUUM_VM_SKU = _env_first("CONTINUUM_VM_SKU", "NMC_VM_SKU", default="standard-a2")
CONTINUUM_VM_OS = _env_first("CONTINUUM_VM_OS", "NMC_VM_OS", default="ubuntu-22.04")
CONTINUUM_VM_PUBLIC_KEY_ID = _env_first("CONTINUUM_VM_PUBLIC_KEY_ID", "NMC_VM_PUBLIC_KEY_ID", default="")
CONTINUUM_VM_INIT_SCRIPT = (os.getenv("CONTINUUM_VM_INIT_SCRIPT") or "").strip()
CONTINUUM_IDE_URL_TEMPLATE = _env_first("REFINER_IDE_URL_TEMPLATE", "CONTINUUM_IDE_URL_TEMPLATE", default="")
CONTINUUM_PREVIEW_URL_TEMPLATE = _env_first("REFINER_PREVIEW_URL_TEMPLATE", "CONTINUUM_PREVIEW_URL_TEMPLATE", default="")
CONTINUUM_IDE_FILE_URL_TEMPLATE = _env_first(
    "REFINER_IDE_FILE_URL_TEMPLATE",
    "CONTINUUM_IDE_FILE_URL_TEMPLATE",
    default="",
)
CONTINUUM_AUTOSCALE_ENABLED = _env_flag("CONTINUUM_AUTOSCALE_ENABLED", True)
CONTINUUM_AUTOSCALE_POLL_SEC = max(1.0, float(os.getenv("CONTINUUM_AUTOSCALE_POLL_SEC", "8")))
CONTINUUM_AUTOSCALE_MIN_REPLICAS = max(0, int(os.getenv("CONTINUUM_AUTOSCALE_MIN_REPLICAS", "1")))
CONTINUUM_AUTOSCALE_MAX_REPLICAS = max(
    CONTINUUM_AUTOSCALE_MIN_REPLICAS,
    int(os.getenv("CONTINUUM_AUTOSCALE_MAX_REPLICAS", "8")),
)
CONTINUUM_AUTOSCALE_BACKLOG_PER_REPLICA = max(1, int(os.getenv("CONTINUUM_AUTOSCALE_BACKLOG_PER_REPLICA", "1")))
CONTINUUM_AUTOSCALE_SCALE_UP_STEP = max(1, int(os.getenv("CONTINUUM_AUTOSCALE_SCALE_UP_STEP", "1")))
CONTINUUM_AUTOSCALE_SCALE_DOWN_STEP = max(1, int(os.getenv("CONTINUUM_AUTOSCALE_SCALE_DOWN_STEP", "1")))
CONTINUUM_AUTOSCALE_IDLE_SEC = max(0.0, float(os.getenv("CONTINUUM_AUTOSCALE_IDLE_SEC", "180")))
CONTINUUM_AUTOSCALE_COOLDOWN_SEC = max(0.0, float(os.getenv("CONTINUUM_AUTOSCALE_COOLDOWN_SEC", "45")))
CONTINUUM_AUTOSCALE_TIMEOUT_SEC = max(1.0, float(os.getenv("CONTINUUM_AUTOSCALE_TIMEOUT_SEC", "10")))
CONTINUUM_AUTOSCALE_NAMESPACE = (os.getenv("CONTINUUM_AUTOSCALE_NAMESPACE") or "refiner").strip() or "refiner"
CONTINUUM_AUTOSCALE_DEPLOYMENT = (os.getenv("CONTINUUM_AUTOSCALE_DEPLOYMENT") or "refiner").strip() or "refiner"
CONTINUUM_AUTOSCALE_HISTORY_MAX = max(120, int(os.getenv("CONTINUUM_AUTOSCALE_HISTORY_MAX", "720")))
if CONTINUUM_API_BASE and "://" not in CONTINUUM_API_BASE:
    CONTINUUM_API_BASE = f"http://{CONTINUUM_API_BASE}"
EDITOR_MAX_BYTES = int(os.getenv("REFINER_EDITOR_MAX_BYTES", "400000"))
EDITOR_MAX_LIST = int(os.getenv("REFINER_EDITOR_MAX_LIST", "400"))
EDITOR_MAX_SCAN = int(os.getenv("REFINER_EDITOR_MAX_SCAN", "2000"))
EDITOR_MAX_DEPTH = int(os.getenv("REFINER_EDITOR_MAX_DEPTH", "6"))
EDITOR_SKIP_DIRS = {
    ".git",
    ".hg",
    ".svn",
    ".idea",
    ".vscode",
    ".venv",
    "venv",
    "node_modules",
    "__pycache__",
    ".pytest_cache",
    ".mypy_cache",
    ".cache",
    "dist",
    "build",
    "coverage",
}

ensure_dir_permissions(JOB_ROOT, mode=0o700)
ensure_dir_permissions(PROJECTS_ROOT, mode=0o700)
ensure_dir_permissions(SECRET_STORE_ROOT, mode=0o700)
ensure_dir_permissions(ASSISTANT_MEMORY_ROOT, mode=0o700)
ensure_dir_permissions(WORKSPACE_ROOT, mode=0o700)
ensure_dir_permissions(TEAM_LEDGER_ROOT, mode=0o700)
audit_logger = AuditLogger(AUDIT_LOG_PATH)


class _SafeFormatDict(dict):
    """Format-map dict that substitutes missing keys with empty strings."""

    def __missing__(self, key: str) -> str:
        return ""


def _continuum_client_config() -> ContinuumClientConfig:
    """Resolve the current Continuum client configuration from runtime globals."""

    return ContinuumClientConfig(
        api_base=CONTINUUM_API_BASE,
        auth_token=CONTINUUM_AUTH_TOKEN,
        timeout=CONTINUUM_TIMEOUT,
        vm_region=CONTINUUM_VM_REGION,
        vm_sku=CONTINUUM_VM_SKU,
        vm_os=CONTINUUM_VM_OS,
        vm_public_key_id=CONTINUUM_VM_PUBLIC_KEY_ID,
        vm_init_script=CONTINUUM_VM_INIT_SCRIPT,
        ide_url_template=CONTINUUM_IDE_URL_TEMPLATE,
        preview_url_template=CONTINUUM_PREVIEW_URL_TEMPLATE,
    )


def _continuum_enabled() -> bool:
    return _continuum_enabled_impl(_continuum_client_config())


def _continuum_ready() -> bool:
    return _continuum_ready_impl(_continuum_client_config())


def _continuum_headers() -> Dict[str, str]:
    return _continuum_headers_impl(_continuum_client_config())


_EXTERNAL_HTTP_EXECUTOR = ThreadPoolExecutor(
    max_workers=EXTERNAL_HTTP_WORKERS,
    thread_name_prefix="refiner-http",
)


def _submit_external_http_request(
    method: str,
    url: str,
    *,
    session_obj: Optional[requests.Session] = None,
    headers: Optional[Dict[str, Any]] = None,
    json_body: Optional[Any] = None,
    data: Optional[Any] = None,
    files: Optional[Dict[str, Any]] = None,
    params: Optional[Any] = None,
    timeout: float = 20.0,
    auth: Optional[Any] = None,
    allow_redirects: Optional[bool] = None,
) -> Future:
    """Submit an outbound HTTP request to the shared non-blocking transport pool."""
    method_upper = str(method or "GET").upper()
    kwargs: Dict[str, Any] = {
        "method": method_upper,
        "url": url,
        "headers": headers,
        "json": json_body,
        "data": data,
        "files": files,
        "params": params,
        "timeout": max(0.001, float(timeout)),
    }
    if auth is not None:
        kwargs["auth"] = auth
    if allow_redirects is not None:
        kwargs["allow_redirects"] = bool(allow_redirects)
    kwargs = {
        key: value
        for key, value in kwargs.items()
        if value is not None or key in {"method", "url", "timeout"}
    }

    request_callable: Callable[..., Any] = requests.request
    if session_obj is not None:
        session_request = getattr(session_obj, "request", None)
        if callable(session_request):
            request_callable = session_request
        else:
            method_callable = getattr(session_obj, method_upper.lower(), None)
            if callable(method_callable):
                def _method_request(**request_kwargs: Any) -> Any:
                    request_kwargs.pop("method", None)
                    request_url = request_kwargs.pop("url")
                    return method_callable(request_url, **request_kwargs)

                request_callable = _method_request
    return _EXTERNAL_HTTP_EXECUTOR.submit(request_callable, **kwargs)


def _await_external_http_response(
    future: Future,
    *,
    url: str,
    timeout: float,
) -> requests.Response:
    """Await one submitted HTTP request with an additional guard window."""
    wait_timeout = max(0.1, float(timeout)) + EXTERNAL_HTTP_FUTURE_GRACE_SEC
    try:
        response = future.result(timeout=wait_timeout)
    except FutureTimeoutError as exc:
        future.cancel()
        raise requests.Timeout(f"external request timed out for {url}") from exc
    if response is None or not hasattr(response, "status_code"):
        raise RuntimeError(f"external request returned unexpected response type for {url}")
    return response


def _external_http_request(
    method: str,
    url: str,
    *,
    session_obj: Optional[requests.Session] = None,
    headers: Optional[Dict[str, Any]] = None,
    json_body: Optional[Any] = None,
    data: Optional[Any] = None,
    files: Optional[Dict[str, Any]] = None,
    params: Optional[Any] = None,
    timeout: float = 20.0,
    auth: Optional[Any] = None,
    allow_redirects: Optional[bool] = None,
) -> requests.Response:
    """Run one outbound HTTP request on the shared executor and await result."""
    future = _submit_external_http_request(
        method,
        url,
        session_obj=session_obj,
        headers=headers,
        json_body=json_body,
        data=data,
        files=files,
        params=params,
        timeout=timeout,
        auth=auth,
        allow_redirects=allow_redirects,
    )
    return _await_external_http_response(future, url=url, timeout=timeout)


def _submit_external_call(func: Callable[..., Any], /, *args: Any, **kwargs: Any) -> Future:
    """Submit one external integration call to the shared non-blocking pool."""
    return _EXTERNAL_HTTP_EXECUTOR.submit(func, *args, **kwargs)


def _await_external_call_result(
    future: Future,
    *,
    operation: str,
    timeout: float,
) -> Any:
    """Await one submitted external call with a bounded guard window."""
    wait_timeout = max(0.05, float(timeout)) + EXTERNAL_HTTP_FUTURE_GRACE_SEC
    try:
        return future.result(timeout=wait_timeout)
    except FutureTimeoutError as exc:
        future.cancel()
        raise TimeoutError(f"{operation} timed out") from exc


def _run_external_call(
    operation: str,
    func: Callable[..., Any],
    /,
    *args: Any,
    timeout: float = 20.0,
    **kwargs: Any,
) -> Any:
    """Execute one external call via the shared pool and await completion."""
    future = _submit_external_call(func, *args, **kwargs)
    return _await_external_call_result(
        future,
        operation=operation,
        timeout=timeout,
    )


def _service_timeout_seconds(service: Any, *, default: float = 5.0) -> float:
    try:
        return max(0.1, float(getattr(service, "timeout", default)))
    except Exception:
        return max(0.1, float(default))


def _log_external_future_exception(future: Future, *, operation: str) -> None:
    try:
        future.result()
    except Exception as exc:
        logger.warning("%s failed: %s", operation, exc)


def _submit_external_background_call(
    operation: str,
    func: Callable[..., Any],
    /,
    *args: Any,
    **kwargs: Any,
) -> None:
    future = _submit_external_call(func, *args, **kwargs)
    future.add_done_callback(
        lambda done_future, op=operation: _log_external_future_exception(done_future, operation=op)
    )


def _http_request_with_retry(
    method: str,
    url: str,
    *,
    session_obj: Optional[requests.Session] = None,
    headers: Optional[Dict[str, Any]] = None,
    json_body: Optional[Dict[str, Any]] = None,
    data: Optional[Any] = None,
    files: Optional[Dict[str, Any]] = None,
    params: Optional[Any] = None,
    timeout: float = 20.0,
    auth: Optional[Any] = None,
    allow_redirects: Optional[bool] = None,
    retries: Optional[int] = None,
    retryable_statuses: Optional[set[int]] = None,
) -> requests.Response:
    """Issue an outbound HTTP request with bounded retries and exponential backoff."""
    attempts = max(1, (EXTERNAL_HTTP_RETRIES if retries is None else int(retries)) + 1)
    retryable = retryable_statuses or {408, 429, 500, 502, 503, 504}
    response: Optional[requests.Response] = None
    for attempt in range(attempts):
        try:
            response = _external_http_request(
                method=method,
                url=url,
                session_obj=session_obj,
                headers=headers,
                json_body=json_body,
                data=data,
                files=files,
                params=params,
                timeout=timeout,
                auth=auth,
                allow_redirects=allow_redirects,
            )
        except requests.RequestException:
            if attempt + 1 >= attempts:
                raise
            delay = min(
                EXTERNAL_HTTP_BACKOFF_MAX,
                max(0.0, EXTERNAL_HTTP_BACKOFF_BASE * (2 ** max(0, attempt))),
            )
            if delay > 0:
                time.sleep(delay)
            continue
        if response.status_code in retryable and attempt + 1 < attempts:
            retry_after = (response.headers.get("Retry-After") or "").strip()
            delay = 0.0
            if retry_after:
                try:
                    delay = max(0.0, float(retry_after))
                except Exception:
                    delay = 0.0
            if delay <= 0:
                delay = min(
                    EXTERNAL_HTTP_BACKOFF_MAX,
                    max(0.0, EXTERNAL_HTTP_BACKOFF_BASE * (2 ** max(0, attempt))),
                )
            if delay > 0:
                time.sleep(delay)
            continue
        return response
    if response is None:
        raise RuntimeError("HTTP request failed before receiving a response.")
    return response


def _format_workspace_template(template: str, values: Dict[str, Any]) -> str:
    if not template:
        return ""
    try:
        return template.format_map(_SafeFormatDict(values))
    except Exception:
        return template


def _workspace_template_vars(job: "Job", vm_data: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    return _workspace_template_vars_impl(
        job,
        config=_continuum_client_config(),
        vm_data=vm_data,
    )


def _continuum_request(
    method: str,
    path: str,
    *,
    json_body: Optional[Dict[str, Any]] = None,
    timeout_sec: Optional[float] = None,
    retries: Optional[int] = None,
) -> requests.Response:
    return _continuum_request_impl(
        _continuum_client_config(),
        http_request=_http_request_with_retry,
        method=method,
        path=path,
        json_body=json_body,
        timeout_sec=timeout_sec,
        retries=retries,
    )


def _continuum_json_payload(response: requests.Response, *, operation: str) -> Dict[str, Any]:
    return _continuum_json_payload_impl(response, operation=operation)


def _friendly_continuum_error(message: Optional[str]) -> str:
    return _friendly_continuum_error_impl(message)


def _editor_allowed(job: "Job") -> bool:
    workflow = (job.payload.get("workflow") or job.workflow or "").strip().lower()
    return workflow in {"project_solver", "project", "topic_research"}


def _editor_root_candidates(job: "Job") -> List[Dict[str, Any]]:
    roots: List[Dict[str, Any]] = []
    workflow = (job.payload.get("workflow") or job.workflow or "").strip().lower()
    job_dir = os.path.join(JOB_ROOT, job.job_id)
    project_root = (job.payload.get("project_root") or job.payload.get("delivery_project_root") or "").strip()
    repo_workspace = ""
    if isinstance(job.repo_info, dict):
        repo_workspace = str(job.repo_info.get("workspace") or "").strip()

    if workflow in {"project_solver", "project"}:
        if project_root and os.path.isdir(project_root):
            roots.append({"id": "project", "label": "Project", "path": project_root})
        if repo_workspace and os.path.isdir(repo_workspace) and repo_workspace != project_root:
            roots.append({"id": "repo", "label": "Repo Workspace", "path": repo_workspace})
    if workflow == "topic_research" and os.path.isdir(job_dir):
        roots.append({"id": "job", "label": "Research Output", "path": job_dir})
    return roots


def _is_under_root(path: str, root: str) -> bool:
    try:
        path_real = os.path.realpath(path)
        root_real = os.path.realpath(root)
        return os.path.commonpath([path_real, root_real]) == root_real
    except Exception:
        return False


def _editor_default_path(job: "Job", root: str) -> Optional[str]:
    workflow = (job.payload.get("workflow") or job.workflow or "").strip().lower()
    output_path = job.output_paths.get("primary") if isinstance(job.output_paths, dict) else None
    if output_path and _is_under_root(output_path, root):
        return os.path.relpath(output_path, root)
    if workflow in {"project_solver", "project"}:
        req_path = (job.payload.get("requirements_path") or "").strip()
        if req_path and _is_under_root(req_path, root):
            return os.path.relpath(req_path, root)
    return _editor_latest_file(root)


def _editor_latest_file(root: str) -> Optional[str]:
    latest_path = None
    latest_mtime = -1.0
    root_real = os.path.realpath(root)
    scanned = 0
    for dirpath, dirs, files in os.walk(root_real):
        rel_dir = os.path.relpath(dirpath, root_real)
        if rel_dir == ".":
            rel_dir = ""
        depth = rel_dir.count(os.sep)
        if depth >= EDITOR_MAX_DEPTH:
            dirs[:] = []
        dirs[:] = [d for d in dirs if d not in EDITOR_SKIP_DIRS]
        for filename in files:
            scanned += 1
            if EDITOR_MAX_SCAN and scanned > EDITOR_MAX_SCAN:
                return latest_path
            full_path = os.path.join(dirpath, filename)
            try:
                stat = os.stat(full_path)
            except Exception:
                continue
            if stat.st_mtime > latest_mtime:
                latest_mtime = stat.st_mtime
                rel_path = os.path.relpath(full_path, root_real)
                latest_path = rel_path.replace("\\", "/")
    return latest_path


def _editor_root_by_id(job: "Job", root_id: str) -> Optional[Dict[str, Any]]:
    for entry in _editor_root_candidates(job):
        if entry.get("id") == root_id:
            return entry
    return None


def _editor_normalize_rel(path: Optional[str]) -> str:
    raw = (path or "").strip().replace("\\", "/")
    if not raw or raw == "." or raw == "/":
        return ""
    raw = raw.lstrip("/")
    normal = os.path.normpath(raw)
    if normal in {".", ""}:
        return ""
    if normal.startswith(".."):
        raise ValueError("invalid path")
    return normal.replace("\\", "/")


def _editor_join(root: str, rel: str) -> str:
    root_real = os.path.realpath(root)
    candidate = os.path.realpath(os.path.join(root_real, rel))
    if os.path.commonpath([root_real, candidate]) != root_real:
        raise ValueError("invalid path")
    return candidate


def _editor_list_dir(root: str, rel: str) -> Tuple[List[Dict[str, Any]], bool]:
    abs_path = _editor_join(root, rel)
    if not os.path.isdir(abs_path):
        raise FileNotFoundError("directory not found")
    entries: List[Dict[str, Any]] = []
    truncated = False
    try:
        names = sorted(os.listdir(abs_path))
    except Exception:
        names = []
    for name in names:
        if name in EDITOR_SKIP_DIRS:
            continue
        full_path = os.path.join(abs_path, name)
        rel_path = os.path.join(rel, name) if rel else name
        rel_path = rel_path.replace("\\", "/")
        try:
            stat = os.stat(full_path)
        except Exception:
            continue
        if os.path.isdir(full_path):
            entries.append({"name": name, "path": rel_path, "type": "dir", "modified": stat.st_mtime})
        else:
            entries.append(
                {
                    "name": name,
                    "path": rel_path,
                    "type": "file",
                    "size": stat.st_size,
                    "modified": stat.st_mtime,
                }
            )
        if EDITOR_MAX_LIST and len(entries) >= EDITOR_MAX_LIST:
            truncated = True
            break
    entries.sort(key=lambda e: (0 if e.get("type") == "dir" else 1, e.get("name") or ""))
    return entries, truncated


def _editor_read_file(root: str, rel: str) -> Dict[str, Any]:
    abs_path = _editor_join(root, rel)
    if not os.path.exists(abs_path):
        raise FileNotFoundError("file not found")
    if not os.path.isfile(abs_path):
        raise IsADirectoryError("path is a directory")
    size = os.path.getsize(abs_path)
    if EDITOR_MAX_BYTES and size > EDITOR_MAX_BYTES:
        raise ValueError("file too large")
    with open(abs_path, "rb") as handle:
        data = handle.read(EDITOR_MAX_BYTES + 1)
    if EDITOR_MAX_BYTES and len(data) > EDITOR_MAX_BYTES:
        raise ValueError("file too large")
    if b"\x00" in data:
        raise ValueError("binary file")
    content = data.decode("utf-8", errors="replace")
    stat = os.stat(abs_path)
    return {"content": content, "size": stat.st_size, "modified": stat.st_mtime}


def _editor_write_file(root: str, rel: str, content: str) -> Dict[str, Any]:
    abs_path = _editor_join(root, rel)
    parent = os.path.dirname(abs_path)
    os.makedirs(parent, exist_ok=True)
    encoded = content.encode("utf-8")
    if EDITOR_MAX_BYTES and len(encoded) > EDITOR_MAX_BYTES:
        raise ValueError("file too large")
    with open(abs_path, "wb") as handle:
        handle.write(encoded)
    stat = os.stat(abs_path)
    return {"size": stat.st_size, "modified": stat.st_mtime}


def _editor_create_file(root: str, rel: str, content: str = "") -> Dict[str, Any]:
    abs_path = _editor_join(root, rel)
    if os.path.exists(abs_path):
        raise ValueError("already_exists")
    parent = os.path.dirname(abs_path)
    os.makedirs(parent, exist_ok=True)
    encoded = content.encode("utf-8")
    if EDITOR_MAX_BYTES and len(encoded) > EDITOR_MAX_BYTES:
        raise ValueError("file too large")
    with open(abs_path, "wb") as handle:
        handle.write(encoded)
    stat = os.stat(abs_path)
    return {"size": stat.st_size, "modified": stat.st_mtime}


def _editor_delete_path(root: str, rel: str, *, force: bool = False) -> Dict[str, Any]:
    abs_path = _editor_join(root, rel)
    if not os.path.exists(abs_path):
        raise FileNotFoundError("not_found")
    if os.path.isdir(abs_path):
        if os.listdir(abs_path):
            if not force:
                raise ValueError("dir_not_empty")
            shutil.rmtree(abs_path)
            return {"deleted": True, "type": "dir", "recursive": True}
        os.rmdir(abs_path)
        return {"deleted": True, "type": "dir"}
    os.remove(abs_path)
    return {"deleted": True, "type": "file"}


def _editor_rename_path(root: str, rel: str, new_rel: str) -> Dict[str, Any]:
    src_path = _editor_join(root, rel)
    dest_path = _editor_join(root, new_rel)
    if not os.path.exists(src_path):
        raise FileNotFoundError("not_found")
    if os.path.exists(dest_path):
        raise ValueError("already_exists")
    os.makedirs(os.path.dirname(dest_path), exist_ok=True)
    os.rename(src_path, dest_path)
    return {"renamed": True, "path": new_rel}


def _editor_create_dir(root: str, rel: str) -> Dict[str, Any]:
    abs_path = _editor_join(root, rel)
    if os.path.exists(abs_path):
        raise ValueError("already_exists")
    os.makedirs(abs_path, exist_ok=False)
    stat = os.stat(abs_path)
    return {"created": True, "type": "dir", "modified": stat.st_mtime}


def _workspace_file_open_url(
    job: "Job",
    root_entry: Dict[str, Any],
    rel_path: str,
) -> Tuple[Optional[str], bool]:
    workspace_env = job.workspace_env if isinstance(job.workspace_env, dict) else {}
    ide_url = str(workspace_env.get("ide_url") or "").strip()
    if not ide_url:
        return None, False
    template = CONTINUUM_IDE_FILE_URL_TEMPLATE
    if template:
        vm_data = workspace_env.get("vm") if isinstance(workspace_env.get("vm"), dict) else {}
        values = _workspace_template_vars(job, vm_data)
        values.update(
            {
                "ide_url": ide_url,
                "path": rel_path,
                "path_url": quote(rel_path),
                "root_id": root_entry.get("id") or "",
                "root_label": root_entry.get("label") or "",
            }
        )
        rendered = _format_workspace_template(template, values)
        if rendered:
            return rendered, True
    return ide_url, False

ESTIMATE_TEXT_EXTS = {
    ".py",
    ".js",
    ".ts",
    ".tsx",
    ".jsx",
    ".java",
    ".go",
    ".rs",
    ".cpp",
    ".cc",
    ".c",
    ".h",
    ".hpp",
    ".cs",
    ".rb",
    ".php",
    ".sh",
    ".sql",
    ".md",
    ".txt",
    ".rst",
    ".json",
    ".yml",
    ".yaml",
    ".toml",
    ".ini",
    ".cfg",
}

ESTIMATE_IGNORED_DIRS = {
    ".git",
    ".hg",
    ".svn",
    ".venv",
    "venv",
    ".research_cache",
    "__pypackages__",
    "site-packages",
    "dist-packages",
    "node_modules",
    "__pycache__",
    ".pytest_cache",
    ".mypy_cache",
    ".tox",
    ".nox",
    ".idea",
    ".vscode",
    "dist",
    "build",
    "test_output",
    "project_solver_output",
    "delivery_pipeline_output",
}

REFUND_ALLOWED_EXTS = {".png", ".jpg", ".jpeg", ".webp", ".gif"}

_estimate_repo_cache: Dict[str, Dict[str, Any]] = {}
_estimate_repo_cache_lock = threading.Lock()
_estimate_calibration_cache: Dict[str, Any] = {"ts": 0.0, "data": {}, "job_count": 0}
_job_output_insights_cache: Dict[str, Dict[str, Any]] = {}
_job_output_insights_cache_lock = threading.Lock()

app = Flask(
    __name__,
    static_folder=os.path.join(BASE_DIR, "web", "static"),
    template_folder=os.path.join(BASE_DIR, "web", "templates"),
)
SECRET_KEY_REQUIRED = _env_flag("REFINER_REQUIRE_SECRET_KEY", False)
_secret_key_env = os.getenv("REFINER_SECRET_KEY")
if SECRET_KEY_REQUIRED and not _secret_key_env:
    raise RuntimeError("REFINER_SECRET_KEY is required for secure session management.")
if not _secret_key_env:
    logger.warning("REFINER_SECRET_KEY not set; using a transient session key.")
app.secret_key = _secret_key_env or os.urandom(32)


@app.context_processor
def inject_template_globals() -> Dict[str, Any]:
    """Expose shared template metadata."""
    return {"app_version": get_version_info()}


def _env_list(name: str) -> List[str]:
    value = os.getenv(name, "")
    items = []
    for entry in value.split(","):
        entry = entry.strip()
        if not entry:
            continue
        items.append(entry.rstrip("/"))
    return items


def _env_list_any(*names: str) -> List[str]:
    items: List[str] = []
    seen = set()
    for name in names:
        for entry in _env_list(name):
            if entry in seen:
                continue
            items.append(entry)
            seen.add(entry)
    return items


def _normalize_samesite(value: Optional[str]) -> str:
    if not value:
        return ""
    cleaned = value.strip().lower()
    if cleaned == "none":
        return "None"
    if cleaned == "lax":
        return "Lax"
    if cleaned == "strict":
        return "Strict"
    return ""


APP_START_TIME = time.time()

METRICS_PATH = (os.getenv("REFINER_METRICS_PATH", "/metrics") or "/metrics").strip()
if not METRICS_PATH.startswith("/"):
    METRICS_PATH = f"/{METRICS_PATH}"

try:
    from prometheus_client import Counter, Gauge, Histogram, generate_latest, CONTENT_TYPE_LATEST

    PROMETHEUS_AVAILABLE = True
except Exception:
    PROMETHEUS_AVAILABLE = False

METRICS_ENABLED = _env_flag("REFINER_METRICS_ENABLED", True) and PROMETHEUS_AVAILABLE

if METRICS_ENABLED:
    REQUEST_COUNT = Counter(
        "refiner_http_requests_total",
        "Total HTTP requests",
        ["method", "path", "status"],
    )
    REQUEST_LATENCY = Histogram(
        "refiner_http_request_duration_seconds",
        "HTTP request duration in seconds",
        ["method", "path"],
    )
    INFLIGHT = Gauge("refiner_http_inflight_requests", "In-flight HTTP requests")
    JOBS_BY_STATUS = Gauge("refiner_jobs_total", "Jobs by status", ["status"])
    JOB_QUEUE_DEPTH = Gauge("refiner_job_queue_depth", "Job queue depth")
    WORKER_COUNT = Gauge("refiner_worker_threads", "Worker threads")
    JOB_ACTION_QUEUE_DEPTH = Gauge("refiner_job_action_queue_depth", "Job action queue depth")
    JOB_ACTION_INFLIGHT = Gauge("refiner_job_action_inflight", "Job action tasks currently running")
    UPTIME = Gauge("refiner_uptime_seconds", "Application uptime in seconds")


CORS_ORIGINS = _env_list("REFINER_CORS_ORIGINS")
CORS_ALLOW_HEADERS = ["Content-Type", "Authorization", "X-Requested-With"]
CORS_ALLOW_METHODS = ["GET", "POST", "DELETE", "OPTIONS"]
CORS_MAX_AGE = int(os.getenv("REFINER_CORS_MAX_AGE", "600"))
API_BASE = os.getenv("REFINER_API_BASE", "").strip().rstrip("/")
COOKIE_DOMAIN = (os.getenv("REFINER_COOKIE_DOMAIN") or "").strip() or None
SESSION_COOKIE_NAME = _env_first("REFINER_SESSION_COOKIE_NAME", default="nm_refiner_session").strip() or "nm_refiner_session"
EXTERNAL_REDIRECT_HOSTS = _env_list("REFINER_EXTERNAL_REDIRECT_HOSTS")

COOKIE_SAMESITE = _normalize_samesite(os.getenv("REFINER_COOKIE_SAMESITE"))
if not COOKIE_SAMESITE:
    COOKIE_SAMESITE = "None" if CORS_ORIGINS else "Lax"
SECURE_COOKIES = _env_flag("REFINER_SECURE_COOKIES", COOKIE_SAMESITE == "None")
ENFORCE_HTTPS = _env_flag("REFINER_ENFORCE_HTTPS", False)
CSRF_ORIGIN_CHECK = _env_flag("REFINER_CSRF_ORIGIN_CHECK", True)
CSP_POLICY = (os.getenv("REFINER_CSP_POLICY") or "").strip()
AUTH_MODE = _env_first("REFINER_AUTH_MODE", "NM_AUTH_MODE", default="local").strip().lower()

OIDC_ENABLED = _env_flag("REFINER_OIDC_ENABLED", _env_flag("NM_OIDC_ENABLED", False))
OIDC_ISSUER = _env_first("REFINER_OIDC_ISSUER", "NM_OIDC_ISSUER")
OIDC_CLIENT_ID = _env_first("REFINER_OIDC_CLIENT_ID", "NM_OIDC_CLIENT_ID")
OIDC_CLIENT_SECRET = _env_first("REFINER_OIDC_CLIENT_SECRET", "NM_OIDC_CLIENT_SECRET")
OIDC_REDIRECT_URI = _env_first("REFINER_OIDC_REDIRECT_URI", "NM_OIDC_REDIRECT_URI", "NM_OIDC_REDIRECT_URL")
OIDC_SCOPE = _env_first("REFINER_OIDC_SCOPE", "NM_OIDC_SCOPE", default="openid email profile")
OIDC_USERNAME_CLAIM = _env_first("REFINER_OIDC_USERNAME_CLAIM", "NM_OIDC_USERNAME_CLAIM", default="email")
OIDC_EMAIL_CLAIM = _env_first("REFINER_OIDC_EMAIL_CLAIM", "NM_OIDC_EMAIL_CLAIM", default="email")
OIDC_GROUPS_CLAIM = _env_first("REFINER_OIDC_GROUPS_CLAIM", "NM_OIDC_GROUPS_CLAIM", default="groups")
OIDC_ADMIN_DOMAINS = _env_list_any("REFINER_OIDC_ADMIN_DOMAINS", "NM_OIDC_ADMIN_DOMAINS")
OIDC_ADMIN_GROUPS = _env_list_any("REFINER_OIDC_ADMIN_GROUPS", "NM_OIDC_ADMIN_GROUPS")
OIDC_DISCOVERY_TTL = int(_env_first("REFINER_OIDC_DISCOVERY_TTL", "NM_OIDC_DISCOVERY_TTL", default="3600"))
OIDC_JWT_LEEWAY = int(_env_first("REFINER_OIDC_JWT_LEEWAY", "NM_OIDC_JWT_LEEWAY", default="120"))
OIDC_SKIP_JWT_VERIFY = _env_flag("REFINER_OIDC_SKIP_JWT_VERIFY", _env_flag("NM_OIDC_SKIP_JWT_VERIFY", False))
OIDC_USE_USERINFO = _env_flag("REFINER_OIDC_USE_USERINFO", _env_flag("NM_OIDC_USE_USERINFO", False))
OIDC_BUTTON_LABEL = _env_first("REFINER_OIDC_BUTTON_LABEL", "NM_OIDC_BUTTON_LABEL", default="Sign in with SSO")
OIDC_REQUIRE_CONFIG = _env_flag("REFINER_OIDC_REQUIRE_CONFIG", _env_flag("NM_OIDC_REQUIRE_CONFIG", True))
OIDC_CLIENT_AUTH = _env_first("REFINER_OIDC_CLIENT_AUTH", "NM_OIDC_CLIENT_AUTH", default="basic").strip().lower()
OIDC_ALLOWED_AUDIENCES = _env_list_any(
    "REFINER_OIDC_ALLOWED_AUDIENCES",
    "REFINER_OIDC_AUDIENCE",
    "NM_OIDC_ALLOWED_AUDIENCES",
    "NM_OIDC_AUDIENCE",
)
OIDC_ALLOWED_REDIRECT_URIS = _env_list_any("REFINER_OIDC_ALLOWED_REDIRECT_URIS", "NM_OIDC_ALLOWED_REDIRECT_URIS")

if AUTH_MODE not in {"local", "oidc", "mixed"}:
    raise RuntimeError("REFINER_AUTH_MODE must be one of local, oidc, mixed.")
if AUTH_MODE == "oidc":
    OIDC_ENABLED = True
if OIDC_ENABLED and OIDC_REQUIRE_CONFIG:
    if not OIDC_ISSUER or not OIDC_CLIENT_ID:
        raise RuntimeError("OIDC enabled but REFINER_OIDC_ISSUER or REFINER_OIDC_CLIENT_ID missing.")
OIDC_EXCHANGE_ENABLED = _env_flag(
    "REFINER_OIDC_EXCHANGE_ENABLED",
    _env_flag("NM_OIDC_EXCHANGE_ENABLED", OIDC_ENABLED),
)
PASSWORD_MIN_LEN = int(os.getenv("REFINER_PASSWORD_MIN_LENGTH", "12"))
SELF_REGISTRATION_ENABLED = _env_flag("REFINER_SELF_REGISTRATION_ENABLED", False)
LOGIN_WINDOW_SEC = int(os.getenv("REFINER_LOGIN_WINDOW_SEC", "300"))
LOGIN_MAX_ATTEMPTS = int(os.getenv("REFINER_LOGIN_MAX_ATTEMPTS", "10"))

_LOGIN_ATTEMPTS: Dict[str, List[float]] = {}

app.config.update(
    SESSION_COOKIE_NAME=SESSION_COOKIE_NAME,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE=COOKIE_SAMESITE,
    SESSION_COOKIE_SECURE=SECURE_COOKIES,
    SESSION_COOKIE_DOMAIN=COOKIE_DOMAIN,
)

def _client_ip() -> str:
    forwarded = request.headers.get("X-Forwarded-For") if request else None
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.remote_addr or ""


def _login_key(username: str) -> str:
    return f"{username}:{_client_ip()}"


def _record_login_attempt(username: str, ok: bool) -> None:
    key = _login_key(username)
    now = time.time()
    attempts = _LOGIN_ATTEMPTS.get(key, [])
    attempts = [t for t in attempts if now - t < LOGIN_WINDOW_SEC]
    if not ok:
        attempts.append(now)
    _LOGIN_ATTEMPTS[key] = attempts


def _login_throttled(username: str) -> bool:
    key = _login_key(username)
    now = time.time()
    attempts = _LOGIN_ATTEMPTS.get(key, [])
    attempts = [t for t in attempts if now - t < LOGIN_WINDOW_SEC]
    _LOGIN_ATTEMPTS[key] = attempts
    return len(attempts) >= LOGIN_MAX_ATTEMPTS


def _audit_event(action: str, *, actor: Optional[str], status: str, details: Optional[Dict[str, object]] = None) -> None:
    safe_details = details or {}
    if has_request_context():
        ip = _client_ip()
        if ip:
            safe_details = dict(safe_details)
            safe_details.setdefault("ip_hash", hash_identifier(ip))
        agent = request.headers.get("User-Agent")
        if agent:
            safe_details = dict(safe_details)
            safe_details.setdefault("user_agent", agent[:200])
    audit_logger.log(action, actor=actor, status=status, details=safe_details)


def _read_audit_entries(limit: int = 60, actions: Optional[List[str]] = None) -> List[Dict[str, Any]]:
    if limit <= 0:
        return []
    action_set = {str(action).strip() for action in (actions or []) if str(action).strip()}
    try:
        lines = deque(maxlen=max(limit * 5, limit))
        if not os.path.exists(AUDIT_LOG_PATH):
            return []
        with open(AUDIT_LOG_PATH, "r", encoding="utf-8") as handle:
            for line in handle:
                if line.strip():
                    lines.append(line)
        entries: List[Dict[str, Any]] = []
        for line in reversed(lines):
            try:
                entry = json.loads(line)
            except Exception:
                continue
            if action_set and entry.get("action") not in action_set:
                continue
            entries.append(entry)
            if len(entries) >= limit:
                break
        return entries
    except Exception:
        return []


def _oidc_role_from_claims(claims: Dict[str, Any]) -> str:
    if not claims:
        return "user"
    email = claims.get(OIDC_EMAIL_CLAIM) or claims.get("email")
    if isinstance(email, str) and OIDC_ADMIN_DOMAINS:
        domain = email.split("@")[-1].lower() if "@" in email else ""
        if domain and domain in {d.lower() for d in OIDC_ADMIN_DOMAINS}:
            return "admin"
    groups = claims.get(OIDC_GROUPS_CLAIM) or claims.get("groups")
    group_values: List[str] = []
    if isinstance(groups, str):
        group_values = [g.strip() for g in groups.split(",") if g.strip()]
    elif isinstance(groups, list):
        group_values = [str(g).strip() for g in groups if str(g).strip()]
    if OIDC_ADMIN_GROUPS and group_values:
        wanted = {g.lower() for g in OIDC_ADMIN_GROUPS}
        if any(g.lower() in wanted for g in group_values):
            return "admin"
    return "user"


def _oidc_username_from_claims(claims: Dict[str, Any]) -> str:
    for key in (OIDC_USERNAME_CLAIM, "preferred_username", "email", "sub"):
        value = claims.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""

SSO_TTL_SECONDS = int(os.getenv("REFINER_SSO_TTL", "300"))
SSO_STORE_MODE = (os.getenv("REFINER_SSO_STORE") or "auto").strip().lower()
SSO_REDIS_URL = (os.getenv("REFINER_SSO_REDIS_URL") or os.getenv("REDIS_URL") or "").strip() or None
SSO_REDIS_PREFIX = (os.getenv("REFINER_SSO_REDIS_PREFIX") or "refiner:sso:").strip() or "refiner:sso:"

class SsoStore:
    """Abstract interface for issuing and consuming one-time SSO tokens."""

    type_name = "base"

    def issue(self, user: str) -> str:
        raise NotImplementedError

    def consume(self, token: str) -> Optional[str]:
        raise NotImplementedError

    def health(self) -> Dict[str, Any]:
        return {"type": self.type_name, "ok": True}


class MemorySsoStore(SsoStore):
    """In-memory SSO token store for single-process deployments."""

    type_name = "memory"

    def __init__(self, ttl_seconds: int):
        self.ttl_seconds = max(30, int(ttl_seconds))
        self._tokens: Dict[str, Dict[str, Any]] = {}
        self._lock = threading.Lock()

    def _prune(self, now: Optional[float] = None) -> None:
        timestamp = now if now is not None else time.time()
        with self._lock:
            expired = [
                token for token, meta in self._tokens.items() if meta.get("expires_at", 0) <= timestamp
            ]
            for token in expired:
                self._tokens.pop(token, None)

    def issue(self, user: str) -> str:
        token = secrets_lib.token_urlsafe(32)
        timestamp = time.time()
        expires_at = timestamp + self.ttl_seconds
        self._prune(timestamp)
        with self._lock:
            self._tokens[token] = {"user": user, "issued_at": timestamp, "expires_at": expires_at}
        return token

    def consume(self, token: str) -> Optional[str]:
        if not token:
            return None
        self._prune()
        with self._lock:
            meta = self._tokens.pop(token, None)
        if not meta:
            return None
        if meta.get("expires_at", 0) <= time.time():
            return None
        return meta.get("user")

    def health(self) -> Dict[str, Any]:
        return {"type": self.type_name, "ok": True}


class RedisSsoStore(SsoStore):
    """Redis-backed SSO token store for multi-instance deployments."""

    type_name = "redis"

    def __init__(self, client: Any, prefix: str, ttl_seconds: int):
        self._client = client
        self._prefix = prefix
        self._ttl_seconds = max(30, int(ttl_seconds))

    def _key(self, token: str) -> str:
        return f"{self._prefix}{token}"

    def issue(self, user: str) -> str:
        token = secrets_lib.token_urlsafe(32)
        self._client.setex(self._key(token), self._ttl_seconds, user)
        return token

    def consume(self, token: str) -> Optional[str]:
        if not token:
            return None
        key = self._key(token)
        if hasattr(self._client, "getdel"):
            user = self._client.getdel(key)
        else:
            script = "local v = redis.call('GET', KEYS[1]); if v then redis.call('DEL', KEYS[1]); end; return v;"
            user = self._client.eval(script, 1, key)
        if user:
            return str(user)
        return None

    def health(self) -> Dict[str, Any]:
        try:
            self._client.ping()
            return {"type": self.type_name, "ok": True}
        except Exception as exc:
            return {"type": self.type_name, "ok": False, "error": str(exc)}


def _init_central_store() -> Optional[Any]:
    try:
        store = create_central_store_from_env()
    except Exception as exc:
        logger.warning("central auth store initialization failed: %s", exc)
        return None
    if store is not None:
        logger.info("central auth store enabled with Postgres backing")
    return store


CENTRAL_STORE = _init_central_store()


def _central_job_store() -> Optional[Any]:
    if CENTRAL_STORE is None:
        return None
    return getattr(CENTRAL_STORE, "jobs", None)


def _init_sso_store() -> Dict[str, Any]:
    status: Dict[str, Any] = {"mode": SSO_STORE_MODE}
    if CENTRAL_STORE is not None:
        status.update({"ok": True, "active_store": "postgres"})
        return {"store": CENTRAL_STORE.sso_tokens, "status": status}
    mode = SSO_STORE_MODE
    if mode not in {"auto", "redis", "memory"}:
        status.update({"ok": False, "error": "invalid_mode"})
        mode = "auto"
    if mode == "memory":
        status.update({"ok": True, "active_store": "memory"})
        return {"store": MemorySsoStore(SSO_TTL_SECONDS), "status": status}
    if mode in {"auto", "redis"}:
        if not SSO_REDIS_URL:
            status.update(
                {
                    "ok": mode != "redis",
                    "active_store": "memory",
                    "error": "redis_url_missing" if mode == "redis" else None,
                }
            )
            return {"store": MemorySsoStore(SSO_TTL_SECONDS), "status": status}
        if redis is None:
            status.update(
                {
                    "ok": False,
                    "active_store": "memory",
                    "error": "redis_library_missing",
                }
            )
            return {"store": MemorySsoStore(SSO_TTL_SECONDS), "status": status}
        try:
            client = redis.Redis.from_url(SSO_REDIS_URL, decode_responses=True)
            client.ping()
            status.update({"ok": True, "active_store": "redis"})
            return {"store": RedisSsoStore(client, SSO_REDIS_PREFIX, SSO_TTL_SECONDS), "status": status}
        except Exception as exc:
            status.update(
                {
                    "ok": mode != "redis",
                    "active_store": "memory",
                    "error": "redis_unavailable",
                    "detail": str(exc),
                }
            )
            return {"store": MemorySsoStore(SSO_TTL_SECONDS), "status": status}
    status.update({"ok": True, "active_store": "memory"})
    return {"store": MemorySsoStore(SSO_TTL_SECONDS), "status": status}


_sso_init = _init_sso_store()
SSO_STORE: SsoStore = _sso_init["store"]
SSO_STORE_STATUS: Dict[str, Any] = _sso_init["status"]
if not SSO_STORE_STATUS.get("ok", True):
    logger.warning("SSO store initialized with warning: %s", SSO_STORE_STATUS)


def _sso_store_health() -> Dict[str, Any]:
    status = dict(SSO_STORE_STATUS)
    store_health = SSO_STORE.health()
    status["active_store"] = status.get("active_store") or store_health.get("type")
    status["store"] = store_health
    status["ok"] = bool(status.get("ok", True)) and bool(store_health.get("ok", True))
    return status


def _issue_sso_token(user: str) -> str:
    return SSO_STORE.issue(user)


def _consume_sso_token(token: str) -> Optional[str]:
    return SSO_STORE.consume(token)


def _safe_next_path(value: Optional[str]) -> str:
    raw = (value or "").strip()
    if not raw:
        return "/"
    if raw.startswith("//"):
        return "/"
    parsed = urlparse(raw)
    if parsed.scheme or parsed.netloc:
        return "/"
    if not raw.startswith("/"):
        return f"/{raw}"
    return raw


def _host_matches_pattern(host: str, pattern: str) -> bool:
    host = (host or "").strip().lower()
    normalized = (pattern or "").strip().lower().lstrip(".")
    if not host or not normalized:
        return False
    return host == normalized or host.endswith("." + normalized)


def _safe_external_redirect(value: Optional[str]) -> str:
    raw = (value or "").strip()
    if not raw:
        return "/"
    parsed = urlparse(raw)
    if not parsed.scheme and not parsed.netloc:
        return _safe_next_path(raw)
    if parsed.scheme not in {"http", "https"}:
        return "/"
    host = (parsed.hostname or "").strip().lower()
    if not host:
        return "/"

    allowed_patterns = [item for item in EXTERNAL_REDIRECT_HOSTS if item and item.strip()]
    if COOKIE_DOMAIN:
        allowed_patterns.append(COOKIE_DOMAIN)
    if has_request_context() and request.host:
        allowed_patterns.append(request.host.split(":", 1)[0])

    if any(_host_matches_pattern(host, pattern) for pattern in allowed_patterns):
        return raw
    return "/"


def _normalise_allowed_roots() -> List[str]:
    roots = []
    for root in RAG_ALLOWED_ROOTS:
        if not root:
            continue
        try:
            roots.append(os.path.abspath(root))
        except Exception:
            continue
    return sorted(set(roots))


def _coerce_rag_sources(payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    return _coerce_rag_sources_impl(payload)


def _build_rag_documents(
    sources: List[Dict[str, Any]],
    *,
    max_docs: int,
    max_doc_bytes: int,
    allowed_roots: Optional[List[str]] = None,
) -> List[Any]:
    return _build_rag_documents_impl(
        sources,
        max_docs=max_docs,
        max_doc_bytes=max_doc_bytes,
        allowed_roots=allowed_roots or _normalise_allowed_roots(),
    )


def _rag_match_citation(match: Any) -> str:
    metadata = match.metadata if isinstance(getattr(match, "metadata", None), dict) else {}
    citation = str(getattr(match, "citation", "") or metadata.get("citation") or "").strip()
    return citation or str(getattr(match, "source", "") or "source")


def _serialize_rag_match(match: Any) -> Dict[str, Any]:
    return _serialize_rag_match_impl(match)


def _render_rag_context(matches: List[Any]) -> str:
    return _render_rag_context_impl(matches)

if _env_flag("REFINER_TRUST_PROXY", False):
    from werkzeug.middleware.proxy_fix import ProxyFix

    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1)


def _ensure_dirs() -> None:
    ensure_dir_permissions(JOB_ROOT, mode=0o700)
    ensure_dir_permissions(PROJECTS_ROOT, mode=0o700)
    ensure_dir_permissions(SECRET_STORE_ROOT, mode=0o700)
    ensure_dir_permissions(ASSISTANT_MEMORY_ROOT, mode=0o700)
    ensure_dir_permissions(WORKSPACE_ROOT, mode=0o700)
    ensure_dir_permissions(TODO_ROOT, mode=0o700)
    ensure_dir_permissions(SCHEDULE_ROOT, mode=0o700)
    ensure_dir_permissions(LEDGER_ROOT, mode=0o700)
    ensure_dir_permissions(TEAM_LEDGER_ROOT, mode=0o700)
    ensure_dir_permissions(RAG_STORE_ROOT, mode=0o700)
    ensure_dir_permissions(MCP_STORE_ROOT, mode=0o700)
    ensure_dir_permissions(SESSIONS_ROOT, mode=0o700)


def _now_iso() -> str:
    return dt.datetime.now(UK_TZ).strftime(UK_DATETIME_FORMAT)


_OIDC_CACHE: Dict[str, Any] = {"ts": 0.0, "config": None, "jwks": None}


def _oidc_discovery() -> Optional[Dict[str, Any]]:
    if not OIDC_ENABLED or not OIDC_ISSUER:
        return None
    now = time.time()
    cached = _OIDC_CACHE.get("config")
    if cached and (now - float(_OIDC_CACHE.get("ts", 0.0)) < OIDC_DISCOVERY_TTL):
        return cached
    url = OIDC_ISSUER.rstrip("/") + "/.well-known/openid-configuration"
    resp = _http_request_with_retry("GET", url, timeout=12, retries=1)
    resp.raise_for_status()
    data = resp.json()
    if not isinstance(data, dict):
        raise RuntimeError("OIDC discovery failed (invalid JSON).")
    issuer = data.get("issuer")
    if issuer and issuer.rstrip("/") != OIDC_ISSUER.rstrip("/"):
        raise RuntimeError("OIDC issuer mismatch.")
    _OIDC_CACHE["config"] = data
    _OIDC_CACHE["ts"] = now
    return data


def _oidc_jwks() -> Optional[Dict[str, Any]]:
    config = _oidc_discovery()
    if not config:
        return None
    jwks = _OIDC_CACHE.get("jwks")
    if jwks and (time.time() - float(_OIDC_CACHE.get("ts", 0.0)) < OIDC_DISCOVERY_TTL):
        return jwks
    jwks_uri = config.get("jwks_uri")
    if not jwks_uri:
        return None
    resp = _http_request_with_retry("GET", jwks_uri, timeout=12, retries=1)
    resp.raise_for_status()
    data = resp.json()
    if isinstance(data, dict):
        _OIDC_CACHE["jwks"] = data
        _OIDC_CACHE["ts"] = time.time()
        return data
    return None


def _b64url_decode(value: str) -> bytes:
    if not value:
        return b""
    padding = "=" * (-len(value) % 4)
    return base64.urlsafe_b64decode(value + padding)


def _parse_jwt(token: str) -> Tuple[Dict[str, Any], Dict[str, Any], bytes, bytes]:
    parts = token.split(".")
    if len(parts) != 3:
        raise RuntimeError("Invalid JWT format.")
    header = json.loads(_b64url_decode(parts[0]).decode("utf-8"))
    payload = json.loads(_b64url_decode(parts[1]).decode("utf-8"))
    signature = _b64url_decode(parts[2])
    signing_input = ".".join(parts[:2]).encode("utf-8")
    return header, payload, signature, signing_input


def _jwk_to_public_key(jwk: Dict[str, Any]):
    if not jwk:
        return None
    if jwk.get("kty") != "RSA":
        return None
    n = jwk.get("n")
    e = jwk.get("e")
    if not n or not e:
        return None
    n_int = int.from_bytes(_b64url_decode(n), "big")
    e_int = int.from_bytes(_b64url_decode(e), "big")
    from cryptography.hazmat.primitives.asymmetric import rsa

    public_numbers = rsa.RSAPublicNumbers(e_int, n_int)
    return public_numbers.public_key()


def _verify_jwt(token: str, *, nonce: Optional[str]) -> Dict[str, Any]:
    header, payload, signature, signing_input = _parse_jwt(token)
    if not OIDC_SKIP_JWT_VERIFY:
        alg = header.get("alg")
        if alg != "RS256":
            raise RuntimeError("Unsupported JWT algorithm.")
        jwks = _oidc_jwks() or {}
        keys = jwks.get("keys") if isinstance(jwks, dict) else None
        if not isinstance(keys, list) or not keys:
            raise RuntimeError("JWKS missing.")
        kid = header.get("kid")
        jwk = None
        if kid:
            jwk = next((k for k in keys if k.get("kid") == kid), None)
        if not jwk:
            jwk = keys[0]
        public_key = _jwk_to_public_key(jwk)
        if not public_key:
            raise RuntimeError("Unsupported JWKS key.")
        from cryptography.hazmat.primitives.asymmetric import padding
        from cryptography.hazmat.primitives import hashes

        public_key.verify(signature, signing_input, padding.PKCS1v15(), hashes.SHA256())
    now = int(time.time())
    iss = payload.get("iss")
    if iss and iss.rstrip("/") != OIDC_ISSUER.rstrip("/"):
        raise RuntimeError("OIDC issuer mismatch.")
    allowed_audiences = {OIDC_CLIENT_ID} if OIDC_CLIENT_ID else set()
    if OIDC_ALLOWED_AUDIENCES:
        allowed_audiences.update({aud.strip() for aud in OIDC_ALLOWED_AUDIENCES if aud and aud.strip()})
    aud = payload.get("aud")
    if isinstance(aud, list):
        if allowed_audiences and not any(item in allowed_audiences for item in aud):
            raise RuntimeError("OIDC audience mismatch.")
    elif aud:
        if allowed_audiences and aud not in allowed_audiences:
            raise RuntimeError("OIDC audience mismatch.")
    exp = payload.get("exp")
    if exp and (now - OIDC_JWT_LEEWAY) > int(exp):
        raise RuntimeError("OIDC token expired.")
    if nonce:
        token_nonce = payload.get("nonce")
        if token_nonce and token_nonce != nonce:
            raise RuntimeError("OIDC nonce mismatch.")
    return payload


def _oidc_maybe_enrich_claims(
    claims: Dict[str, Any],
    access_token: Optional[str],
    *,
    config: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    if not claims or not access_token:
        return claims
    if not (OIDC_USE_USERINFO or not claims.get(OIDC_EMAIL_CLAIM)):
        return claims
    if config is None:
        config = _oidc_discovery()
    userinfo_endpoint = config.get("userinfo_endpoint") if isinstance(config, dict) else None
    if not userinfo_endpoint:
        return claims
    try:
        info_resp = _http_request_with_retry(
            "GET",
            userinfo_endpoint,
            headers={"Authorization": f"Bearer {access_token}"},
            timeout=12,
            retries=1,
        )
    except Exception as exc:
        logger.debug("OIDC userinfo request failed: %s", exc)
        return claims
    if info_resp.status_code >= 400:
        return claims
    try:
        info_data = info_resp.json()
    except Exception:
        return claims
    if isinstance(info_data, dict):
        claims.update(info_data)
    return claims


def _oidc_redirect_uri() -> str:
    if OIDC_REDIRECT_URI:
        return OIDC_REDIRECT_URI
    if SITE_BASE:
        return f"{SITE_BASE.rstrip('/')}/oidc/callback"
    if has_request_context():
        return request.host_url.rstrip("/") + "/oidc/callback"
    return "/oidc/callback"


def _oidc_allowed_redirect_uris() -> List[str]:
    candidates: List[str] = []
    if OIDC_REDIRECT_URI:
        candidates.append(OIDC_REDIRECT_URI)
    if SITE_BASE:
        candidates.append(f"{SITE_BASE.rstrip('/')}/oidc/callback")
    if has_request_context():
        candidates.append(request.host_url.rstrip("/") + "/oidc/callback")
    candidates.extend([uri for uri in OIDC_ALLOWED_REDIRECT_URIS if uri])
    seen: set[str] = set()
    cleaned: List[str] = []
    for uri in candidates:
        value = uri.strip().rstrip("/")
        if value and value not in seen:
            cleaned.append(value)
            seen.add(value)
    return cleaned


def _oidc_is_redirect_allowed(redirect_uri: str) -> bool:
    if not redirect_uri:
        return False
    candidate = redirect_uri.strip().rstrip("/")
    allowed = _oidc_allowed_redirect_uris()
    return candidate in allowed


def _parse_timestamp(value: Optional[str]) -> Optional[dt.datetime]:
    if not value or not isinstance(value, str):
        return None
    cleaned = value.strip()
    if not cleaned:
        return None
    for fmt in (UK_DATETIME_FORMAT, "%d/%m/%Y %H:%M"):
        try:
            parsed = dt.datetime.strptime(cleaned, fmt)
            return parsed.replace(tzinfo=UK_TZ)
        except ValueError:
            continue
    try:
        if cleaned.endswith("Z"):
            parsed = dt.datetime.strptime(cleaned, "%Y-%m-%dT%H:%M:%SZ")
            return parsed.replace(tzinfo=dt.timezone.utc)
        parsed = dt.datetime.fromisoformat(cleaned.replace("Z", "+00:00"))
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=dt.timezone.utc)
        return parsed
    except ValueError:
        return None


def _timestamp_sort_key(value: Optional[str]) -> float:
    parsed = _parse_timestamp(value)
    if not parsed:
        return 0.0
    return parsed.timestamp()


def _normalise_timestamp(value: Optional[str]) -> Optional[str]:
    parsed = _parse_timestamp(value)
    if not parsed:
        return value
    return parsed.astimezone(UK_TZ).strftime(UK_DATETIME_FORMAT)


def _slugify(value: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9_-]+", "-", (value or "").strip())
    cleaned = cleaned.strip("-")
    return cleaned or "project"


SECRET_NAME_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]{0,63}$")
USERNAME_RE = re.compile(r"^[A-Za-z0-9_\\-]{3,32}$")
EMAIL_RE = re.compile(r"^[^@\\s]+@[^@\\s]+\\.[^@\\s]+$")

GUARDRAIL_PATTERNS = {
    "violence": [
        r"\\b(make|build|assemble|design)\\b.*\\b(bomb|explosive|grenade)\\b",
        r"\\b(weapon|firearm|gun)\\b",
        r"\\bassassinate\\b",
    ],
    "malware": [
        r"\\b(ransomware|keylogger|malware|virus|trojan|rootkit)\\b",
        r"\\b(phishing|credential\\s+harvest)\\b",
    ],
    "illegal": [
        r"\\b(counterfeit|forgery|fake\\s+id)\\b",
        r"\\bcredit\\s+card\\s+fraud\\b",
        r"\\bdrug\\s+trafficking\\b",
    ],
    "self_harm": [
        r"\\b(suicide|self-harm|self harm)\\b",
    ],
}


class SecretStore:
    """Encrypted-at-rest credential store scoped per user."""

    def __init__(self, path: str):
        self.path = path
        self.lock = threading.RLock()
        self.data: Dict[str, Dict[str, str]] = {}
        self._fernet: Optional[object] = None
        if SECRET_STORE_KEY:
            if not Fernet:
                if SECRET_STORE_REQUIRE_ENCRYPTION:
                    raise RuntimeError("Secret store encryption required but cryptography is unavailable.")
                logger.warning("Secret store encryption key provided but cryptography is unavailable.")
            else:
                try:
                    key = SECRET_STORE_KEY.encode("utf-8")
                    try:
                        self._fernet = Fernet(key)
                    except Exception:
                        derived = base64.urlsafe_b64encode(hashlib.sha256(key).digest())
                        self._fernet = Fernet(derived)
                except Exception as exc:
                    if SECRET_STORE_REQUIRE_ENCRYPTION:
                        raise RuntimeError("Secret store encryption key invalid.") from exc
                    logger.warning("Secret store encryption disabled due to key error: %s", exc)
        self._load()

    def _load(self) -> None:
        if not os.path.exists(self.path):
            return
        try:
            with open(self.path, "r", encoding="utf-8") as handle:
                data = json.load(handle)
            if isinstance(data, dict):
                self.data = data
        except Exception:
            self.data = {}

    def _encrypt_value(self, value: str) -> Dict[str, str]:
        if not self._fernet:
            return {"value": value}
        token = self._fernet.encrypt(value.encode("utf-8")).decode("utf-8")
        return {"value": token, "enc": "fernet"}

    def _decrypt_entry(self, entry: Dict[str, str]) -> Optional[str]:
        if not entry:
            return None
        enc = entry.get("enc")
        value = entry.get("value") if isinstance(entry, dict) else None
        if not enc:
            if SECRET_STORE_KEY and not SECRET_STORE_ALLOW_PLAINTEXT:
                return None
            return value
        if enc == "fernet":
            if not self._fernet:
                return None
            try:
                return self._fernet.decrypt(value.encode("utf-8")).decode("utf-8")
            except Exception:
                return None
        return None

    def _write(self) -> None:
        tmp = f"{self.path}.tmp"
        try:
            os.makedirs(os.path.dirname(self.path), exist_ok=True)
            with open(tmp, "w", encoding="utf-8") as handle:
                json.dump(self.data, handle, indent=2)
            os.replace(tmp, self.path)
            try:
                os.chmod(self.path, 0o600)
            except Exception:
                pass
        finally:
            if os.path.exists(tmp):
                try:
                    os.remove(tmp)
                except Exception:
                    pass

    def list_masked(self) -> List[Dict[str, str]]:
        with self.lock:
            items = []
            for name, entry in sorted(self.data.items()):
                value = entry.get("value") if isinstance(entry, dict) else ""
                masked = self._mask(value)
                updated_at = entry.get("updated_at") if isinstance(entry, dict) else None
                items.append({"name": name, "masked": masked, "updated_at": updated_at})
            return items

    def get(self, name: str) -> Optional[str]:
        with self.lock:
            entry = self.data.get(name)
            if not isinstance(entry, dict):
                return None
            return self._decrypt_entry(entry)

    def set(self, name: str, value: str) -> None:
        with self.lock:
            payload = self._encrypt_value(value)
            payload["updated_at"] = _now_iso()
            self.data[name] = payload
            self._write()

    def delete(self, name: str) -> bool:
        with self.lock:
            if name not in self.data:
                return False
            self.data.pop(name, None)
            self._write()
            return True

    def get_env(self) -> Dict[str, str]:
        with self.lock:
            env = {}
            for name, entry in self.data.items():
                if isinstance(entry, dict) and entry.get("value"):
                    value = self._decrypt_entry(entry)
                    if value:
                        env[name] = value
            return env

    @staticmethod
    def _mask(value: Optional[str]) -> str:
        if not value:
            return "not set"
        tail = value[-4:] if len(value) >= 4 else value
        return f"***{tail}"


class UserStore:
    """Persistent local user/role registry with password hashing."""

    def __init__(self, path: str):
        self.path = path
        self.lock = threading.RLock()
        self.users: Dict[str, Dict[str, str]] = {}
        self._load()

    def _load(self) -> None:
        if not os.path.exists(self.path):
            return
        try:
            with open(self.path, "r", encoding="utf-8") as handle:
                data = json.load(handle)
            if isinstance(data, dict):
                self.users = data
        except Exception:
            self.users = {}

    def _write(self) -> None:
        tmp = f"{self.path}.tmp"
        try:
            with open(tmp, "w", encoding="utf-8") as handle:
                json.dump(self.users, handle, indent=2)
            os.replace(tmp, self.path)
            try:
                os.chmod(self.path, 0o600)
            except Exception:
                pass
        finally:
            if os.path.exists(tmp):
                try:
                    os.remove(tmp)
                except Exception:
                    pass

    def has_users(self) -> bool:
        with self.lock:
            return bool(self.users)

    def count_users(self) -> int:
        with self.lock:
            return len(self.users)

    def ensure_admin_from_env(self) -> None:
        admin_user = (os.getenv("REFINER_ADMIN_USER") or "").strip()
        admin_pass = (os.getenv("REFINER_ADMIN_PASS") or "").strip()
        admin_email = (os.getenv("REFINER_ADMIN_EMAIL") or "").strip()
        if not admin_user or not admin_pass:
            return
        with self.lock:
            if self.users:
                return
            self.users[admin_user] = {
                "password": generate_password_hash(admin_pass),
                "created_at": _now_iso(),
                "role": "admin",
            }
            if admin_email:
                self.users[admin_user]["email"] = admin_email
            self._write()

    def create_user(self, username: str, password: str, role: str = "user", email: Optional[str] = None) -> None:
        with self.lock:
            self.users[username] = {
                "password": generate_password_hash(password),
                "created_at": _now_iso(),
                "role": role,
            }
            if email:
                self.users[username]["email"] = email
            self._write()

    def upsert_external_user(
        self,
        username: str,
        *,
        role: str = "user",
        email: Optional[str] = None,
        provider: str = "oidc",
        subject: Optional[str] = None,
    ) -> None:
        with self.lock:
            entry = self.users.get(username, {})
            if "created_at" not in entry:
                entry["created_at"] = _now_iso()
            entry["role"] = role
            entry["external"] = True
            entry["provider"] = provider
            if subject:
                entry["subject"] = subject
            if email:
                entry["email"] = email
            self.users[username] = entry
            self._write()

    def set_email(self, username: str, email: Optional[str]) -> bool:
        with self.lock:
            entry = self.users.get(username)
            if not entry:
                return False
            if email:
                entry["email"] = email
            else:
                entry.pop("email", None)
            entry["updated_at"] = _now_iso()
            self._write()
            return True

    def get_email(self, username: str) -> Optional[str]:
        with self.lock:
            entry = self.users.get(username) or {}
            return entry.get("email")

    def verify(self, username: str, password: str) -> bool:
        with self.lock:
            entry = self.users.get(username)
            if not entry:
                return False
            return check_password_hash(entry.get("password") or "", password)

    def get_role(self, username: str) -> Optional[str]:
        with self.lock:
            entry = self.users.get(username) or {}
            return entry.get("role")

    def get_metadata(self, username: str) -> Dict[str, Any]:
        with self.lock:
            entry = self.users.get(username) or {}
            metadata = entry.get("metadata")
            return dict(metadata) if isinstance(metadata, dict) else {}

    def set_metadata(self, username: str, metadata: Optional[Dict[str, Any]]) -> bool:
        with self.lock:
            entry = self.users.get(username)
            if not entry:
                return False
            entry["metadata"] = dict(metadata or {})
            entry["updated_at"] = _now_iso()
            self._write()
            return True


class AccessStore:
    """Store for team/project membership and role-based access metadata."""

    def __init__(self, path: str):
        self.path = path
        self.lock = threading.RLock()
        self.data: Dict[str, Any] = {"version": 1, "teams": {}, "projects": {}}
        self._load()

    def _load(self) -> None:
        if not os.path.exists(self.path):
            return
        try:
            with open(self.path, "r", encoding="utf-8") as handle:
                data = json.load(handle)
            if isinstance(data, dict):
                self.data.update(data)
        except Exception:
            self.data = {"version": 1, "teams": {}, "projects": {}}

    def _write(self) -> None:
        tmp = f"{self.path}.tmp"
        try:
            with open(tmp, "w", encoding="utf-8") as handle:
                json.dump(self.data, handle, indent=2)
            os.replace(tmp, self.path)
            try:
                os.chmod(self.path, 0o600)
            except Exception:
                pass
        finally:
            if os.path.exists(tmp):
                try:
                    os.remove(tmp)
                except Exception:
                    pass

    @staticmethod
    def _normalise_users(values: Optional[object]) -> List[str]:
        if values is None:
            return []
        if isinstance(values, str):
            values = [item.strip() for item in values.split(",") if item.strip()]
        if not isinstance(values, list):
            return []
        seen = set()
        cleaned: List[str] = []
        for value in values:
            name = str(value or "").strip()
            if not name or name in seen:
                continue
            cleaned.append(name)
            seen.add(name)
        return cleaned

    @staticmethod
    def _normalise_permissions(value: Optional[Dict[str, Any]]) -> Dict[str, List[str]]:
        if not isinstance(value, dict):
            return {"read": [], "write": [], "grant": []}
        read = AccessStore._normalise_users(value.get("read") if isinstance(value.get("read"), list) else value.get("read"))
        write = AccessStore._normalise_users(value.get("write") if isinstance(value.get("write"), list) else value.get("write"))
        grant = AccessStore._normalise_users(value.get("grant") if isinstance(value.get("grant"), list) else value.get("grant"))
        return {"read": read, "write": write, "grant": grant}

    def list_teams(self) -> List[Dict[str, Any]]:
        with self.lock:
            return list(self.data.get("teams", {}).values())

    def list_projects(self) -> List[Dict[str, Any]]:
        with self.lock:
            return list(self.data.get("projects", {}).values())

    def get_team(self, team_id: str) -> Optional[Dict[str, Any]]:
        if not team_id:
            return None
        with self.lock:
            return self.data.get("teams", {}).get(team_id)

    def get_project(self, project_id: str) -> Optional[Dict[str, Any]]:
        if not project_id:
            return None
        with self.lock:
            return self.data.get("projects", {}).get(project_id)

    def create_team(
        self,
        name: str,
        *,
        parent_id: Optional[str] = None,
        leaders: Optional[List[str]] = None,
        members: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        team_id = uuid.uuid4().hex
        entry = {
            "id": team_id,
            "name": str(name or "").strip() or f"Team {team_id[:6]}",
            "parent_id": parent_id,
            "leaders": self._normalise_users(leaders),
            "members": self._normalise_users(members),
            "created_at": _now_iso(),
            "updated_at": _now_iso(),
        }
        with self.lock:
            self.data.setdefault("teams", {})[team_id] = entry
            self._write()
        return entry

    def update_team(
        self,
        team_id: str,
        *,
        name: Optional[str] = None,
        parent_id: Optional[str] = None,
        leaders: Optional[List[str]] = None,
        members: Optional[List[str]] = None,
    ) -> Optional[Dict[str, Any]]:
        with self.lock:
            team = self.data.get("teams", {}).get(team_id)
            if not team:
                return None
            if name is not None:
                team["name"] = str(name).strip() or team.get("name")
            if parent_id is not None:
                team["parent_id"] = parent_id or None
            if leaders is not None:
                team["leaders"] = self._normalise_users(leaders)
            if members is not None:
                team["members"] = self._normalise_users(members)
            team["updated_at"] = _now_iso()
            self._write()
            return dict(team)

    def delete_team(self, team_id: str) -> bool:
        with self.lock:
            projects = self.data.get("projects", {})
            for project in projects.values():
                if project.get("team_id") == team_id:
                    return False
            if team_id not in self.data.get("teams", {}):
                return False
            self.data["teams"].pop(team_id, None)
            self._write()
            return True

    def create_project(
        self,
        name: str,
        *,
        team_id: Optional[str] = None,
        leaders: Optional[List[str]] = None,
        contributors: Optional[List[str]] = None,
        viewers: Optional[List[str]] = None,
        permissions: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        if team_id and not _team_display_record(team_id):
            raise ValueError("Unknown team_id")
        project_id = uuid.uuid4().hex
        entry = {
            "id": project_id,
            "name": str(name or "").strip() or f"Project {project_id[:6]}",
            "team_id": team_id,
            "leaders": self._normalise_users(leaders),
            "contributors": self._normalise_users(contributors),
            "viewers": self._normalise_users(viewers),
            "permissions": self._normalise_permissions(permissions),
            "created_at": _now_iso(),
            "updated_at": _now_iso(),
        }
        with self.lock:
            self.data.setdefault("projects", {})[project_id] = entry
            self._write()
        return entry

    def update_project(
        self,
        project_id: str,
        *,
        name: Optional[str] = None,
        team_id: Optional[str] = None,
        leaders: Optional[List[str]] = None,
        contributors: Optional[List[str]] = None,
        viewers: Optional[List[str]] = None,
        permissions: Optional[Dict[str, Any]] = None,
    ) -> Optional[Dict[str, Any]]:
        with self.lock:
            project = self.data.get("projects", {}).get(project_id)
            if not project:
                return None
            if name is not None:
                project["name"] = str(name).strip() or project.get("name")
            if team_id is not None:
                if team_id and not _team_display_record(team_id):
                    raise ValueError("Unknown team_id")
                project["team_id"] = team_id or None
            if leaders is not None:
                project["leaders"] = self._normalise_users(leaders)
            if contributors is not None:
                project["contributors"] = self._normalise_users(contributors)
            if viewers is not None:
                project["viewers"] = self._normalise_users(viewers)
            if permissions is not None:
                project["permissions"] = self._normalise_permissions(permissions)
            project["updated_at"] = _now_iso()
            self._write()
            return dict(project)

    def delete_project(self, project_id: str) -> bool:
        with self.lock:
            if project_id not in self.data.get("projects", {}):
                return False
            self.data["projects"].pop(project_id, None)
            self._write()
            return True

    def _team_chain(self, team_id: Optional[str]) -> List[Dict[str, Any]]:
        chain: List[Dict[str, Any]] = []
        visited = set()
        current = team_id
        while current:
            if current in visited:
                break
            visited.add(current)
            team = self.get_team(current)
            if not team:
                break
            chain.append(team)
            current = team.get("parent_id")
        return chain

    def project_team(self, project_id: str) -> Optional[Dict[str, Any]]:
        project = self.get_project(project_id)
        if not project:
            return None
        team_id = project.get("team_id")
        return _team_display_record(team_id) if team_id else None

    def _team_role(self, user: str, team_id: Optional[str]) -> Optional[str]:
        if not user or not team_id:
            return None
        team_chain = self._team_chain(team_id)
        for team in team_chain:
            if user in (team.get("leaders") or []):
                return "leader"
        for team in team_chain:
            if user in (team.get("members") or []):
                return "member"
        return None

    def project_capabilities(self, user: str, project_id: Optional[str]) -> Dict[str, bool]:
        capabilities = {"read": False, "write": False, "grant": False}
        if not user or not project_id:
            return capabilities
        project = self.get_project(project_id)
        if not project:
            return capabilities
        user = str(user).strip()
        permissions = project.get("permissions") if isinstance(project.get("permissions"), dict) else {}
        permissions = self._normalise_permissions(permissions)
        if user in (permissions.get("grant") or []):
            return {"read": True, "write": True, "grant": True}
        if user in (permissions.get("write") or []):
            capabilities["read"] = True
            capabilities["write"] = True
        if user in (permissions.get("read") or []):
            capabilities["read"] = True
        if user in (project.get("leaders") or []):
            return {"read": True, "write": True, "grant": True}
        if user in (project.get("contributors") or []):
            return {"read": True, "write": True, "grant": False}
        if user in (project.get("viewers") or []):
            return {"read": True, "write": False, "grant": False}
        team_role = _team_role_for_user(user, project.get("team_id"))
        if team_role == "leader":
            return {"read": True, "write": True, "grant": True}
        if team_role == "member":
            return {"read": True, "write": True, "grant": False}
        return capabilities

    def can_create_project(self, user: str, team_id: Optional[str]) -> bool:
        if not user or not team_id:
            return False
        if _team_role_for_user(user, team_id) == "leader":
            return True
        for project in self.list_projects():
            if project.get("team_id") != team_id:
                continue
            caps = self.project_capabilities(user, project.get("id"))
            if caps.get("grant"):
                return True
        return False

    def project_role(self, user: str, project_id: Optional[str]) -> Optional[str]:
        if not user or not project_id:
            return None
        project = self.get_project(project_id)
        if not project:
            return None
        user = str(user).strip()
        if user in (project.get("leaders") or []):
            return "leader"
        if user in (project.get("contributors") or []):
            return "contributor"
        if user in (project.get("viewers") or []):
            return "viewer"
        permissions = project.get("permissions") if isinstance(project.get("permissions"), dict) else {}
        permissions = self._normalise_permissions(permissions)
        if user in (permissions.get("grant") or []):
            return "grant"
        if user in (permissions.get("write") or []):
            return "writer"
        if user in (permissions.get("read") or []):
            return "reader"
        team_role = _team_role_for_user(user, project.get("team_id"))
        if team_role == "leader":
            return "leader"
        if team_role == "member":
            return "contributor"
        return None

    def can_view_project(self, user: str, project_id: Optional[str]) -> bool:
        caps = self.project_capabilities(user, project_id)
        return caps.get("read", False)

    def can_submit_project(self, user: str, project_id: Optional[str]) -> bool:
        caps = self.project_capabilities(user, project_id)
        return caps.get("write", False)

    def can_manage_project(self, user: str, project_id: Optional[str]) -> bool:
        caps = self.project_capabilities(user, project_id)
        return caps.get("grant", False)

    def projects_for_user(self, user: str, include_viewers: bool = True) -> List[Dict[str, Any]]:
        projects = self.list_projects()
        visible: List[Dict[str, Any]] = []
        for project in projects:
            caps = self.project_capabilities(user, project.get("id"))
            if not caps.get("read"):
                continue
            role = self.project_role(user, project.get("id"))
            if not include_viewers and not caps.get("write") and role == "viewer":
                continue
            entry = dict(project)
            entry["role"] = role
            entry["capabilities"] = caps
            team = _team_display_record(project.get("team_id")) if project.get("team_id") else None
            if team:
                entry["team_name"] = team.get("name")
            visible.append(entry)
        return visible

    def tree_for_user(self, user: str, include_viewers: bool = True) -> List[Dict[str, Any]]:
        teams = {team.get("id"): dict(team) for team in self.list_teams() if team.get("id")}
        projects = self.projects_for_user(user, include_viewers=include_viewers)
        projects_by_team: Dict[Optional[str], List[Dict[str, Any]]] = {}
        for project in projects:
            projects_by_team.setdefault(project.get("team_id"), []).append(project)
        children: Dict[Optional[str], List[Dict[str, Any]]] = {}
        for team in teams.values():
            parent_id = team.get("parent_id")
            children.setdefault(parent_id, []).append(team)
        for team_list in children.values():
            team_list.sort(key=lambda t: t.get("name") or "")

        def build_node(team: Dict[str, Any]) -> Dict[str, Any]:
            node = dict(team)
            node["projects"] = sorted(
                projects_by_team.get(team.get("id"), []),
                key=lambda p: p.get("name") or "",
            )
            node["children"] = [build_node(child) for child in children.get(team.get("id"), [])]
            return node

        roots = children.get(None, []) + children.get("", [])
        tree = [build_node(team) for team in roots]
        # Include projects without a team
        unassigned = projects_by_team.get(None) or []
        if unassigned:
            tree.append({"id": None, "name": "Unassigned", "projects": unassigned, "children": []})
        return tree

    def tree_all(self) -> List[Dict[str, Any]]:
        teams = {team.get("id"): dict(team) for team in self.list_teams() if team.get("id")}
        projects = [dict(project) for project in self.list_projects()]
        projects_by_team: Dict[Optional[str], List[Dict[str, Any]]] = {}
        for project in projects:
            projects_by_team.setdefault(project.get("team_id"), []).append(project)
        children: Dict[Optional[str], List[Dict[str, Any]]] = {}
        for team in teams.values():
            parent_id = team.get("parent_id")
            children.setdefault(parent_id, []).append(team)
        for team_list in children.values():
            team_list.sort(key=lambda t: t.get("name") or "")

        def build_node(team: Dict[str, Any]) -> Dict[str, Any]:
            node = dict(team)
            node["projects"] = sorted(
                projects_by_team.get(team.get("id"), []),
                key=lambda p: p.get("name") or "",
            )
            node["children"] = [build_node(child) for child in children.get(team.get("id"), [])]
            return node

        roots = children.get(None, []) + children.get("", [])
        tree = [build_node(team) for team in roots]
        unassigned = projects_by_team.get(None) or []
        if unassigned:
            tree.append({"id": None, "name": "Unassigned", "projects": unassigned, "children": []})
        return tree


class VoiceTokenStore:
    """Persistent registry for API voice tokens and associated users."""

    def __init__(self, path: str):
        self.path = path
        self.lock = threading.RLock()
        self.data: Dict[str, Any] = {"version": 1, "tokens": []}
        self._load()

    def _load(self) -> None:
        if not os.path.exists(self.path):
            return
        try:
            with open(self.path, "r", encoding="utf-8") as handle:
                data = json.load(handle)
            if isinstance(data, dict):
                tokens = data.get("tokens")
                if isinstance(tokens, list):
                    self.data = {"version": int(data.get("version") or 1), "tokens": tokens}
        except Exception:
            self.data = {"version": 1, "tokens": []}

    def _write(self) -> None:
        _write_json_atomic(self.path, self.data)

    @staticmethod
    def _hash_token(token: str) -> str:
        return hashlib.sha256(token.encode("utf-8")).hexdigest()

    def issue(self, user: str, label: Optional[str] = None) -> Dict[str, Any]:
        token = secrets_lib.token_urlsafe(32)
        entry = {
            "id": uuid.uuid4().hex,
            "hash": self._hash_token(token),
            "user": user,
            "label": (label or "").strip() or None,
            "created_at": _now_iso(),
            "last_used_at": None,
            "disabled": False,
        }
        with self.lock:
            tokens = self.data.get("tokens")
            if not isinstance(tokens, list):
                tokens = []
            tokens.append(entry)
            self.data["tokens"] = tokens
            self._write()
        return {"token": token, **entry}

    def verify(self, token: str) -> Optional[str]:
        if not token:
            return None
        token_hash = self._hash_token(token)
        with self.lock:
            tokens = self.data.get("tokens")
            if not isinstance(tokens, list):
                return None
            for entry in tokens:
                if not isinstance(entry, dict):
                    continue
                if entry.get("hash") == token_hash and not entry.get("disabled"):
                    entry["last_used_at"] = _now_iso()
                    self._write()
                    return str(entry.get("user") or "").strip() or None
        return None

    def list_tokens(self, user: Optional[str] = None) -> List[Dict[str, Any]]:
        with self.lock:
            tokens = list(self.data.get("tokens") or [])
        filtered: List[Dict[str, Any]] = []
        for entry in tokens:
            if not isinstance(entry, dict):
                continue
            if user and entry.get("user") != user:
                continue
            filtered.append(
                {
                    "id": entry.get("id"),
                    "user": entry.get("user"),
                    "label": entry.get("label"),
                    "created_at": entry.get("created_at"),
                    "last_used_at": entry.get("last_used_at"),
                    "disabled": bool(entry.get("disabled")),
                }
            )
        filtered.sort(key=lambda t: _timestamp_sort_key(t.get("created_at")), reverse=True)
        return filtered

    def revoke(self, token_id: str) -> bool:
        if not token_id:
            return False
        updated = False
        with self.lock:
            tokens = self.data.get("tokens")
            if not isinstance(tokens, list):
                return False
            for entry in tokens:
                if not isinstance(entry, dict):
                    continue
                if entry.get("id") == token_id and not entry.get("disabled"):
                    entry["disabled"] = True
                    entry["disabled_at"] = _now_iso()
                    updated = True
            if updated:
                self._write()
        return updated


class TodoStore:
    """Per-user Thought Inbox persistence with local triage and claim tracking.

    The original implementation stored a flat list of free-form strings. The
    upgraded store keeps the same surface area but enriches each item with
    deterministic metadata inspired by always-on assistant runtimes:

    - duplicate captures collapse into one item via a stable fingerprint,
    - route hints map thoughts onto existing Refiner assistant/job endpoints,
    - stale claims are automatically recovered so "next when idle" cannot get
      permanently wedged, and
    - lightweight local search works without needing an external memory system.
    """

    def __init__(self, root: str, *, claim_ttl_sec: int = TODO_CLAIM_TTL_SEC, retention_days: int = TODO_RETENTION_DAYS):
        self.root = root
        self.lock = threading.RLock()
        self.claim_ttl_sec = max(60, int(claim_ttl_sec))
        self.retention_days = max(0, int(retention_days))
        ensure_dir_permissions(root, mode=0o700)

    def _safe_user(self, user: str) -> str:
        return re.sub(r"[^A-Za-z0-9_.-]+", "_", user or "unknown")

    def _path(self, user: str) -> str:
        return os.path.join(self.root, f"{self._safe_user(user)}.json")

    def _load(self, user: str) -> Dict[str, Any]:
        path = self._path(user)
        if not os.path.exists(path):
            return {"version": 2, "items": []}
        try:
            with open(path, "r", encoding="utf-8") as handle:
                data = json.load(handle)
            if not isinstance(data, dict):
                return {"version": 2, "items": []}
            if not isinstance(data.get("items"), list):
                data["items"] = []
            if "version" not in data:
                data["version"] = 2
            return data
        except Exception:
            return {"version": 2, "items": []}

    def _write(self, user: str, data: Dict[str, Any]) -> None:
        _write_json_atomic(self._path(user), data)

    def _coerce_text_list(self, value: Any) -> List[str]:
        if isinstance(value, str):
            value = [item.strip() for item in value.split(",") if item.strip()]
        if not isinstance(value, list):
            return []
        return [str(item).strip() for item in value if str(item or "").strip()]

    def _refresh_text_fields(self, item: Dict[str, Any]) -> bool:
        """Backfill or recompute the derived fields tied directly to thought text."""
        changed = False
        cleaned_text = normalize_thought_text(str(item.get("text") or ""))
        if cleaned_text != item.get("text"):
            item["text"] = cleaned_text
            changed = True
        kind = inbox_infer_kind(cleaned_text, source=item.get("source"))
        if item.get("kind") != kind:
            item["kind"] = kind
            changed = True
        priority = inbox_infer_priority(
            cleaned_text,
            kind=kind,
            defer_until_idle=bool(item.get("defer_until_idle")),
        )
        if item.get("priority") != priority:
            item["priority"] = priority
            changed = True
        keywords = inbox_extract_keywords(cleaned_text)
        if item.get("keywords") != keywords:
            item["keywords"] = keywords
            changed = True
        fingerprint = inbox_build_fingerprint(cleaned_text)
        if item.get("fingerprint") != fingerprint:
            item["fingerprint"] = fingerprint
            changed = True
        return changed

    def _hydrate_item(self, item: Dict[str, Any]) -> bool:
        """Normalize a stored item into the current schema in-place."""
        changed = False
        if not isinstance(item, dict):
            return False
        if not item.get("id"):
            item["id"] = uuid.uuid4().hex
            changed = True
        if not item.get("status"):
            item["status"] = "todo"
            changed = True
        if "defer_until_idle" not in item:
            item["defer_until_idle"] = False
            changed = True
        if "created_at" not in item:
            item["created_at"] = _now_iso()
            changed = True
        if "updated_at" not in item:
            item["updated_at"] = item.get("created_at") or _now_iso()
            changed = True
        if "first_captured_at" not in item:
            item["first_captured_at"] = item.get("created_at") or _now_iso()
            changed = True
        if "last_captured_at" not in item:
            item["last_captured_at"] = item.get("updated_at") or item.get("created_at") or _now_iso()
            changed = True
        if "occurrences" not in item or int(item.get("occurrences") or 0) <= 0:
            item["occurrences"] = 1
            changed = True
        if "source_history" not in item:
            item["source_history"] = self._coerce_text_list(item.get("source_history"))
            if item.get("source"):
                item["source_history"] = [str(item.get("source")).strip()]
            changed = True
        else:
            coerced = self._coerce_text_list(item.get("source_history"))
            if coerced != item.get("source_history"):
                item["source_history"] = coerced
                changed = True
        if "device_history" not in item:
            item["device_history"] = self._coerce_text_list(item.get("device_history"))
            if item.get("device"):
                item["device_history"] = [str(item.get("device")).strip()]
            changed = True
        else:
            coerced = self._coerce_text_list(item.get("device_history"))
            if coerced != item.get("device_history"):
                item["device_history"] = coerced
                changed = True
        if "execution_state" not in item:
            item["execution_state"] = "ready" if str(item.get("status") or "todo").lower() == "todo" else "completed"
            changed = True
        if "tags" in item:
            coerced_tags = self._coerce_text_list(item.get("tags"))
            if coerced_tags != item.get("tags"):
                item["tags"] = coerced_tags
                changed = True
        if "links" in item and not isinstance(item.get("links"), dict):
            item.pop("links", None)
            changed = True
        changed = self._refresh_text_fields(item) or changed
        return changed

    def _maintenance_locked(self, data: Dict[str, Any]) -> bool:
        """Repair stale items and prune terminal records when retention is enabled."""
        changed = False
        items = data.get("items")
        if not isinstance(items, list):
            data["items"] = []
            return True
        active_items: List[Dict[str, Any]] = []
        cutoff_ts = time.time() - (self.retention_days * 86400) if self.retention_days > 0 else None
        now_iso = _now_iso()
        for raw_item in items:
            if not isinstance(raw_item, dict):
                changed = True
                continue
            item = raw_item
            changed = self._hydrate_item(item) or changed
            claim_expires_at = item.get("claim_expires_at")
            execution_state = str(item.get("execution_state") or "").strip().lower()
            if execution_state in {"claimed", "processing"} and _timestamp_sort_key(claim_expires_at) > 0:
                if _timestamp_sort_key(claim_expires_at) <= time.time():
                    item["execution_state"] = "ready"
                    item["updated_at"] = now_iso
                    item.pop("claim_expires_at", None)
                    item.pop("claimed_at", None)
                    changed = True
            status = str(item.get("status") or "todo").lower()
            updated_ts = _timestamp_sort_key(item.get("updated_at") or item.get("created_at"))
            if cutoff_ts is not None and status in {"done", "archived"} and updated_ts and updated_ts < cutoff_ts:
                changed = True
                continue
            active_items.append(item)
        if len(active_items) != len(items):
            data["items"] = active_items
            changed = True
        data["version"] = 2
        return changed

    def _priority_weight(self, item: Dict[str, Any]) -> int:
        return {"high": 3, "medium": 2, "low": 1}.get(str(item.get("priority") or "low").lower(), 1)

    def _item_sort_key(self, item: Dict[str, Any]) -> Tuple[int, int, int, float, float]:
        status = str(item.get("status") or "todo").lower()
        status_rank = {"todo": 0, "done": 1, "archived": 2}.get(status, 3)
        occurrences = max(1, int(item.get("occurrences") or 1))
        updated = _timestamp_sort_key(item.get("last_captured_at") or item.get("updated_at") or item.get("created_at"))
        created = _timestamp_sort_key(item.get("created_at"))
        return (
            status_rank,
            -self._priority_weight(item),
            -occurrences,
            -updated,
            -created,
        )

    def _next_sort_key(self, item: Dict[str, Any]) -> Tuple[int, int, float, float]:
        occurrences = max(1, int(item.get("occurrences") or 1))
        created = _timestamp_sort_key(item.get("first_captured_at") or item.get("created_at"))
        updated = _timestamp_sort_key(item.get("last_captured_at") or item.get("updated_at") or item.get("created_at"))
        return (
            -self._priority_weight(item),
            -occurrences,
            created or updated,
            updated,
        )

    def _is_ready_item(self, item: Dict[str, Any]) -> bool:
        status = str(item.get("status") or "todo").lower()
        execution_state = str(item.get("execution_state") or "ready").lower()
        available_after = item.get("available_after")
        if status != "todo":
            return False
        if execution_state in {"claimed", "processing"}:
            return False
        if available_after and _timestamp_sort_key(available_after) > time.time():
            return False
        return True

    def _next_candidate_locked(self, data: Dict[str, Any], *, idle_only: bool = False) -> Optional[Dict[str, Any]]:
        """Select the next eligible item without mutating claim state."""
        items = data.get("items")
        if not isinstance(items, list):
            return None
        candidates = [
            item
            for item in items
            if isinstance(item, dict)
            and self._is_ready_item(item)
            and (not idle_only or bool(item.get("defer_until_idle")))
        ]
        if not candidates:
            return None
        candidates.sort(key=self._next_sort_key)
        return candidates[0]

    def get_item(self, user: str, todo_id: str) -> Optional[Dict[str, Any]]:
        if not todo_id:
            return None
        with self.lock:
            data = self._load(user)
            changed = self._maintenance_locked(data)
            for item in data.get("items", []):
                if isinstance(item, dict) and item.get("id") == todo_id:
                    if changed:
                        data["updated_at"] = _now_iso()
                        self._write(user, data)
                    return dict(item)
            if changed:
                data["updated_at"] = _now_iso()
                self._write(user, data)
            return None

    def list_items(
        self,
        user: str,
        *,
        statuses: Optional[List[str]] = None,
        limit: Optional[int] = None,
        query: Optional[str] = None,
        ready_only: bool = False,
        include_routes: bool = False,
    ) -> List[Dict[str, Any]]:
        with self.lock:
            data = self._load(user)
            changed = self._maintenance_locked(data)
            items = [dict(item) for item in data.get("items", []) if isinstance(item, dict)]
            if changed:
                data["updated_at"] = _now_iso()
                self._write(user, data)
        if statuses:
            wanted = {s.strip().lower() for s in statuses if s and str(s).strip()}
            if wanted:
                items = [item for item in items if str(item.get("status") or "todo").lower() in wanted]
        if ready_only:
            items = [item for item in items if self._is_ready_item(item)]
        query_text = str(query or "").strip()
        if query_text:
            scored_items: List[Dict[str, Any]] = []
            for item in items:
                score = score_query_match(item, query_text)
                if score <= 0:
                    continue
                item["_query_score"] = score
                scored_items.append(item)
            scored_items.sort(
                key=lambda item: (
                    -float(item.get("_query_score") or 0.0),
                    *self._item_sort_key(item),
                )
            )
            items = scored_items
        else:
            items.sort(key=self._item_sort_key)
        if include_routes:
            for item in items:
                item["route"] = build_route_suggestion(item)
        for item in items:
            item.pop("_query_score", None)
        if limit is not None and limit >= 0:
            items = items[:limit]
        return items

    def add_item(
        self,
        user: str,
        text: str,
        *,
        source: Optional[str] = None,
        device: Optional[str] = None,
        meta: Optional[Dict[str, Any]] = None,
        defer_until_idle: bool = True,
        available_after: Optional[str] = None,
    ) -> Dict[str, Any]:
        now = _now_iso()
        item = build_thought_item(
            text,
            now_iso=now,
            source=source,
            device=device,
            meta=meta,
            defer_until_idle=defer_until_idle,
        )
        if available_after:
            item["available_after"] = available_after
        with self.lock:
            data = self._load(user)
            self._maintenance_locked(data)
            items = data.get("items")
            if not isinstance(items, list):
                items = []
            duplicate = next(
                (
                    existing
                    for existing in items
                    if isinstance(existing, dict)
                    and str(existing.get("status") or "todo").lower() == "todo"
                    and existing.get("fingerprint") == item.get("fingerprint")
                ),
                None,
            )
            if duplicate is not None:
                merged = merge_duplicate_capture(
                    duplicate,
                    text=text,
                    now_iso=now,
                    source=source,
                    device=device,
                    meta=meta,
                    defer_until_idle=defer_until_idle,
                )
                if available_after:
                    merged["available_after"] = available_after
                duplicate.clear()
                duplicate.update(merged)
                item = dict(duplicate)
            else:
                items.append(item)
            data["items"] = items
            data["version"] = 2
            data["updated_at"] = now
            self._write(user, data)
        return dict(item)

    def peek_next_item(self, user: str, *, idle_only: bool = False) -> Optional[Dict[str, Any]]:
        """Return the next eligible thought without claiming it."""
        with self.lock:
            data = self._load(user)
            changed = self._maintenance_locked(data)
            item = self._next_candidate_locked(data, idle_only=idle_only)
            if item is None:
                if changed:
                    data["updated_at"] = _now_iso()
                    self._write(user, data)
                return None
            if changed:
                data["updated_at"] = _now_iso()
                self._write(user, data)
            result = dict(item)
        result["route"] = build_route_suggestion(result)
        return result

    def claim_next_item(self, user: str, *, idle_only: bool = False) -> Optional[Dict[str, Any]]:
        """Claim the next eligible thought so concurrent clients cannot race it."""
        with self.lock:
            data = self._load(user)
            changed = self._maintenance_locked(data)
            claimed = self._next_candidate_locked(data, idle_only=idle_only)
            if claimed is None:
                if changed:
                    data["updated_at"] = _now_iso()
                    self._write(user, data)
                return None
            now = _now_iso()
            expiry = dt.datetime.now(UK_TZ) + dt.timedelta(seconds=self.claim_ttl_sec)
            claimed["execution_state"] = "claimed"
            claimed["claimed_at"] = now
            claimed["claim_expires_at"] = expiry.strftime(UK_DATETIME_FORMAT)
            claimed["updated_at"] = now
            data["updated_at"] = now
            self._write(user, data)
            result = dict(claimed)
        result["route"] = build_route_suggestion(result)
        return result

    def _append_link_locked(self, item: Dict[str, Any], bucket: str, value: str, *, limit: int = 12) -> bool:
        bucket_name = str(bucket or "").strip()
        link_value = str(value or "").strip()
        if not bucket_name or not link_value:
            return False
        links = item.get("links")
        if not isinstance(links, dict):
            links = {}
            item["links"] = links
        existing = links.get(bucket_name)
        values: List[str] = []
        if isinstance(existing, list):
            values = [str(entry).strip() for entry in existing if str(entry or "").strip()]
        elif existing:
            values = [str(existing).strip()]
        if link_value in values:
            return False
        values.append(link_value)
        links[bucket_name] = values[-limit:] if limit > 0 else values
        return True

    def append_link(self, user: str, todo_id: str, bucket: str, value: str) -> Optional[Dict[str, Any]]:
        if not todo_id:
            return None
        with self.lock:
            data = self._load(user)
            changed = self._maintenance_locked(data)
            target = None
            for item in data.get("items", []):
                if not isinstance(item, dict) or item.get("id") != todo_id:
                    continue
                if self._append_link_locked(item, bucket, value):
                    now = _now_iso()
                    item["updated_at"] = now
                    data["updated_at"] = now
                    self._write(user, data)
                elif changed:
                    data["updated_at"] = _now_iso()
                    self._write(user, data)
                target = dict(item)
                break
            return target

    def claim_item(self, user: str, todo_id: str) -> Optional[Dict[str, Any]]:
        if not todo_id:
            return None
        with self.lock:
            data = self._load(user)
            changed = self._maintenance_locked(data)
            target = None
            for item in data.get("items", []):
                if not isinstance(item, dict) or item.get("id") != todo_id:
                    continue
                if not self._is_ready_item(item):
                    break
                now = _now_iso()
                expiry = dt.datetime.now(UK_TZ) + dt.timedelta(seconds=self.claim_ttl_sec)
                item["execution_state"] = "claimed"
                item["claimed_at"] = now
                item["claim_expires_at"] = expiry.strftime(UK_DATETIME_FORMAT)
                item["updated_at"] = now
                data["updated_at"] = now
                self._write(user, data)
                target = dict(item)
                break
            if not target and changed:
                data["updated_at"] = _now_iso()
                self._write(user, data)
            return target

    def begin_processing(self, user: str, todo_id: str) -> Optional[Dict[str, Any]]:
        if not todo_id:
            return None
        with self.lock:
            data = self._load(user)
            changed = self._maintenance_locked(data)
            target = None
            for item in data.get("items", []):
                if not isinstance(item, dict) or item.get("id") != todo_id:
                    continue
                if str(item.get("status") or "todo").lower() != "todo":
                    break
                now = _now_iso()
                expiry = dt.datetime.now(UK_TZ) + dt.timedelta(seconds=self.claim_ttl_sec)
                item["execution_state"] = "processing"
                item["claimed_at"] = item.get("claimed_at") or now
                item["claim_expires_at"] = expiry.strftime(UK_DATETIME_FORMAT)
                item["updated_at"] = now
                data["updated_at"] = now
                self._write(user, data)
                target = dict(item)
                break
            if not target and changed:
                data["updated_at"] = _now_iso()
                self._write(user, data)
            return target

    def release_execution(
        self,
        user: str,
        todo_id: str,
        *,
        execution_state: str = "ready",
        error: Optional[str] = None,
        result: Optional[Dict[str, Any]] = None,
    ) -> Optional[Dict[str, Any]]:
        if not todo_id:
            return None
        next_state = str(execution_state or "ready").strip().lower() or "ready"
        if next_state not in {"ready", "failed", "cancelled"}:
            next_state = "ready"
        with self.lock:
            data = self._load(user)
            changed = self._maintenance_locked(data)
            target = None
            for item in data.get("items", []):
                if not isinstance(item, dict) or item.get("id") != todo_id:
                    continue
                now = _now_iso()
                item["execution_state"] = next_state
                item.pop("claim_expires_at", None)
                item.pop("claimed_at", None)
                item["updated_at"] = now
                if error:
                    item["last_error"] = str(error)
                    item["last_execution_at"] = now
                elif next_state == "ready":
                    item.pop("last_error", None)
                if result is not None:
                    item["last_result"] = result
                    item["last_execution_at"] = now
                data["updated_at"] = now
                self._write(user, data)
                target = dict(item)
                break
            if not target and changed:
                data["updated_at"] = _now_iso()
                self._write(user, data)
            return target

    def complete_execution(
        self,
        user: str,
        todo_id: str,
        result: Optional[Dict[str, Any]] = None,
        *,
        status: str = "done",
        links: Optional[Dict[str, Any]] = None,
    ) -> Optional[Dict[str, Any]]:
        if not todo_id:
            return None
        next_status = str(status or "done").strip().lower() or "done"
        if next_status not in {"todo", "done", "archived"}:
            next_status = "done"
        with self.lock:
            data = self._load(user)
            changed = self._maintenance_locked(data)
            target = None
            for item in data.get("items", []):
                if not isinstance(item, dict) or item.get("id") != todo_id:
                    continue
                now = _now_iso()
                item["status"] = next_status
                item["execution_state"] = "completed"
                item["updated_at"] = now
                item["last_execution_at"] = now
                item.pop("claim_expires_at", None)
                item.pop("claimed_at", None)
                item.pop("last_error", None)
                if result is not None:
                    item["last_result"] = result
                if isinstance(links, dict):
                    for bucket, raw_values in links.items():
                        values = raw_values if isinstance(raw_values, list) else [raw_values]
                        for raw_value in values:
                            self._append_link_locked(item, str(bucket or ""), str(raw_value or ""))
                data["updated_at"] = now
                self._write(user, data)
                target = dict(item)
                break
            if not target and changed:
                data["updated_at"] = _now_iso()
                self._write(user, data)
            return target

    def fail_execution(
        self,
        user: str,
        todo_id: str,
        error: str,
        *,
        result: Optional[Dict[str, Any]] = None,
    ) -> Optional[Dict[str, Any]]:
        if not todo_id:
            return None
        return self.release_execution(
            user,
            todo_id,
            execution_state="failed",
            error=error,
            result=result,
        )

    def update_item(self, user: str, todo_id: str, updates: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        if not todo_id:
            return None
        allowed_keys = {
            "text",
            "status",
            "notes",
            "tags",
            "priority",
            "defer_until_idle",
            "available_after",
            "execution_state",
        }
        cleaned = {k: v for k, v in updates.items() if k in allowed_keys}
        if not cleaned:
            return None
        with self.lock:
            data = self._load(user)
            self._maintenance_locked(data)
            items = data.get("items")
            if not isinstance(items, list):
                items = []
            target = None
            for item in items:
                if not isinstance(item, dict):
                    continue
                if item.get("id") == todo_id:
                    for key, value in cleaned.items():
                        if key == "text":
                            value = normalize_thought_text(str(value or ""))
                        if key == "tags":
                            value = self._coerce_text_list(value)
                        item[key] = value
                    if "text" in cleaned or "defer_until_idle" in cleaned:
                        self._refresh_text_fields(item)
                    status = str(item.get("status") or "todo").lower()
                    if status != "todo":
                        item["execution_state"] = "completed"
                        item.pop("claim_expires_at", None)
                        item.pop("claimed_at", None)
                    elif str(item.get("execution_state") or "").lower() in {"", "completed", "failed", "cancelled"}:
                        item["execution_state"] = "ready"
                    item["updated_at"] = _now_iso()
                    target = dict(item)
                    break
            if target:
                data["items"] = items
                data["version"] = 2
                data["updated_at"] = _now_iso()
                self._write(user, data)
            return target

    def delete_item(self, user: str, todo_id: str) -> bool:
        if not todo_id:
            return False
        with self.lock:
            data = self._load(user)
            items = data.get("items")
            if not isinstance(items, list):
                items = []
            next_items = [item for item in items if isinstance(item, dict) and item.get("id") != todo_id]
            if len(next_items) == len(items):
                return False
            data["items"] = next_items
            data["updated_at"] = _now_iso()
            self._write(user, data)
        return True


class ScheduleStore:
    """Persistent schedule queue for deferred TODO execution."""

    def __init__(self, root: str, *, claim_ttl_sec: int = TODO_SCHEDULE_CLAIM_TTL_SEC):
        self.root = root
        self.path = os.path.join(root, "schedules.json")
        self.lock = threading.RLock()
        self.claim_ttl_sec = max(30, int(claim_ttl_sec))
        ensure_dir_permissions(root, mode=0o700)

    def _load(self) -> Dict[str, Any]:
        if not os.path.exists(self.path):
            return {"version": 1, "items": []}
        try:
            with open(self.path, "r", encoding="utf-8") as handle:
                data = json.load(handle)
            if not isinstance(data, dict):
                return {"version": 1, "items": []}
            if not isinstance(data.get("items"), list):
                data["items"] = []
            if "version" not in data:
                data["version"] = 1
            return data
        except Exception:
            return {"version": 1, "items": []}

    def _write(self, data: Dict[str, Any]) -> None:
        _write_json_atomic(self.path, data)

    def _hydrate_item(self, item: Dict[str, Any]) -> bool:
        changed = False
        if not isinstance(item, dict):
            return False
        if not item.get("id"):
            item["id"] = uuid.uuid4().hex
            changed = True
        for key in ("user", "todo_id", "status"):
            if key not in item:
                item[key] = ""
                changed = True
        if not item.get("created_at"):
            item["created_at"] = _now_iso()
            changed = True
        if not item.get("updated_at"):
            item["updated_at"] = item.get("created_at") or _now_iso()
            changed = True
        run_at = _normalise_timestamp(item.get("run_at")) if item.get("run_at") else None
        if run_at and run_at != item.get("run_at"):
            item["run_at"] = run_at
            changed = True
        if item.get("route_override") is not None and not isinstance(item.get("route_override"), dict):
            item.pop("route_override", None)
            changed = True
        return changed

    def _maintenance_locked(self, data: Dict[str, Any]) -> bool:
        changed = False
        items = data.get("items")
        if not isinstance(items, list):
            data["items"] = []
            return True
        now_ts = time.time()
        now_iso = _now_iso()
        for item in items:
            if not isinstance(item, dict):
                continue
            changed = self._hydrate_item(item) or changed
            status = str(item.get("status") or "scheduled").strip().lower() or "scheduled"
            claim_expires_at = item.get("claim_expires_at")
            if status == "dispatching" and _timestamp_sort_key(claim_expires_at) and _timestamp_sort_key(claim_expires_at) <= now_ts:
                item["status"] = "scheduled"
                item["updated_at"] = now_iso
                item.pop("claim_expires_at", None)
                changed = True
        data["version"] = 1
        return changed

    def _item_sort_key(self, item: Dict[str, Any]) -> Tuple[int, float, float]:
        status = str(item.get("status") or "scheduled").strip().lower() or "scheduled"
        status_rank = {
            "scheduled": 0,
            "dispatching": 1,
            "queued": 2,
            "running": 3,
            "failed": 4,
            "completed": 5,
            "cancelled": 6,
        }.get(status, 9)
        run_at = _timestamp_sort_key(item.get("run_at") or item.get("created_at"))
        updated = _timestamp_sort_key(item.get("updated_at") or item.get("created_at"))
        return (status_rank, run_at or 0.0, -updated)

    def create(
        self,
        *,
        user: str,
        todo_id: str,
        run_at: str,
        route_override: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        schedule = {
            "id": uuid.uuid4().hex,
            "user": str(user or "").strip(),
            "todo_id": str(todo_id or "").strip(),
            "run_at": _normalise_timestamp(run_at) or run_at,
            "created_at": _now_iso(),
            "updated_at": _now_iso(),
            "status": "scheduled",
        }
        if isinstance(route_override, dict) and route_override:
            schedule["route_override"] = dict(route_override)
        with self.lock:
            data = self._load()
            self._maintenance_locked(data)
            items = data.get("items")
            if not isinstance(items, list):
                items = []
            items.append(schedule)
            data["items"] = items
            data["updated_at"] = _now_iso()
            self._write(data)
        return dict(schedule)

    def list_items(
        self,
        *,
        user: Optional[str] = None,
        todo_id: Optional[str] = None,
        statuses: Optional[List[str]] = None,
        limit: Optional[int] = None,
    ) -> List[Dict[str, Any]]:
        with self.lock:
            data = self._load()
            changed = self._maintenance_locked(data)
            items = [dict(item) for item in data.get("items", []) if isinstance(item, dict)]
            if changed:
                data["updated_at"] = _now_iso()
                self._write(data)
        if user:
            items = [item for item in items if item.get("user") == user]
        if todo_id:
            items = [item for item in items if item.get("todo_id") == todo_id]
        if statuses:
            wanted = {str(status or "").strip().lower() for status in statuses if str(status or "").strip()}
            items = [item for item in items if str(item.get("status") or "").strip().lower() in wanted]
        items.sort(key=self._item_sort_key)
        if limit is not None and limit >= 0:
            items = items[:limit]
        return items

    def list_due(self, *, limit: int = 10) -> List[Dict[str, Any]]:
        now_ts = time.time()
        due = [
            item
            for item in self.list_items(statuses=["scheduled"])
            if _timestamp_sort_key(item.get("run_at") or item.get("created_at")) <= now_ts
        ]
        return due[: max(1, int(limit))]

    def list_active(self, *, limit: int = 50) -> List[Dict[str, Any]]:
        return self.list_items(statuses=["dispatching", "queued", "running"], limit=limit)

    def get_item(self, schedule_id: str, *, user: Optional[str] = None) -> Optional[Dict[str, Any]]:
        if not schedule_id:
            return None
        items = self.list_items(user=user)
        for item in items:
            if item.get("id") == schedule_id:
                return item
        return None

    def _update_item_locked(self, data: Dict[str, Any], schedule_id: str, mutator) -> Optional[Dict[str, Any]]:
        items = data.get("items")
        if not isinstance(items, list):
            return None
        for item in items:
            if not isinstance(item, dict) or item.get("id") != schedule_id:
                continue
            mutator(item)
            item["updated_at"] = _now_iso()
            data["updated_at"] = item["updated_at"]
            self._write(data)
            return dict(item)
        return None

    def mark_dispatching(self, schedule_id: str) -> Optional[Dict[str, Any]]:
        with self.lock:
            data = self._load()
            self._maintenance_locked(data)

            def _mutate(item: Dict[str, Any]) -> None:
                item["status"] = "dispatching"
                item["claim_expires_at"] = (
                    dt.datetime.now(UK_TZ) + dt.timedelta(seconds=self.claim_ttl_sec)
                ).strftime(UK_DATETIME_FORMAT)

            return self._update_item_locked(data, schedule_id, _mutate)

    def mark_queued(self, schedule_id: str, subtask_id: str) -> Optional[Dict[str, Any]]:
        with self.lock:
            data = self._load()
            self._maintenance_locked(data)

            def _mutate(item: Dict[str, Any]) -> None:
                now = _now_iso()
                item["status"] = "queued"
                item["subtask_id"] = subtask_id
                item["queued_at"] = item.get("queued_at") or now
                item["claim_expires_at"] = (
                    dt.datetime.now(UK_TZ) + dt.timedelta(seconds=self.claim_ttl_sec)
                ).strftime(UK_DATETIME_FORMAT)

            return self._update_item_locked(data, schedule_id, _mutate)

    def mark_running(self, schedule_id: str, subtask_id: Optional[str] = None) -> Optional[Dict[str, Any]]:
        with self.lock:
            data = self._load()
            self._maintenance_locked(data)

            def _mutate(item: Dict[str, Any]) -> None:
                now = _now_iso()
                item["status"] = "running"
                item["started_at"] = item.get("started_at") or now
                if subtask_id:
                    item["subtask_id"] = subtask_id
                item["claim_expires_at"] = (
                    dt.datetime.now(UK_TZ) + dt.timedelta(seconds=self.claim_ttl_sec)
                ).strftime(UK_DATETIME_FORMAT)

            return self._update_item_locked(data, schedule_id, _mutate)

    def complete(self, schedule_id: str, result: Optional[Dict[str, Any]] = None) -> Optional[Dict[str, Any]]:
        with self.lock:
            data = self._load()
            self._maintenance_locked(data)

            def _mutate(item: Dict[str, Any]) -> None:
                item["status"] = "completed"
                item["finished_at"] = _now_iso()
                item.pop("claim_expires_at", None)
                item.pop("error", None)
                if result is not None:
                    item["result"] = result

            return self._update_item_locked(data, schedule_id, _mutate)

    def fail(self, schedule_id: str, error: str, *, result: Optional[Dict[str, Any]] = None) -> Optional[Dict[str, Any]]:
        with self.lock:
            data = self._load()
            self._maintenance_locked(data)

            def _mutate(item: Dict[str, Any]) -> None:
                item["status"] = "failed"
                item["finished_at"] = _now_iso()
                item["error"] = str(error or "schedule_failed")
                item.pop("claim_expires_at", None)
                if result is not None:
                    item["result"] = result

            return self._update_item_locked(data, schedule_id, _mutate)

    def cancel(self, schedule_id: str, *, reason: Optional[str] = None) -> Optional[Dict[str, Any]]:
        with self.lock:
            data = self._load()
            self._maintenance_locked(data)

            def _mutate(item: Dict[str, Any]) -> None:
                item["status"] = "cancelled"
                item["finished_at"] = _now_iso()
                item.pop("claim_expires_at", None)
                if reason:
                    item["error"] = str(reason)

            return self._update_item_locked(data, schedule_id, _mutate)


class PostgresTodoStore(TodoStore):
    """Todo store that persists Thought Inbox metadata in the central Postgres store."""

    def __init__(
        self,
        root: str,
        document_store: PostgresTodoDocumentStore,
        *,
        claim_ttl_sec: int = TODO_CLAIM_TTL_SEC,
        retention_days: int = TODO_RETENTION_DAYS,
    ):
        super().__init__(root, claim_ttl_sec=claim_ttl_sec, retention_days=retention_days)
        self.document_store = document_store

    def _load(self, user: str) -> Dict[str, Any]:
        cleaned = str(user or "").strip()
        if cleaned:
            try:
                data = self.document_store.load_user(cleaned)
                if isinstance(data, dict):
                    return data
            except Exception as exc:
                logger.debug("todo metadata load fell back to disk for %s: %s", cleaned, exc)
        data = super()._load(user)
        if cleaned and (os.path.exists(self._path(user)) or data.get("items")):
            try:
                self.document_store.write_user(cleaned, data)
            except Exception as exc:
                logger.debug("todo metadata backfill skipped for %s: %s", cleaned, exc)
        return data

    def _write(self, user: str, data: Dict[str, Any]) -> None:
        cleaned = str(user or "").strip()
        if cleaned:
            try:
                self.document_store.write_user(cleaned, data)
                return
            except Exception as exc:
                logger.warning("todo metadata write fell back to disk for %s: %s", cleaned, exc)
        super()._write(user, data)


class PostgresScheduleStore(ScheduleStore):
    """Schedule store that persists deferred TODO metadata in central Postgres."""

    def __init__(
        self,
        root: str,
        document_store: PostgresScheduleDocumentStore,
        *,
        claim_ttl_sec: int = TODO_SCHEDULE_CLAIM_TTL_SEC,
    ):
        super().__init__(root, claim_ttl_sec=claim_ttl_sec)
        self.document_store = document_store

    def _load(self) -> Dict[str, Any]:
        try:
            data = self.document_store.load()
            if isinstance(data, dict):
                return data
        except Exception as exc:
            logger.debug("schedule metadata load fell back to disk: %s", exc)
        data = super()._load()
        if os.path.exists(self.path) or data.get("items"):
            try:
                self.document_store.write(data)
            except Exception as exc:
                logger.debug("schedule metadata backfill skipped: %s", exc)
        return data

    def _write(self, data: Dict[str, Any]) -> None:
        try:
            self.document_store.write(data)
            return
        except Exception as exc:
            logger.warning("schedule metadata write fell back to disk: %s", exc)
        super()._write(data)


class SessionHistoryStore:
    """Persistent append-only history of conversational room events."""

    def __init__(self, root: str, max_events: int = SESSION_HISTORY_MAX):
        self.root = root
        self.max_events = max_events
        self.lock = threading.RLock()
        ensure_dir_permissions(root, mode=0o700)

    def _safe_room(self, room_id: str) -> str:
        room_id = str(room_id or "").strip()
        if not room_id:
            return ""
        return re.sub(r"[^A-Za-z0-9_.-]+", "_", room_id)

    def _room_path(self, room_id: str) -> str:
        safe = self._safe_room(room_id) or "room"
        return os.path.join(self.root, f"{safe}.json")

    def load(self, room_id: str) -> Optional[Dict[str, Any]]:
        path = self._room_path(room_id)
        if not os.path.exists(path):
            return None
        try:
            with open(path, "r", encoding="utf-8") as handle:
                data = json.load(handle)
            return data if isinstance(data, dict) else None
        except Exception:
            return None

    def write(self, room_id: str, data: Dict[str, Any]) -> None:
        path = self._room_path(room_id)
        _write_json_atomic(path, data)

    def persist_snapshot(self, room_id: str, snapshot: Dict[str, Any]) -> None:
        if not room_id or not isinstance(snapshot, dict):
            return
        with self.lock:
            data = self.load(room_id) or {
                "room_id": room_id,
                "created_at": snapshot.get("created_at") or _now_iso(),
                "updated_at": snapshot.get("updated_at") or _now_iso(),
                "events": [],
            }
            if snapshot.get("job_id") and not data.get("job_id"):
                data["job_id"] = snapshot.get("job_id")
            if snapshot.get("project_id") and not data.get("project_id"):
                data["project_id"] = snapshot.get("project_id")
            if snapshot.get("created_by") and not data.get("created_by"):
                data["created_by"] = snapshot.get("created_by")
            data["snapshot"] = dict(snapshot)
            data["updated_at"] = snapshot.get("updated_at") or _now_iso()
            self.write(room_id, data)

    def append_event(self, room_id: str, event: Dict[str, Any]) -> None:
        if not room_id:
            return
        with self.lock:
            data = self.load(room_id) or {
                "room_id": room_id,
                "created_at": _now_iso(),
                "updated_at": _now_iso(),
                "events": [],
            }
            if event.get("job_id"):
                data["job_id"] = event.get("job_id")
            if event.get("project_id"):
                data["project_id"] = event.get("project_id")
            if event.get("user") and not data.get("created_by") and event.get("type") == "created":
                data["created_by"] = event.get("user")
            events = data.get("events")
            if not isinstance(events, list):
                events = []
            events.append(event)
            if self.max_events and len(events) > self.max_events:
                events = events[-self.max_events :]
            data["events"] = events
            data["updated_at"] = event.get("ts") or _now_iso()
            self.write(room_id, data)

    def list_rooms(self, limit: int = 50, tail: int = 5) -> List[Dict[str, Any]]:
        rooms: List[Dict[str, Any]] = []
        try:
            entries = [name for name in os.listdir(self.root) if name.endswith(".json")]
        except Exception:
            entries = []
        for filename in entries:
            path = os.path.join(self.root, filename)
            try:
                with open(path, "r", encoding="utf-8") as handle:
                    data = json.load(handle)
            except Exception:
                continue
            if not isinstance(data, dict):
                continue
            events = data.get("events") if isinstance(data.get("events"), list) else []
            last_event = events[-1] if events else None
            tail_events = events[-tail:] if tail and events else []
            rooms.append(
                {
                    "room_id": data.get("room_id") or os.path.splitext(filename)[0],
                    "job_id": data.get("job_id"),
                    "project_id": data.get("project_id"),
                    "created_by": data.get("created_by"),
                    "created_at": data.get("created_at"),
                    "updated_at": data.get("updated_at"),
                    "events_count": len(events),
                    "last_event": last_event,
                    "events_tail": tail_events,
                }
            )
        rooms.sort(key=lambda r: _timestamp_sort_key(r.get("updated_at")), reverse=True)
        if limit and len(rooms) > limit:
            rooms = rooms[:limit]
        return rooms


class PostgresSessionHistoryStore(SessionHistoryStore):
    """Session history store backed by the central Postgres room registry."""

    def __init__(
        self,
        root: str,
        room_store: PostgresSessionRoomStore,
        max_events: int = SESSION_HISTORY_MAX,
    ):
        super().__init__(root, max_events=max_events)
        self.room_store = room_store

    def load(self, room_id: str) -> Optional[Dict[str, Any]]:
        cleaned = str(room_id or "").strip()
        if cleaned:
            try:
                data = self.room_store.load_room(cleaned)
                if isinstance(data, dict):
                    return data
            except Exception as exc:
                logger.debug("session history load fell back to disk for %s: %s", cleaned, exc)
        data = super().load(room_id)
        if cleaned and data:
            try:
                self.room_store.write_room(cleaned, data)
            except Exception as exc:
                logger.debug("session history backfill skipped for %s: %s", cleaned, exc)
        return data

    def write(self, room_id: str, data: Dict[str, Any]) -> None:
        cleaned = str(room_id or "").strip()
        if cleaned:
            try:
                self.room_store.write_room(cleaned, data)
                return
            except Exception as exc:
                logger.warning("session history write fell back to disk for %s: %s", cleaned, exc)
        super().write(room_id, data)

    def persist_snapshot(self, room_id: str, snapshot: Dict[str, Any]) -> None:
        cleaned = str(room_id or "").strip()
        if cleaned and isinstance(snapshot, dict):
            try:
                self.room_store.persist_snapshot(cleaned, snapshot)
                return
            except Exception as exc:
                logger.warning("session snapshot write fell back to disk for %s: %s", cleaned, exc)
        super().persist_snapshot(room_id, snapshot)

    def list_rooms(self, limit: int = 50, tail: int = 5) -> List[Dict[str, Any]]:
        try:
            rooms = self.room_store.list_rooms(limit=limit, tail=tail)
            if rooms:
                return rooms
        except Exception as exc:
            logger.debug("session room listing fell back to disk: %s", exc)
        rooms = super().list_rooms(limit=limit, tail=tail)
        for room in rooms:
            room_id = str(room.get("room_id") or "").strip()
            if not room_id:
                continue
            data = super().load(room_id)
            if not data:
                continue
            try:
                self.room_store.write_room(room_id, data)
            except Exception as exc:
                logger.debug("session history room backfill skipped for %s: %s", room_id, exc)
        return rooms


class WorkspaceSession:
    """In-memory workspace collaboration session model."""

    def __init__(self, room_id: str, job_id: str, project_id: Optional[str], created_by: str):
        self.room_id = room_id
        self.session_id = room_id
        self.job_id = job_id
        self.project_id = project_id
        self.created_by = created_by
        self.created_at = _now_iso()
        self.updated_at = self.created_at
        self.participants: Dict[str, Dict[str, Any]] = {}
        self.listeners: List[queue.Queue] = []
        self.lock = threading.RLock()

    def add_listener(self) -> queue.Queue:
        q: queue.Queue = queue.Queue()
        with self.lock:
            self.listeners.append(q)
        return q

    def remove_listener(self, q: queue.Queue) -> None:
        with self.lock:
            if q in self.listeners:
                self.listeners.remove(q)

    def _notify(self, event: str, payload: Dict[str, Any]) -> None:
        with self.lock:
            listeners = list(self.listeners)
        for listener in listeners:
            try:
                listener.put_nowait({"event": event, "payload": payload})
            except queue.Full:
                continue

    def _record_event(self, event_type: str, user: Optional[str] = None, detail: Optional[Dict[str, Any]] = None) -> None:
        payload = {
            "ts": _now_iso(),
            "type": event_type,
            "user": user,
            "job_id": self.job_id,
            "project_id": self.project_id,
        }
        if detail:
            payload["detail"] = detail
        session_history.append_event(self.room_id, payload)

    def _persist_snapshot(self, snapshot: Optional[Dict[str, Any]] = None) -> None:
        try:
            session_history.persist_snapshot(self.room_id, snapshot or self.snapshot())
        except Exception:
            logger.debug("session snapshot persistence skipped for room %s", self.room_id, exc_info=True)

    def join(self, user: str, role: Optional[str] = None) -> None:
        now = _now_iso()
        with self.lock:
            entry = self.participants.get(user) or {"user": user, "joined_at": now}
            entry["last_seen"] = now
            if role:
                entry["role"] = role
            self.participants[user] = entry
            self.updated_at = now
            snapshot = self.snapshot()
        self._persist_snapshot(snapshot)
        self._notify("presence", snapshot)
        self._record_event("join", user=user, detail={"role": role})

    def leave(self, user: str) -> None:
        with self.lock:
            if user in self.participants:
                self.participants.pop(user, None)
                self.updated_at = _now_iso()
                snapshot = self.snapshot()
            else:
                return
        self._persist_snapshot(snapshot)
        self._notify("presence", snapshot)
        self._record_event("leave", user=user)

    def heartbeat(self, user: str) -> None:
        now = _now_iso()
        with self.lock:
            entry = self.participants.get(user)
            if not entry:
                return
            entry["last_seen"] = now
            self.participants[user] = entry
            self.updated_at = now
            snapshot = self.snapshot()
        self._persist_snapshot(snapshot)

    def snapshot(self) -> Dict[str, Any]:
        with self.lock:
            participants = sorted(self.participants.values(), key=lambda p: p.get("user") or "")
            return {
                "session_id": self.session_id,
                "room_id": self.room_id,
                "job_id": self.job_id,
                "project_id": self.project_id,
                "created_by": self.created_by,
                "created_at": self.created_at,
                "updated_at": self.updated_at,
                "participants": participants,
            }


class SessionStore:
    """TTL-based in-memory manager for active workspace sessions."""

    def __init__(self, ttl_sec: int = SESSION_TTL_SEC):
        self.sessions: Dict[str, WorkspaceSession] = {}
        self.lock = threading.RLock()
        self.ttl_sec = ttl_sec

    def _now_ts(self) -> float:
        return time.time()

    def _parse_ts(self, value: Optional[str]) -> float:
        if not value:
            return 0.0
        parsed = _parse_timestamp(value)
        if not parsed:
            return 0.0
        return parsed.timestamp()

    def prune(self) -> None:
        cutoff = self._now_ts() - float(self.ttl_sec or 0)
        if cutoff <= 0:
            return
        with self.lock:
            for session_id, session in list(self.sessions.items()):
                updated = self._parse_ts(session.updated_at)
                if updated and updated < cutoff and not session.participants:
                    self.sessions.pop(session_id, None)

    def get(self, session_id: str) -> Optional[WorkspaceSession]:
        if not session_id:
            return None
        with self.lock:
            session = self.sessions.get(session_id)
        if session:
            return session
        session = self._load_from_history(session_id)
        if session:
            with self.lock:
                self.sessions[session_id] = session
        return session

    def _load_from_history(self, room_id: str) -> Optional[WorkspaceSession]:
        data = session_history.load(room_id)
        if not data:
            return None
        snapshot = data.get("snapshot") if isinstance(data.get("snapshot"), dict) else {}
        job_id = str(data.get("job_id") or snapshot.get("job_id") or "").strip()
        project_id = data.get("project_id") or snapshot.get("project_id")
        created_by = data.get("created_by") or snapshot.get("created_by") or "system"
        session = WorkspaceSession(room_id, job_id, project_id, created_by)
        session.created_at = snapshot.get("created_at") or data.get("created_at") or session.created_at
        session.updated_at = snapshot.get("updated_at") or data.get("updated_at") or session.updated_at
        participants = snapshot.get("participants") if isinstance(snapshot.get("participants"), list) else []
        cutoff = self._now_ts() - float(self.ttl_sec or 0)
        hydrated: Dict[str, Dict[str, Any]] = {}
        for entry in participants:
            if not isinstance(entry, dict):
                continue
            username = str(entry.get("user") or "").strip()
            if not username:
                continue
            last_seen = self._parse_ts(entry.get("last_seen") or entry.get("joined_at"))
            if cutoff > 0 and last_seen and last_seen < cutoff:
                continue
            hydrated[username] = dict(entry)
        session.participants = hydrated
        return session

    def get_or_create(
        self,
        job_id: str,
        project_id: Optional[str],
        user: str,
        role: Optional[str],
        room_id: Optional[str] = None,
    ) -> WorkspaceSession:
        self.prune()
        room_id = (room_id or "").strip() or None
        with self.lock:
            if room_id:
                existing = self.sessions.get(room_id)
                if not existing:
                    existing = self._load_from_history(room_id)
                    if existing:
                        self.sessions[room_id] = existing
                if existing:
                    if job_id and job_id != existing.job_id:
                        existing.job_id = job_id
                        existing.project_id = project_id
                        existing.updated_at = _now_iso()
                        existing._record_event("job_bind", user=user, detail={"job_id": job_id})
                    existing.join(user, role)
                    return existing
            for session in self.sessions.values():
                if session.job_id == job_id and not room_id:
                    session.join(user, role)
                    return session
            session_id = room_id or uuid.uuid4().hex
            session = WorkspaceSession(session_id, job_id, project_id, user)
            self.sessions[session_id] = session
        session._record_event("created", user=user, detail={"job_id": job_id})
        session.join(user, role)
        return session

    def join(self, session_id: str, user: str, role: Optional[str]) -> Optional[WorkspaceSession]:
        session = self.get(session_id)
        if not session:
            session = self._load_from_history(session_id)
            if not session:
                return None
            with self.lock:
                self.sessions[session_id] = session
        session.join(user, role)
        return session

    def leave(self, session_id: str, user: str) -> None:
        session = self.get(session_id)
        if not session:
            return
        session.leave(user)

class TokenLedger:
    """Ledger for token balances, reservations, usage, and cashout events."""

    def __init__(self, root: str):
        self.root = root
        self.lock = threading.RLock()
        os.makedirs(root, exist_ok=True)

    def _safe_user(self, user: str) -> str:
        return re.sub(r"[^A-Za-z0-9_\\-]+", "_", user or "unknown")

    def _ledger_path(self, user: str) -> str:
        return os.path.join(self.root, f"{self._safe_user(user)}.jsonl")

    def _summary_path(self, user: str) -> str:
        return os.path.join(self.root, f"{self._safe_user(user)}.summary.json")

    def _default_summary(self) -> Dict[str, Any]:
        return {
            "version": 1,
            "balance": 0,
            "paid_balance": 0,
            "free_balance": 0,
            "last_topup_tokens": 0,
            "last_topup_at": None,
            "updated_at": None,
            "spent_total": 0,
            "cashout_total": 0,
            "shortfall_total": 0,
            "free_grant_total": 0,
        }

    def _load_summary(self, user: str) -> Dict[str, Any]:
        path = self._summary_path(user)
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as handle:
                    data = json.load(handle)
                if isinstance(data, dict):
                    summary = self._default_summary()
                    summary.update(data)
                    if "paid_balance" not in data:
                        summary["paid_balance"] = int(summary.get("balance") or 0)
                        summary["free_balance"] = int(summary.get("free_balance") or 0)
                    return summary
            except Exception:
                pass
        return self._rebuild_summary(user)

    def _write_summary(self, user: str, summary: Dict[str, Any]) -> None:
        _write_json_atomic(self._summary_path(user), summary)

    def _rebuild_summary(self, user: str) -> Dict[str, Any]:
        summary = self._default_summary()
        path = self._ledger_path(user)
        if os.path.exists(path):
            try:
                with open(path, "r", encoding="utf-8") as handle:
                    for line in handle:
                        line = line.strip()
                        if not line:
                            continue
                        try:
                            entry = json.loads(line)
                        except Exception:
                            continue
                        meta = entry.get("meta") or {}
                        delta = int(entry.get("delta") or 0)
                        paid_after = meta.get("paid_after")
                        free_after = meta.get("free_after")
                        if paid_after is not None or free_after is not None:
                            if paid_after is not None:
                                summary["paid_balance"] = int(paid_after or 0)
                            if free_after is not None:
                                summary["free_balance"] = int(free_after or 0)
                        else:
                            if entry.get("type") == "grant":
                                summary["free_balance"] += max(0, delta)
                            else:
                                summary["paid_balance"] = max(0, summary["paid_balance"] + delta)
                        summary["balance"] = max(0, summary["paid_balance"] + summary["free_balance"])
                        etype = entry.get("type")
                        if etype == "topup":
                            summary["last_topup_tokens"] = int(entry.get("meta", {}).get("tokens") or abs(delta) or 0)
                            summary["last_topup_at"] = entry.get("ts")
                        if etype == "sync":
                            capacity = entry.get("meta", {}).get("capacity")
                            if capacity is not None:
                                summary["last_topup_tokens"] = int(capacity or 0)
                                summary["last_topup_at"] = entry.get("ts")
                        if etype == "debit":
                            used = meta.get("used_total")
                            if used is None:
                                used = abs(delta)
                            summary["spent_total"] += int(used or 0)
                            shortfall = int(entry.get("meta", {}).get("shortfall") or 0)
                            summary["shortfall_total"] += shortfall
                        if etype == "cashout":
                            summary["cashout_total"] += abs(delta)
                        if etype == "grant":
                            summary["free_grant_total"] += abs(delta)
                        summary["updated_at"] = entry.get("ts")
            except Exception:
                pass
        self._write_summary(user, summary)
        return summary

    def get_summary(self, user: str) -> Dict[str, Any]:
        with self.lock:
            return self._load_summary(user)

    def record(
        self,
        user: str,
        entry_type: str,
        delta: int,
        meta: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        meta = meta or {}
        with self.lock:
            summary = self._load_summary(user)
            paid_balance = int(summary.get("paid_balance") or summary.get("balance") or 0)
            free_balance = int(summary.get("free_balance") or 0)
            balance = paid_balance + free_balance
            requested_delta = int(delta or 0)
            new_paid = paid_balance
            new_free = free_balance
            shortfall = 0
            if entry_type == "topup":
                if requested_delta > 0:
                    new_paid += requested_delta
                else:
                    requested_delta = 0
            elif entry_type == "refund":
                if requested_delta > 0:
                    new_paid += requested_delta
                else:
                    requested_delta = 0
            elif entry_type == "grant":
                if requested_delta > 0:
                    new_free += requested_delta
                else:
                    requested_delta = 0
            elif entry_type == "cashout":
                if requested_delta >= 0:
                    requested_delta = -abs(requested_delta)
                desired = abs(requested_delta)
                paid_used = min(new_paid, desired)
                new_paid -= paid_used
                shortfall = desired - paid_used
                if shortfall:
                    meta["shortfall"] = shortfall
                requested_delta = -(paid_used)
                meta["paid_used"] = paid_used
                meta["free_used"] = 0
                meta["used_total"] = paid_used
            elif entry_type == "debit":
                if requested_delta >= 0:
                    requested_delta = -abs(requested_delta or 0)
                desired = abs(requested_delta)
                free_used = min(new_free, desired)
                new_free -= free_used
                remaining = desired - free_used
                paid_used = min(new_paid, remaining)
                new_paid -= paid_used
                shortfall = remaining - paid_used
                if shortfall:
                    meta["shortfall"] = shortfall
                meta["free_used"] = free_used
                meta["paid_used"] = paid_used
                meta["used_total"] = free_used + paid_used
                requested_delta = -(free_used + paid_used)
            elif entry_type in {"reserve", "release"}:
                requested_delta = 0
            elif entry_type == "sync":
                target_paid = meta.get("target_paid_balance")
                target_free = meta.get("target_free_balance")
                target_balance = meta.get("target_balance")
                if target_paid is not None or target_free is not None:
                    if target_paid is not None:
                        new_paid = max(0, int(target_paid or 0))
                    if target_free is not None:
                        new_free = max(0, int(target_free or 0))
                else:
                    if target_balance is None:
                        target_balance = balance + requested_delta
                    try:
                        target_balance = int(float(target_balance))
                    except Exception:
                        target_balance = balance
                    target_balance = max(0, target_balance)
                    if target_balance >= new_free:
                        new_paid = target_balance - new_free
                    else:
                        new_free = target_balance
                        new_paid = 0
                requested_delta = (new_paid + new_free) - balance
            if requested_delta == 0 and entry_type not in {"reserve", "release", "sync"}:
                entry_type = "adjust"
            new_balance = max(0, new_paid + new_free)
            meta["paid_after"] = new_paid
            meta["free_after"] = new_free
            entry = {
                "ts": _now_iso(),
                "type": entry_type,
                "user": user,
                "delta": requested_delta,
                "balance_after": new_balance,
                "meta": meta,
            }
            path = self._ledger_path(user)
            os.makedirs(os.path.dirname(path), exist_ok=True)
            with open(path, "a", encoding="utf-8") as handle:
                handle.write(json.dumps(entry) + "\n")

            summary["balance"] = new_balance
            summary["paid_balance"] = new_paid
            summary["free_balance"] = new_free
            if entry_type == "topup":
                summary["last_topup_tokens"] = int(meta.get("tokens") or abs(requested_delta) or 0)
                summary["last_topup_at"] = entry["ts"]
            if entry_type == "sync" and meta.get("capacity") is not None:
                summary["last_topup_tokens"] = int(meta.get("capacity") or 0)
                summary["last_topup_at"] = meta.get("capacity_ts") or entry["ts"]
            if entry_type == "debit":
                used_total = int(meta.get("used_total") or abs(requested_delta) or 0)
                summary["spent_total"] = int(summary.get("spent_total") or 0) + used_total
                summary["shortfall_total"] = int(summary.get("shortfall_total") or 0) + int(meta.get("shortfall") or 0)
            if entry_type == "cashout":
                summary["cashout_total"] = int(summary.get("cashout_total") or 0) + abs(requested_delta)
            if entry_type == "grant":
                summary["free_grant_total"] = int(summary.get("free_grant_total") or 0) + abs(requested_delta)
            summary["updated_at"] = entry["ts"]
            self._write_summary(user, summary)
            entry["shortfall"] = shortfall
            return entry

    def list_entries(self, user: str, limit: int = 50) -> List[Dict[str, Any]]:
        path = self._ledger_path(user)
        if not os.path.exists(path):
            return []
        try:
            lines = deque(maxlen=limit)
            with open(path, "r", encoding="utf-8") as handle:
                for line in handle:
                    if line.strip():
                        lines.append(line)
            entries = []
            for line in reversed(lines):
                try:
                    entries.append(json.loads(line))
                except Exception:
                    continue
            return entries
        except Exception:
            return []


def _default_token_usage_bucket() -> Dict[str, Any]:
    return {
        "prompt": 0,
        "completion": 0,
        "total": 0,
        "chargeable_total": 0,
        "non_chargeable_total": 0,
        "cached": None,
        "cost": None,
        "events": 0,
        "last_event_at": None,
    }


def _default_token_usage_metrics() -> Dict[str, Any]:
    payload = _default_token_usage_bucket()
    payload["by_category"] = {}
    payload["by_credential_source"] = {}
    payload["by_model"] = {}
    return payload


def _normalize_token_cost(value: Any) -> Optional[Dict[str, Any]]:
    if not isinstance(value, dict):
        return None
    amount = value.get("amount")
    currency = str(value.get("currency") or "").strip().upper() or None
    try:
        amount_value = float(amount)
    except Exception:
        amount_value = None
    if amount_value is None:
        return {"amount": None, "currency": currency} if currency else None
    return {"amount": round(amount_value, 8), "currency": currency}


def _normalize_token_usage_bucket(value: Any) -> Dict[str, Any]:
    bucket = _default_token_usage_bucket()
    if not isinstance(value, dict):
        return bucket
    bucket["prompt"] = _safe_int(value.get("prompt"), 0)
    bucket["completion"] = _safe_int(value.get("completion"), 0)
    bucket["total"] = _safe_int(value.get("total"), 0)
    bucket["chargeable_total"] = _safe_int(value.get("chargeable_total"), 0)
    bucket["non_chargeable_total"] = _safe_int(value.get("non_chargeable_total"), 0)
    cached = value.get("cached")
    bucket["cached"] = _safe_int(cached) if cached not in (None, "") else None
    bucket["cost"] = _normalize_token_cost(value.get("cost"))
    bucket["events"] = _safe_int(value.get("events"), 0)
    bucket["last_event_at"] = str(value.get("last_event_at") or "").strip() or None
    return bucket


def _normalize_token_usage_metrics(value: Any) -> Dict[str, Any]:
    metrics = _normalize_token_usage_bucket(value)
    metrics["by_category"] = {}
    metrics["by_credential_source"] = {}
    metrics["by_model"] = {}
    if isinstance(value, dict):
        by_category = value.get("by_category")
        if isinstance(by_category, dict):
            for key, item in by_category.items():
                label = str(key or "").strip()
                if label:
                    metrics["by_category"][label] = _normalize_token_usage_bucket(item)
        by_credential_source = value.get("by_credential_source")
        if isinstance(by_credential_source, dict):
            for key, item in by_credential_source.items():
                label = str(key or "").strip()
                if label:
                    metrics["by_credential_source"][label] = _normalize_token_usage_bucket(item)
        by_model = value.get("by_model")
        if isinstance(by_model, dict):
            for key, item in by_model.items():
                label = str(key or "").strip()
                if label:
                    bucket = _normalize_token_usage_bucket(item)
                    if isinstance(item, dict):
                        provider = str(item.get("provider") or "").strip()
                        model = str(item.get("model") or "").strip()
                        if provider:
                            bucket["provider"] = provider
                        if model:
                            bucket["model"] = model
                    metrics["by_model"][label] = bucket
    return metrics


def _event_ts_to_iso(value: Any) -> Optional[str]:
    if isinstance(value, str) and value.strip():
        return value.strip()
    try:
        ts = float(value)
    except Exception:
        return None
    return dt.datetime.fromtimestamp(ts, dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _merge_token_cost(total: Optional[Dict[str, Any]], incoming: Any) -> Optional[Dict[str, Any]]:
    next_cost = _normalize_token_cost(incoming)
    if not next_cost:
        return total
    if not total:
        return next_cost
    current = _normalize_token_cost(total) or {"amount": None, "currency": next_cost.get("currency")}
    current_currency = current.get("currency")
    next_currency = next_cost.get("currency")
    if current_currency and next_currency and current_currency != next_currency:
        return {"amount": None, "currency": "MIXED"}
    if current.get("amount") is None or next_cost.get("amount") is None:
        return {"amount": None, "currency": current_currency or next_currency}
    return {
        "amount": round(float(current.get("amount") or 0.0) + float(next_cost.get("amount") or 0.0), 8),
        "currency": current_currency or next_currency,
    }


def _merge_token_usage_bucket(bucket: Dict[str, Any], usage: Any, *, cost: Any = None, event_at: Optional[str] = None) -> Dict[str, Any]:
    extras: Dict[str, Any] = {}
    if isinstance(bucket, dict):
        for key in ("by_category", "by_credential_source", "by_model", "provider", "model"):
            if key in bucket:
                extras[key] = bucket.get(key)
    target = _normalize_token_usage_bucket(bucket)
    payload = usage if isinstance(usage, dict) else {}
    prompt = _safe_int(payload.get("prompt"), 0)
    completion = _safe_int(payload.get("completion"), 0)
    total = payload.get("total")
    total_value = _safe_int(total) if total not in (None, "") else None
    if total_value is None and (prompt or completion):
        total_value = prompt + completion
    cached = payload.get("cached")
    cached_value = _safe_int(cached) if cached not in (None, "") else None
    target["prompt"] = _safe_int(target.get("prompt"), 0) + prompt
    target["completion"] = _safe_int(target.get("completion"), 0) + completion
    target["total"] = _safe_int(target.get("total"), 0) + int(total_value or 0)
    if cached_value is not None:
        target["cached"] = _safe_int(target.get("cached"), 0) + cached_value
    target["events"] = _safe_int(target.get("events"), 0) + 1
    target["cost"] = _merge_token_cost(target.get("cost"), cost)
    if event_at:
        target["last_event_at"] = event_at
    if bool(payload.get("__chargeable__", False)):
        target["chargeable_total"] = _safe_int(target.get("chargeable_total"), 0) + int(total_value or 0)
    else:
        target["non_chargeable_total"] = _safe_int(target.get("non_chargeable_total"), 0) + int(total_value or 0)
    target.update(extras)
    return target


def _normalize_credential_source(value: Any) -> str:
    cleaned = str(value or "").strip().lower()
    if cleaned in {"user", "user_key", "userkey"}:
        return "user_key"
    if cleaned in {"service", "service_key", "servicekey"}:
        return "service_key"
    if cleaned == "local":
        return "local"
    return "unknown"


def _coerce_chargeable_flag(value: Any, *, provider: Optional[str], credential_source: str) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    cleaned = str(value or "").strip().lower()
    if cleaned in {"1", "true", "yes", "on", "y"}:
        return True
    if cleaned in {"0", "false", "no", "off", "n"}:
        return False
    if credential_source in {"user_key", "local"}:
        return False
    return _normalize_llm_provider_type(provider) not in {"ollama"}


def _merge_token_usage_event(job: "Job", data: Dict[str, Any]) -> None:
    token_usage = _normalize_token_usage_metrics(job.metrics.get("token_usage") if isinstance(job.metrics, dict) else {})
    usage = data.get("usage") if isinstance(data.get("usage"), dict) else {}
    event_at = _event_ts_to_iso(data.get("at") or data.get("ts"))
    cost = data.get("cost")
    provider = str(data.get("provider") or "unknown").strip() or "unknown"
    model = str(data.get("model") or "unknown").strip() or "unknown"
    credential_source = _normalize_credential_source(data.get("credential_source"))
    chargeable = _coerce_chargeable_flag(
        data.get("chargeable"),
        provider=provider,
        credential_source=credential_source,
    )
    usage_payload = dict(usage)
    usage_payload["__chargeable__"] = chargeable
    # Keep the flat totals and the grouped breakdowns derived from the exact same
    # event so UI summaries and token settlement never drift apart.
    token_usage = _merge_token_usage_bucket(token_usage, usage_payload, cost=cost, event_at=event_at)

    category = str(data.get("category") or "llm").strip() or "llm"
    token_usage["by_category"][category] = _merge_token_usage_bucket(
        token_usage["by_category"].get(category),
        usage_payload,
        cost=cost,
        event_at=event_at,
    )

    token_usage["by_credential_source"][credential_source] = _merge_token_usage_bucket(
        token_usage["by_credential_source"].get(credential_source),
        usage_payload,
        cost=cost,
        event_at=event_at,
    )

    model_key = f"{provider}/{model}"
    bucket = _merge_token_usage_bucket(
        token_usage["by_model"].get(model_key),
        usage_payload,
        cost=cost,
        event_at=event_at,
    )
    bucket["provider"] = provider
    bucket["model"] = model
    token_usage["by_model"][model_key] = bucket

    job.metrics["token_usage"] = token_usage
    job.token_actual = int(token_usage.get("total") or 0)


def _llm_request_telemetry_store() -> Optional[Any]:
    if CENTRAL_STORE is None:
        return None
    return getattr(CENTRAL_STORE, "llm_request_telemetry", None)


def _record_llm_request_telemetry_subject(subject: Optional[str], data: Dict[str, Any], *, scope: str = "user") -> bool:
    cleaned_subject = str(subject or "").strip()
    store = _llm_request_telemetry_store()
    if not cleaned_subject or store is None or not isinstance(data, dict):
        return False
    try:
        store.record(scope, cleaned_subject, data)
        return True
    except Exception as exc:
        logger.debug("llm request telemetry persistence failed for %s/%s: %s", scope, cleaned_subject, exc)
        return False


def _record_llm_request_telemetry(
    user: Optional[str],
    data: Dict[str, Any],
    *,
    scope: str = "user",
    team_id: Optional[str] = None,
) -> bool:
    # Persist a user-scoped record plus an optional team roll-up so operators can
    # inspect either personal or shared usage without duplicating raw job logs.
    recorded = _record_llm_request_telemetry_subject(user, data, scope=scope)
    cleaned_team_id = str(team_id or "").strip()
    if cleaned_team_id:
        recorded = _record_llm_request_telemetry_subject(cleaned_team_id, data, scope="team") or recorded
    return recorded


def _current_team_telemetry_subject() -> Optional[str]:
    profile = _current_auth_profile()
    active_team = profile.get("active_team") if isinstance(profile, dict) else None
    if isinstance(active_team, dict):
        cleaned = str(active_team.get("team_id") or active_team.get("id") or "").strip()
        return cleaned or None
    return None


def _handle_inprocess_llm_provider_event(event: Dict[str, Any]) -> None:
    if not isinstance(event, dict) or event.get("type") != "llm_request":
        return
    # Assistant/playground requests run inside the web process, so they do not
    # pass through the job stdout event reader. Persist those here instead.
    _record_llm_request_telemetry(_current_user(), event, team_id=_current_team_telemetry_subject())


register_event_callback(_handle_inprocess_llm_provider_event)


@dataclass
class Stage:
    """One lifecycle stage entry for a background job."""

    name: str
    status: str
    started_at: Optional[str] = None
    finished_at: Optional[str] = None
    message: Optional[str] = None


@dataclass
class Job:
    """Runtime/persistent state for an asynchronous Refiner job."""

    job_id: str
    payload: Dict[str, Any]
    project_name: str = ""
    owner: str = ""
    status: str = "queued"
    workflow: str = "project_solver"
    progress: int = 0
    created_at: str = field(default_factory=_now_iso)
    updated_at: str = field(default_factory=_now_iso)
    started_at: Optional[str] = None
    finished_at: Optional[str] = None
    exit_code: Optional[int] = None
    restart_count: int = 0
    log_path: str = ""
    events_path: str = ""
    output_paths: Dict[str, str] = field(default_factory=dict)
    metrics: Dict[str, Any] = field(default_factory=dict)
    stages: List[Stage] = field(default_factory=list)
    process: Optional[subprocess.Popen] = None
    pid: Optional[int] = None
    stop_requested: bool = False
    notify_email: Optional[str] = None
    notified_at: Optional[str] = None
    notified_via: Optional[str] = None
    notification_error: Optional[str] = None
    token_estimate: Optional[int] = None
    token_reserved: int = 0
    token_reserved_user: int = 0
    token_reserved_team: int = 0
    token_actual: int = 0
    token_debited: int = 0
    token_debited_user: int = 0
    token_debited_team: int = 0
    token_shortfall: int = 0
    token_shortfall_user: int = 0
    token_shortfall_team: int = 0
    token_status: str = "none"
    token_source: str = "none"
    transfer_request: Optional[Dict[str, Any]] = None
    repo_info: Dict[str, Any] = field(default_factory=dict)
    workspace_env: Dict[str, Any] = field(default_factory=dict)
    refunds: List[Dict[str, Any]] = field(default_factory=list)
    archived: bool = False
    archived_at: Optional[str] = None
    meta_path: str = ""
    log_buffer: Deque[Dict[str, Any]] = field(default_factory=lambda: deque(maxlen=2000))
    log_listeners: List[queue.Queue] = field(default_factory=list)
    lock: threading.RLock = field(default_factory=threading.RLock)
    _last_persist_ts: float = field(default=0.0, repr=False)

    def __post_init__(self) -> None:
        payload_workflow = self.payload.get("workflow")
        if payload_workflow:
            self.workflow = payload_workflow
        if not self.project_name:
            self.project_name = self._derive_project_name(self.payload)
        if not self.owner:
            self.owner = self.payload.get("owner") or self.owner
        if not self.notify_email:
            candidate = self.payload.get("notify_email") or self.payload.get("notification_email") or ""
            if isinstance(candidate, str):
                candidate = candidate.strip()
            else:
                candidate = ""
            self.notify_email = candidate or None
        if not self.metrics:
            self.metrics = {
                "token_usage": _default_token_usage_metrics(),
                "errors": 0,
                "resolved": 0,
                "warnings": 0,
                "queue_wait_sec": None,
                "runtime_sec": None,
            }
        else:
            self.metrics["token_usage"] = _normalize_token_usage_metrics(self.metrics.get("token_usage"))
            self.metrics.setdefault("errors", 0)
            self.metrics.setdefault("resolved", 0)
            self.metrics.setdefault("warnings", 0)
            self.metrics.setdefault("queue_wait_sec", None)
            self.metrics.setdefault("runtime_sec", None)
        if not self.meta_path:
            self.meta_path = os.path.join(JOB_ROOT, self.job_id, JOB_META_FILENAME)
        if self.token_estimate is None:
            self.token_estimate = None
        self.token_reserved = int(self.token_reserved or 0)
        self.token_reserved_user = int(self.token_reserved_user or 0)
        self.token_reserved_team = int(self.token_reserved_team or 0)
        self.token_actual = int(self.token_actual or 0)
        self.token_debited = int(self.token_debited or 0)
        self.token_debited_user = int(self.token_debited_user or 0)
        self.token_debited_team = int(self.token_debited_team or 0)
        self.token_shortfall = int(self.token_shortfall or 0)
        self.token_shortfall_user = int(self.token_shortfall_user or 0)
        self.token_shortfall_team = int(self.token_shortfall_team or 0)
        if not self.token_status:
            self.token_status = "none"
        if not self.token_source:
            self.token_source = "none"
        if self.token_reserved and self.token_reserved_user == 0 and self.token_reserved_team == 0:
            self.token_reserved_user = self.token_reserved
        if self.token_debited and self.token_debited_user == 0 and self.token_debited_team == 0:
            self.token_debited_user = self.token_debited
        if self.token_shortfall and self.token_shortfall_user == 0 and self.token_shortfall_team == 0:
            self.token_shortfall_user = self.token_shortfall
        if self.token_reserved != self.token_reserved_user + self.token_reserved_team:
            self.token_reserved = self.token_reserved_user + self.token_reserved_team
        if self.token_debited != self.token_debited_user + self.token_debited_team:
            self.token_debited = self.token_debited_user + self.token_debited_team
        if self.token_shortfall != self.token_shortfall_user + self.token_shortfall_team:
            self.token_shortfall = self.token_shortfall_user + self.token_shortfall_team
        if self.transfer_request is not None and not isinstance(self.transfer_request, dict):
            self.transfer_request = None
        if self.refunds is None:
            self.refunds = []
        if self.archived is None:
            self.archived = False
        if not self.archived:
            self.archived_at = None
        if self.workspace_env is None or not isinstance(self.workspace_env, dict):
            self.workspace_env = {}

    @staticmethod
    def _derive_project_name(payload: Dict[str, Any]) -> str:
        name = (payload.get("project_name") or "").strip()
        if name:
            return name
        root = (payload.get("project_root") or payload.get("delivery_project_root") or "").strip()
        if root:
            return os.path.basename(root.rstrip("/")) or root
        space = (payload.get("space") or "").strip()
        if space:
            return f"Space {space}"
        projects = (payload.get("projects") or "").strip()
        if projects:
            return f"Projects {projects}"
        topic = (payload.get("topic_source") or payload.get("topic_research") or "").strip()
        if topic:
            return topic[:60] + ("…" if len(topic) > 60 else "")
        return "Untitled"

    def set_status(self, status: str) -> None:
        with self.lock:
            self.status = status
            self.updated_at = _now_iso()
        self.persist(force=True)

    def set_progress(self, progress: int) -> None:
        with self.lock:
            self.progress = max(0, min(100, int(progress)))
            self.updated_at = _now_iso()
        self.persist()

    def add_listener(self) -> queue.Queue:
        q: queue.Queue = queue.Queue()
        with self.lock:
            self.log_listeners.append(q)
        return q

    def remove_listener(self, q: queue.Queue) -> None:
        with self.lock:
            if q in self.log_listeners:
                self.log_listeners.remove(q)

    def append_log(self, line: str, stream: str = "stdout") -> None:
        entry = {"ts": _now_iso(), "line": line.rstrip("\n"), "stream": stream}
        with self.lock:
            self.log_buffer.append(entry)
            listeners = list(self.log_listeners)
        for listener in listeners:
            try:
                listener.put_nowait(entry)
            except queue.Full:
                continue

    def get_log_tail(self, count: int) -> List[Dict[str, Any]]:
        with self.lock:
            buffer = list(self.log_buffer)
        if buffer:
            if count <= 0:
                return buffer
            return buffer[-count:]
        if self.log_path and os.path.exists(self.log_path):
            return _read_log_tail(self.log_path, count)
        return buffer

    def update_stage(self, name: str, status: str, message: Optional[str] = None) -> None:
        now = _now_iso()
        with self.lock:
            for stage in self.stages:
                if stage.name == name:
                    stage.status = status
                    if status == "running" and not stage.started_at:
                        stage.started_at = now
                    if status in {"completed", "failed", "skipped", "blocked"}:
                        stage.finished_at = now
                    if message:
                        stage.message = message
                    self.updated_at = now
                    return
            stage = Stage(name=name, status=status, started_at=now if status == "running" else None)
            if status in {"completed", "failed", "skipped", "blocked"}:
                stage.finished_at = now
            if message:
                stage.message = message
            self.stages.append(stage)
            self.updated_at = now
        self.persist()

    def to_persisted_dict(self) -> Dict[str, Any]:
        project_id = None
        team_id = None
        if isinstance(self.payload, dict):
            project_id = self.payload.get("project_id") or self.payload.get("project")
            team_id = self.payload.get("team_id")
        with self.lock:
            data = {
                "version": JOB_META_VERSION,
                "id": self.job_id,
                "workflow": self.workflow,
                "project_name": self.project_name,
                "project_id": project_id,
                "team_id": team_id,
                "owner": self.owner,
                "status": self.status,
                "progress": self.progress,
                "created_at": self.created_at,
                "updated_at": self.updated_at,
                "started_at": self.started_at,
                "finished_at": self.finished_at,
                "exit_code": self.exit_code,
                "restart_count": self.restart_count,
                "output_paths": self.output_paths,
                "metrics": dict(self.metrics),
                "stages": [stage.__dict__ for stage in self.stages],
                "repo_info": self.repo_info,
                "workspace_env": self.workspace_env,
                "refunds": list(self.refunds),
                "transfer_request": self.transfer_request,
                "archived": bool(self.archived),
                "archived_at": self.archived_at,
                "payload": _sanitize_payload(self.payload),
                "notification": {
                    "email": self.notify_email,
                    "sent_at": self.notified_at,
                    "method": self.notified_via,
                    "error": self.notification_error,
                },
                "tokens": {
                    "estimate": self.token_estimate,
                    "reserved": self.token_reserved,
                    "reserved_user": self.token_reserved_user,
                    "reserved_team": self.token_reserved_team,
                    "actual": self.token_actual,
                    "debited": self.token_debited,
                    "debited_user": self.token_debited_user,
                    "debited_team": self.token_debited_team,
                    "shortfall": self.token_shortfall,
                    "shortfall_user": self.token_shortfall_user,
                    "shortfall_team": self.token_shortfall_team,
                    "status": self.token_status,
                    "source": self.token_source,
                },
            }
        return data

    def persist(self, force: bool = False) -> None:
        if not self.meta_path:
            self.meta_path = os.path.join(JOB_ROOT, self.job_id, JOB_META_FILENAME)
        now = time.time()
        if not force and now - self._last_persist_ts < 1.0:
            return
        payload = self.to_persisted_dict()
        persisted = False
        try:
            _write_json_atomic(self.meta_path, payload)
            persisted = True
        except Exception:
            pass
        central_job_store = _central_job_store()
        if central_job_store is not None:
            try:
                central_job_store.upsert(payload)
                persisted = True
            except Exception as exc:
                logger.debug("job metadata sync skipped for %s: %s", self.job_id, exc)
        if persisted:
            self._last_persist_ts = now

    @classmethod
    def from_persisted(cls, data: Dict[str, Any], meta_path: str) -> "Job":
        job_id = data.get("id") or data.get("job_id") or os.path.basename(os.path.dirname(meta_path))
        payload = data.get("payload") if isinstance(data.get("payload"), dict) else {}
        job_dir = os.path.dirname(meta_path)
        log_path = os.path.join(job_dir, "job.log")
        events_path = os.path.join(job_dir, "events.jsonl")
        stages_data = data.get("stages") or []
        stages: List[Stage] = []
        for entry in stages_data:
            if not isinstance(entry, dict):
                continue
            stages.append(
                Stage(
                    name=str(entry.get("name") or "stage"),
                    status=str(entry.get("status") or "unknown"),
                    started_at=_normalise_timestamp(entry.get("started_at")),
                    finished_at=_normalise_timestamp(entry.get("finished_at")),
                    message=entry.get("message"),
                )
            )
        notification = data.get("notification") if isinstance(data.get("notification"), dict) else {}
        tokens = data.get("tokens") if isinstance(data.get("tokens"), dict) else {}
        metrics = data.get("metrics") if isinstance(data.get("metrics"), dict) else {}
        output_paths = data.get("output_paths") if isinstance(data.get("output_paths"), dict) else {}
        repo_info = data.get("repo_info") if isinstance(data.get("repo_info"), dict) else {}
        workspace_env = data.get("workspace_env") if isinstance(data.get("workspace_env"), dict) else {}
        refunds = data.get("refunds") if isinstance(data.get("refunds"), list) else []
        transfer_request = data.get("transfer_request") if isinstance(data.get("transfer_request"), dict) else None
        archived = data.get("archived")
        archived_at = _normalise_timestamp(data.get("archived_at") or None)
        if refunds:
            for refund in refunds:
                if not isinstance(refund, dict):
                    continue
                for key in ("requested_at", "screened_at", "decided_at", "settled_at"):
                    if key in refund:
                        refund[key] = _normalise_timestamp(refund.get(key))
                history = refund.get("history")
                if isinstance(history, list):
                    for entry in history:
                        if isinstance(entry, dict) and entry.get("at"):
                            entry["at"] = _normalise_timestamp(entry.get("at"))
                admin_decision = refund.get("admin_decision")
                if isinstance(admin_decision, dict) and admin_decision.get("decided_at"):
                    admin_decision["decided_at"] = _normalise_timestamp(admin_decision.get("decided_at"))
                screening = refund.get("llm_screening")
                if isinstance(screening, dict) and screening.get("screened_at"):
                    screening["screened_at"] = _normalise_timestamp(screening.get("screened_at"))
        job = cls(
            job_id=job_id,
            payload=payload,
            project_name=data.get("project_name") or "",
            owner=data.get("owner") or "",
            status=data.get("status") or "queued",
            workflow=data.get("workflow") or "project_solver",
            progress=int(data.get("progress") or 0),
            created_at=_normalise_timestamp(data.get("created_at")) or _now_iso(),
            updated_at=_normalise_timestamp(data.get("updated_at")) or _now_iso(),
            started_at=_normalise_timestamp(data.get("started_at")),
            finished_at=_normalise_timestamp(data.get("finished_at")),
            exit_code=data.get("exit_code"),
            restart_count=int(data.get("restart_count") or 0),
            log_path=log_path,
            events_path=events_path,
            output_paths=output_paths,
            metrics=metrics,
            stages=stages,
            repo_info=repo_info,
            workspace_env=workspace_env,
            notify_email=notification.get("email") or data.get("notify_email"),
            notified_at=_normalise_timestamp(notification.get("sent_at")),
            notified_via=notification.get("method"),
            notification_error=notification.get("error"),
            token_estimate=tokens.get("estimate"),
            token_reserved=int(tokens.get("reserved") or 0),
            token_reserved_user=int(tokens.get("reserved_user") or 0),
            token_reserved_team=int(tokens.get("reserved_team") or 0),
            token_actual=int(tokens.get("actual") or 0),
            token_debited=int(tokens.get("debited") or 0),
            token_debited_user=int(tokens.get("debited_user") or 0),
            token_debited_team=int(tokens.get("debited_team") or 0),
            token_shortfall=int(tokens.get("shortfall") or 0),
            token_shortfall_user=int(tokens.get("shortfall_user") or 0),
            token_shortfall_team=int(tokens.get("shortfall_team") or 0),
            token_status=tokens.get("status") or "none",
            token_source=tokens.get("source") or "none",
            transfer_request=transfer_request,
            refunds=refunds,
            archived=bool(archived) if archived is not None else False,
            archived_at=archived_at,
            meta_path=meta_path,
        )
        return job

    def to_dict(self, include_logs: bool = False, log_tail: int = DEFAULT_TAIL) -> Dict[str, Any]:
        project_id = None
        team_id = None
        if isinstance(self.payload, dict):
            project_id = self.payload.get("project_id") or self.payload.get("project")
            team_id = self.payload.get("team_id")
        with self.lock:
            data = {
                "id": self.job_id,
                "workflow": self.workflow,
                "project_name": self.project_name,
                "project_id": project_id,
                "team_id": team_id,
                "owner": self.owner,
                "status": self.status,
                "progress": self.progress,
                "created_at": self.created_at,
                "updated_at": self.updated_at,
                "started_at": self.started_at,
                "finished_at": self.finished_at,
                "exit_code": self.exit_code,
                "restart_count": self.restart_count,
                "output_paths": self.output_paths,
                "metrics": dict(self.metrics),
                "stages": [stage.__dict__ for stage in self.stages],
                "pid": self.pid,
                "repo_info": self.repo_info,
                "workspace_env": self.workspace_env,
                "refunds": list(self.refunds),
                "transfer_request": self.transfer_request,
                "archived": bool(self.archived),
                "archived_at": self.archived_at,
                "tokens": {
                    "estimate": self.token_estimate,
                    "reserved": self.token_reserved,
                    "reserved_user": self.token_reserved_user,
                    "reserved_team": self.token_reserved_team,
                    "actual": self.token_actual,
                    "debited": self.token_debited,
                    "debited_user": self.token_debited_user,
                    "debited_team": self.token_debited_team,
                    "shortfall": self.token_shortfall,
                    "shortfall_user": self.token_shortfall_user,
                    "shortfall_team": self.token_shortfall_team,
                    "status": self.token_status,
                    "source": self.token_source,
                },
            }
        if include_logs:
            data["logs"] = self.get_log_tail(log_tail)
        return data


class JobActionExecutionError(RuntimeError):
    """Typed failure for background actions with stable API error semantics."""

    def __init__(self, code: str, details: Optional[str] = None, status_code: int = 400):
        super().__init__(details or code)
        self.code = code
        self.details = details or code
        self.status_code = int(status_code)

    def to_payload(self) -> Dict[str, Any]:
        return {
            "error": self.code,
            "details": self.details,
            "status_code": self.status_code,
        }


@dataclass
class JobActionTask:
    """Queued side-action task associated with a job."""

    task_id: str
    job_id: str
    owner: str
    action: str
    payload: Dict[str, Any] = field(default_factory=dict)
    status: str = "queued"
    created_at: str = field(default_factory=_now_iso)
    started_at: Optional[str] = None
    finished_at: Optional[str] = None
    timeout_sec: float = JOB_ACTION_TIMEOUT_SEC
    cancel_requested: bool = False
    error: Optional[str] = None
    result: Optional[Dict[str, Any]] = None

    def is_terminal(self) -> bool:
        return self.status in {"completed", "failed", "cancelled"}

    def to_dict(self, include_result: bool = True) -> Dict[str, Any]:
        payload: Dict[str, Any] = {
            "task_id": self.task_id,
            "job_id": self.job_id,
            "owner": self.owner,
            "action": self.action,
            "status": self.status,
            "created_at": self.created_at,
            "started_at": self.started_at,
            "finished_at": self.finished_at,
            "timeout_sec": self.timeout_sec,
            "cancel_requested": bool(self.cancel_requested),
            "error": self.error,
        }
        if include_result and self.result is not None:
            payload["result"] = self.result
        return payload


class JobActionManager:
    """Owner-aware execution pool for non-critical long-running API side actions."""

    def __init__(
        self,
        *,
        workers: int = JOB_ACTION_WORKERS,
        max_queue: int = JOB_ACTION_MAX_QUEUE,
        task_ttl_sec: int = JOB_ACTION_TASK_TTL_SEC,
        max_outstanding_per_owner: int = JOB_ACTION_MAX_OUTSTANDING_PER_OWNER,
        max_inflight_per_owner: int = JOB_ACTION_MAX_INFLIGHT_PER_OWNER,
    ):
        self.max_queue = max(1, int(max_queue))
        self.task_ttl_sec = max(60, int(task_ttl_sec))
        self.lock = threading.RLock()
        self._condition = threading.Condition(self.lock)
        self.tasks: Dict[str, JobActionTask] = {}
        self.job_task_ids: Dict[str, Deque[str]] = {}
        self.owner_task_ids: Dict[str, Deque[str]] = {}
        self._ready_task_ids: Dict[str, Deque[str]] = {}
        self._ready_owner_order: Deque[str] = deque()
        self.workers: List[threading.Thread] = []
        self._inflight = 0
        self._inflight_by_owner: Dict[str, int] = {}
        self.worker_count = max(1, int(workers))
        self.max_outstanding_per_owner = max(1, min(int(max_outstanding_per_owner), self.max_queue))
        self.max_inflight_per_owner = max(1, min(int(max_inflight_per_owner), self.worker_count))
        self._ready_count = 0
        for idx in range(self.worker_count):
            worker = threading.Thread(target=self._worker_loop, args=(idx,), daemon=True)
            worker.start()
            self.workers.append(worker)

    @staticmethod
    def _owner_key(owner: Optional[str]) -> str:
        cleaned = str(owner or "").strip()
        return cleaned or "__anonymous__"

    def queue_depth(self) -> int:
        with self.lock:
            return int(self._ready_count)

    def inflight(self) -> int:
        with self.lock:
            return int(self._inflight)

    def active_owner_count(self) -> int:
        with self.lock:
            owners = set(self._ready_task_ids)
            owners.update(self._inflight_by_owner)
            return len(owners)

    def snapshot(self) -> Dict[str, int]:
        with self.lock:
            return {
                "workers": len(self.workers),
                "queue_depth": int(self._ready_count),
                "inflight": int(self._inflight),
                "queue_capacity": self.max_queue,
                "max_outstanding_per_owner": self.max_outstanding_per_owner,
                "max_inflight_per_owner": self.max_inflight_per_owner,
                "active_owner_count": self.active_owner_count(),
            }

    def _owner_outstanding_locked(self, owner: str) -> int:
        outstanding = 0
        for task_id in self.owner_task_ids.get(owner) or []:
            task = self.tasks.get(task_id)
            if task and not task.is_terminal():
                outstanding += 1
        return outstanding

    def _enqueue_ready_locked(self, owner_key: str, task_id: str) -> None:
        ready_ids = self._ready_task_ids.get(owner_key)
        if ready_ids is None:
            ready_ids = deque()
            self._ready_task_ids[owner_key] = ready_ids
            self._ready_owner_order.append(owner_key)
        ready_ids.append(task_id)
        self._ready_count += 1

    def _remove_ready_locked(self, owner_key: str, task_id: str) -> bool:
        ready_ids = self._ready_task_ids.get(owner_key)
        if not ready_ids:
            return False
        try:
            ready_ids.remove(task_id)
        except ValueError:
            return False
        self._ready_count = max(0, self._ready_count - 1)
        if not ready_ids:
            self._ready_task_ids.pop(owner_key, None)
            try:
                self._ready_owner_order.remove(owner_key)
            except ValueError:
                pass
        return True

    def _ready_owner_count_locked(self) -> int:
        return sum(1 for owner_key in self._ready_owner_order if self._ready_task_ids.get(owner_key))

    def _next_ready_task_id_locked(self) -> Optional[str]:
        owner_slots = len(self._ready_owner_order)
        if owner_slots <= 0:
            return None
        ready_owner_count = self._ready_owner_count_locked()
        for _ in range(owner_slots):
            owner_key = self._ready_owner_order.popleft()
            ready_ids = self._ready_task_ids.get(owner_key)
            if not ready_ids:
                self._ready_task_ids.pop(owner_key, None)
                continue
            owner_inflight = int(self._inflight_by_owner.get(owner_key) or 0)
            competing_owners_ready = ready_owner_count > 1
            if competing_owners_ready and owner_inflight >= self.max_inflight_per_owner:
                self._ready_owner_order.append(owner_key)
                continue
            task_id = ready_ids.popleft()
            self._ready_count = max(0, self._ready_count - 1)
            if ready_ids:
                self._ready_owner_order.append(owner_key)
            else:
                self._ready_task_ids.pop(owner_key, None)
            return task_id
        return None

    def submit(
        self,
        *,
        job_id: str,
        owner: str,
        action: str,
        payload: Optional[Dict[str, Any]] = None,
        timeout_sec: Optional[float] = None,
    ) -> JobActionTask:
        task = JobActionTask(
            task_id=uuid.uuid4().hex,
            job_id=job_id,
            owner=owner,
            action=action,
            payload=payload if isinstance(payload, dict) else {},
            timeout_sec=max(1.0, float(timeout_sec or JOB_ACTION_TIMEOUT_SEC)),
        )
        with self._condition:
            self._purge_expired_locked()
            if self._ready_count >= self.max_queue:
                raise queue.Full()
            if self._owner_outstanding_locked(task.owner) >= self.max_outstanding_per_owner:
                raise queue.Full()
            self.tasks[task.task_id] = task
            task_ids = self.job_task_ids.get(job_id)
            if task_ids is None:
                task_ids = deque(maxlen=100)
                self.job_task_ids[job_id] = task_ids
            task_ids.appendleft(task.task_id)
            owner_ids = self.owner_task_ids.get(task.owner)
            if owner_ids is None:
                owner_ids = deque(maxlen=100)
                self.owner_task_ids[task.owner] = owner_ids
            owner_ids.appendleft(task.task_id)
            self._enqueue_ready_locked(self._owner_key(task.owner), task.task_id)
            self._condition.notify_all()
        return task

    def get_task(self, task_id: str) -> Optional[JobActionTask]:
        with self.lock:
            self._purge_expired_locked()
            return self.tasks.get(task_id)

    def list_for_job(self, job_id: str, *, limit: int = 20, include_results: bool = False) -> List[Dict[str, Any]]:
        limit = max(1, min(int(limit), 200))
        with self.lock:
            self._purge_expired_locked()
            task_ids = list(self.job_task_ids.get(job_id) or [])
            results: List[Dict[str, Any]] = []
            for task_id in task_ids:
                task = self.tasks.get(task_id)
                if not task:
                    continue
                results.append(task.to_dict(include_result=include_results))
                if len(results) >= limit:
                    break
            return results

    def cancel(self, task_id: str, *, job_id: Optional[str] = None) -> bool:
        with self._condition:
            task = self.tasks.get(task_id)
            if not task:
                return False
            if job_id and task.job_id != job_id:
                return False
            if task.is_terminal():
                return task.status == "cancelled"
            task.cancel_requested = True
            if task.status == "queued":
                task.status = "cancelled"
                task.finished_at = _now_iso()
                task.error = "cancelled_by_user"
                task.result = {
                    "error": "cancelled",
                    "details": "Cancelled before execution started.",
                    "status_code": 409,
                }
                self._remove_ready_locked(self._owner_key(task.owner), task.task_id)
                self._condition.notify_all()
            return True

    def _purge_expired_locked(self) -> None:
        now_ts = time.time()
        expired: List[str] = []
        for task_id, task in self.tasks.items():
            if not task.is_terminal():
                continue
            anchor_ts = _timestamp_sort_key(task.finished_at or task.created_at)
            if anchor_ts <= 0:
                continue
            if now_ts - anchor_ts >= self.task_ttl_sec:
                expired.append(task_id)
        if not expired:
            return
        for task_id in expired:
            task = self.tasks.pop(task_id, None)
            if not task:
                continue
            task_ids = self.job_task_ids.get(task.job_id)
            if task_ids and task_id in task_ids:
                task_ids.remove(task_id)
                if not task_ids:
                    self.job_task_ids.pop(task.job_id, None)
            owner_ids = self.owner_task_ids.get(task.owner)
            if owner_ids and task_id in owner_ids:
                owner_ids.remove(task_id)
                if not owner_ids:
                    self.owner_task_ids.pop(task.owner, None)
            self._remove_ready_locked(self._owner_key(task.owner), task_id)

    def _worker_loop(self, worker_id: int) -> None:
        while True:
            task: Optional[JobActionTask] = None
            with self._condition:
                while task is None:
                    task_id = self._next_ready_task_id_locked()
                    if not task_id:
                        self._condition.wait()
                        continue
                    candidate = self.tasks.get(task_id)
                    if not candidate:
                        continue
                    if candidate.status == "cancelled" or (candidate.cancel_requested and candidate.status == "queued"):
                        candidate.status = "cancelled"
                        candidate.finished_at = _now_iso()
                        candidate.error = candidate.error or "cancelled_before_execution"
                        candidate.result = candidate.result or {
                            "error": "cancelled",
                            "details": "Cancelled before execution started.",
                            "status_code": 409,
                        }
                        continue
                    candidate.status = "running"
                    candidate.started_at = _now_iso()
                    self._inflight += 1
                    owner_key = self._owner_key(candidate.owner)
                    self._inflight_by_owner[owner_key] = int(self._inflight_by_owner.get(owner_key) or 0) + 1
                    task = candidate
            try:
                result = _execute_job_action_task(task)
                with self.lock:
                    if task.cancel_requested:
                        task.status = "cancelled"
                        task.error = "cancelled_during_execution"
                        task.result = {
                            "error": "cancelled",
                            "details": "Cancellation was requested while the action was running.",
                            "status_code": 409,
                        }
                    else:
                        task.status = "completed"
                        task.result = result if isinstance(result, dict) else {}
                        task.error = None
                    task.finished_at = _now_iso()
            except JobActionExecutionError as exc:
                with self.lock:
                    task.status = "cancelled" if task.cancel_requested else "failed"
                    task.error = exc.code
                    task.result = exc.to_payload()
                    task.finished_at = _now_iso()
            except Exception as exc:
                with self.lock:
                    task.status = "cancelled" if task.cancel_requested else "failed"
                    task.error = str(exc)
                    task.result = {"error": "job_action_failed", "details": str(exc), "status_code": 500}
                    task.finished_at = _now_iso()
            finally:
                with self._condition:
                    self._inflight = max(0, self._inflight - 1)
                    owner_key = self._owner_key(task.owner)
                    owner_inflight = int(self._inflight_by_owner.get(owner_key) or 0)
                    if owner_inflight <= 1:
                        self._inflight_by_owner.pop(owner_key, None)
                    else:
                        self._inflight_by_owner[owner_key] = owner_inflight - 1
                    self._purge_expired_locked()
                    self._condition.notify_all()


class SubtaskExecutionError(RuntimeError):
    """Typed failure for generic background subtasks with stable API semantics."""

    def __init__(self, code: str, details: Optional[str] = None, status_code: int = 400):
        super().__init__(details or code)
        self.code = code
        self.details = details or code
        self.status_code = int(status_code)

    def to_payload(self) -> Dict[str, Any]:
        return {
            "error": self.code,
            "details": self.details,
            "status_code": self.status_code,
        }


@dataclass
class Subtask:
    """Queued background subtask associated with a user and optional scope."""

    task_id: str
    owner: str
    action: str
    payload: Dict[str, Any] = field(default_factory=dict)
    scope_type: str = "user"
    scope_id: Optional[str] = None
    status: str = "queued"
    created_at: str = field(default_factory=_now_iso)
    started_at: Optional[str] = None
    finished_at: Optional[str] = None
    timeout_sec: float = SUBTASK_TIMEOUT_SEC
    cancel_requested: bool = False
    error: Optional[str] = None
    result: Optional[Dict[str, Any]] = None

    def is_terminal(self) -> bool:
        return self.status in {"completed", "failed", "cancelled"}

    def to_dict(self, include_result: bool = True) -> Dict[str, Any]:
        payload: Dict[str, Any] = {
            "task_id": self.task_id,
            "owner": self.owner,
            "action": self.action,
            "scope_type": self.scope_type,
            "scope_id": self.scope_id,
            "status": self.status,
            "created_at": self.created_at,
            "started_at": self.started_at,
            "finished_at": self.finished_at,
            "timeout_sec": self.timeout_sec,
            "cancel_requested": bool(self.cancel_requested),
            "error": self.error,
        }
        if include_result and self.result is not None:
            payload["result"] = self.result
        return payload


class SubtaskManager:
    """Bounded async execution pool for generic assistant/job orchestration subtasks."""

    def __init__(
        self,
        *,
        workers: int = SUBTASK_WORKERS,
        max_queue: int = SUBTASK_MAX_QUEUE,
        task_ttl_sec: int = SUBTASK_TASK_TTL_SEC,
        max_outstanding_per_owner: int = SUBTASK_MAX_OUTSTANDING_PER_OWNER,
        max_inflight_per_owner: int = SUBTASK_MAX_INFLIGHT_PER_OWNER,
    ):
        self.max_queue = max(1, int(max_queue))
        self.task_ttl_sec = max(60, int(task_ttl_sec))
        self.lock = threading.RLock()
        self._condition = threading.Condition(self.lock)
        self.tasks: Dict[str, Subtask] = {}
        self.owner_task_ids: Dict[str, Deque[str]] = {}
        self.scope_task_ids: Dict[str, Deque[str]] = {}
        self._ready_task_ids: Dict[str, Deque[str]] = {}
        self._ready_owner_order: Deque[str] = deque()
        self.workers: List[threading.Thread] = []
        self._inflight = 0
        self._inflight_by_owner: Dict[str, int] = {}
        self.worker_count = max(1, int(workers))
        self.max_outstanding_per_owner = max(1, int(max_outstanding_per_owner))
        self.max_inflight_per_owner = max(1, min(int(max_inflight_per_owner), self.worker_count))
        self._ready_count = 0
        for idx in range(self.worker_count):
            worker = threading.Thread(target=self._worker_loop, args=(idx,), daemon=True)
            worker.start()
            self.workers.append(worker)

    @staticmethod
    def _scope_key(scope_type: Optional[str], scope_id: Optional[str]) -> Optional[str]:
        scope_type = str(scope_type or "").strip().lower()
        scope_id = str(scope_id or "").strip()
        if not scope_type or not scope_id:
            return None
        return f"{scope_type}:{scope_id}"

    @staticmethod
    def _owner_key(owner: Optional[str]) -> str:
        cleaned = str(owner or "").strip()
        return cleaned or "__anonymous__"

    def queue_depth(self) -> int:
        with self.lock:
            return int(self._ready_count)

    def inflight(self) -> int:
        with self.lock:
            return int(self._inflight)

    def _owner_outstanding_locked(self, owner: str) -> int:
        outstanding = 0
        for task_id in self.owner_task_ids.get(owner) or []:
            task = self.tasks.get(task_id)
            if task and not task.is_terminal():
                outstanding += 1
        return outstanding

    def _enqueue_ready_locked(self, owner_key: str, task_id: str) -> None:
        ready_ids = self._ready_task_ids.get(owner_key)
        if ready_ids is None:
            ready_ids = deque()
            self._ready_task_ids[owner_key] = ready_ids
            self._ready_owner_order.append(owner_key)
        ready_ids.append(task_id)
        self._ready_count += 1

    def _remove_ready_locked(self, owner_key: str, task_id: str) -> bool:
        ready_ids = self._ready_task_ids.get(owner_key)
        if not ready_ids:
            return False
        try:
            ready_ids.remove(task_id)
        except ValueError:
            return False
        self._ready_count = max(0, self._ready_count - 1)
        if not ready_ids:
            self._ready_task_ids.pop(owner_key, None)
            try:
                self._ready_owner_order.remove(owner_key)
            except ValueError:
                pass
        return True

    def _ready_owner_count_locked(self) -> int:
        return sum(1 for owner_key in self._ready_owner_order if self._ready_task_ids.get(owner_key))

    def _next_ready_task_id_locked(self) -> Optional[str]:
        owner_slots = len(self._ready_owner_order)
        if owner_slots <= 0:
            return None
        ready_owner_count = self._ready_owner_count_locked()
        for _ in range(owner_slots):
            owner_key = self._ready_owner_order.popleft()
            ready_ids = self._ready_task_ids.get(owner_key)
            if not ready_ids:
                self._ready_task_ids.pop(owner_key, None)
                continue
            owner_inflight = int(self._inflight_by_owner.get(owner_key) or 0)
            competing_owners_ready = ready_owner_count > 1
            if competing_owners_ready and owner_inflight >= self.max_inflight_per_owner:
                self._ready_owner_order.append(owner_key)
                continue
            task_id = ready_ids.popleft()
            self._ready_count = max(0, self._ready_count - 1)
            if ready_ids:
                self._ready_owner_order.append(owner_key)
            else:
                self._ready_task_ids.pop(owner_key, None)
            return task_id
        return None

    def submit(
        self,
        *,
        owner: str,
        action: str,
        payload: Optional[Dict[str, Any]] = None,
        scope_type: str = "user",
        scope_id: Optional[str] = None,
        timeout_sec: Optional[float] = None,
    ) -> Subtask:
        task = Subtask(
            task_id=uuid.uuid4().hex,
            owner=str(owner or "").strip(),
            action=str(action or "").strip(),
            payload=payload if isinstance(payload, dict) else {},
            scope_type=str(scope_type or "user").strip().lower() or "user",
            scope_id=str(scope_id).strip() if scope_id is not None else None,
            timeout_sec=max(1.0, float(timeout_sec or SUBTASK_TIMEOUT_SEC)),
        )
        with self._condition:
            self._purge_expired_locked()
            if self._ready_count >= self.max_queue:
                raise queue.Full()
            if self._owner_outstanding_locked(task.owner) >= self.max_outstanding_per_owner:
                raise queue.Full()
            self.tasks[task.task_id] = task
            owner_ids = self.owner_task_ids.get(task.owner)
            if owner_ids is None:
                owner_ids = deque(maxlen=200)
                self.owner_task_ids[task.owner] = owner_ids
            owner_ids.appendleft(task.task_id)
            scope_key = self._scope_key(task.scope_type, task.scope_id)
            if scope_key:
                scope_ids = self.scope_task_ids.get(scope_key)
                if scope_ids is None:
                    scope_ids = deque(maxlen=200)
                    self.scope_task_ids[scope_key] = scope_ids
                scope_ids.appendleft(task.task_id)
            self._enqueue_ready_locked(self._owner_key(task.owner), task.task_id)
            self._condition.notify_all()
        return task

    def get_task(self, task_id: str) -> Optional[Subtask]:
        with self.lock:
            self._purge_expired_locked()
            return self.tasks.get(task_id)

    def list_for_owner(self, owner: str, *, limit: int = 20, include_results: bool = False) -> List[Dict[str, Any]]:
        limit = max(1, min(int(limit), 200))
        with self.lock:
            self._purge_expired_locked()
            task_ids = list(self.owner_task_ids.get(owner) or [])
            results: List[Dict[str, Any]] = []
            for task_id in task_ids:
                task = self.tasks.get(task_id)
                if not task:
                    continue
                results.append(task.to_dict(include_result=include_results))
                if len(results) >= limit:
                    break
            return results

    def list_for_scope(
        self,
        scope_type: str,
        scope_id: str,
        *,
        limit: int = 20,
        include_results: bool = False,
    ) -> List[Dict[str, Any]]:
        limit = max(1, min(int(limit), 200))
        scope_key = self._scope_key(scope_type, scope_id)
        if not scope_key:
            return []
        with self.lock:
            self._purge_expired_locked()
            task_ids = list(self.scope_task_ids.get(scope_key) or [])
            results: List[Dict[str, Any]] = []
            for task_id in task_ids:
                task = self.tasks.get(task_id)
                if not task:
                    continue
                results.append(task.to_dict(include_result=include_results))
                if len(results) >= limit:
                    break
            return results

    def cancel(self, task_id: str, *, owner: Optional[str] = None) -> bool:
        with self.lock:
            task = self.tasks.get(task_id)
            if not task:
                return False
            if owner and task.owner != owner:
                return False
            if task.is_terminal():
                return task.status == "cancelled"
            task.cancel_requested = True
            if task.status == "queued":
                task.status = "cancelled"
                task.finished_at = _now_iso()
                task.error = "cancelled_by_user"
                task.result = {
                    "error": "cancelled",
                    "details": "Cancelled before execution started.",
                    "status_code": 409,
                }
                self._remove_ready_locked(self._owner_key(task.owner), task.task_id)
                _mark_subtask_cancelled_side_effects(task)
                self._condition.notify_all()
            return True

    def _purge_expired_locked(self) -> None:
        now_ts = time.time()
        expired: List[str] = []
        for task_id, task in self.tasks.items():
            if not task.is_terminal():
                continue
            anchor_ts = _timestamp_sort_key(task.finished_at or task.created_at)
            if anchor_ts <= 0:
                continue
            if now_ts - anchor_ts >= self.task_ttl_sec:
                expired.append(task_id)
        if not expired:
            return
        for task_id in expired:
            task = self.tasks.pop(task_id, None)
            if not task:
                continue
            self._remove_ready_locked(self._owner_key(task.owner), task_id)
            owner_ids = self.owner_task_ids.get(task.owner)
            if owner_ids and task_id in owner_ids:
                owner_ids.remove(task_id)
                if not owner_ids:
                    self.owner_task_ids.pop(task.owner, None)
            scope_key = self._scope_key(task.scope_type, task.scope_id)
            if scope_key:
                scope_ids = self.scope_task_ids.get(scope_key)
                if scope_ids and task_id in scope_ids:
                    scope_ids.remove(task_id)
                    if not scope_ids:
                        self.scope_task_ids.pop(scope_key, None)

    def _worker_loop(self, worker_id: int) -> None:
        while True:
            task: Optional[Subtask] = None
            with self._condition:
                while task is None:
                    task_id = self._next_ready_task_id_locked()
                    if not task_id:
                        self._condition.wait()
                        continue
                    candidate = self.tasks.get(task_id)
                    if not candidate:
                        continue
                    if candidate.status == "cancelled" or (candidate.cancel_requested and candidate.status == "queued"):
                        candidate.status = "cancelled"
                        candidate.finished_at = _now_iso()
                        candidate.error = candidate.error or "cancelled_before_execution"
                        _mark_subtask_cancelled_side_effects(candidate)
                        continue
                    candidate.status = "running"
                    candidate.started_at = _now_iso()
                    self._inflight += 1
                    owner_key = self._owner_key(candidate.owner)
                    self._inflight_by_owner[owner_key] = int(self._inflight_by_owner.get(owner_key) or 0) + 1
                    task = candidate
            try:
                result = _execute_subtask(task)
                with self.lock:
                    task.status = "completed"
                    task.result = result if isinstance(result, dict) else {}
                    if task.cancel_requested:
                        task.result = dict(task.result or {})
                        task.result["cancel_requested"] = True
                    task.error = None
                    task.finished_at = _now_iso()
            except SubtaskExecutionError as exc:
                with self.lock:
                    task.status = "cancelled" if exc.code == "cancelled" else "failed"
                    task.error = exc.code
                    task.result = exc.to_payload()
                    task.finished_at = _now_iso()
            except Exception as exc:
                with self.lock:
                    task.status = "failed"
                    task.error = str(exc)
                    task.result = {"error": "subtask_failed", "details": str(exc), "status_code": 500}
                    task.finished_at = _now_iso()
            finally:
                with self._condition:
                    self._inflight = max(0, self._inflight - 1)
                    owner_key = self._owner_key(task.owner)
                    owner_inflight = int(self._inflight_by_owner.get(owner_key) or 0)
                    if owner_inflight <= 1:
                        self._inflight_by_owner.pop(owner_key, None)
                    else:
                        self._inflight_by_owner[owner_key] = owner_inflight - 1
                    self._purge_expired_locked()
                    self._condition.notify_all()


class TodoScheduler:
    """Daemon scheduler that dispatches due TODO schedules into the subtask queue."""

    def __init__(
        self,
        *,
        schedule_store: ScheduleStore,
        todo_store: TodoStore,
        subtask_manager: SubtaskManager,
        poll_sec: float = TODO_SCHEDULER_POLL_SEC,
        execution_timeout_sec: float = TODO_SCHEDULE_EXECUTION_TIMEOUT_SEC,
        orphan_ttl_sec: int = SUBTASK_ORPHAN_TTL_SEC,
    ):
        self.schedule_store = schedule_store
        self.todo_store = todo_store
        self.subtask_manager = subtask_manager
        self.poll_sec = max(1.0, float(poll_sec))
        self.execution_timeout_sec = max(1.0, float(execution_timeout_sec))
        self.orphan_ttl_sec = max(60, int(orphan_ttl_sec))
        self._stop_event = threading.Event()
        self._thread: Optional[threading.Thread] = None

    def start(self) -> None:
        if self._thread and self._thread.is_alive():
            return
        self._stop_event.clear()
        self._thread = threading.Thread(target=self._loop, name="todo-scheduler", daemon=True)
        self._thread.start()

    def stop(self, timeout: float = 2.0) -> None:
        self._stop_event.set()
        if self._thread and self._thread.is_alive():
            self._thread.join(timeout=max(0.1, float(timeout)))

    def status(self) -> Dict[str, Any]:
        thread = self._thread
        return {
            "enabled": bool(TODO_SCHEDULER_ENABLED),
            "running": bool(thread and thread.is_alive()),
            "poll_sec": self.poll_sec,
            "queue_depth": self.subtask_manager.queue_depth(),
            "inflight": self.subtask_manager.inflight(),
        }

    def _loop(self) -> None:
        while not self._stop_event.is_set():
            try:
                self.run_once()
            except Exception:
                logger.exception("TODO scheduler cycle failed")
            self._stop_event.wait(self.poll_sec)

    def run_once(self, *, dispatch_limit: int = 8, active_limit: int = 64) -> None:
        self._reconcile_active_schedules(limit=active_limit)
        self._dispatch_due_schedules(limit=dispatch_limit)

    def _dispatch_due_schedules(self, *, limit: int) -> None:
        schedules = self.schedule_store.list_due(limit=max(1, int(limit)))
        now_ts = time.time()
        for schedule in schedules:
            schedule_id = str(schedule.get("id") or "").strip()
            user = str(schedule.get("user") or "").strip()
            todo_id = str(schedule.get("todo_id") or "").strip()
            if not schedule_id or not user or not todo_id:
                continue
            todo = self.todo_store.get_item(user, todo_id)
            if not todo:
                self.schedule_store.fail(
                    schedule_id,
                    "todo_not_found",
                    result={"error": "todo_not_found", "todo_id": todo_id},
                )
                _audit_event("todo_schedule_dispatch", actor=user, status="failed", details={"schedule_id": schedule_id, "todo_id": todo_id, "error": "todo_not_found"})
                continue
            if str(todo.get("status") or "todo").strip().lower() != "todo":
                self.schedule_store.cancel(schedule_id, reason="todo_not_pending")
                _audit_event("todo_schedule_dispatch", actor=user, status="cancelled", details={"schedule_id": schedule_id, "todo_id": todo_id, "reason": "todo_not_pending"})
                continue
            available_after = todo.get("available_after")
            if available_after and _timestamp_sort_key(available_after) > now_ts:
                continue
            if str(todo.get("execution_state") or "ready").strip().lower() in {"claimed", "processing"}:
                continue
            claimed = self.todo_store.claim_item(user, todo_id)
            if not claimed:
                continue
            self.schedule_store.mark_dispatching(schedule_id)
            try:
                task = self.subtask_manager.submit(
                    owner=user,
                    action="todo_execute",
                    payload={
                        "todo_id": todo_id,
                        "schedule_id": schedule_id,
                        "route_override": schedule.get("route_override"),
                        "_todo_store": self.todo_store,
                    },
                    scope_type="schedule",
                    scope_id=schedule_id,
                    timeout_sec=self.execution_timeout_sec,
                )
            except queue.Full:
                self.schedule_store.fail(
                    schedule_id,
                    "subtask_capacity_unavailable",
                    result={"error": "subtask_capacity_unavailable", "todo_id": todo_id},
                )
                self.todo_store.release_execution(user, todo_id, execution_state="ready", error="subtask_capacity_unavailable")
                _audit_event("todo_schedule_dispatch", actor=user, status="failed", details={"schedule_id": schedule_id, "todo_id": todo_id, "error": "subtask_capacity_unavailable"})
                continue
            except Exception as exc:
                self.schedule_store.fail(
                    schedule_id,
                    "schedule_dispatch_failed",
                    result={"error": "schedule_dispatch_failed", "details": str(exc), "todo_id": todo_id},
                )
                self.todo_store.release_execution(user, todo_id, execution_state="ready", error="schedule_dispatch_failed")
                _audit_event("todo_schedule_dispatch", actor=user, status="failed", details={"schedule_id": schedule_id, "todo_id": todo_id, "error": str(exc)})
                continue
            self.todo_store.append_link(user, todo_id, "schedules", schedule_id)
            self.todo_store.append_link(user, todo_id, "subtasks", task.task_id)
            self.schedule_store.mark_queued(schedule_id, task.task_id)
            _audit_event("todo_schedule_dispatch", actor=user, status="success", details={"schedule_id": schedule_id, "todo_id": todo_id, "subtask_id": task.task_id})

    def _reconcile_active_schedules(self, *, limit: int) -> None:
        active_schedules = self.schedule_store.list_active(limit=max(1, int(limit)))
        now_ts = time.time()
        for schedule in active_schedules:
            schedule_id = str(schedule.get("id") or "").strip()
            user = str(schedule.get("user") or "").strip()
            todo_id = str(schedule.get("todo_id") or "").strip()
            subtask_id = str(schedule.get("subtask_id") or "").strip()
            status = str(schedule.get("status") or "scheduled").strip().lower()
            if not schedule_id or not user or not todo_id:
                continue
            if status == "dispatching" and not subtask_id:
                updated_ts = _timestamp_sort_key(schedule.get("updated_at") or schedule.get("created_at"))
                if updated_ts and now_ts - updated_ts >= self.orphan_ttl_sec:
                    self.schedule_store.fail(schedule_id, "subtask_orphaned")
                    self.todo_store.release_execution(user, todo_id, execution_state="ready", error="subtask_orphaned")
                continue
            if not subtask_id:
                continue
            task = self.subtask_manager.get_task(subtask_id)
            if not task:
                updated_ts = _timestamp_sort_key(schedule.get("updated_at") or schedule.get("created_at"))
                if updated_ts and now_ts - updated_ts >= self.orphan_ttl_sec:
                    self.schedule_store.fail(schedule_id, "subtask_orphaned")
                    self.todo_store.release_execution(user, todo_id, execution_state="ready", error="subtask_orphaned")
                continue
            if task.status == "queued":
                self.schedule_store.mark_queued(schedule_id, subtask_id)
                continue
            if task.status == "running":
                self.schedule_store.mark_running(schedule_id, subtask_id)
                continue
            if task.status == "completed":
                self.schedule_store.complete(schedule_id, task.result or {})
                _audit_event("todo_schedule_complete", actor=user, status="success", details={"schedule_id": schedule_id, "todo_id": todo_id, "subtask_id": subtask_id})
                continue
            if task.status == "cancelled":
                self.schedule_store.cancel(schedule_id, reason=task.error or "cancelled")
                todo = self.todo_store.get_item(user, todo_id)
                if todo and str(todo.get("status") or "todo").strip().lower() == "todo":
                    self.todo_store.release_execution(user, todo_id, execution_state="ready", error=task.error or "cancelled")
                _audit_event("todo_schedule_complete", actor=user, status="cancelled", details={"schedule_id": schedule_id, "todo_id": todo_id, "subtask_id": subtask_id})
                continue
            if task.status == "failed":
                self.schedule_store.fail(schedule_id, task.error or "subtask_failed", result=task.result or {})
                _audit_event("todo_schedule_complete", actor=user, status="failed", details={"schedule_id": schedule_id, "todo_id": todo_id, "subtask_id": subtask_id, "error": task.error or "subtask_failed"})


class LLMTelemetryJanitor:
    """Background janitor that trims old Postgres LLM telemetry buckets."""

    def __init__(
        self,
        *,
        telemetry_store_factory: Callable[[], Optional[Any]],
        retention_hours: int = LLM_TELEMETRY_RETENTION_HOURS,
        poll_sec: float = LLM_TELEMETRY_PRUNE_POLL_SEC,
    ):
        self.telemetry_store_factory = telemetry_store_factory
        self.retention_hours = max(0, int(retention_hours))
        self.poll_sec = max(60.0, float(poll_sec))
        self._stop_event = threading.Event()
        self._thread: Optional[threading.Thread] = None
        self._lock = threading.Lock()
        self._last_run_at: Optional[str] = None
        self._last_removed = 0
        self._last_error: Optional[str] = None

    def start(self) -> None:
        if self._thread and self._thread.is_alive():
            return
        if not self.retention_hours or self.telemetry_store_factory() is None:
            return
        self._stop_event.clear()
        self._thread = threading.Thread(target=self._loop, name="llm-telemetry-janitor", daemon=True)
        self._thread.start()

    def stop(self, timeout: float = 2.0) -> None:
        self._stop_event.set()
        if self._thread and self._thread.is_alive():
            self._thread.join(timeout=max(0.1, float(timeout)))

    def status(self) -> Dict[str, Any]:
        store = self.telemetry_store_factory()
        thread = self._thread
        with self._lock:
            last_run_at = self._last_run_at
            last_removed = self._last_removed
            last_error = self._last_error
        return {
            "enabled": bool(self.retention_hours > 0),
            "available": bool(store is not None),
            "running": bool(thread and thread.is_alive()),
            "retention_hours": self.retention_hours,
            "poll_sec": self.poll_sec,
            "last_run_at": last_run_at,
            "last_removed": last_removed,
            "last_error": last_error,
            "aggregate_store": "postgres" if store is not None else "disabled",
            # Raw event lines remain in per-job events.jsonl under shared job storage.
            "raw_event_stream": "job_events_jsonl",
        }

    def _loop(self) -> None:
        while not self._stop_event.is_set():
            try:
                self.run_once()
            except Exception:
                logger.exception("LLM telemetry janitor cycle failed")
            self._stop_event.wait(self.poll_sec)

    def run_once(self) -> int:
        store = self.telemetry_store_factory()
        if store is None or self.retention_hours <= 0:
            with self._lock:
                self._last_run_at = _now_iso()
                self._last_removed = 0
                self._last_error = None if self.retention_hours <= 0 else "telemetry_store_unavailable"
            return 0
        removed = 0
        error: Optional[str] = None
        try:
            removed = max(0, int(store.prune_older_than(self.retention_hours)))
        except Exception as exc:
            error = str(exc)
            with self._lock:
                self._last_run_at = _now_iso()
                self._last_removed = 0
                self._last_error = error
            raise
        with self._lock:
            self._last_run_at = _now_iso()
            self._last_removed = removed
            self._last_error = error
        return removed


class JobManager:
    """Background job queue manager for workflow execution processes."""

    def __init__(self, workers: int = DEFAULT_WORKERS):
        self.jobs: Dict[str, Job] = {}
        self.queue: queue.Queue = queue.Queue()
        self.lock = threading.Lock()
        self.workers: List[threading.Thread] = []
        _ensure_dirs()
        self._load_jobs_from_disk()
        self._cleanup_old_jobs()
        for idx in range(max(1, workers)):
            t = threading.Thread(target=self._worker_loop, args=(idx,), daemon=True)
            t.start()
            self.workers.append(t)

    def submit_job(self, payload: Dict[str, Any], owner: str) -> Job:
        job_id = uuid.uuid4().hex
        job_dir = os.path.join(JOB_ROOT, job_id)
        ensure_dir_permissions(job_dir, mode=0o700)
        log_path = os.path.join(job_dir, "job.log")
        events_path = os.path.join(job_dir, "events.jsonl")
        project_id = None
        if isinstance(payload, dict):
            _apply_user_job_defaults(payload, owner)
            project_id = payload.get("project_id") or payload.get("project")
            if project_id:
                project = access_store.get_project(project_id)
                if project:
                    project_name = project.get("name")
                    if project_name:
                        payload["project_name"] = project_name
                    if not payload.get("team_id"):
                        payload["team_id"] = project.get("team_id")
        job = Job(job_id=job_id, payload=payload, owner=owner, log_path=log_path, events_path=events_path)
        job.output_paths = self._resolve_output_paths(job)
        job.meta_path = os.path.join(job_dir, JOB_META_FILENAME)
        for path in (log_path, events_path, job.meta_path):
            try:
                if not os.path.exists(path):
                    with open(path, "a", encoding="utf-8"):
                        pass
                ensure_file_permissions(path, mode=0o600)
            except Exception:
                pass
        notify_email = payload.get("notify_email") or payload.get("notification_email") or ""
        if isinstance(notify_email, str):
            notify_email = notify_email.strip()
        else:
            notify_email = ""
        if notify_email and EMAIL_RE.match(notify_email):
            job.notify_email = notify_email
        job.persist(force=True)
        _audit_event(
            "job_submit",
            actor=owner,
            status="success",
            details={"job_id": job_id, "workflow": job.workflow},
        )
        with self.lock:
            self.jobs[job_id] = job
        self.queue.put(job_id)
        _notify_continuum_autoscaler()
        return job

    @staticmethod
    def _owner_key(owner: Optional[str]) -> str:
        cleaned = str(owner or "").strip()
        return cleaned or "__anonymous__"

    @staticmethod
    def _owner_label(owner_key: str) -> str:
        return "anonymous" if owner_key == "__anonymous__" else owner_key

    @staticmethod
    def _increment_owner_count(counts: Dict[str, int], owner: Optional[str]) -> None:
        owner_key = JobManager._owner_key(owner)
        counts[owner_key] = int(counts.get(owner_key) or 0) + 1

    @staticmethod
    def _top_owner_counts(counts: Dict[str, int], *, limit: int = 3) -> List[Dict[str, Any]]:
        cleaned_limit = max(1, int(limit))
        ranked = sorted(counts.items(), key=lambda item: (-int(item[1] or 0), item[0]))
        return [
            {"owner": JobManager._owner_label(owner_key), "count": int(count or 0)}
            for owner_key, count in ranked[:cleaned_limit]
        ]

    @staticmethod
    def _owner_ratio(total: int, ranked: List[Dict[str, Any]]) -> float:
        if total <= 0 or not ranked:
            return 0.0
        top_count = max(0, int(ranked[0].get("count") or 0))
        return round(top_count / float(total), 3)

    def queue_snapshot(self, *, top_limit: int = 3) -> Dict[str, Any]:
        """Return a read-only view of queue depth and owner distribution.

        The core job queue is still FIFO. This snapshot exists so admin and
        workers telemetry can show whether queue pressure is concentrated under
        one owner before we consider owner-aware admission for the core queue.
        """

        with self.lock:
            queue_depth = max(0, int(self.queue.qsize()))
            jobs_snapshot = list(self.jobs.values())
            workers = max(1, len(self.workers))

        queued = 0
        running = 0
        paused = 0
        queued_by_owner: Dict[str, int] = {}
        running_by_owner: Dict[str, int] = {}
        paused_by_owner: Dict[str, int] = {}
        active_by_owner: Dict[str, int] = {}

        for job in jobs_snapshot:
            status = (getattr(job, "status", "") or "").strip().lower()
            if status not in {"queued", "running", "paused"}:
                continue
            owner = getattr(job, "owner", "")
            self._increment_owner_count(active_by_owner, owner)
            if status == "queued":
                queued += 1
                self._increment_owner_count(queued_by_owner, owner)
            elif status == "running":
                running += 1
                self._increment_owner_count(running_by_owner, owner)
            elif status == "paused":
                paused += 1
                self._increment_owner_count(paused_by_owner, owner)

        top_queued_owners = self._top_owner_counts(queued_by_owner, limit=top_limit)
        top_running_owners = self._top_owner_counts(running_by_owner, limit=top_limit)
        top_active_owners = self._top_owner_counts(active_by_owner, limit=top_limit)

        return {
            "queue_depth": queue_depth,
            "queued": queued,
            "running": running,
            "paused": paused,
            "workers": workers,
            "queued_owner_count": len(queued_by_owner),
            "running_owner_count": len(running_by_owner),
            "paused_owner_count": len(paused_by_owner),
            "active_owner_count": len(active_by_owner),
            "top_queued_owners": top_queued_owners,
            "top_running_owners": top_running_owners,
            "top_active_owners": top_active_owners,
            "queued_owner_skew_ratio": self._owner_ratio(queued, top_queued_owners),
            "running_owner_skew_ratio": self._owner_ratio(running, top_running_owners),
            "active_owner_skew_ratio": self._owner_ratio(queued + running + paused, top_active_owners),
            "single_owner_queue_pressure": bool(queued > workers and len(queued_by_owner) == 1),
        }

    def get_job(self, job_id: str, owner: Optional[str] = None) -> Optional[Job]:
        with self.lock:
            job = self.jobs.get(job_id)
        if not job:
            return None
        if owner and job.owner != owner:
            return None
        return job

    def list_jobs(self, status: Optional[str] = None, owner: Optional[str] = None) -> List[Job]:
        with self.lock:
            jobs = list(self.jobs.values())
        if owner:
            jobs = [job for job in jobs if job.owner == owner]
        if status:
            jobs = [job for job in jobs if job.status == status]
        return sorted(jobs, key=lambda j: _timestamp_sort_key(j.created_at), reverse=True)

    @staticmethod
    def _job_team_id(job: Job) -> Optional[str]:
        if not isinstance(job.payload, dict):
            return None
        return job.payload.get("team_id")

    @staticmethod
    def _effective_team_id(job: Job) -> Optional[str]:
        if not isinstance(job.payload, dict):
            return None
        if job.payload.get("token_scope") == "personal":
            return None
        return job.payload.get("team_id")

    @staticmethod
    def _reserved_split(job: Job) -> Tuple[int, int]:
        reserved_user = int(getattr(job, "token_reserved_user", 0) or 0)
        reserved_team = int(getattr(job, "token_reserved_team", 0) or 0)
        if reserved_user == 0 and reserved_team == 0:
            legacy = int(getattr(job, "token_reserved", 0) or 0)
            if legacy:
                reserved_user = legacy
        return reserved_user, reserved_team

    @staticmethod
    def _split_usage(run_total: int, reserved_user: int, reserved_team: int, team_id: Optional[str]) -> Tuple[int, int]:
        if run_total <= 0:
            return 0, 0
        total_reserved = reserved_user + reserved_team
        if total_reserved > 0:
            user_share = int(round(run_total * (reserved_user / total_reserved)))
            user_share = max(0, min(user_share, run_total))
            team_share = run_total - user_share
            return user_share, team_share
        if team_id:
            return 0, run_total
        return run_total, 0

    def reserved_tokens(
        self,
        owner: Optional[str] = None,
        *,
        team_id: Optional[str] = None,
        source: Optional[str] = None,
    ) -> int:
        with self.lock:
            jobs = list(self.jobs.values())
        total = 0
        for job in jobs:
            if owner and job.owner != owner:
                continue
            if team_id and self._effective_team_id(job) != team_id:
                continue
            if job.status not in {"queued", "running", "paused"}:
                continue
            if source == "user":
                reserved_user, _ = self._reserved_split(job)
                if reserved_user:
                    total += reserved_user
                continue
            if source == "team":
                _, reserved_team = self._reserved_split(job)
                if reserved_team:
                    total += reserved_team
                continue
            if job.token_reserved:
                total += int(job.token_reserved or 0)
        return total

    def in_use_tokens(
        self,
        owner: Optional[str] = None,
        *,
        team_id: Optional[str] = None,
        source: Optional[str] = None,
    ) -> int:
        with self.lock:
            jobs = list(self.jobs.values())
        total = 0
        for job in jobs:
            if owner and job.owner != owner:
                continue
            if team_id and self._effective_team_id(job) != team_id:
                continue
            if job.status not in {"running", "paused"}:
                continue
            usage = job.metrics.get("token_usage") if isinstance(job.metrics, dict) else {}
            run_total = usage.get("total") if isinstance(usage, dict) else None
            if not run_total:
                continue
            run_total = int(run_total or 0)
            if source is None:
                total += run_total
                continue
            reserved_user, reserved_team = self._reserved_split(job)
            user_share, team_share = self._split_usage(run_total, reserved_user, reserved_team, self._effective_team_id(job))
            if source == "user":
                total += user_share
            elif source == "team":
                total += team_share
        return total

    def reconcile_tokens(self) -> None:
        with self.lock:
            jobs = list(self.jobs.values())
        for job in jobs:
            if job.status in {"completed", "failed", "stopped"} and job.token_status != "settled":
                self._settle_tokens(job)

    def _read_job_meta(self, path: str) -> Optional[Dict[str, Any]]:
        try:
            with open(path, "r", encoding="utf-8") as handle:
                data = json.load(handle)
            return data if isinstance(data, dict) else None
        except Exception:
            return None

    def _remember_loaded_job(self, job: Job) -> None:
        existing = self.jobs.get(job.job_id)
        if existing and _timestamp_sort_key(existing.updated_at) > _timestamp_sort_key(job.updated_at):
            return
        self.jobs[job.job_id] = job

    def _load_jobs_from_central_store(self) -> None:
        central_job_store = _central_job_store()
        if central_job_store is None:
            return
        try:
            rows = central_job_store.list_jobs(limit=10000)
        except Exception as exc:
            logger.debug("central job metadata load skipped: %s", exc)
            return
        for data in rows:
            if not isinstance(data, dict):
                continue
            job_id = str(data.get("id") or data.get("job_id") or "").strip()
            if not job_id:
                continue
            meta_path = os.path.join(JOB_ROOT, job_id, JOB_META_FILENAME)
            try:
                job = Job.from_persisted(data, meta_path)
            except Exception:
                continue
            self._remember_loaded_job(job)

    def _load_jobs_from_disk(self) -> None:
        self._load_jobs_from_central_store()
        if not os.path.isdir(JOB_ROOT):
            return
        central_job_store = _central_job_store()
        skip_dirs = {"projects", "secrets", "workspaces"}
        for entry in os.listdir(JOB_ROOT):
            if entry in skip_dirs:
                continue
            job_dir = os.path.join(JOB_ROOT, entry)
            if not os.path.isdir(job_dir):
                continue
            meta_path = os.path.join(job_dir, JOB_META_FILENAME)
            if not os.path.exists(meta_path):
                continue
            data = self._read_job_meta(meta_path)
            if not data:
                continue
            try:
                job = Job.from_persisted(data, meta_path)
            except Exception:
                continue
            self._remember_loaded_job(job)
            if central_job_store is not None:
                try:
                    central_job_store.upsert(job.to_persisted_dict())
                except Exception as exc:
                    logger.debug("central job metadata backfill skipped for %s: %s", job.job_id, exc)

        for job in self.jobs.values():
            if job.status in {"queued", "running", "paused"}:
                job.status = "stopped"
                job.updated_at = _now_iso()
                if not job.finished_at:
                    job.finished_at = job.updated_at
                job.persist(force=True)

    def _cleanup_old_jobs(self) -> None:
        if JOB_RETENTION_DAYS <= 0:
            return
        cutoff = time.time() - JOB_RETENTION_DAYS * 86400
        removed = 0
        central_job_store = _central_job_store()
        for job in list(self.jobs.values()):
            if job.status not in {"completed", "failed", "stopped"}:
                continue
            finished = _parse_timestamp(job.finished_at) if job.finished_at else None
            ts = finished.timestamp() if finished else None
            if ts is None:
                try:
                    ts = os.path.getmtime(os.path.join(JOB_ROOT, job.job_id))
                except Exception:
                    ts = None
            if ts is None or ts >= cutoff:
                continue
            try:
                job_dir = os.path.join(JOB_ROOT, job.job_id)
                shutil.rmtree(job_dir, ignore_errors=True)
                with self.lock:
                    self.jobs.pop(job.job_id, None)
                if central_job_store is not None:
                    try:
                        central_job_store.delete(job.job_id)
                    except Exception as exc:
                        logger.debug("central job metadata cleanup skipped for %s: %s", job.job_id, exc)
                removed += 1
            except Exception:
                continue
        if removed:
            _audit_event("job_retention_cleanup", actor=None, status="success", details={"removed": removed})

    def pause_job(self, job_id: str) -> bool:
        job = self.get_job(job_id)
        if not job:
            return False
        if not job.process or job.status != "running":
            job.set_status("paused")
            return True
        if os.name != "nt":
            try:
                os.kill(job.process.pid, signal.SIGSTOP)
                job.set_status("paused")
                job.append_log("Paused by user")
                return True
            except Exception:
                return False
        return False

    def resume_job(self, job_id: str) -> bool:
        job = self.get_job(job_id)
        if not job:
            return False
        if job.process and job.process.poll() is not None:
            job.process = None
            job.pid = None
        if not job.process:
            if not isinstance(job.payload, dict):
                job.payload = {}
            job.payload["include_global_requirements"] = True
            if job.status == "completed":
                self._apply_project_limits(job)
                summary = _completion_summary_from_output(job.output_paths.get("primary"))
                if isinstance(summary, dict):
                    steps_applied = _safe_int(summary.get("steps_applied"))
                    needs_more = bool(summary.get("needs_more_iterations"))
                    max_steps_reached = bool(summary.get("max_steps_reached"))
                    iterations_exhausted = summary.get("iterations_exhausted_sources") or []
                    if max_steps_reached:
                        current_steps = _safe_int(job.payload.get("project_max_steps"))
                        bump = max(25, int(current_steps * 0.5)) if current_steps else 50
                        job.payload["project_max_steps"] = max(current_steps + bump, 50)
                        job.append_log(f"Auto-bump: project_max_steps -> {job.payload['project_max_steps']}")
                    if isinstance(iterations_exhausted, list) and iterations_exhausted:
                        current_iters = _safe_int(job.payload.get("project_iterations"))
                        bump = max(1, int(current_iters * 0.5)) if current_iters else 2
                        job.payload["project_iterations"] = max(current_iters + bump, 2)
                        job.append_log(f"Auto-bump: project_iterations -> {job.payload['project_iterations']}")
                    if steps_applied <= 0 and needs_more and not max_steps_reached and not iterations_exhausted:
                        current_tokens = _safe_int(job.payload.get("llm_max_tokens"), DEFAULT_LLM_MAX_TOKENS)
                        cap = RESUME_LLM_MAX_TOKENS_CAP if RESUME_LLM_MAX_TOKENS_CAP > 0 else current_tokens
                        target = max(current_tokens * 2, DEFAULT_LLM_MAX_TOKENS)
                        if cap:
                            target = min(target, cap)
                        if target > current_tokens:
                            job.payload["llm_max_tokens"] = target
                            job.append_log(f"Auto-bump: llm_max_tokens -> {target}")
                with job.lock:
                    job.stages = [stage for stage in job.stages if stage.name != "finalize"]
                    execute_stage = None
                    for stage in job.stages:
                        if stage.name == "execute":
                            execute_stage = stage
                            break
                    if execute_stage:
                        execute_stage.status = "queued"
                        execute_stage.started_at = None
                        execute_stage.finished_at = None
                        execute_stage.message = None
                    else:
                        job.stages.append(Stage(name="execute", status="queued"))
                    job.updated_at = _now_iso()
                job.persist(force=True)
            job.set_status("queued")
            self.queue.put(job_id)
            _notify_continuum_autoscaler()
            return True
        if os.name != "nt":
            try:
                os.kill(job.process.pid, signal.SIGCONT)
                job.set_status("running")
                job.append_log("Resumed by user")
                return True
            except Exception:
                return False
        return False

    def stop_job(self, job_id: str) -> bool:
        job = self.get_job(job_id)
        if not job:
            return False
        job.stop_requested = True
        if job.process and job.process.poll() is None:
            try:
                job.process.terminate()
                job.append_log("Stop requested")
                return True
            except Exception:
                return False
        job.set_status("stopped")
        job.finished_at = _now_iso()
        job.persist(force=True)
        self._settle_tokens(job)
        return True

    def _derive_project_limits(self, job: "Job") -> Optional[Dict[str, int]]:
        payload = job.payload if isinstance(job.payload, dict) else {}
        workflow = payload.get("workflow") or job.workflow
        if workflow not in {"project_solver", "project"}:
            return None
        req_count = 0
        req_text = payload.get("requirements_text") if isinstance(payload.get("requirements_text"), str) else ""
        if req_text:
            req_count = _count_req_lines(req_text)
        if req_count <= 0:
            req_path = payload.get("requirements_path")
            if isinstance(req_path, str) and req_path:
                req_count = _count_req_lines(_read_file_limited(req_path, 250_000))
        if payload.get("include_global_requirements"):
            req_count += _global_requirements_count()
        iterations = payload.get("project_iterations")
        max_steps = payload.get("project_max_steps")
        try:
            iterations = int(iterations) if iterations else 0
        except Exception:
            iterations = 0
        try:
            max_steps = int(max_steps) if max_steps else 0
        except Exception:
            max_steps = 0
        if iterations <= 0:
            iterations = req_count if req_count > 0 else 3
        if max_steps <= 0:
            max_steps = max(25, req_count * 4) if req_count > 0 else 25
        return {"project_iterations": iterations, "project_max_steps": max_steps}

    def _apply_project_limits(self, job: "Job") -> None:
        limits = self._derive_project_limits(job)
        if not limits:
            return
        if not isinstance(job.payload, dict):
            job.payload = {}
        job.payload.update(limits)

    def _wait_for_process(self, job: Job, timeout: float = 4.0) -> None:
        if not job.process:
            return
        try:
            job.process.wait(timeout=timeout)
        except Exception:
            try:
                job.process.kill()
                job.process.wait(timeout=1.0)
            except Exception:
                pass

    def set_archived(self, job_id: str, archived: bool) -> bool:
        job = self.get_job(job_id)
        if not job:
            return False
        job.archived = bool(archived)
        job.archived_at = _now_iso() if job.archived else None
        job.updated_at = _now_iso()
        job.persist(force=True)
        return True

    def delete_job(self, job_id: str, owner: Optional[str] = None, stop_if_active: bool = False) -> bool:
        job = self.get_job(job_id, owner=owner)
        if not job:
            return False
        if job.status in {"queued", "running", "paused"}:
            if not stop_if_active:
                return False
            self.stop_job(job_id)
            self._wait_for_process(job)
        with self.lock:
            self.jobs.pop(job_id, None)
        job_dir = os.path.join(JOB_ROOT, job_id)
        try:
            shutil.rmtree(job_dir)
        except FileNotFoundError:
            pass
        except Exception:
            return False
        return True

    def restart_job(self, job_id: str) -> bool:
        job = self.get_job(job_id)
        if not job:
            return False
        if job.status in {"running", "paused"}:
            return False
        if not isinstance(job.payload, dict):
            job.payload = {}
        job.payload["include_global_requirements"] = True
        self._apply_project_limits(job)
        job.restart_count += 1
        job.stop_requested = False
        job.exit_code = None
        job.started_at = None
        job.finished_at = None
        job.progress = 0
        job.stages = []
        job.token_actual = 0
        job.token_debited = 0
        job.token_debited_user = 0
        job.token_debited_team = 0
        job.token_shortfall = 0
        job.token_shortfall_user = 0
        job.token_shortfall_team = 0
        if job.token_reserved:
            job.token_status = "reserved"
        job.metrics = {
            "token_usage": _default_token_usage_metrics(),
            "errors": 0,
            "resolved": 0,
            "warnings": 0,
            "queue_wait_sec": None,
            "runtime_sec": None,
        }
        job.append_log(f"--- restart #{job.restart_count} ---")
        job.set_status("queued")
        self.queue.put(job_id)
        _notify_continuum_autoscaler()
        return True

    def _worker_loop(self, worker_id: int) -> None:
        while True:
            job_id = self.queue.get()
            if job_id is None:
                self.queue.task_done()
                break
            job = self.get_job(job_id)
            if not job:
                self.queue.task_done()
                continue
            if job.status in {"stopped", "completed", "failed"}:
                self.queue.task_done()
                continue
            if job.status == "paused":
                self.queue.put(job_id)
                self.queue.task_done()
                time.sleep(0.5)
                continue
            self._run_job(job)
            self.queue.task_done()

    def _run_job(self, job: Job) -> None:
        job.set_status("running")
        job.started_at = _now_iso()
        job.token_status = "running" if job.token_reserved else job.token_status
        job.persist()
        job.metrics["queue_wait_sec"] = self._compute_wait_seconds(job)
        job.set_progress(5)
        try:
            self._prepare_repo(job)
            self._guardrail_check(job)
        except Exception as exc:
            job.append_log(f"Job blocked: {exc}")
            job.exit_code = 1
            job.set_status("failed")
            job.set_progress(100)
            job.finished_at = _now_iso()
            job.persist(force=True)
            self._settle_tokens(job)
            self._maybe_notify(job)
            return
        try:
            command = self._build_command(job)
        except Exception as exc:
            job.append_log(f"Failed to build command: {exc}")
            job.exit_code = 1
            job.set_status("failed")
            job.set_progress(100)
            job.finished_at = _now_iso()
            job.persist(force=True)
            self._settle_tokens(job)
            self._maybe_notify(job)
            return
        job.append_log("Starting job: " + " ".join(command))
        try:
            use_defaults = job.payload.get("use_default_secrets", True)
            job_secrets = job.payload.get("job_secrets") or {}
            owner_settings = _user_settings(job.owner)
            owner_role = _user_role(job.owner)
            owner_secret_env = _get_secret_store(job.owner).get_env() if job.owner else {}
            env = _build_job_runtime_env(
                job.owner,
                settings=owner_settings,
                role=owner_role,
                secret_env=owner_secret_env,
                use_default_secrets=bool(use_defaults),
                job_secrets=job_secrets,
            )
            env["REFINER_EMIT_USAGE_EVENTS"] = "1"
            env["REFINER_EVENTS_FILE"] = job.events_path
            if not _user_can_use_shared_llm_credentials(job.owner, role=owner_role):
                env["REFINER_AI_INCLUDE_CONFIGURED_CANDIDATES"] = "0"
            policy_mode = (
                str(job.payload.get("solver_command_policy_mode") or _settings_defaults_for_user(job.owner).get("command_policy_mode") or "standard")
                .strip()
                .lower()
            )
            if policy_mode in {"standard", "strict"}:
                env["REFINER_SOLVER_COMMAND_POLICY_MODE"] = policy_mode
            job.process = subprocess.Popen(
                command,
                cwd=BASE_DIR,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                bufsize=1,
                env=env,
            )
            job.pid = job.process.pid
        except Exception as exc:
            job.append_log(f"Failed to start process: {exc}")
            job.exit_code = 1
            job.set_status("failed")
            job.set_progress(100)
            job.finished_at = _now_iso()
            job.persist(force=True)
            self._settle_tokens(job)
            self._maybe_notify(job)
            return

        assert job.process.stdout is not None
        for line in job.process.stdout:
            if not line:
                continue
            if self._handle_event_line(job, line):
                continue
            self._update_metrics(job, line)
            job.append_log(line)

        job.exit_code = job.process.wait()
        job.process = None
        job.pid = None
        job.finished_at = _now_iso()
        job.metrics["runtime_sec"] = self._compute_runtime_seconds(job)
        completion_reason = None
        completion_summary = None
        solver_incomplete = False
        if job.workflow in {"project_solver", "project"}:
            completion_summary = _completion_summary_from_output(job.output_paths.get("primary"))
            completion_reason = _completion_reason_from_summary(completion_summary)
            solver_incomplete = bool(
                isinstance(completion_summary, dict)
                and completion_summary.get("needs_more_iterations")
            )
        if job.exit_code == 0 or (job.exit_code == 2 and solver_incomplete):
            try:
                self._finalize_repo(job)
            except Exception as exc:
                job.append_log(f"Repo finalize failed: {exc}")
        if job.stop_requested:
            job.set_status("stopped")
        elif job.exit_code == 0 and not solver_incomplete:
            job.set_status("completed")
        else:
            job.set_status("failed")
        job.set_progress(100)
        if completion_reason and job.status in {"completed", "failed"}:
            job.update_stage("finalize", completion_reason)
        job.append_log(f"Job finished with exit code {job.exit_code}")
        job.persist(force=True)
        self._settle_tokens(job)
        if job.status == "completed" and not completion_reason and job.token_shortfall > 0:
            job.update_stage("finalize", "tokens")
        self._maybe_notify(job)

    def _handle_event_line(self, job: Job, line: str) -> bool:
        if not line.startswith("__RAG_EVENT__ "):
            return False
        payload = line[len("__RAG_EVENT__ "):].strip()
        try:
            data = json.loads(payload)
        except json.JSONDecodeError:
            return False
        event_type = data.get("type")
        if event_type == "workflow_selected":
            workflow = data.get("workflow")
            if workflow:
                job.workflow = workflow
        elif event_type == "stage":
            stage_name = data.get("stage")
            status = data.get("status")
            progress = data.get("progress")
            message = data.get("message")
            if stage_name and status:
                job.update_stage(stage_name, status, message=message)
            if isinstance(progress, int):
                job.set_progress(progress)
        elif event_type == "workflow_complete":
            status = data.get("status")
            if status and job.status == "running":
                job.update_stage("finalize", status)
        elif event_type == "llm_request":
            _record_llm_request_telemetry(job.owner, data, team_id=self._effective_team_id(job))
        elif event_type == "token_usage":
            _merge_token_usage_event(job, data)
            job.persist()
        return True

    def _update_metrics(self, job: Job, line: str) -> None:
        if "ERROR" in line or "Traceback" in line:
            job.metrics["errors"] += 1
        if "WARNING" in line:
            job.metrics["warnings"] += 1
        if re.search(r"\brecover(?:ed|y)\b|\bretry(?:ing|ed)\b|\bfallback\b", line, re.IGNORECASE):
            job.metrics["resolved"] += 1

    def _reserve_tokens(self, job: Job, estimate: int) -> None:
        job.token_estimate = int(estimate or 0)
        reserve_estimate = _chargeable_token_estimate(job.owner, job.payload, job.token_estimate)
        token_scope = None
        if isinstance(job.payload, dict):
            token_scope = job.payload.get("token_scope")
        team_id = None if token_scope == "personal" else self._job_team_id(job)
        team_available = 0
        if team_id:
            team_snapshot = _team_token_snapshot(team_id)
            team_available = int(team_snapshot.get("available") or 0)
        remaining = int(reserve_estimate or 0)
        team_reserved = min(remaining, team_available) if team_id else 0
        remaining = max(0, remaining - team_reserved)
        user_reserved = remaining
        job.token_reserved_team = int(team_reserved)
        job.token_reserved_user = int(user_reserved)
        job.token_reserved = int(team_reserved + user_reserved)
        if team_reserved and user_reserved:
            job.token_source = "split"
        elif team_reserved:
            job.token_source = "team"
        elif user_reserved:
            job.token_source = "user"
        else:
            job.token_source = "none"
        job.token_status = "reserved" if job.token_reserved > 0 else "none"
        if team_reserved and user_reserved:
            job.append_log(
                f"Reserved {job.token_reserved} tokens for estimate (team {team_reserved}, personal {user_reserved})."
            )
        elif team_reserved:
            job.append_log(f"Reserved {job.token_reserved} team tokens for estimate.")
        elif user_reserved:
            job.append_log(f"Reserved {job.token_reserved} personal tokens for estimate.")
        elif job.token_estimate > 0 and reserve_estimate <= 0:
            job.append_log(
                "Estimated token usage tracked, but no platform tokens were reserved because only "
                "user-supplied or local AI providers are currently selected."
            )
        else:
            job.append_log("No tokens reserved for estimate.")
        base_meta = {"job_id": job.job_id, "estimate": job.token_estimate, "workflow": job.workflow}
        if team_reserved and team_id:
            _record_token_event(
                "team",
                team_id,
                "reserve",
                0,
                {**base_meta, "reserved": team_reserved, "team_id": team_id, "source": "team"},
                request_id=_chain_request_id("job-reserve", job.job_id, "team"),
            )
        if user_reserved:
            _record_token_event(
                "user",
                job.owner,
                "reserve",
                0,
                {**base_meta, "reserved": user_reserved, "team_id": team_id, "source": "user"},
                request_id=_chain_request_id("job-reserve", job.job_id, "user"),
            )
        job.persist(force=True)

    def _release_tokens(self, job: Job, reason: str = "release") -> None:
        if job.token_reserved <= 0 and not job.token_reserved_user and not job.token_reserved_team:
            return
        team_id = self._effective_team_id(job)
        reserved_user, reserved_team = self._reserved_split(job)
        if reserved_team and team_id:
            _record_token_event(
                "team",
                team_id,
                "release",
                0,
                {"job_id": job.job_id, "reserved": reserved_team, "reason": reason, "team_id": team_id},
                request_id=_chain_request_id("job-release", job.job_id, reason, "team"),
            )
        if reserved_user:
            _record_token_event(
                "user",
                job.owner,
                "release",
                0,
                {"job_id": job.job_id, "reserved": reserved_user, "reason": reason, "team_id": team_id},
                request_id=_chain_request_id("job-release", job.job_id, reason, "user"),
            )
        job.token_reserved = 0
        job.token_reserved_user = 0
        job.token_reserved_team = 0
        job.persist(force=True)

    def _settle_tokens(self, job: Job) -> None:
        if job.token_status == "settled":
            return
        token_usage = job.metrics.get("token_usage", {}) if isinstance(job.metrics, dict) else {}
        actual = int(token_usage.get("total") or 0)
        chargeable_actual = int(token_usage.get("chargeable_total") or 0)
        job.token_actual = actual
        team_id = self._effective_team_id(job)
        reserved_user, reserved_team = self._reserved_split(job)
        if job.token_reserved:
            self._release_tokens(job, reason="settle")
        if chargeable_actual > 0:
            user_share, team_share = self._split_usage(chargeable_actual, reserved_user, reserved_team, team_id)
            base_meta = {"job_id": job.job_id, "estimate": job.token_estimate, "workflow": job.workflow}
            team_shortfall = 0
            team_debited = 0
            if team_id and team_share > 0:
                entry_team = _record_token_event(
                    "team",
                    team_id,
                    "debit",
                    -team_share,
                    {**base_meta, "team_id": team_id, "source": "team"},
                    request_id=_chain_request_id("job-debit", job.job_id, "team"),
                )
                team_shortfall = int(entry_team.get("shortfall") or 0)
                team_debited = team_share - team_shortfall
            user_request = int(user_share + team_shortfall)
            user_shortfall = 0
            user_debited = 0
            if user_request > 0:
                entry_user = _record_token_event(
                    "user",
                    job.owner,
                    "debit",
                    -user_request,
                    {**base_meta, "team_id": team_id, "source": "user"},
                    request_id=_chain_request_id("job-debit", job.job_id, "user"),
                )
                user_shortfall = int(entry_user.get("shortfall") or 0)
                user_debited = user_request - user_shortfall
            total_shortfall = user_shortfall if user_request > 0 else team_shortfall
            job.token_debited_user = user_debited
            job.token_debited_team = team_debited
            job.token_debited = team_debited + user_debited
            job.token_shortfall_user = user_shortfall
            job.token_shortfall_team = team_shortfall
            job.token_shortfall = total_shortfall
            if team_debited and user_debited:
                job.token_source = "split"
            elif team_debited:
                job.token_source = "team"
            elif user_debited:
                job.token_source = "user"
            else:
                job.token_source = "none"
            if team_debited or user_debited or total_shortfall:
                job.append_log(
                    "Debited "
                    f"{job.token_debited} tokens (team {team_debited}, personal {user_debited}, shortfall {total_shortfall})."
                )
        elif actual > 0:
            job.append_log(
                "Tracked token usage settled without a platform debit because the job used only "
                "user-supplied or local AI providers."
            )
        job.token_status = "settled"
        job.persist(force=True)
        _notify_portal_usage(job)

    def _maybe_notify(self, job: Job) -> None:
        if job.status not in {"completed", "failed"}:
            return
        if job.notified_at or job.notification_error:
            return
        if _is_user_active(job.owner):
            return
        recipient = job.notify_email or user_store.get_email(job.owner)
        if not recipient:
            job.notification_error = "No notification email configured."
            job.persist(force=True)
            return
        if not EMAIL_RE.match(recipient):
            job.notification_error = "Invalid notification email."
            job.persist(force=True)
            return
        subject = f"Refiner job {job.status}: {job.project_name or job.job_id[:8]}"
        body = _job_notification_body(job)
        error = _send_email(recipient, subject, body)
        if error:
            job.notification_error = error
            job.append_log(error)
        else:
            job.notified_at = _now_iso()
            job.notified_via = "email"
            job.append_log(f"Notification email sent to {recipient}")
        job.persist(force=True)

    def _build_command(self, job: Job) -> List[str]:
        payload = job.payload
        workflow = payload.get("workflow") or "project_solver"
        if workflow in {"project_solver", "project"} and not payload.get("llm_max_tokens"):
            payload["llm_max_tokens"] = DEFAULT_LLM_MAX_TOKENS
        command = [
            os.getenv("REFINER_PYTHON", sys.executable),
            "-m",
            os.getenv("REFINER_RUNNER_MODULE", "refiner.run_refiner"),
            "--log-file",
            job.log_path,
            "--emit-events",
            "--events-file",
            job.events_path,
        ]
        if payload.get("verbose", True):
            command.append("--verbose")
        if payload.get("debug"):
            command.append("--debug")

        self._add_llm_args(command, payload)

        if workflow == "project_solver" or workflow == "project":
            project_root = self._resolve_project_root(job)
            command += ["--project", project_root]
            requirements_path = self._resolve_requirements(job)
            if requirements_path:
                command += ["--requirements", requirements_path]
            if payload.get("project_run"):
                command.append("--project-run")
            if payload.get("project_max_steps"):
                command += ["--project-max-steps", str(payload.get("project_max_steps"))]
            if payload.get("project_iterations"):
                command += ["--project-iterations", str(payload.get("project_iterations"))]
            if payload.get("project_output_dir"):
                command += ["--project-output-dir", str(payload.get("project_output_dir"))]
            if payload.get("codingagent"):
                command += ["--codingagent", payload.get("codingagent")]
            if payload.get("codingagent_fallback"):
                command += ["--codingagent-fallback", payload.get("codingagent_fallback")]
            if payload.get("codingagent_model"):
                command += ["--codingagent-model", payload.get("codingagent_model")]
            if payload.get("codingagent_reasoning_effort"):
                command += ["--codingagent-reasoning-effort", payload.get("codingagent_reasoning_effort")]
            if job.output_paths.get("primary"):
                command += ["--output", job.output_paths["primary"]]
        elif workflow == "topic_research":
            source_path = self._resolve_topic_source(job)
            command += ["--topic-research", source_path]
            if job.output_paths.get("primary"):
                command += ["--output", job.output_paths["primary"]]
            if payload.get("max_iterations"):
                command += ["--max-iterations", str(payload.get("max_iterations"))]
            for ctx in payload.get("context_sources", []) or []:
                command += ["--context", ctx]
            if payload.get("references_output"):
                command += ["--references-output", payload.get("references_output")]
        elif workflow == "jira_analysis":
            command.append("--analyze-jira")
            if payload.get("projects"):
                command += ["--projects", payload.get("projects")]
            if payload.get("jql"):
                command += ["--jql", payload.get("jql")]
            if payload.get("action_plan"):
                command.append("--action-plan")
            if payload.get("dry_run"):
                command.append("--dry-run")
            if payload.get("post_comments"):
                command.append("--post-comments")
            if payload.get("post_target"):
                command += ["--post-target", payload.get("post_target")]
            if job.output_paths.get("primary"):
                command += ["--output", job.output_paths["primary"]]
        elif workflow == "confluence_analysis":
            command.append("--analyze-confluence")
            if payload.get("space"):
                command += ["--space", payload.get("space")]
            if payload.get("use_rovo"):
                command.append("--use-rovo")
            if payload.get("action_plan"):
                command.append("--action-plan")
            if payload.get("dry_run"):
                command.append("--dry-run")
            if payload.get("post_comments"):
                command.append("--post-comments")
            if payload.get("post_target"):
                command += ["--post-target", payload.get("post_target")]
            if job.output_paths.get("primary"):
                command += ["--output", job.output_paths["primary"]]
        elif workflow == "delivery_pipeline":
            command.append("--delivery")
            project_root = self._resolve_project_root(job)
            command += ["--project", project_root]
            if payload.get("delivery_config"):
                command += ["--delivery-config", payload.get("delivery_config")]
            if payload.get("delivery_run"):
                command.append("--delivery-run")
            if payload.get("delivery_allow_unfinished"):
                command.append("--delivery-allow-unfinished")
            if payload.get("delivery_enable_interim"):
                command.append("--delivery-enable-interim")
            if payload.get("delivery_project_solution"):
                command += ["--delivery-project-solution", payload.get("delivery_project_solution")]
            if job.output_paths.get("primary"):
                command += ["--output", job.output_paths["primary"]]
        elif workflow == "jira_stats":
            pass

        if payload.get("disable_jira"):
            command.append("--disable-jira")
        if payload.get("disable_confluence"):
            command.append("--disable-confluence")

        extra_args = payload.get("extra_args")
        if isinstance(extra_args, str) and extra_args.strip():
            command += shlex.split(extra_args)
        return command

    def _resolve_output_paths(self, job: Job) -> Dict[str, str]:
        job_dir = os.path.join(JOB_ROOT, job.job_id)
        workflow = job.payload.get("workflow") or "project_solver"
        override = job.payload.get("output_path") or job.payload.get("output")
        if override:
            return {"primary": override}
        if workflow in {"project_solver", "project"}:
            return {"primary": os.path.join(job_dir, "project_solution.json")}
        if workflow == "topic_research":
            return {"primary": job.payload.get("topic_output") or os.path.join(job_dir, "researched_document.md")}
        if workflow == "jira_analysis":
            return {"primary": job.payload.get("jira_output") or os.path.join(job_dir, "jira_report.html")}
        if workflow == "confluence_analysis":
            return {"primary": job.payload.get("confluence_output") or os.path.join(job_dir, "confluence_report.html")}
        if workflow == "delivery_pipeline":
            return {"primary": job.payload.get("pipeline_output") or os.path.join(job_dir, "pipeline_report.json")}
        return {}

    def _resolve_project_root(self, job: Job) -> str:
        project_root = job.payload.get("project_root") or job.payload.get("delivery_project_root")
        project_name = job.payload.get("project_name")
        create_project = bool(job.payload.get("create_project"))
        if project_root:
            if create_project and not os.path.exists(project_root):
                os.makedirs(project_root, exist_ok=True)
            return project_root
        if project_name:
            slug = _slugify(project_name)
            project_root = os.path.join(PROJECTS_ROOT, f"{slug}-{job.job_id[:8]}")
            os.makedirs(project_root, exist_ok=True)
            return project_root
        raise ValueError("Project root or project name is required for this workflow")

    def _prepare_repo(self, job: Job) -> None:
        payload = job.payload
        repo_input = (payload.get("repo_url") or payload.get("repo") or "").strip()
        if not repo_input:
            if self._maybe_create_starter_repo(job):
                return
            return
        workflow = payload.get("workflow") or "project_solver"
        if workflow not in {"project_solver", "project", "delivery_pipeline"}:
            return
        owner, repo = self._parse_repo_input(repo_input)
        if not owner or not repo:
            raise ValueError("Invalid GitHub repo input. Use owner/repo or full URL.")
        fork_org = (payload.get("fork_org") or DEFAULT_FORK_ORG).strip()
        token = self._get_github_token(job)
        skip_fork = bool(payload.get("skip_fork")) or owner == fork_org
        fork_repo = None
        fork = None
        if skip_fork:
            job.append_log(f"Using existing repo without fork: {owner}/{repo}")
            default_branch = payload.get("repo_branch") or "main"
            fork_org = owner
            fork_repo = repo
        else:
            if not token:
                raise ValueError("Missing GitHub token. Add GITHUB_TOKEN in Credentials Vault or per-job secrets.")
            fork_repo = self._build_fork_repo_name(repo, job.owner or "user", job.project_name or repo)
            job.append_log(f"Preparing GitHub workspace for {owner}/{repo} -> {fork_org}/{fork_repo}")
            fork = self._ensure_fork(owner, repo, fork_org, fork_repo, token, job)
            default_branch = fork.get("default_branch") or payload.get("repo_branch") or "main"
        workspace = os.path.join(WORKSPACE_ROOT, job.job_id, repo)
        if os.path.exists(workspace):
            try:
                shutil.rmtree(workspace)
            except Exception:
                pass
        os.makedirs(os.path.dirname(workspace), exist_ok=True)

        fork_name = fork.get("name") if fork else fork_repo
        clone_url = f"https://github.com/{fork_org}/{fork_name}.git"
        self._git_clone(clone_url, workspace, default_branch, token, job)
        branch_name = (payload.get("work_branch") or f"refiner/{job.job_id[:8]}").strip()
        self._git_checkout(workspace, branch_name, job)

        project_subdir = (payload.get("repo_subdir") or "").strip()
        project_root = os.path.join(workspace, project_subdir) if project_subdir else workspace
        if not os.path.isdir(project_root):
            raise ValueError(f"Project subdir does not exist: {project_subdir or '.'}")

        requirements_rel = (payload.get("requirements_relpath") or "").strip()
        if requirements_rel:
            req_path = os.path.join(workspace, requirements_rel)
            if not os.path.exists(req_path):
                raise ValueError(f"Requirements path not found: {requirements_rel}")
            payload["requirements_path"] = req_path
        elif payload.get("requirements_text"):
            req_path = os.path.join(workspace, "requirements.md")
            with open(req_path, "w", encoding="utf-8") as handle:
                handle.write(payload.get("requirements_text"))
            payload["requirements_path"] = req_path

        payload["project_root"] = project_root
        job.repo_info = {
            "source": "github",
            "owner": owner,
            "repo": repo,
            "fork_org": fork_org,
            "fork_repo": fork_name,
            "branch": branch_name,
            "workspace": workspace,
            "clone_url": clone_url,
            "skip_fork": skip_fork,
        }
        if not job.project_name:
            job.project_name = repo

    def _guardrail_check(self, job: Job) -> None:
        payload = job.payload
        text = (payload.get("requirements_text") or "").strip()
        if text:
            reason = _guardrail_scan(text)
            if reason:
                raise ValueError(f"Guardrail blocked requirements: {reason}")
            return
        req_path = (payload.get("requirements_path") or "").strip()
        if req_path and os.path.exists(req_path):
            content = _read_file_limited(req_path, REQUIREMENTS_MAX_BYTES)
            reason = _guardrail_scan(content)
            if reason:
                raise ValueError(f"Guardrail blocked requirements file: {reason}")

    def _maybe_create_starter_repo(self, job: Job) -> bool:
        payload = job.payload
        workflow = payload.get("workflow") or "project_solver"
        if workflow not in {"project_solver", "project"}:
            return False
        if payload.get("project_root"):
            return False
        requirements_text = (payload.get("requirements_text") or "").strip()
        requirements_path = (payload.get("requirements_path") or "").strip()
        if not requirements_text and not requirements_path:
            return False
        project_name = job.project_name or payload.get("project_name") or ""
        if not project_name:
            raise ValueError("Project name is required to create a starter repo.")

        fork_org = (payload.get("fork_org") or DEFAULT_FORK_ORG).strip()
        token = self._get_github_token(job)
        if not token:
            raise ValueError("Missing GitHub token. Add GITHUB_TOKEN in Credentials Vault or per-job secrets.")

        repo_name = self._build_fork_repo_name(_slugify(project_name), job.owner or "user", project_name)
        job.append_log(f"Creating starter repo {fork_org}/{repo_name}")
        repo_data = self._ensure_repo_exists(fork_org, repo_name, token, job, payload)
        default_branch = repo_data.get("default_branch") or "main"
        workspace = os.path.join(WORKSPACE_ROOT, job.job_id, repo_name)
        if os.path.exists(workspace):
            try:
                shutil.rmtree(workspace)
            except Exception:
                pass
        os.makedirs(os.path.dirname(workspace), exist_ok=True)

        clone_url = f"https://github.com/{fork_org}/{repo_name}.git"
        self._git_clone(clone_url, workspace, default_branch, token, job)
        branch_name = (payload.get("work_branch") or f"refiner/{job.job_id[:8]}").strip()
        self._git_checkout(workspace, branch_name, job)

        requirements_rel = (payload.get("requirements_relpath") or "requirements.md").strip()
        req_path = os.path.join(workspace, requirements_rel)
        os.makedirs(os.path.dirname(req_path), exist_ok=True)
        if requirements_text:
            with open(req_path, "w", encoding="utf-8") as handle:
                handle.write(requirements_text)
        else:
            if not os.path.exists(requirements_path):
                raise ValueError("Requirements path not found on server.")
            shutil.copyfile(requirements_path, req_path)
        payload["requirements_path"] = req_path
        payload["project_root"] = workspace

        job.repo_info = {
            "source": "starter",
            "owner": fork_org,
            "repo": repo_name,
            "fork_org": fork_org,
            "fork_repo": repo_name,
            "branch": branch_name,
            "workspace": workspace,
            "clone_url": clone_url,
        }
        if not job.project_name:
            job.project_name = project_name
        return True

    def _finalize_repo(self, job: Job) -> None:
        if not job.repo_info:
            return
        workspace = job.repo_info.get("workspace")
        branch = job.repo_info.get("branch")
        if not workspace or not branch:
            return
        if not self._git_has_changes(workspace, job):
            job.append_log("No git changes detected; skipping commit/push.")
            return
        author_name = job.payload.get("git_author_name") or "refiner-bot"
        author_email = job.payload.get("git_author_email") or "automation@neuralmimicry.ai"
        commit_message = job.payload.get("commit_message") or f"Refiner updates ({job.job_id[:8]})"
        self._git_config(workspace, author_name, author_email, job)
        self._git_add_all(workspace, job)
        self._git_commit(workspace, commit_message, job)
        token = self._get_github_token(job)
        self._git_push(workspace, branch, token, job)
        fork_org = job.repo_info.get("fork_org")
        fork_repo = job.repo_info.get("fork_repo") or job.repo_info.get("repo")
        if fork_org and fork_repo:
            repo_url = f"https://github.com/{fork_org}/{fork_repo}"
            job.append_log(f"Repo ready: {repo_url} (branch: {branch})")
            job.repo_info["repo_url"] = repo_url

    def _get_github_token(self, job: Job) -> Optional[str]:
        job_secrets = job.payload.get("job_secrets") or []
        candidates = ["GITHUB_TOKEN", "GH_TOKEN", "GITHUB_PAT"]
        if isinstance(job_secrets, list):
            for entry in job_secrets:
                if not isinstance(entry, dict):
                    continue
                name = (entry.get("name") or "").strip()
                if name in candidates:
                    return (entry.get("value") or "").strip()
        elif isinstance(job_secrets, dict):
            for key in candidates:
                if key in job_secrets:
                    return str(job_secrets.get(key) or "").strip()
        if job.owner:
            env = _get_secret_store(job.owner).get_env()
            for key in candidates:
                if env.get(key):
                    return env.get(key)
        return None

    @staticmethod
    def _parse_repo_input(value: str) -> Optional[tuple]:
        value = value.strip()
        if value.startswith("http://") or value.startswith("https://"):
            match = re.match(r"https?://[^/]+/(?P<owner>[^/]+)/(?P<repo>[^/]+?)(?:\\.git)?$", value)
            if not match:
                return None, None
            return match.group("owner"), match.group("repo")
        if value.count("/") == 1:
            owner, repo = value.split("/", 1)
            repo = repo.strip()
            if repo.endswith(".git"):
                repo = repo[:-4]
            return owner.strip(), repo
        return None, None

    def _ensure_fork(
        self,
        owner: str,
        repo: str,
        fork_org: str,
        fork_repo: str,
        token: str,
        job: Job,
    ) -> Dict[str, Any]:
        headers = {"Authorization": f"token {token}", "Accept": "application/vnd.github+json"}
        fork_url = f"https://api.github.com/repos/{fork_org}/{fork_repo}"
        parent_full_name = f"{owner}/{repo}"
        resp = _http_request_with_retry("GET", fork_url, headers=headers, timeout=20)
        if resp.status_code == 200:
            data = resp.json()
            parent = data.get("parent") or {}
            if parent.get("full_name") and parent.get("full_name") != parent_full_name:
                raise ValueError(f"Existing fork {fork_org}/{fork_repo} is not based on {parent_full_name}.")
            return data
        create_url = f"https://api.github.com/repos/{owner}/{repo}/forks"
        payload = {"organization": fork_org}
        create_resp = _http_request_with_retry(
            "POST",
            create_url,
            headers=headers,
            json_body=payload,
            timeout=20,
        )
        if create_resp.status_code not in (202, 201):
            raise ValueError(f"Fork failed: {create_resp.status_code} {create_resp.text}")
        for _ in range(20):
            time.sleep(2)
            resp = _http_request_with_retry("GET", fork_url, headers=headers, timeout=20, retries=0)
            if resp.status_code == 200:
                return resp.json()
        # Fork may exist under default name; attempt rename if needed
        default_fork_url = f"https://api.github.com/repos/{fork_org}/{repo}"
        resp = _http_request_with_retry("GET", default_fork_url, headers=headers, timeout=20)
        if resp.status_code == 200:
            data = resp.json()
            parent = data.get("parent") or {}
            if parent.get("full_name") != parent_full_name:
                raise ValueError(f"Existing fork {fork_org}/{repo} is not based on {parent_full_name}.")
            if fork_repo != repo:
                rename_url = f"https://api.github.com/repos/{fork_org}/{repo}"
                rename_resp = _http_request_with_retry(
                    "PATCH",
                    rename_url,
                    headers=headers,
                    json_body={"name": fork_repo},
                    timeout=20,
                )
                if rename_resp.status_code not in (200, 201):
                    raise ValueError(f"Fork rename failed: {rename_resp.status_code} {rename_resp.text}")
                return rename_resp.json()
            return data
        raise ValueError("Fork not ready yet. Try again shortly.")

    def _ensure_repo_exists(
        self,
        org: str,
        repo_name: str,
        token: str,
        job: Job,
        payload: Dict[str, Any],
    ) -> Dict[str, Any]:
        headers = {"Authorization": f"token {token}", "Accept": "application/vnd.github+json"}
        repo_url = f"https://api.github.com/repos/{org}/{repo_name}"
        resp = _http_request_with_retry("GET", repo_url, headers=headers, timeout=20)
        if resp.status_code == 200:
            return resp.json()
        private_flag = payload.get("repo_private")
        if private_flag is None:
            private_flag = DEFAULT_REPO_PRIVATE
        create_url = f"https://api.github.com/orgs/{org}/repos"
        body = {
            "name": repo_name,
            "private": bool(private_flag),
            "auto_init": True,
            "description": payload.get("repo_description") or "Refiner starter repo",
        }
        create_resp = _http_request_with_retry(
            "POST",
            create_url,
            headers=headers,
            json_body=body,
            timeout=20,
        )
        if create_resp.status_code not in (201, 202):
            raise ValueError(f"Repo create failed: {create_resp.status_code} {create_resp.text}")
        return create_resp.json()

    @staticmethod
    def _build_fork_repo_name(repo: str, owner: str, project_name: str) -> str:
        repo_slug = _slugify(repo)
        owner_slug = _slugify(owner)
        project_slug = _slugify(project_name)
        if project_slug and project_slug != repo_slug:
            name = f"{repo_slug}-{owner_slug}-{project_slug}"
        else:
            name = f"{repo_slug}-{owner_slug}"
        if len(name) > 90:
            trim = name[:90]
            return trim.rstrip("-")
        return name

    def _git_env(self, token: Optional[str], askpass_dir: str) -> Dict[str, str]:
        env = os.environ.copy()
        env["GIT_TERMINAL_PROMPT"] = "0"
        if token:
            os.makedirs(askpass_dir, exist_ok=True)
            askpass_path = os.path.join(askpass_dir, ".git-askpass.sh")
            script = (
                "#!/bin/sh\n"
                "case \"$1\" in\n"
                "*Username*) echo \"x-access-token\" ;;\n"
                "*Password*) echo \"" + token.replace('"', '\\"') + "\" ;;\n"
                "*) echo \"" + token.replace('"', '\\"') + "\" ;;\n"
                "esac\n"
            )
            with open(askpass_path, "w", encoding="utf-8") as handle:
                handle.write(script)
            try:
                os.chmod(askpass_path, 0o700)
            except Exception:
                pass
            env["GIT_ASKPASS"] = askpass_path
        return env

    def _git_run(self, command: List[str], cwd: str, job: Job, token: Optional[str] = None) -> subprocess.CompletedProcess:
        askpass_dir = os.path.dirname(job.log_path) if job.log_path else cwd
        env = self._git_env(token, askpass_dir)
        job.append_log(f"git: {' '.join(command)}")
        return subprocess.run(command, cwd=cwd, capture_output=True, text=True, env=env, check=False)

    def _git_clone(self, clone_url: str, workspace: str, branch: str, token: Optional[str], job: Job) -> None:
        result = self._git_run(
            ["git", "clone", "--depth", "1", "--branch", branch, clone_url, workspace],
            cwd=BASE_DIR,
            job=job,
            token=token,
        )
        if result.returncode != 0:
            raise ValueError(f"git clone failed: {result.stderr.strip()}")

    def _git_checkout(self, workspace: str, branch: str, job: Job) -> None:
        result = self._git_run(["git", "checkout", "-B", branch], cwd=workspace, job=job)
        if result.returncode != 0:
            raise ValueError(f"git checkout failed: {result.stderr.strip()}")

    def _git_has_changes(self, workspace: str, job: Job) -> bool:
        result = self._git_run(["git", "status", "--porcelain"], cwd=workspace, job=job)
        return bool(result.stdout.strip())

    def _git_config(self, workspace: str, name: str, email: str, job: Job) -> None:
        self._git_run(["git", "config", "user.name", name], cwd=workspace, job=job)
        self._git_run(["git", "config", "user.email", email], cwd=workspace, job=job)

    def _git_add_all(self, workspace: str, job: Job) -> None:
        result = self._git_run(["git", "add", "-A"], cwd=workspace, job=job)
        if result.returncode != 0:
            raise ValueError(f"git add failed: {result.stderr.strip()}")

    def _git_commit(self, workspace: str, message: str, job: Job) -> None:
        result = self._git_run(["git", "commit", "-m", message], cwd=workspace, job=job)
        if result.returncode != 0:
            stderr = result.stderr.strip()
            if "nothing to commit" in stderr.lower():
                return
            raise ValueError(f"git commit failed: {stderr}")

    def _git_push(self, workspace: str, branch: str, token: Optional[str], job: Job) -> None:
        result = self._git_run(["git", "push", "origin", branch], cwd=workspace, job=job, token=token)
        if result.returncode != 0:
            raise ValueError(f"git push failed: {result.stderr.strip()}")

    def _resolve_requirements(self, job: Job) -> Optional[str]:
        requirements_path = job.payload.get("requirements_path")
        requirements_text = job.payload.get("requirements_text")
        if requirements_path:
            return requirements_path
        if requirements_text:
            job_dir = os.path.join(JOB_ROOT, job.job_id)
            path = os.path.join(job_dir, "requirements.md")
            with open(path, "w", encoding="utf-8") as handle:
                handle.write(requirements_text)
            return path
        return None

    def _resolve_topic_source(self, job: Job) -> str:
        source = job.payload.get("topic_source") or job.payload.get("topic_research")
        if not source:
            raise ValueError("Topic source is required for topic research")
        if source.startswith("http://") or source.startswith("https://"):
            return source
        job_dir = os.path.join(JOB_ROOT, job.job_id)
        path = os.path.join(job_dir, "topic.txt")
        with open(path, "w", encoding="utf-8") as handle:
            handle.write(source)
        return path

    def _add_llm_args(self, command: List[str], payload: Dict[str, Any]) -> None:
        if payload.get("llm_provider"):
            command += ["--llm-provider", payload.get("llm_provider")]
        if payload.get("llm_model"):
            command += ["--llm-model", payload.get("llm_model")]
        if payload.get("llm_reasoning_effort"):
            command += ["--llm-reasoning-effort", payload.get("llm_reasoning_effort")]
        if payload.get("fallback_llm_provider"):
            command += ["--fallback-llm-provider", payload.get("fallback_llm_provider")]
        if payload.get("fallback_llm_model"):
            command += ["--fallback-llm-model", payload.get("fallback_llm_model")]
        if payload.get("ollama_base_url"):
            command += ["--ollama-base-url", payload.get("ollama_base_url")]
        if payload.get("llm_max_tokens"):
            command += ["--llm-max-tokens", str(payload.get("llm_max_tokens"))]
        if payload.get("llm_chunk_size"):
            command += ["--llm-chunk-size", str(payload.get("llm_chunk_size"))]
        if payload.get("llm_temperature") is not None:
            command += ["--llm-temperature", str(payload.get("llm_temperature"))]
        if payload.get("llm_timeout"):
            command += ["--llm-timeout", str(payload.get("llm_timeout"))]
        if payload.get("llm_inter_request_gap"):
            command += ["--llm-inter-request-gap", str(payload.get("llm_inter_request_gap"))]

    def _compute_wait_seconds(self, job: Job) -> Optional[float]:
        if not job.started_at:
            return None
        created = _parse_timestamp(job.created_at)
        started = _parse_timestamp(job.started_at)
        if not created or not started:
            return None
        return started.timestamp() - created.timestamp()

    def _compute_runtime_seconds(self, job: Job) -> Optional[float]:
        if not (job.started_at and job.finished_at):
            return None
        started = _parse_timestamp(job.started_at)
        finished = _parse_timestamp(job.finished_at)
        if not started or not finished:
            return None
        return finished.timestamp() - started.timestamp()

    @staticmethod
    def _coerce_int(value: Optional[str]) -> Optional[int]:
        if value is None or value == "?":
            return None
        try:
            return int(value)
        except Exception:
            return None


_JOB_QUEUE_OWNER_LIST_FIELDS = (
    "top_queued_owners",
    "top_running_owners",
    "top_active_owners",
)


def _serialise_job_queue_snapshot(
    snapshot: Optional[Dict[str, Any]],
    *,
    include_owner_lists: bool = False,
) -> Dict[str, Any]:
    if not isinstance(snapshot, dict):
        return {}
    payload = dict(snapshot)
    if not include_owner_lists:
        for field in _JOB_QUEUE_OWNER_LIST_FIELDS:
            payload.pop(field, None)
    return payload


def _job_queue_snapshot(
    manager_obj: Any,
    *,
    top_limit: int = 3,
    include_owner_lists: bool = False,
) -> Dict[str, Any]:
    snapshotter = getattr(manager_obj, "queue_snapshot", None)
    if callable(snapshotter):
        try:
            snapshot = snapshotter(top_limit=top_limit)
        except TypeError:
            snapshot = snapshotter()
        return _serialise_job_queue_snapshot(snapshot, include_owner_lists=include_owner_lists)

    queue_depth = max(0, int(manager_obj.queue.qsize()))
    with manager_obj.lock:
        jobs_snapshot = list(manager_obj.jobs.values())
        workers = max(1, len(manager_obj.workers))
    queued = 0
    running = 0
    paused = 0
    for job in jobs_snapshot:
        status = (getattr(job, "status", "") or "").strip().lower()
        if status == "queued":
            queued += 1
        elif status == "running":
            running += 1
        elif status == "paused":
            paused += 1
    return {
        "queue_depth": queue_depth,
        "queued": queued,
        "running": running,
        "paused": paused,
        "workers": workers,
    }


class ContinuumQueueAutoscaler(_ContinuumQueueAutoscalerImpl):
    """Compatibility wrapper bound to the Refiner runtime helper surface."""

    def __init__(
        self,
        manager: JobManager,
        *,
        enabled: bool = CONTINUUM_AUTOSCALE_ENABLED,
        poll_sec: float = CONTINUUM_AUTOSCALE_POLL_SEC,
        min_replicas: int = CONTINUUM_AUTOSCALE_MIN_REPLICAS,
        max_replicas: int = CONTINUUM_AUTOSCALE_MAX_REPLICAS,
        backlog_per_replica: int = CONTINUUM_AUTOSCALE_BACKLOG_PER_REPLICA,
        scale_up_step: int = CONTINUUM_AUTOSCALE_SCALE_UP_STEP,
        scale_down_step: int = CONTINUUM_AUTOSCALE_SCALE_DOWN_STEP,
        idle_sec: float = CONTINUUM_AUTOSCALE_IDLE_SEC,
        cooldown_sec: float = CONTINUUM_AUTOSCALE_COOLDOWN_SEC,
        timeout_sec: float = CONTINUUM_AUTOSCALE_TIMEOUT_SEC,
        namespace: str = CONTINUUM_AUTOSCALE_NAMESPACE,
        deployment: str = CONTINUUM_AUTOSCALE_DEPLOYMENT,
    ):
        super().__init__(
            manager,
            enabled=enabled,
            poll_sec=poll_sec,
            min_replicas=min_replicas,
            max_replicas=max_replicas,
            backlog_per_replica=backlog_per_replica,
            scale_up_step=scale_up_step,
            scale_down_step=scale_down_step,
            idle_sec=idle_sec,
            cooldown_sec=cooldown_sec,
            timeout_sec=timeout_sec,
            namespace=namespace,
            deployment=deployment,
            history_max=CONTINUUM_AUTOSCALE_HISTORY_MAX,
            continuum_enabled=lambda: _continuum_enabled(),
            continuum_request=lambda method, path, **kwargs: _continuum_request(method, path, **kwargs),
            continuum_json_payload=lambda response, *, operation: _continuum_json_payload(response, operation=operation),
            friendly_continuum_error=lambda message: _friendly_continuum_error(message),
            now_iso=lambda: _now_iso(),
            logger=logger,
            job_queue_snapshot=lambda manager, top_limit=3, include_owner_lists=False: _job_queue_snapshot(
                manager,
                top_limit=top_limit,
                include_owner_lists=include_owner_lists,
            ),
        )


continuum_autoscaler: Optional[ContinuumQueueAutoscaler] = None


def _notify_continuum_autoscaler() -> None:
    autoscaler = continuum_autoscaler
    if autoscaler:
        autoscaler.notify_queue_change()


def _continuum_autoscaler_status() -> Dict[str, Any]:
    autoscaler = continuum_autoscaler
    if not autoscaler:
        return {"enabled": False, "continuum_configured": _continuum_enabled(), "running": False}
    return autoscaler.status()


def _continuum_cluster_snapshot(timeout_sec: float) -> Optional[Dict[str, Any]]:
    return _continuum_cluster_snapshot_impl(
        timeout_sec=timeout_sec,
        continuum_enabled=lambda: _continuum_enabled(),
        continuum_request=lambda method, path, **kwargs: _continuum_request(method, path, **kwargs),
        continuum_json_payload=lambda response, *, operation: _continuum_json_payload(response, operation=operation),
        safe_int=lambda value, default=0: _safe_int(value, default),
    )


def _workers_telemetry_payload(
    *,
    limit: int = 180,
    refresh: bool = False,
    include_cluster: bool = False,
) -> Dict[str, Any]:
    return _workers_telemetry_payload_impl(
        autoscaler=continuum_autoscaler,
        limit=limit,
        refresh=refresh,
        include_cluster=include_cluster,
        continuum_enabled=lambda: _continuum_enabled(),
        friendly_continuum_error=lambda message: _friendly_continuum_error(message),
        continuum_cluster_snapshot=lambda timeout: _continuum_cluster_snapshot(timeout),
        now_iso=lambda: _now_iso(),
        logger=logger,
        serialise_job_queue_snapshot=lambda snapshot, include_owner_lists=False: _serialise_job_queue_snapshot(
            snapshot,
            include_owner_lists=include_owner_lists,
        ),
    )


access_store = AccessStore(ACCESS_STORE_PATH)
manager = JobManager()
job_action_manager = JobActionManager()
continuum_autoscaler = ContinuumQueueAutoscaler(manager)
continuum_autoscaler.start()
user_store = CENTRAL_STORE.users if CENTRAL_STORE is not None else UserStore(USERS_PATH)
user_store.ensure_admin_from_env()
if CENTRAL_STORE is not None:
    CENTRAL_STORE.bootstrap_from_env((os.getenv("REFINER_ADMIN_USER") or "").strip())


def _rehydrate_job_manager_from_disk() -> None:
    """Best-effort startup pass to ensure persisted jobs are visible in memory."""
    before = len(manager.jobs)
    try:
        manager._load_jobs_from_disk()
    except Exception as exc:
        logger.warning("job manager disk rehydrate failed for %s: %s", JOB_ROOT, exc)
        return
    after = len(manager.jobs)
    if after != before:
        logger.info("job manager rehydrated from disk: %s -> %s (%s)", before, after, JOB_ROOT)

_secret_stores: Dict[str, SecretStore] = {}
_user_activity: Dict[str, float] = {}
_user_activity_lock = threading.Lock()
token_ledger = CENTRAL_STORE.user_ledger if CENTRAL_STORE is not None else TokenLedger(LEDGER_ROOT)
team_token_ledger = CENTRAL_STORE.team_ledger if CENTRAL_STORE is not None else TokenLedger(TEAM_LEDGER_ROOT)
rag_store = RagStore(RAG_STORE_ROOT)
mcp_store = PostgresMCPServerStore(CENTRAL_STORE.pool) if CENTRAL_STORE is not None else MCPServerStore(MCP_STORE_ROOT)
session_history = (
    PostgresSessionHistoryStore(SESSIONS_ROOT, CENTRAL_STORE.session_rooms)
    if CENTRAL_STORE is not None
    else SessionHistoryStore(SESSIONS_ROOT)
)
session_store = SessionStore()
voice_token_store = CENTRAL_STORE.voice_tokens if CENTRAL_STORE is not None else VoiceTokenStore(VOICE_TOKEN_PATH)
todo_store = (
    PostgresTodoStore(TODO_ROOT, CENTRAL_STORE.todo_documents)
    if CENTRAL_STORE is not None
    else TodoStore(TODO_ROOT)
)
schedule_store = (
    PostgresScheduleStore(SCHEDULE_ROOT, CENTRAL_STORE.schedule_documents)
    if CENTRAL_STORE is not None
    else ScheduleStore(SCHEDULE_ROOT)
)
subtask_manager = SubtaskManager()
todo_scheduler = TodoScheduler(
    schedule_store=schedule_store,
    todo_store=todo_store,
    subtask_manager=subtask_manager,
)
llm_telemetry_janitor = LLMTelemetryJanitor(telemetry_store_factory=_llm_request_telemetry_store)
ai_model_inventory_monitor = AIModelInventoryMonitor(config_path=os.path.join(BASE_DIR, "config.json"))


def _llm_telemetry_retention_status() -> Dict[str, Any]:
    return llm_telemetry_janitor.status()


def _chain_enabled() -> bool:
    return bool(NMCHAIN and NMCHAIN.enabled)


def _customers_enabled() -> bool:
    return bool(CUSTOMERS_SERVICE and CUSTOMERS_SERVICE.base_url)


def _billing_enabled() -> bool:
    return bool(BILLING_SERVICE and BILLING_SERVICE.base_url)


def _current_request_auth_headers() -> Tuple[Optional[str], Optional[str]]:
    if not has_request_context():
        return None, None
    return (
        request.headers.get("Authorization") or request.headers.get("authorization"),
        request.headers.get("Cookie"),
    )


def _customers_profile_prefetch(
    *,
    authorization: Optional[str],
    cookie_header: Optional[str],
) -> Optional[Future]:
    """Submit a one-time customers profile lookup for this request context."""
    if not has_request_context() or not _customers_enabled():
        return None
    existing = getattr(g, "customers_profile_future", None)
    if isinstance(existing, Future):
        return existing
    if not authorization and not cookie_header:
        return None
    future = _submit_external_call(
        CUSTOMERS_SERVICE.profile_from_headers,
        authorization=authorization,
        cookie_header=cookie_header,
    )
    g.customers_profile_future = future
    return future


def _customers_profile_from_prefetch(
    *,
    wait_timeout: float = CUSTOMERS_PROFILE_PREFETCH_WAIT_SEC,
) -> Optional[Dict[str, Any]]:
    """Read the prefetched customers profile result without blocking indefinitely."""
    if not has_request_context():
        return None
    future = getattr(g, "customers_profile_future", None)
    if not isinstance(future, Future):
        return None
    try:
        payload = _await_external_call_result(
            future,
            operation="customers profile prefetch",
            timeout=max(0.05, float(wait_timeout)),
        )
    except TimeoutError:
        return None
    except Exception as exc:
        logger.warning("customers profile prefetch failed: %s", exc)
        return None
    if isinstance(payload, dict):
        return payload
    return None


def _normalize_identity_groups(value: Any) -> List[str]:
    if isinstance(value, str):
        raw_values = [item.strip() for item in value.split(",") if item.strip()]
    elif isinstance(value, (list, tuple, set)):
        raw_values = [str(item).strip() for item in value if str(item).strip()]
    else:
        raw_values = []
    normalized: List[str] = []
    seen = set()
    for item in raw_values:
        lowered = item.lower()
        if lowered in seen:
            continue
        seen.add(lowered)
        normalized.append(lowered)
    return normalized


def _normalize_identity_active_team(value: Any) -> Optional[Dict[str, Any]]:
    if isinstance(value, dict):
        payload = dict(value)
        team_id = str(payload.get("team_id") or payload.get("id") or "").strip()
        if not team_id:
            return None
        payload["team_id"] = team_id
        return payload
    if isinstance(value, str):
        team_id = value.strip()
        if team_id:
            return {"team_id": team_id}
    return None


SERVICE_ACCESS_NONE = "none"
SERVICE_ACCESS_REQUEST = "request"
SERVICE_ACCESS_OBSERVE = "observe"
SERVICE_ACCESS_USE = "use"
SERVICE_ACCESS_CONTROL = "control"

SERVICE_ACCESS_ORDER = {
    SERVICE_ACCESS_NONE: 0,
    SERVICE_ACCESS_REQUEST: 1,
    SERVICE_ACCESS_OBSERVE: 2,
    SERVICE_ACCESS_USE: 3,
    SERVICE_ACCESS_CONTROL: 4,
}

SERVICE_ACCESS_PUBLIC_DEFAULTS = {
    "refiner": SERVICE_ACCESS_REQUEST,
    "billing": SERVICE_ACCESS_NONE,
    "continuum": SERVICE_ACCESS_OBSERVE,
    "tracey": SERVICE_ACCESS_OBSERVE,
    "aarnn": SERVICE_ACCESS_REQUEST,
    "webots": SERVICE_ACCESS_REQUEST,
    "gail": SERVICE_ACCESS_NONE,
    "nmstt": SERVICE_ACCESS_NONE,
    "conductor": SERVICE_ACCESS_NONE,
    "customers": SERVICE_ACCESS_NONE,
    "nmchain": SERVICE_ACCESS_NONE,
}

SERVICE_ACCESS_AUTHENTICATED_DEFAULTS = {
    "refiner": SERVICE_ACCESS_USE,
    "billing": SERVICE_ACCESS_USE,
    "aarnn": SERVICE_ACCESS_REQUEST,
    "webots": SERVICE_ACCESS_REQUEST,
}


def _normalize_service_access_level(value: Any, *, default: str = SERVICE_ACCESS_NONE) -> str:
    cleaned = str(value or default).strip().lower() or default
    if cleaned not in SERVICE_ACCESS_ORDER:
        return default
    return cleaned


def _service_access_at_least(current: Any, required: Any) -> bool:
    current_level = _normalize_service_access_level(current)
    required_level = _normalize_service_access_level(required)
    return SERVICE_ACCESS_ORDER[current_level] >= SERVICE_ACCESS_ORDER[required_level]


def _max_service_access_level(*values: Any) -> str:
    best = SERVICE_ACCESS_NONE
    for value in values:
        candidate = _normalize_service_access_level(value)
        if SERVICE_ACCESS_ORDER[candidate] > SERVICE_ACCESS_ORDER[best]:
            best = candidate
    return best


def _normalize_service_access_entry(
    service_key: Any,
    raw_entry: Any,
    *,
    default_public_level: str = SERVICE_ACCESS_NONE,
    default_access_level: str = SERVICE_ACCESS_NONE,
) -> Optional[Dict[str, Any]]:
    cleaned_service_key = str(service_key or "").strip().lower()
    if not cleaned_service_key:
        return None
    if isinstance(raw_entry, dict):
        payload = dict(raw_entry)
    else:
        payload = {}
        if raw_entry is not None:
            payload["access_level"] = raw_entry
    public_access_level = _normalize_service_access_level(
        payload.get("public_access_level") or default_public_level,
        default=SERVICE_ACCESS_NONE,
    )
    access_level = _normalize_service_access_level(
        payload.get("access_level") or payload.get("level") or default_access_level,
        default=SERVICE_ACCESS_NONE,
    )
    visible_access_level = _normalize_service_access_level(
        payload.get("visible_access_level"),
        default=_max_service_access_level(access_level, public_access_level),
    )
    can_request = payload.get("can_request")
    can_observe = payload.get("can_observe")
    can_use = payload.get("can_use")
    can_control = payload.get("can_control")
    return {
        **payload,
        "service_key": cleaned_service_key,
        "access_level": access_level,
        "public_access_level": public_access_level,
        "visible_access_level": visible_access_level,
        "visible": bool(payload.get("visible")) or visible_access_level != SERVICE_ACCESS_NONE,
        "can_request": can_request if isinstance(can_request, bool) else _service_access_at_least(visible_access_level, SERVICE_ACCESS_REQUEST),
        "can_observe": can_observe if isinstance(can_observe, bool) else _service_access_at_least(visible_access_level, SERVICE_ACCESS_OBSERVE),
        "can_use": can_use if isinstance(can_use, bool) else _service_access_at_least(access_level, SERVICE_ACCESS_USE),
        "can_control": can_control if isinstance(can_control, bool) else _service_access_at_least(access_level, SERVICE_ACCESS_CONTROL),
    }


def _default_identity_service_access(payload: Optional[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    identity = dict(payload or {})
    authenticated = bool(identity.get("authenticated", True))
    identity_type = str(identity.get("identity_type") or "").strip().lower() or None
    default_role = "service_account" if identity_type == "service_account" else "user"
    role = str(identity.get("role") or default_role).strip().lower() or default_role
    groups = _normalize_identity_groups(identity.get("groups"))
    is_service_account = identity_type == "service_account" or role == "service_account"
    is_admin = role == "admin" or "admin" in groups
    resolved: Dict[str, Dict[str, Any]] = {}
    for service_key, public_access_level in SERVICE_ACCESS_PUBLIC_DEFAULTS.items():
        granted_access_level = SERVICE_ACCESS_NONE
        if is_admin:
            granted_access_level = SERVICE_ACCESS_CONTROL
        elif authenticated and not is_service_account:
            granted_access_level = _normalize_service_access_level(
                SERVICE_ACCESS_AUTHENTICATED_DEFAULTS.get(service_key),
                default=SERVICE_ACCESS_NONE,
            )
        entry = _normalize_service_access_entry(
            service_key,
            {},
            default_public_level=public_access_level,
            default_access_level=granted_access_level,
        )
        if entry:
            resolved[service_key] = entry
    return resolved


def _resolve_identity_service_access(payload: Optional[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    identity = dict(payload or {})
    resolved = _default_identity_service_access(identity)
    raw_service_access = identity.get("service_access")
    if isinstance(raw_service_access, dict):
        items = list(raw_service_access.items())
    elif isinstance(raw_service_access, list):
        items = []
        for entry in raw_service_access:
            if not isinstance(entry, dict):
                continue
            service_key = entry.get("service_key") or entry.get("key")
            items.append((service_key, entry))
    else:
        items = []
    for service_key, raw_entry in items:
        cleaned_service_key = str(service_key or "").strip().lower()
        if not cleaned_service_key:
            continue
        normalized = _normalize_service_access_entry(
            cleaned_service_key,
            raw_entry,
            default_public_level=_normalize_service_access_level(
                (resolved.get(cleaned_service_key) or {}).get("public_access_level")
                or SERVICE_ACCESS_PUBLIC_DEFAULTS.get(cleaned_service_key),
                default=SERVICE_ACCESS_NONE,
            ),
            default_access_level=(resolved.get(cleaned_service_key) or {}).get("access_level", SERVICE_ACCESS_NONE),
        )
        if normalized:
            resolved[cleaned_service_key] = normalized
    return resolved


def _identity_service_access_entry(identity: Optional[Dict[str, Any]], service_key: str) -> Dict[str, Any]:
    cleaned_service_key = str(service_key or "").strip().lower()
    if not cleaned_service_key:
        return {
            "service_key": "",
            "access_level": SERVICE_ACCESS_NONE,
            "public_access_level": SERVICE_ACCESS_NONE,
            "visible_access_level": SERVICE_ACCESS_NONE,
            "visible": False,
            "can_request": False,
            "can_observe": False,
            "can_use": False,
            "can_control": False,
        }
    services = _resolve_identity_service_access(identity)
    entry = services.get(cleaned_service_key)
    if entry:
        return entry
    fallback = _normalize_service_access_entry(
        cleaned_service_key,
        {},
        default_public_level=SERVICE_ACCESS_PUBLIC_DEFAULTS.get(cleaned_service_key, SERVICE_ACCESS_NONE),
        default_access_level=SERVICE_ACCESS_NONE,
    )
    return fallback or {
        "service_key": cleaned_service_key,
        "access_level": SERVICE_ACCESS_NONE,
        "public_access_level": SERVICE_ACCESS_NONE,
        "visible_access_level": SERVICE_ACCESS_NONE,
        "visible": False,
        "can_request": False,
        "can_observe": False,
        "can_use": False,
        "can_control": False,
    }


def _identity_has_service_access(identity: Optional[Dict[str, Any]], service_key: str, minimum: str) -> bool:
    entry = _identity_service_access_entry(identity, service_key)
    required = _normalize_service_access_level(minimum)
    if required == SERVICE_ACCESS_REQUEST:
        return bool(entry.get("can_request"))
    if required == SERVICE_ACCESS_OBSERVE:
        return bool(entry.get("can_observe"))
    if required == SERVICE_ACCESS_USE:
        return bool(entry.get("can_use"))
    if required == SERVICE_ACCESS_CONTROL:
        return bool(entry.get("can_control"))
    return False


def _normalize_auth_identity(payload: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not isinstance(payload, dict):
        return None
    user = str(payload.get("user") or "").strip()
    if not user:
        return None
    identity_type = str(payload.get("identity_type") or "").strip().lower() or None
    default_role = "service_account" if identity_type == "service_account" else (user_store.get_role(user) or "user")
    role = str(payload.get("role") or default_role).strip().lower() or default_role
    groups = _normalize_identity_groups(payload.get("groups"))
    if role and role != "service_account" and role not in groups:
        groups.insert(0, role)
    group_memberships = payload.get("group_memberships") if isinstance(payload.get("group_memberships"), list) else []
    manageable_groups = _normalize_identity_groups(payload.get("manageable_groups"))
    visible_groups = _normalize_identity_groups(payload.get("visible_groups")) or list(groups)
    active_team = _normalize_identity_active_team(payload.get("active_team"))
    team_tree = payload.get("team_tree") if isinstance(payload.get("team_tree"), list) else []
    try:
        team_count = int(payload.get("team_count") or (1 if active_team else 0))
    except Exception:
        team_count = 1 if active_team else 0
    try:
        pending_invitation_count = int(payload.get("pending_invitation_count") or 0)
    except Exception:
        pending_invitation_count = 0
    service_access = _resolve_identity_service_access(
        {
            **payload,
            "authenticated": bool(payload.get("authenticated", True)),
            "role": role,
            "groups": groups,
        }
    )
    visible_services = sorted(
        {
            str(item).strip().lower()
            for item in payload.get("visible_services", [])
            if str(item).strip()
        }
        | {
            service_key
            for service_key, entry in service_access.items()
            if bool(entry.get("visible"))
        }
    )
    return {
        **payload,
        "authenticated": bool(payload.get("authenticated", True)),
        "identity_type": identity_type,
        "user": user,
        "role": role,
        "groups": groups,
        "group_memberships": group_memberships,
        "manageable_groups": manageable_groups,
        "visible_groups": visible_groups,
        "can_manage_access": bool(payload.get("can_manage_access")) or bool(manageable_groups),
        "active_team": active_team,
        "team_count": max(0, team_count),
        "pending_invitation_count": max(0, pending_invitation_count),
        "team_tree": team_tree,
        "service_access": service_access,
        "visible_services": visible_services,
        "is_admin": role == "admin" or "admin" in groups,
    }


def _set_request_identity(identity: Optional[Dict[str, Any]], *, resolved: bool = False) -> None:
    if not has_request_context():
        return
    if not identity:
        g.auth_identity = None
        g.auth_user = None
        g.auth_user_role = None
        g.auth_user_groups = []
        g.auth_active_team = None
        if resolved:
            g.auth_user_resolved = True
        return
    g.auth_identity = identity
    g.auth_user = identity.get("user")
    g.auth_user_role = identity.get("role")
    g.auth_user_groups = identity.get("groups") if isinstance(identity.get("groups"), list) else []
    g.auth_active_team = identity.get("active_team")
    if resolved:
        g.auth_user_resolved = True


def _current_auth_identity() -> Optional[Dict[str, Any]]:
    if not has_request_context():
        return None
    identity = getattr(g, "auth_identity", None)
    return identity if isinstance(identity, dict) else None


def _current_auth_profile() -> Optional[Dict[str, Any]]:
    if not has_request_context():
        return None
    if getattr(g, "auth_profile_resolved", False):
        profile = getattr(g, "auth_profile", None)
        return profile if isinstance(profile, dict) else None
    identity = _current_auth_identity()
    if not identity and _current_user():
        identity = _current_auth_identity()
    if not _customers_enabled() or not identity or not identity.get("user"):
        g.auth_profile = identity
        g.auth_profile_resolved = True
        return identity
    authorization, cookie_header = _current_request_auth_headers()
    prefetch_future = _customers_profile_prefetch(authorization=authorization, cookie_header=cookie_header)
    payload = _customers_profile_from_prefetch(wait_timeout=CUSTOMERS_PROFILE_PREFETCH_WAIT_SEC)
    if payload is None and prefetch_future is None and (authorization or cookie_header):
        try:
            payload = _run_external_call(
                "customers profile lookup",
                CUSTOMERS_SERVICE.profile_from_headers,
                authorization=authorization,
                cookie_header=cookie_header,
                timeout=_service_timeout_seconds(CUSTOMERS_SERVICE, default=max(1.0, CUSTOMERS_TIMEOUT)),
            )
        except Exception as exc:
            logger.warning("customers profile lookup failed: %s", exc)
            payload = {}
    profile = _normalize_auth_identity(payload if isinstance(payload, dict) else {}) or identity
    g.auth_profile = profile
    g.auth_profile_resolved = True
    if profile:
        _set_request_identity(profile, resolved=False)
    return profile


def _user_metadata(user: Optional[str]) -> Dict[str, Any]:
    cleaned = str(user or "").strip()
    if not cleaned:
        return {}
    try:
        metadata = user_store.get_metadata(cleaned)
    except Exception:
        metadata = {}
    return dict(metadata) if isinstance(metadata, dict) else {}


def _user_settings(user: Optional[str]) -> Dict[str, Any]:
    cleaned = str(user or "").strip()
    if not cleaned:
        return default_user_settings()
    return settings_from_metadata(_user_metadata(cleaned))


def _settings_defaults_for_user(user: Optional[str]) -> Dict[str, Any]:
    return llm_defaults_from_settings(_user_settings(user))


def _build_job_runtime_env(
    owner: Optional[str],
    *,
    settings: Optional[Dict[str, Any]] = None,
    role: Optional[str] = None,
    secret_env: Optional[Dict[str, str]] = None,
    process_env: Optional[Dict[str, str]] = None,
    use_default_secrets: bool = True,
    job_secrets: Any = None,
) -> Dict[str, str]:
    resolved_settings = settings if settings is not None else _user_settings(owner)
    resolved_role = role if role is not None else _user_role(owner)
    resolved_secret_env = (
        dict(secret_env)
        if secret_env is not None
        else (_get_secret_store(owner).get_env() if owner else {})
    )
    return _build_job_runtime_env_impl(
        owner,
        settings=resolved_settings,
        role=resolved_role,
        secret_env=resolved_secret_env,
        process_env=process_env,
        use_default_secrets=use_default_secrets,
        job_secrets=job_secrets,
    )


def _update_user_settings(user: str, raw_settings: Any) -> Dict[str, Any]:
    cleaned = str(user or "").strip()
    if not cleaned:
        raise SettingsValidationError(["user is required"])
    current_metadata = _user_metadata(cleaned)
    current_settings = settings_from_metadata(current_metadata)
    merged = validate_settings_patch(raw_settings, current=current_settings)
    updated_metadata = metadata_with_settings(current_metadata, merged, updated_at=_now_iso())
    if not user_store.set_metadata(cleaned, updated_metadata):
        raise KeyError(cleaned)
    return merged


def _local_identity_payload(user: str, *, include_directory: bool = False) -> Dict[str, Any]:
    role = str(user_store.get_role(user) or "user").strip().lower() or "user"
    base_payload: Dict[str, Any] = {
        "authenticated": True,
        "user": user,
        "role": role,
        "groups": [role] if role else [],
        "group_memberships": [],
        "manageable_groups": ["admin"] if role == "admin" else [],
        "visible_groups": [role] if role else [],
        "can_manage_access": role == "admin",
    }
    service_access = _resolve_identity_service_access(base_payload)
    payload: Dict[str, Any] = {
        **base_payload,
        "email": user_store.get_email(user),
        "active_team": None,
        "team_count": 0,
        "pending_invitation_count": 0,
        "is_admin": role == "admin",
        "service_access": service_access,
        "visible_services": sorted(
            service_key
            for service_key, entry in service_access.items()
            if bool(entry.get("visible"))
        ),
        "settings": _user_settings(user),
    }
    if include_directory:
        payload["team_tree"] = access_store.tree_all() if payload["is_admin"] else access_store.tree_for_user(user)
    return payload


def _refiner_auth_config_payload() -> Dict[str, Any]:
    has_users = user_store.has_users()
    local_auth_available = AUTH_MODE != "oidc"
    return {
        "service": "refiner",
        "auth_mode": AUTH_MODE,
        "oidc_enabled": OIDC_ENABLED,
        "oidc_button_label": OIDC_BUTTON_LABEL,
        "password_min_length": PASSWORD_MIN_LEN,
        "has_users": has_users,
        "setup_allowed": local_auth_available,
        "setup_available": local_auth_available and not has_users,
        "local_login_enabled": local_auth_available and has_users,
        "self_registration_enabled": local_auth_available and has_users and SELF_REGISTRATION_ENABLED,
        "team_provisioning_available": False,
        "totp_supported": False,
        "passkeys_supported": False,
    }


def _team_tree_role(team_id: Optional[str], nodes: Optional[List[Dict[str, Any]]]) -> Optional[str]:
    cleaned_team_id = str(team_id or "").strip()
    if not cleaned_team_id or not isinstance(nodes, list):
        return None
    for node in nodes:
        if not isinstance(node, dict):
            continue
        if str(node.get("id") or "").strip() == cleaned_team_id:
            membership_role = str(node.get("membership_role") or "").strip().lower()
            if membership_role == "owner":
                return "leader"
            if membership_role == "member":
                return "member"
        nested = _team_tree_role(cleaned_team_id, node.get("children") if isinstance(node.get("children"), list) else [])
        if nested:
            return nested
    return None


def _team_role_for_user(user: Optional[str], team_id: Optional[str]) -> Optional[str]:
    cleaned_user = str(user or "").strip()
    cleaned_team_id = str(team_id or "").strip()
    if not cleaned_user or not cleaned_team_id:
        return None
    profile = _current_auth_profile()
    if isinstance(profile, dict) and str(profile.get("user") or "").strip() == cleaned_user:
        membership_role = _team_tree_role(cleaned_team_id, profile.get("team_tree"))
        if membership_role:
            return membership_role
    return access_store._team_role(cleaned_user, cleaned_team_id)


def _user_role(user: Optional[str]) -> Optional[str]:
    if not user:
        return None
    identity = _current_auth_profile() or _current_auth_identity()
    if isinstance(identity, dict) and str(identity.get("user") or "").strip() == user:
        role = identity.get("role")
        if role:
            return str(role).strip().lower() or None
    return user_store.get_role(user)


def _forward_proxy_headers(*, include_auth: bool = True, include_cookie: bool = True) -> Dict[str, str]:
    headers: Dict[str, str] = {}
    for key in ("Accept", "Content-Type", "User-Agent", "X-Requested-With", "Origin"):
        value = request.headers.get(key)
        if value:
            headers[key] = value
    if include_auth:
        auth = request.headers.get("Authorization") or request.headers.get("authorization")
        if auth:
            headers["Authorization"] = auth
    if include_cookie:
        cookie_header = request.headers.get("Cookie")
        if cookie_header:
            headers["Cookie"] = cookie_header
    forwarded_for = request.headers.get("X-Forwarded-For")
    if forwarded_for:
        headers["X-Forwarded-For"] = forwarded_for
    elif request.remote_addr:
        headers["X-Forwarded-For"] = request.remote_addr
    headers["X-Forwarded-Host"] = request.host
    headers["X-Forwarded-Proto"] = request.headers.get("X-Forwarded-Proto") or ("https" if _is_secure_request() else "http")
    return headers


def _proxy_service_request(base_url: str, timeout: float, *, path: Optional[str] = None) -> Response:
    target_path = path if path is not None else request.path
    url = base_url.rstrip("/") + target_path
    try:
        upstream = _external_http_request(
            method=request.method,
            url=url,
            params=request.args,
            data=request.get_data(cache=True) if request.method not in {"GET", "HEAD"} else None,
            headers=_forward_proxy_headers(),
            allow_redirects=False,
            timeout=timeout,
        )
    except requests.RequestException as exc:
        logger.warning("upstream proxy request failed for %s: %s", url, exc)
        return jsonify({"error": "upstream_unavailable", "details": str(exc)}), 503

    response = Response(upstream.content, status=upstream.status_code)
    excluded = {
        "connection",
        "content-encoding",
        "content-length",
        "keep-alive",
        "proxy-authenticate",
        "proxy-authorization",
        "te",
        "trailer",
        "transfer-encoding",
        "upgrade",
    }
    for header, value in upstream.headers.items():
        if header.lower() in excluded:
            continue
        if header.lower() == "set-cookie":
            response.headers.add(header, value)
        else:
            response.headers[header] = value
    return response


def _customer_team_member_lists(detail_payload: Dict[str, Any]) -> Dict[str, List[str]]:
    members = detail_payload.get("members") if isinstance(detail_payload.get("members"), list) else []
    leaders: List[str] = []
    contributors: List[str] = []
    for membership in members:
        if not isinstance(membership, dict):
            continue
        username = str(membership.get("username") or "").strip()
        if not username:
            continue
        membership_role = str(membership.get("membership_role") or "").strip().lower()
        if membership_role == "owner":
            leaders.append(username)
        else:
            contributors.append(username)
    return {
        "leaders": sorted(set(leaders)),
        "members": sorted(set(contributors)),
    }


def _enrich_customer_team_tree(nodes: List[Dict[str, Any]], detail_map: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    enriched: List[Dict[str, Any]] = []
    for node in nodes:
        if not isinstance(node, dict):
            continue
        entry = dict(node)
        team_id = str(entry.get("id") or "").strip()
        details = detail_map.get(team_id, {})
        if details:
            entry["leaders"] = details.get("leaders") or []
            entry["members"] = details.get("members") or []
        children = entry.get("children") if isinstance(entry.get("children"), list) else []
        entry["children"] = _enrich_customer_team_tree(children, detail_map)
        enriched.append(entry)
    return enriched


def _customer_access_tree() -> List[Dict[str, Any]]:
    if not has_request_context():
        return []
    if getattr(g, "customers_access_tree_resolved", False):
        tree = getattr(g, "customers_access_tree", None)
        return tree if isinstance(tree, list) else []
    if not _customers_enabled():
        g.customers_access_tree = []
        g.customers_access_tree_resolved = True
        return []
    authorization, cookie_header = _current_request_auth_headers()
    customers_timeout = _service_timeout_seconds(CUSTOMERS_SERVICE, default=max(1.0, CUSTOMERS_TIMEOUT))
    try:
        payload = _run_external_call(
            "customers teams lookup",
            CUSTOMERS_SERVICE.teams_from_headers,
            authorization=authorization,
            cookie_header=cookie_header,
            timeout=customers_timeout,
        )
    except Exception as exc:
        logger.warning("customers teams lookup failed: %s", exc)
        payload = {}
    tree = payload.get("tree") if isinstance(payload.get("tree"), list) else []
    teams = payload.get("teams") if isinstance(payload.get("teams"), list) else []
    detail_map: Dict[str, Dict[str, Any]] = {}
    team_ids: List[str] = []
    for team in teams:
        if not isinstance(team, dict):
            continue
        team_id = str(team.get("id") or "").strip()
        if not team_id:
            continue
        team_ids.append(team_id)
        detail_map[team_id] = {}

    max_parallel = max(1, min(CUSTOMERS_TEAM_DETAIL_MAX_WORKERS, len(team_ids)))
    for start in range(0, len(team_ids), max_parallel):
        batch_ids = team_ids[start : start + max_parallel]
        futures: Dict[Future, str] = {
            _submit_external_call(
                CUSTOMERS_SERVICE.team_detail_from_headers,
                team_id,
                authorization=authorization,
                cookie_header=cookie_header,
            ): team_id
            for team_id in batch_ids
        }
        try:
            for future in as_completed(futures, timeout=customers_timeout + EXTERNAL_HTTP_FUTURE_GRACE_SEC):
                team_id = futures[future]
                try:
                    detail_payload = future.result()
                except Exception as exc:
                    logger.warning("customers team detail lookup failed for %s: %s", team_id, exc)
                    detail_payload = {}
                detail_map[team_id] = _customer_team_member_lists(
                    detail_payload if isinstance(detail_payload, dict) else {}
                )
        except FutureTimeoutError:
            for future, team_id in futures.items():
                if not future.done():
                    future.cancel()
                    logger.warning("customers team detail lookup timed out for %s", team_id)
    enriched_tree = _enrich_customer_team_tree(tree, detail_map)
    g.customers_access_tree = enriched_tree
    g.customers_access_tree_resolved = True
    return enriched_tree


def _find_team_in_tree(team_id: Optional[str], nodes: Optional[List[Dict[str, Any]]]) -> Optional[Dict[str, Any]]:
    cleaned_team_id = str(team_id or "").strip()
    if not cleaned_team_id or not isinstance(nodes, list):
        return None
    for node in nodes:
        if not isinstance(node, dict):
            continue
        if str(node.get("id") or "").strip() == cleaned_team_id:
            return node
        nested = _find_team_in_tree(cleaned_team_id, node.get("children") if isinstance(node.get("children"), list) else [])
        if nested:
            return nested
    return None


def _customer_team_record(team_id: Optional[str]) -> Optional[Dict[str, Any]]:
    return _find_team_in_tree(team_id, _customer_access_tree())


def _team_display_record(team_id: Optional[str]) -> Optional[Dict[str, Any]]:
    cleaned_team_id = str(team_id or "").strip()
    if not cleaned_team_id:
        return None
    return access_store.get_team(cleaned_team_id) or _customer_team_record(cleaned_team_id)


def _chain_request_id(prefix: str, *parts: Any) -> str:
    values = [str(prefix).strip()]
    for part in parts:
        if part in (None, ""):
            continue
        normalized = re.sub(r"[^A-Za-z0-9_.:@\\-]+", "-", str(part).strip())
        if normalized:
            values.append(normalized)
    return ":".join(values)


def _local_account_summary(scope: str, account_id: str) -> Dict[str, Any]:
    ledger = team_token_ledger if scope == "team" else token_ledger
    return ledger.get_summary(account_id)


def _chain_account_summary(scope: str, account_id: str) -> Dict[str, Any]:
    if not _chain_enabled():
        raise NmChainError("nmchain not configured")
    snapshot = _run_external_call(
        "nmchain account snapshot",
        NMCHAIN.account_snapshot,
        scope,
        account_id,
        timeout=_service_timeout_seconds(NMCHAIN, default=10.0),
    )
    return {
        "balance": int(snapshot.get("balance") or 0),
        "paid_balance": int(snapshot.get("paid_balance") or snapshot.get("balance") or 0),
        "free_balance": int(snapshot.get("free_balance") or 0),
        "last_topup_tokens": int(snapshot.get("last_topup_tokens") or snapshot.get("capacity") or 0),
        "last_topup_at": snapshot.get("last_topup_at"),
        "updated_at": snapshot.get("updated_at"),
        "spent_total": int(snapshot.get("spent_total") or 0),
        "cashout_total": int(snapshot.get("cashout_total") or 0),
        "shortfall_total": int(snapshot.get("shortfall_total") or 0),
        "free_grant_total": int(snapshot.get("free_grant_total") or 0),
        "reserved": int(snapshot.get("reserved") or 0),
        "available": int(snapshot.get("available") or 0),
        "identity": snapshot.get("identity") if isinstance(snapshot.get("identity"), dict) else None,
    }


def _account_summary(scope: str, account_id: str) -> Dict[str, Any]:
    if _billing_enabled():
        try:
            return _run_external_call(
                "billing account snapshot",
                BILLING_SERVICE.account_snapshot,
                scope,
                account_id,
                timeout=_service_timeout_seconds(BILLING_SERVICE, default=10.0),
            )
        except Exception as exc:
            logger.warning("billing summary fetch failed for %s/%s: %s", scope, account_id, exc)
    if CENTRAL_STORE is not None:
        return _local_account_summary(scope, account_id)
    if _chain_enabled():
        try:
            return _chain_account_summary(scope, account_id)
        except Exception as exc:
            logger.warning("nmchain summary fetch failed for %s/%s: %s", scope, account_id, exc)
    return _local_account_summary(scope, account_id)


def _account_entries(scope: str, account_id: str, limit: int = 50) -> List[Dict[str, Any]]:
    limit_val = max(1, min(int(limit), 500))
    if _billing_enabled():
        try:
            payload = _run_external_call(
                "billing ledger entries",
                BILLING_SERVICE.ledger_entries,
                scope,
                account_id,
                limit=limit_val,
                timeout=_service_timeout_seconds(BILLING_SERVICE, default=10.0),
            )
            entries = payload.get("entries")
            if isinstance(entries, list):
                return entries
        except Exception as exc:
            logger.warning("billing ledger fetch failed for %s/%s: %s", scope, account_id, exc)
    if CENTRAL_STORE is not None:
        ledger = team_token_ledger if scope == "team" else token_ledger
        return ledger.list_entries(account_id, limit=limit_val)
    if _chain_enabled():
        try:
            payload = _run_external_call(
                "nmchain ledger entries",
                NMCHAIN.ledger_entries,
                scope,
                account_id,
                limit=limit_val,
                timeout=_service_timeout_seconds(NMCHAIN, default=10.0),
            )
            entries = payload.get("entries")
            if isinstance(entries, list):
                return entries
        except Exception as exc:
            logger.warning("nmchain ledger fetch failed for %s/%s: %s", scope, account_id, exc)
    ledger = team_token_ledger if scope == "team" else token_ledger
    return ledger.list_entries(account_id, limit=limit_val)


def _record_token_event(
    scope: str,
    account_id: str,
    entry_type: str,
    delta: int,
    meta: Optional[Dict[str, Any]] = None,
    *,
    request_id: Optional[str] = None,
) -> Dict[str, Any]:
    meta_dict = dict(meta or {})
    if _billing_enabled():
        try:
            result = _run_external_call(
                "billing token event",
                BILLING_SERVICE.apply_token,
                scope,
                account_id,
                entry_type=entry_type,
                delta=int(delta),
                request_id=request_id,
                meta=meta_dict,
                timeout=_service_timeout_seconds(BILLING_SERVICE, default=10.0),
            )
            entry = result.get("entry")
            if isinstance(entry, dict):
                return entry
            snapshot = result.get("snapshot") if isinstance(result.get("snapshot"), dict) else {}
            return {
                "ts": _now_iso(),
                "type": entry_type,
                "user": account_id,
                "delta": int(delta),
                "balance_after": int(snapshot.get("balance") or 0),
                "meta": meta_dict,
                "shortfall": 0,
            }
        except Exception as exc:
            logger.warning("billing token event failed for %s/%s: %s", scope, account_id, exc)
    if CENTRAL_STORE is not None:
        ledger = team_token_ledger if scope == "team" else token_ledger
        return ledger.record(account_id, entry_type, int(delta), meta_dict, request_id=request_id)
    if _chain_enabled():
        result = _run_external_call(
            "nmchain token event",
            NMCHAIN.apply_token,
            scope,
            account_id,
            entry_type=entry_type,
            delta=int(delta),
            request_id=request_id,
            meta=meta_dict,
            timeout=_service_timeout_seconds(NMCHAIN, default=10.0),
        )
        entry = result.get("entry")
        if isinstance(entry, dict):
            return entry
        snapshot = result.get("snapshot") if isinstance(result.get("snapshot"), dict) else {}
        return {
            "ts": _now_iso(),
            "type": entry_type,
            "user": account_id,
            "delta": int(delta),
            "balance_after": int(snapshot.get("balance") or 0),
            "meta": meta_dict,
            "shortfall": 0,
        }
    ledger = team_token_ledger if scope == "team" else token_ledger
    return ledger.record(account_id, entry_type, int(delta), meta_dict)


def _capture_payment_event(
    user: str,
    tokens: int,
    meta: Optional[Dict[str, Any]] = None,
    *,
    request_id: Optional[str] = None,
) -> Dict[str, Any]:
    meta_dict = dict(meta or {})
    if _billing_enabled():
        try:
            result = _run_external_call(
                "billing payment capture",
                BILLING_SERVICE.capture_payment,
                user,
                tokens=int(tokens),
                amount_minor=_safe_int(meta_dict.get("amount_minor")),
                currency=str(meta_dict.get("currency") or "").strip() or None,
                provider=str(meta_dict.get("payment_provider") or meta_dict.get("provider") or "").strip() or None,
                payment_id=str(meta_dict.get("payment_id") or meta_dict.get("btc_txid") or "").strip() or None,
                checkout_flow=str(meta_dict.get("checkout_flow") or "").strip() or None,
                request_id=request_id,
                meta=meta_dict,
                timeout=_service_timeout_seconds(BILLING_SERVICE, default=10.0),
            )
            entry = result.get("entry")
            if isinstance(entry, dict):
                return entry
            snapshot = result.get("snapshot") if isinstance(result.get("snapshot"), dict) else {}
            return {
                "ts": _now_iso(),
                "type": "topup",
                "user": user,
                "delta": int(tokens),
                "balance_after": int(snapshot.get("balance") or 0),
                "meta": meta_dict,
                "shortfall": 0,
            }
        except Exception as exc:
            logger.warning("billing payment capture failed for %s: %s", user, exc)
    if CENTRAL_STORE is not None:
        return token_ledger.record(user, "topup", int(tokens), meta_dict, request_id=request_id)
    if _chain_enabled():
        result = _run_external_call(
            "nmchain payment capture",
            NMCHAIN.capture_payment,
            user,
            tokens=int(tokens),
            amount_minor=_safe_int(meta_dict.get("amount_minor")),
            currency=str(meta_dict.get("currency") or "").strip() or None,
            provider=str(meta_dict.get("payment_provider") or meta_dict.get("provider") or "").strip() or None,
            payment_id=str(meta_dict.get("payment_id") or meta_dict.get("btc_txid") or "").strip() or None,
            checkout_flow=str(meta_dict.get("checkout_flow") or "").strip() or None,
            request_id=request_id,
            meta=meta_dict,
            timeout=_service_timeout_seconds(NMCHAIN, default=10.0),
        )
        entry = result.get("entry")
        if isinstance(entry, dict):
            return entry
        snapshot = result.get("snapshot") if isinstance(result.get("snapshot"), dict) else {}
        return {
            "ts": _now_iso(),
            "type": "topup",
            "user": user,
            "delta": int(tokens),
            "balance_after": int(snapshot.get("balance") or 0),
            "meta": meta_dict,
            "shortfall": 0,
        }
    return token_ledger.record(user, "topup", int(tokens), meta_dict)


def _record_identity_event(
    user: str,
    *,
    role: Optional[str] = None,
    email: Optional[str] = None,
    provider: Optional[str] = None,
    subject: Optional[str] = None,
    request_id: Optional[str] = None,
    meta: Optional[Dict[str, Any]] = None,
) -> None:
    if not _chain_enabled():
        return
    details = dict(meta or {})
    identity = _current_auth_profile() or _current_auth_identity()
    if isinstance(identity, dict) and str(identity.get("user") or "").strip() == str(user or "").strip():
        for key in ("groups", "active_team", "team_count", "pending_invitation_count"):
            if key not in details and key in identity:
                details[key] = identity.get(key)
    elif role and "groups" not in details:
        details["groups"] = [str(role).strip().lower()]
    _submit_external_background_call(
        "nmchain identity upsert",
        NMCHAIN.upsert_identity,
        user,
        role=role,
        email=email,
        provider=provider,
        subject=subject,
        request_id=request_id,
        meta=details,
    )


def _record_login_event(
    user: str,
    *,
    auth_mode: str,
    provider: Optional[str] = None,
    session_id: Optional[str] = None,
    meta: Optional[Dict[str, Any]] = None,
) -> None:
    if not _chain_enabled():
        return
    details = dict(meta or {})
    if provider and "provider" not in details:
        details["provider"] = provider
    remote_addr = request.remote_addr if has_request_context() else None
    _submit_external_background_call(
        "nmchain login observe",
        NMCHAIN.observe_login,
        user,
        system=REFINER_CHAIN_SYSTEM,
        auth_mode=auth_mode,
        session_id=session_id,
        remote_addr=remote_addr,
        meta=details,
    )
if STT_LEARNING_ENABLED:
    try:
        stt_learning_store: Optional[SttLearningStore] = SttLearningStore(
            STT_LEARNING_ROOT,
            seed_paths=STT_KB_LOCAL_PATHS,
            seed_urls=STT_KB_SEED_URLS,
            allow_network=STT_LEARNING_ALLOW_NETWORK,
            prompt_terms=STT_LEARNING_PROMPT_TERMS,
            learn_min_count=STT_LEARNING_MIN_COUNT,
            max_memory_docs=STT_LEARNING_MAX_MEMORY_DOCS,
        )
    except Exception as exc:
        logger.warning("STT learning disabled after init failure: %s", exc)
        stt_learning_store = None
else:
    stt_learning_store = None


def _get_secret_store(user: str) -> SecretStore:
    key = user or "default"
    store = _secret_stores.get(key)
    if store:
        return store
    path = os.path.join(SECRET_STORE_ROOT, f"{key}.json")
    store = SecretStore(path)
    _secret_stores[key] = store
    return store


def _get_github_api_token(user: Optional[str]) -> Optional[str]:
    if not user:
        return None
    env = _get_secret_store(user).get_env()
    for key in ("GITHUB_TOKEN", "GH_TOKEN", "GITHUB_PAT"):
        if env.get(key):
            return env.get(key)
    return None


MCP_SERVER_NAME_RE = re.compile(r"^[A-Za-z0-9_.:-]{1,64}$")
MCP_HEADER_NAME_RE = re.compile(r"^[A-Za-z0-9-]{1,64}$")


def _mcp_internal_metadata(server: Optional[MCPServerConfig]) -> Dict[str, Any]:
    if not server or not isinstance(server.metadata, dict):
        return {}
    internal = server.metadata.get("_refiner")
    return dict(internal) if isinstance(internal, dict) else {}


def _mcp_secret_ref_name(server_name: str, suffix: str) -> str:
    safe_name = re.sub(r"[^A-Za-z0-9]+", "_", str(server_name or "").upper()).strip("_") or "SERVER"
    digest = hashlib.sha256(str(server_name or "").encode("utf-8")).hexdigest()[:10].upper()
    suffix_clean = re.sub(r"[^A-Za-z0-9]+", "_", str(suffix or "").upper()).strip("_") or "SECRET"
    budget = max(8, 63 - len("MCP__") - len(digest) - len(suffix_clean))
    return f"MCP_{safe_name[:budget]}_{digest}_{suffix_clean}"


def _mcp_secret_value(user: str, secret_ref: Optional[str]) -> Optional[str]:
    ref = str(secret_ref or "").strip()
    if not ref:
        return None
    if not SECRET_NAME_RE.match(ref):
        raise ValueError("secret reference must match the secret-name policy")
    return _get_secret_store(user).get(ref)


def _mcp_load_headers_secret(user: str, secret_ref: Optional[str]) -> Dict[str, str]:
    raw = _mcp_secret_value(user, secret_ref)
    if raw is None:
        return {}
    try:
        payload = json.loads(raw)
    except Exception as exc:
        raise ValueError(f"headers secret '{secret_ref}' does not contain valid JSON") from exc
    if not isinstance(payload, dict):
        raise ValueError(f"headers secret '{secret_ref}' must decode to an object")
    return _validate_mcp_headers(payload)


def _validate_mcp_headers(headers: Any) -> Dict[str, str]:
    if headers is None:
        return {}
    if not isinstance(headers, dict):
        raise ValueError("headers must be an object")
    normalized: Dict[str, str] = {}
    for key, value in headers.items():
        name = str(key or "").strip()
        if not MCP_HEADER_NAME_RE.match(name):
            raise ValueError(f"header name '{name or key}' is invalid")
        if value is None:
            continue
        text = str(value)
        if "\r" in text or "\n" in text:
            raise ValueError(f"header '{name}' contains an invalid newline")
        if len(text) > 4096:
            raise ValueError(f"header '{name}' exceeds 4096 characters")
        normalized[name] = text
    return normalized


def _mcp_runtime_patch(action: str, *, status: str, duration_ms: int, error: Optional[str] = None, **extra: Any) -> Dict[str, Any]:
    patch: Dict[str, Any] = {
        "last_action": action,
        "last_status": status,
        "last_checked_at": _now_iso(),
        "last_duration_ms": int(max(0, duration_ms)),
    }
    if status == "success":
        patch["last_success_at"] = patch["last_checked_at"]
        patch["last_error"] = None
    else:
        patch["last_failure_at"] = patch["last_checked_at"]
        patch["last_error"] = error or "mcp_request_failed"
    for key, value in extra.items():
        patch[key] = value
    return patch


def _mcp_public_metadata(server: Optional[MCPServerConfig]) -> Dict[str, Any]:
    if not server or not isinstance(server.metadata, dict):
        return {}
    return {key: value for key, value in server.metadata.items() if key != "_refiner"}


def _mcp_masked_server(user: str, server: MCPServerConfig) -> Dict[str, Any]:
    masked = server.masked()
    if masked.get("headers"):
        return masked
    if not server.headers_secret_ref:
        return masked
    try:
        header_keys = sorted(_mcp_load_headers_secret(user, server.headers_secret_ref).keys())
    except Exception:
        return masked
    masked["headers"] = {key: "***" for key in header_keys}
    return masked


def _mcp_resolve_server(user: str, server: MCPServerConfig) -> MCPServerConfig:
    headers = {}
    if isinstance(server.headers, dict):
        headers.update(_validate_mcp_headers(server.headers))
    if server.headers_secret_ref:
        headers.update(_mcp_load_headers_secret(user, server.headers_secret_ref))
    auth_token = server.auth_token
    if server.auth_type != "none" and server.auth_secret_ref:
        auth_token = _mcp_secret_value(user, server.auth_secret_ref)
        if auth_token is None:
            raise ValueError(f"configured auth secret '{server.auth_secret_ref}' is missing")
    return MCPServerConfig(
        name=server.name,
        base_url=server.base_url,
        auth_type=server.auth_type,
        auth_token=auth_token,
        headers=headers or None,
        timeout=int(server.timeout or 20),
        auth_secret_ref=server.auth_secret_ref,
        headers_secret_ref=server.headers_secret_ref,
        metadata=dict(server.metadata or {}),
        runtime=dict(server.runtime or {}),
        created_at=server.created_at,
        updated_at=server.updated_at,
    )


def _mcp_cleanup_managed_secrets(user: str, server: Optional[MCPServerConfig]) -> None:
    internal = _mcp_internal_metadata(server)
    managed = internal.get("managed_secret_refs") if isinstance(internal.get("managed_secret_refs"), dict) else {}
    store = _get_secret_store(user)
    for ref in managed.values():
        cleaned = str(ref or "").strip()
        if cleaned and SECRET_NAME_RE.match(cleaned):
            store.delete(cleaned)


def _prepare_mcp_server_config(
    user: str,
    payload: Dict[str, Any],
    *,
    current: Optional[MCPServerConfig] = None,
) -> MCPServerConfig:
    name = str(payload.get("name") or (current.name if current else "")).strip()
    if not MCP_SERVER_NAME_RE.match(name):
        raise ValueError("name must use 1-64 characters from A-Z, a-z, 0-9, dot, underscore, colon, or dash")

    base_url = str(payload.get("base_url") or (current.base_url if current else "")).strip()
    parsed = urlparse(base_url)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise ValueError("base_url must be a valid http or https URL")
    if not url_allowed(base_url):
        raise ValueError("base_url is blocked by the current URL safety policy")

    auth_type = str(payload.get("auth_type") or (current.auth_type if current else "bearer")).strip().lower() or "bearer"
    if auth_type not in {"bearer", "oauth", "none"}:
        raise ValueError("auth_type must be bearer, oauth, or none")

    timeout_value = payload.get("timeout", current.timeout if current else 20)
    timeout = _safe_int(timeout_value, 20) or 20
    if timeout < 1 or timeout > 120:
        raise ValueError("timeout must be between 1 and 120 seconds")

    metadata = _mcp_public_metadata(current)
    if "metadata" in payload:
        if payload.get("metadata") is None:
            metadata = {}
        elif not isinstance(payload.get("metadata"), dict):
            raise ValueError("metadata must be an object")
        else:
            metadata = dict(payload.get("metadata") or {})

    runtime = dict(current.runtime or {}) if current and isinstance(current.runtime, dict) else {}
    runtime["last_configured_at"] = _now_iso()

    internal = _mcp_internal_metadata(current)
    # Only refs created by Refiner are tracked here so delete/update can clean
    # them up safely without removing user-managed shared secret entries.
    managed_refs = internal.get("managed_secret_refs") if isinstance(internal.get("managed_secret_refs"), dict) else {}
    managed_refs = dict(managed_refs or {})
    store = _get_secret_store(user)

    if auth_type == "none":
        if current and current.auth_secret_ref and managed_refs.get("auth") == current.auth_secret_ref:
            store.delete(current.auth_secret_ref)
        auth_secret_ref = None
        managed_refs.pop("auth", None)
    else:
        auth_secret_ref = current.auth_secret_ref if current else None
        if "auth_token" in payload and "auth_secret_ref" in payload and str(payload.get("auth_token") or "").strip():
            raise ValueError("provide auth_token or auth_secret_ref, not both")
        if "auth_token" in payload:
            raw_token = str(payload.get("auth_token") or "").strip()
            if raw_token:
                ref = managed_refs.get("auth") or auth_secret_ref or _mcp_secret_ref_name(name, "AUTH")
                store.set(ref, raw_token)
                auth_secret_ref = ref
                managed_refs["auth"] = ref
            else:
                if current and current.auth_secret_ref and managed_refs.get("auth") == current.auth_secret_ref:
                    store.delete(current.auth_secret_ref)
                auth_secret_ref = None
                managed_refs.pop("auth", None)
        elif "auth_secret_ref" in payload:
            ref = str(payload.get("auth_secret_ref") or "").strip() or None
            if ref:
                if not SECRET_NAME_RE.match(ref):
                    raise ValueError("auth_secret_ref does not match the secret-name policy")
                if _mcp_secret_value(user, ref) is None:
                    raise ValueError(f"auth secret '{ref}' was not found")
            if current and current.auth_secret_ref and managed_refs.get("auth") == current.auth_secret_ref and current.auth_secret_ref != ref:
                store.delete(current.auth_secret_ref)
            auth_secret_ref = ref
            managed_refs.pop("auth", None)

    raw_headers = payload.get("headers") if "headers" in payload else None
    if "headers" in payload and raw_headers is not None and not isinstance(raw_headers, dict):
        raise ValueError("headers must be an object")
    if "headers" in payload and "headers_secret_ref" in payload and isinstance(raw_headers, dict) and raw_headers:
        raise ValueError("provide headers or headers_secret_ref, not both")
    headers_secret_ref = current.headers_secret_ref if current else None
    header_keys: List[str] = []
    if "headers" in payload:
        normalized_headers = _validate_mcp_headers(raw_headers or {})
        if normalized_headers:
            ref = managed_refs.get("headers") or headers_secret_ref or _mcp_secret_ref_name(name, "HEADERS")
            store.set(ref, json.dumps(normalized_headers, sort_keys=True))
            headers_secret_ref = ref
            managed_refs["headers"] = ref
            header_keys = sorted(normalized_headers.keys())
        else:
            if current and current.headers_secret_ref and managed_refs.get("headers") == current.headers_secret_ref:
                store.delete(current.headers_secret_ref)
            headers_secret_ref = None
            managed_refs.pop("headers", None)
    elif "headers_secret_ref" in payload:
        ref = str(payload.get("headers_secret_ref") or "").strip() or None
        if ref:
            if not SECRET_NAME_RE.match(ref):
                raise ValueError("headers_secret_ref does not match the secret-name policy")
            header_keys = sorted(_mcp_load_headers_secret(user, ref).keys())
        if current and current.headers_secret_ref and managed_refs.get("headers") == current.headers_secret_ref and current.headers_secret_ref != ref:
            store.delete(current.headers_secret_ref)
        headers_secret_ref = ref
        managed_refs.pop("headers", None)
    elif current and current.headers_secret_ref:
        try:
            header_keys = sorted(_mcp_load_headers_secret(user, current.headers_secret_ref).keys())
        except Exception:
            header_keys = current.header_keys()

    if header_keys:
        internal["header_keys"] = header_keys
    else:
        internal.pop("header_keys", None)
    if managed_refs:
        internal["managed_secret_refs"] = managed_refs
    else:
        internal.pop("managed_secret_refs", None)
    if internal:
        metadata["_refiner"] = internal
    else:
        metadata.pop("_refiner", None)

    return MCPServerConfig(
        name=name,
        base_url=base_url,
        auth_type=auth_type,
        timeout=timeout,
        auth_secret_ref=auth_secret_ref,
        headers_secret_ref=headers_secret_ref,
        metadata=metadata,
        runtime=runtime,
        created_at=current.created_at if current else None,
    )


def _mcp_result_items(result: Dict[str, Any], key: str) -> List[Any]:
    if not isinstance(result, dict):
        return []
    for candidate in (result.get("result"), result):
        if isinstance(candidate, dict):
            items = candidate.get(key)
            if isinstance(items, list):
                return items
    return []


def _mcp_execute(
    user: str,
    server_name: str,
    action: str,
    operation: Any,
    *,
    audit_details: Optional[Dict[str, Any]] = None,
    runtime_from_result: Optional[Any] = None,
) -> Dict[str, Any]:
    server = mcp_store.get_server(user, server_name)
    if not server:
        raise KeyError("not_found")
    resolved = _mcp_resolve_server(user, server)
    start = time.time()
    base_details = {
        "server": server_name,
        "host": urlparse(server.base_url).hostname or "",
    }
    if isinstance(audit_details, dict):
        base_details.update(audit_details)
    try:
        client = MCPClient(resolved)
        client.initialize({"name": "refiner"})
        result = operation(client)
    except Exception as exc:
        duration_ms = int((time.time() - start) * 1000)
        mcp_store.update_runtime(
            user,
            server_name,
            _mcp_runtime_patch(action, status="failed", duration_ms=duration_ms, error=str(exc)),
        )
        _audit_event(f"mcp_{action}", actor=user, status="failed", details={**base_details, "error": str(exc)})
        raise
    duration_ms = int((time.time() - start) * 1000)
    runtime_patch = _mcp_runtime_patch(action, status="success", duration_ms=duration_ms)
    if callable(runtime_from_result):
        try:
            extra = runtime_from_result(result)
        except Exception:
            extra = {}
        if isinstance(extra, dict):
            runtime_patch.update(extra)
    mcp_store.update_runtime(user, server_name, runtime_patch)
    _audit_event(f"mcp_{action}", actor=user, status="success", details=base_details)
    return result


def _assistant_execute_atlassian_action(user: str, action_request: Dict[str, Any]) -> Dict[str, Any]:
    start = time.time()
    product = str(action_request.get("product") or "").strip().lower()
    action = str(action_request.get("action") or "").strip().lower()
    instance = str(action_request.get("instance") or "").strip()
    preview = bool(action_request.get("preview"))
    details = {
        "product": product,
        "action": action,
        "instance": instance,
        "preview": preview,
    }
    try:
        result = _execute_atlassian_action_impl(action_request)
    except Exception as exc:
        details["error"] = str(exc)
        details["duration_ms"] = int((time.time() - start) * 1000)
        _audit_event("assistant_atlassian_action", actor=user, status="failed", details=details)
        raise
    details["duration_ms"] = int((time.time() - start) * 1000)
    resolved_instance = str(result.get("instance") or "").strip()
    if resolved_instance:
        details["instance"] = resolved_instance
    _audit_event("assistant_atlassian_action", actor=user, status="success", details=details)
    return result


def _load_llm_config() -> Dict[str, Any]:
    path = os.path.join(BASE_DIR, "config.json")
    try:
        with open(path, "r", encoding="utf-8") as handle:
            data = json.load(handle)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _resolve_llm_settings(
    user: str,
    provider_hint: Optional[str] = None,
    model_hint: Optional[str] = None,
) -> Dict[str, Any]:
    user_settings = _user_settings(user)
    user_role = _user_role(user)
    secret_env = _get_secret_store(user).get_env()
    env = _effective_llm_env_for_user(
        secret_env,
        process_env=os.environ,
        settings=user_settings,
        user=user,
        role=user_role,
    )
    cfg = _load_llm_config()
    providers = cfg.get("llm_providers") if isinstance(cfg.get("llm_providers"), list) else []
    defaults = _settings_defaults_for_user(user)

    provider_type = None
    cleaned_provider_hint = str(provider_hint or "").strip() or None
    cleaned_model_hint = str(model_hint or "").strip() or None
    model = None
    base_url = None

    preferred_provider = cleaned_provider_hint or defaults.get("provider")
    if preferred_provider:
        match = next((p for p in providers if p.get("name") == preferred_provider), None)
        if match:
            candidate_type = _normalize_llm_provider_type(match.get("type") or preferred_provider) or preferred_provider
            if _provider_has_accessible_credentials(candidate_type, env):
                provider_type = candidate_type
                model = cleaned_model_hint or match.get("model") or defaults.get("model")
                base_url = match.get("base_url")
        else:
            candidate_type = _normalize_llm_provider_type(preferred_provider) or preferred_provider
            if _provider_has_accessible_credentials(candidate_type, env):
                provider_type = candidate_type
                model = cleaned_model_hint or defaults.get("model")

    if not provider_type:
        configured = _accessible_configured_llm_provider(providers, env)
        if configured:
            provider_type = configured.get("provider")
            model = configured.get("model") or defaults.get("model")
            base_url = configured.get("base_url")
        elif env.get("OPENAI_API_KEY"):
            provider_type = "openai"
            model = cleaned_model_hint or defaults.get("model")
        elif env.get("NVIDIA_API_KEY"):
            provider_type = "nvidia"
            model = cleaned_model_hint or defaults.get("model")
        elif _gemini_credential_from_env(env):
            provider_type = "gemini"
            model = cleaned_model_hint or defaults.get("model")
        else:
            provider_type = "ollama"
            model = cleaned_model_hint or defaults.get("model")

    provider_type = _normalize_llm_provider_type(provider_type) or provider_type

    if provider_type == "openai":
        api_key = env.get("OPENAI_API_KEY")
    elif provider_type == "nvidia":
        api_key = env.get("NVIDIA_API_KEY")
    elif provider_type == "gemini":
        api_key = _gemini_credential_from_env(env)
    else:
        api_key = None

    base_url = _provider_base_url(provider_type, base_url, env)
    billing = _provider_billing_metadata(
        provider_type,
        settings=user_settings,
        user=user,
        role=user_role,
        secret_env=secret_env,
        process_env=os.environ,
    ) or {}

    return {
        "provider": provider_type,
        "model": model,
        "base_url": base_url,
        "api_key": api_key,
        "credential_source": billing.get("credential_source"),
        "chargeable": bool(billing.get("chargeable", False)),
        "reasoning_effort": defaults.get("reasoning_effort"),
        "assistant_profile": defaults.get("assistant_profile"),
        "assistant_use_memory": defaults.get("assistant_use_memory"),
        "command_policy_mode": defaults.get("command_policy_mode"),
        "show_solver_replay": defaults.get("show_solver_replay"),
        "settings": user_settings,
    }


def _normalize_llm_provider_type(provider_type: Optional[str]) -> Optional[str]:
    lowered = str(provider_type or "").strip().lower()
    if lowered in {"gpt", "chatgpt"}:
        return "openai"
    if lowered == "google":
        return "gemini"
    if lowered in {"nim", "nvidia_nim"}:
        return "nvidia"
    return lowered or None


def _api_key_for_provider_type(provider_type: Optional[str], env: Dict[str, str]) -> Optional[str]:
    lowered = _normalize_llm_provider_type(provider_type)
    if lowered == "openai":
        return env.get("OPENAI_API_KEY")
    if lowered == "nvidia":
        return env.get("NVIDIA_API_KEY")
    if lowered == "gemini":
        return _gemini_credential_from_env(env)
    return None


def _fallback_llm_settings(user: str, primary: Dict[str, Any]) -> Dict[str, Any]:
    user_settings = _user_settings(user)
    user_role = _user_role(user)
    secret_env = _get_secret_store(user).get_env()
    env = _effective_llm_env_for_user(
        secret_env,
        process_env=os.environ,
        settings=user_settings,
        user=user,
        role=user_role,
    )
    cfg = _load_llm_config()
    providers = cfg.get("llm_providers") if isinstance(cfg.get("llm_providers"), list) else []
    primary_provider = (
        _normalize_llm_provider_type(primary.get("provider"))
        or str(primary.get("provider") or "").strip().lower()
    )
    primary_model = str(primary.get("model") or "").strip()

    fallback = _accessible_configured_llm_provider(
        providers,
        env,
        exclude_provider=primary_provider,
        exclude_model=primary_model,
    )
    if fallback:
        billing = _provider_billing_metadata(
            fallback.get("provider"),
            settings=user_settings,
            user=user,
            role=user_role,
            secret_env=secret_env,
            process_env=os.environ,
        ) or {}
        return {
            "provider": fallback.get("provider"),
            "model": fallback.get("model"),
            "base_url": fallback.get("base_url"),
            "api_key": fallback.get("api_key"),
            "credential_source": billing.get("credential_source"),
            "chargeable": bool(billing.get("chargeable", False)),
        }
    return {
        "provider": None,
        "model": None,
        "base_url": None,
        "api_key": None,
        "credential_source": None,
        "chargeable": False,
    }


def _build_request_llm_provider(
    user: str,
    settings: Dict[str, Any],
    *,
    workflow: str,
    role: str = "assistant",
) -> Optional[LLMProvider]:
    fallback = _fallback_llm_settings(user, settings)
    user_role = _user_role(user)
    user_settings = _user_settings(user)
    secret_env = _get_secret_store(user).get_env()
    include_configured = None if _user_can_use_shared_llm_credentials(user, role=user_role) else False
    provider = build_workflow_provider(
        workflow=workflow,
        role=role,
        preferred_provider=settings.get("provider"),
        preferred_model=settings.get("model"),
        preferred_api_key=settings.get("api_key"),
        fallback_provider=fallback.get("provider"),
        fallback_model=fallback.get("model"),
        fallback_api_key=fallback.get("api_key"),
        base_url=settings.get("base_url") or fallback.get("base_url"),
        include_configured=include_configured,
        provider_factory=get_provider,
        config_path=os.path.join(BASE_DIR, "config.json"),
        inter_request_gap=0.0,
    )
    provider_billing_map = _provider_billing_map(
        settings=user_settings,
        user=user,
        role=user_role,
        secret_env=secret_env,
        process_env=os.environ,
    )
    if provider is not None:
        initial_provider = _normalize_llm_provider_type(settings.get("provider")) or getattr(provider, "name", None)
        initial_metadata = provider_billing_map.get(_normalize_llm_provider_type(initial_provider))
        _attach_provider_billing_metadata(
            provider,
            billing_map=provider_billing_map,
            initial_metadata=initial_metadata,
        )
    return provider


def _attach_provider_billing_metadata(
    provider: Any,
    *,
    billing_map: Dict[str, Dict[str, Any]],
    initial_metadata: Optional[Dict[str, Any]] = None,
) -> None:
    if provider is None:
        return
    try:
        setattr(provider, "billing_metadata_map", dict(billing_map))
    except Exception as exc:
        logger.debug("unable to attach provider billing metadata map: %s", exc)
        return
    if not initial_metadata:
        return
    try:
        setattr(provider, "billing_metadata", dict(initial_metadata))
    except Exception as exc:
        logger.debug("unable to attach provider billing metadata: %s", exc)


def _ai_orchestration_status(*, probe_engines: bool = False, candidate_limit: int = 12) -> Dict[str, Any]:
    """Return orchestration registry status without breaking health/admin routes."""

    try:
        payload = orchestration_status(
            config_path=os.path.join(BASE_DIR, "config.json"),
            include_metrics=True,
            probe_engines=probe_engines,
            candidate_limit=candidate_limit,
        )
        model_inventory = payload.get("model_inventory") if isinstance(payload.get("model_inventory"), dict) else {}
        model_inventory["monitor"] = ai_model_inventory_monitor.status()
        payload["model_inventory"] = model_inventory
        return payload
    except Exception as exc:
        logger.warning("ai orchestration status unavailable: %s", exc)
        return {
            "enabled": False,
            "degraded": True,
            "error": "ai_orchestration_status_unavailable",
            "details": str(exc),
        }


def _apply_user_job_defaults(payload: Dict[str, Any], owner: str) -> None:
    if not isinstance(payload, dict):
        return
    defaults = _settings_defaults_for_user(owner)
    if not payload.get("llm_provider") and defaults.get("provider"):
        payload["llm_provider"] = defaults.get("provider")
    if not payload.get("llm_model") and defaults.get("model"):
        payload["llm_model"] = defaults.get("model")
    if not payload.get("llm_reasoning_effort") and defaults.get("reasoning_effort"):
        payload["llm_reasoning_effort"] = defaults.get("reasoning_effort")
    if payload.get("codingagent") and not payload.get("codingagent_reasoning_effort") and defaults.get("reasoning_effort"):
        payload["codingagent_reasoning_effort"] = defaults.get("reasoning_effort")
    if not payload.get("solver_command_policy_mode") and defaults.get("command_policy_mode"):
        payload["solver_command_policy_mode"] = defaults.get("command_policy_mode")


def _guardrail_scan(text: str) -> Optional[str]:
    if not text:
        return None
    lower = text.lower()
    for category, patterns in GUARDRAIL_PATTERNS.items():
        for pat in patterns:
            if re.search(pat, lower, flags=re.IGNORECASE | re.DOTALL):
                return f"Detected {category} content via pattern: {pat}"
    return None


def _read_file_limited(path: str, max_bytes: int) -> str:
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as handle:
            return handle.read(max_bytes)
    except Exception:
        return ""


def _read_json_file(path: str) -> Optional[Dict[str, Any]]:
    try:
        with open(path, "r", encoding="utf-8") as handle:
            return json.load(handle)
    except Exception:
        return None


def _iso_from_mtime(path: str) -> Optional[str]:
    try:
        mtime = os.path.getmtime(path)
    except Exception:
        return None
    return dt.datetime.fromtimestamp(mtime, UK_TZ).strftime(UK_DATETIME_FORMAT)


def _mtime_dt(path: str) -> Optional[dt.datetime]:
    try:
        mtime = os.path.getmtime(path)
    except Exception:
        return None
    return dt.datetime.fromtimestamp(mtime, UK_TZ)


def _safe_int(value: Any, default: int = 0) -> int:
    try:
        if value is None:
            return default
        return int(value)
    except Exception:
        return default


def _safe_float(value: Any, default: float = 0.0) -> float:
    try:
        if value is None:
            return default
        return float(value)
    except Exception:
        return default


def _is_truthy(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "on", "y"}
    return False


def _requirements_progress_from_solution(solution: Dict[str, Any]) -> Dict[str, Any]:
    total = 0
    completed = 0
    in_progress = 0
    remaining = 0
    source = "none"
    trace = solution.get("requirement_traceability") if isinstance(solution, dict) else None
    summary = trace.get("summary") if isinstance(trace, dict) else None
    if isinstance(summary, dict):
        total = _safe_int(summary.get("total"))
        completed = _safe_int(summary.get("with_changes"))
        in_progress = _safe_int(summary.get("planned_only"))
        remaining = _safe_int(summary.get("unmapped"))
        if total <= 0:
            total = completed + in_progress + remaining
        if total > 0:
            remaining = max(total - completed - in_progress, 0)
        source = "traceability"
    elif isinstance(trace, dict):
        reqs = trace.get("requirements")
        if isinstance(reqs, list) and reqs:
            for entry in reqs:
                if not isinstance(entry, dict):
                    continue
                status = str(entry.get("status") or "").lower()
                total += 1
                if status == "covered":
                    completed += 1
                elif status == "planned":
                    in_progress += 1
                else:
                    remaining += 1
            source = "traceability"
    if total == 0:
        register = solution.get("requirements_register") if isinstance(solution, dict) else None
        if isinstance(register, dict):
            reqs = register.get("requirements")
            if isinstance(reqs, list) and reqs:
                total = len([entry for entry in reqs if entry is not None])
                remaining = total
                source = "register"
    return {
        "total": total,
        "completed": completed,
        "in_progress": in_progress,
        "remaining": remaining,
        "source": source,
    }


def _requirements_progress_from_requirements_file(path: str) -> Optional[Dict[str, Any]]:
    if not path or not os.path.exists(path):
        return None
    text = _read_file_limited(path, 250_000)
    if not text:
        return None
    total = _count_req_lines(text)
    if total <= 0:
        return None
    return {
        "total": total,
        "completed": 0,
        "in_progress": 0,
        "remaining": total,
        "source": "requirements_file",
    }


def _normalize_summary_text(text: str, max_len: int = 240) -> str:
    if not text:
        return ""
    cleaned = " ".join(text.strip().split())
    if len(cleaned) <= max_len:
        return cleaned
    return cleaned[: max_len - 1].rstrip() + "…"


def _requirements_summary_from_register(solution: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    register = solution.get("requirements_register") if isinstance(solution, dict) else None
    if not isinstance(register, dict):
        return None
    reqs = register.get("requirements")
    if not isinstance(reqs, list) or not reqs:
        return None
    items = []
    for entry in reqs:
        if not isinstance(entry, dict):
            continue
        req_id = entry.get("id") or ""
        title = entry.get("title") or ""
        desc = entry.get("description") or ""
        source = entry.get("source")
        title = _normalize_summary_text(str(title)) if title else ""
        desc = _normalize_summary_text(str(desc), max_len=280) if desc else ""
        if title and desc and title.lower() in desc.lower():
            text = desc
        elif title and desc:
            text = f"{title}. {desc}"
        else:
            text = title or desc
        if text:
            items.append({"id": req_id, "text": text, "source": source})
    if not items:
        return None
    return {"items": items, "source": "register", "total": len(items)}


def _requirements_summary_from_text(text: str, max_items: int = 12) -> Dict[str, Any]:
    paragraphs = [chunk.strip() for chunk in re.split(r"\n\s*\n", text) if chunk.strip()]
    lead = paragraphs[0] if paragraphs else ""
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    bullet_re = re.compile(r"^(?:[-*+]\s+|\d+[.)]\s+)")
    bullets = [bullet_re.sub("", line).strip() for line in lines if bullet_re.match(line)]
    items = []
    if bullets:
        for line in bullets[:max_items]:
            if line:
                items.append({"id": "", "text": _normalize_summary_text(line, max_len=280)})
    else:
        sentences = []
        for para in paragraphs:
            for sentence in re.split(r"(?<=[.!?])\s+", para):
                cleaned = sentence.strip()
                if len(cleaned) >= 24:
                    sentences.append(cleaned)
        for sentence in sentences[:max_items]:
            items.append({"id": "", "text": _normalize_summary_text(sentence, max_len=280)})
    return {
        "summary": _normalize_summary_text(lead, max_len=320),
        "items": items,
        "source": "requirements_file",
        "total": len(items),
    }


def _completion_summary_from_output(output_path: Optional[str]) -> Optional[Dict[str, Any]]:
    if not output_path or not os.path.exists(output_path):
        return None
    data = _read_json_file(output_path)
    if not isinstance(data, dict):
        return None
    summary = data.get("completion_summary")
    if isinstance(summary, dict):
        return summary
    return None


def _completion_reason_from_summary(summary: Optional[Dict[str, Any]]) -> Optional[str]:
    if not isinstance(summary, dict):
        return None
    exhausted = summary.get("iterations_exhausted_sources")
    if summary.get("max_steps_reached"):
        return "steps"
    if isinstance(exhausted, list) and exhausted:
        return "iterations"
    if bool(summary.get("needs_more_iterations")):
        return "incomplete"
    return None


def _completion_reason_from_output(output_path: Optional[str]) -> Optional[str]:
    return _completion_reason_from_summary(_completion_summary_from_output(output_path))


def _job_output_insights_from_output(output_path: Optional[str]) -> Dict[str, Any]:
    """Return cached solver detail fields extracted from a job output JSON file."""

    if not output_path or not os.path.exists(output_path):
        return {}
    try:
        mtime = os.path.getmtime(output_path)
    except Exception:
        return {}

    with _job_output_insights_cache_lock:
        cached = _job_output_insights_cache.get(output_path)
        if cached and cached.get("mtime") == mtime:
            payload = cached.get("data")
            return dict(payload) if isinstance(payload, dict) else {}

    payload: Dict[str, Any] = {}
    data = _read_json_file(output_path)
    if isinstance(data, dict):
        completion_summary = data.get("completion_summary")
        if isinstance(completion_summary, dict):
            payload["completion_summary"] = completion_summary
        solver_replay_analysis = data.get("solver_replay_analysis")
        if isinstance(solver_replay_analysis, dict):
            payload["solver_replay_analysis"] = solver_replay_analysis

    with _job_output_insights_cache_lock:
        _job_output_insights_cache[output_path] = {"mtime": mtime, "data": dict(payload)}
    return payload


_GLOBAL_REQUIREMENTS_CACHE: Optional[List[Dict[str, str]]] = None
_GLOBAL_REQUIREMENTS_TITLES: Optional[List[str]] = None


def _load_global_requirements() -> List[Dict[str, str]]:
    global _GLOBAL_REQUIREMENTS_CACHE
    if _GLOBAL_REQUIREMENTS_CACHE is not None:
        return _GLOBAL_REQUIREMENTS_CACHE
    items: List[Dict[str, str]] = []
    try:
        from refiner.project_solver import GLOBAL_REQUIREMENTS
    except Exception:
        _GLOBAL_REQUIREMENTS_CACHE = items
        return items
    if isinstance(GLOBAL_REQUIREMENTS, list):
        for entry in GLOBAL_REQUIREMENTS:
            if not isinstance(entry, dict):
                continue
            title = str(entry.get("title") or "").strip()
            desc = str(entry.get("description") or "").strip()
            if not title and not desc:
                continue
            items.append({"title": title, "description": desc})
    _GLOBAL_REQUIREMENTS_CACHE = items
    return items


def _global_requirements_titles() -> List[str]:
    global _GLOBAL_REQUIREMENTS_TITLES
    if _GLOBAL_REQUIREMENTS_TITLES is not None:
        return _GLOBAL_REQUIREMENTS_TITLES
    titles: List[str] = []
    for entry in _load_global_requirements():
        title = str(entry.get("title") or "").strip()
        if title:
            titles.append(title.lower())
    titles = sorted(set(titles), key=len, reverse=True)
    _GLOBAL_REQUIREMENTS_TITLES = titles
    return titles


def _global_requirements_count() -> int:
    return len(_load_global_requirements())


def _normalise_requirement_text(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().lower()


def _global_requirements_summary_items(max_items: int = 0) -> List[Dict[str, str]]:
    items: List[Dict[str, str]] = []
    for idx, entry in enumerate(_load_global_requirements(), start=1):
        title = entry.get("title") or ""
        desc = entry.get("description") or ""
        if title and desc and title.lower() in desc.lower():
            text = desc
        elif title and desc:
            text = f"{title}. {desc}"
        else:
            text = title or desc
        text = _normalize_summary_text(text, max_len=280) if text else ""
        if not text:
            continue
        items.append({"id": f"GLOBAL-{idx:02d}", "text": text, "source": ["global"]})
        if max_items and len(items) >= max_items:
            break
    return items


def _append_global_requirements_summary(summary: Dict[str, Any], redact: bool = False) -> Dict[str, Any]:
    items = summary.get("items")
    if not isinstance(items, list):
        items = []
    existing = {_normalise_requirement_text(item.get("text")) for item in items if isinstance(item, dict)}
    redacted_any = False
    for item in _global_requirements_summary_items():
        norm = _normalise_requirement_text(item.get("text"))
        if not norm or norm in existing:
            continue
        if redact:
            redacted = dict(item)
            redacted["text"] = "[redacted]"
            items.append(redacted)
            redacted_any = True
        else:
            items.append(item)
        existing.add(norm)
    summary["items"] = items
    summary["total"] = len(items)
    source = summary.get("source") or "requirements_file"
    if "global" not in str(source):
        summary["source"] = f"{source}+global"
    if redact and redacted_any:
        summary["redacted"] = True
    return summary


def _is_admin_user(user: Optional[str]) -> bool:
    if not user:
        return False
    identity = _current_auth_profile() or _current_auth_identity()
    if isinstance(identity, dict) and str(identity.get("user") or "").strip() == user:
        if _identity_has_service_access(identity, "refiner", SERVICE_ACCESS_CONTROL):
            return True
        return bool(identity.get("is_admin"))
    return _user_role(user) == "admin"


def _job_project_id(job: "Job") -> Optional[str]:
    if not job or not isinstance(job.payload, dict):
        return None
    return job.payload.get("project_id") or job.payload.get("project")


def _job_team_id(job: "Job") -> Optional[str]:
    if not job or not isinstance(job.payload, dict):
        return None
    return job.payload.get("team_id")


def _job_role_for_user(user: Optional[str], job: "Job") -> Optional[str]:
    if not user or not job:
        return None
    if _is_admin_user(user):
        return "admin"
    if job.owner == user:
        return "owner"
    project_id = _job_project_id(job)
    if not project_id:
        team_id = _job_team_id(job)
        if team_id:
            team_role = _team_role_for_user(user, team_id)
            if team_role == "leader":
                return "leader"
            if team_role == "member":
                return "contributor"
        return None
    return access_store.project_role(user, project_id)


def _job_capabilities_for_user(user: Optional[str], job: "Job") -> Dict[str, bool]:
    if not user or not job:
        return {"read": False, "write": False, "grant": False}
    if _is_admin_user(user):
        return {"read": True, "write": True, "grant": True}
    if job.owner == user:
        return {"read": True, "write": True, "grant": True}
    project_id = _job_project_id(job)
    if project_id:
        return access_store.project_capabilities(user, project_id)
    team_id = _job_team_id(job)
    if team_id:
        team_role = _team_role_for_user(user, team_id)
        if team_role == "leader":
            return {"read": True, "write": True, "grant": True}
        if team_role == "member":
            return {"read": True, "write": True, "grant": False}
    return {"read": False, "write": False, "grant": False}


def _can_view_job(user: Optional[str], job: "Job") -> bool:
    caps = _job_capabilities_for_user(user, job)
    if caps.get("read"):
        return True
    transfer = getattr(job, "transfer_request", None)
    if isinstance(transfer, dict) and transfer.get("status") == "pending":
        team_id = transfer.get("team_id")
        if _is_admin_user(user) or _team_role_for_user(user, team_id) == "leader":
            return True
    return False


def _can_manage_job(user: Optional[str], job: "Job") -> bool:
    caps = _job_capabilities_for_user(user, job)
    return bool(caps.get("write"))


def _augment_job_dict_for_user(data: Dict[str, Any], user: Optional[str], job: "Job") -> Dict[str, Any]:
    if not isinstance(data, dict):
        return data
    project_id = _job_project_id(job)
    team_id = _job_team_id(job)
    data["project_id"] = project_id
    data["team_id"] = team_id
    role = _job_role_for_user(user, job)
    if role:
        data["project_role"] = role
    data["project_capabilities"] = _job_capabilities_for_user(user, job)
    if project_id:
        project = access_store.get_project(project_id)
        if project:
            data["project_name"] = project.get("name") or data.get("project_name")
            team = _team_display_record(project.get("team_id")) if project.get("team_id") else None
            if team:
                data["team_name"] = team.get("name")
    elif team_id:
        team = access_store.get_team(team_id)
        if team:
            data["team_name"] = team.get("name")
    return data


def _is_global_summary_item(item: Dict[str, Any]) -> bool:
    req_id = str(item.get("id") or "").strip().upper()
    if req_id.startswith("GLOBAL-"):
        return True
    source = item.get("source")
    if isinstance(source, list):
        for entry in source:
            if "global" in str(entry).lower():
                return True
    elif isinstance(source, str) and "global" in source.lower():
        return True
    return False


def _redact_global_requirements_summary(summary: Dict[str, Any], is_admin: bool) -> Dict[str, Any]:
    if is_admin or not isinstance(summary, dict):
        return summary
    items = summary.get("items")
    if not isinstance(items, list) or not items:
        return summary
    redacted_any = False
    redacted_items: List[Dict[str, Any]] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        redacted = dict(item)
        if _is_global_summary_item(item):
            redacted["text"] = "[redacted]"
            redacted_any = True
        redacted.pop("source", None)
        redacted_items.append(redacted)
    summary["items"] = redacted_items
    summary["total"] = len(redacted_items)
    if redacted_any:
        summary["redacted"] = True
    return summary


def _redact_requirement_titles_in_line(line: str) -> str:
    if "REQ-" not in line or ":" not in line:
        return line
    titles = _global_requirements_titles()
    if not titles:
        return line
    parts = line.split("; ")
    redacted_parts: List[str] = []
    for part in parts:
        match = re.match(r"(REQ-\d+\s*):\s*(.+)", part.strip(), re.IGNORECASE)
        if not match:
            redacted_parts.append(part)
            continue
        prefix = match.group(1)
        title = match.group(2).strip().lower()
        if any(t in title for t in titles):
            redacted_parts.append(f"{prefix} [redacted]")
        else:
            redacted_parts.append(part)
    return "; ".join(redacted_parts)


def _redact_requirements_table_row(line: str) -> str:
    stripped = line.strip()
    if not stripped.startswith("|") or "global" not in stripped.lower():
        return line
    parts = [p.strip() for p in stripped.strip("|").split("|")]
    if len(parts) < 2:
        return line
    req_id = parts[0]
    if not re.match(r"REQ-\d+", req_id, re.IGNORECASE):
        return line
    source = parts[-1] or "global"
    redacted_parts = [req_id] + ["[redacted]" for _ in parts[1:-1]] + [source]
    return "| " + " | ".join(redacted_parts) + " |"


def _redact_global_requirement_titles_in_line(line: str) -> str:
    titles = _global_requirements_titles()
    if not titles:
        return line
    lowered = line.lower()
    updated = line
    for title in titles:
        if title and title in lowered:
            updated = re.sub(re.escape(title), "[redacted]", updated, flags=re.IGNORECASE)
            lowered = updated.lower()
    return updated


def _redact_global_requirement_line(line: str, is_admin: bool) -> str:
    if is_admin or not line:
        return line
    redacted = _redact_requirements_table_row(line)
    redacted = _redact_requirement_titles_in_line(redacted)
    redacted = _redact_global_requirement_titles_in_line(redacted)
    return redacted


def _redact_log_entries(entries: List[Dict[str, Any]], is_admin: bool) -> List[Dict[str, Any]]:
    if is_admin:
        return entries
    cleaned: List[Dict[str, Any]] = []
    for entry in entries or []:
        if not isinstance(entry, dict):
            continue
        line = entry.get("line") or ""
        redacted_line = _redact_global_requirement_line(line, is_admin)
        if redacted_line == line:
            cleaned.append(entry)
            continue
        updated = dict(entry)
        updated["line"] = redacted_line
        cleaned.append(updated)
    return cleaned


UK_WORD_MAP = {
    "color": "colour",
    "colors": "colours",
    "colorful": "colourful",
    "favorite": "favourite",
    "favorites": "favourites",
    "behavior": "behaviour",
    "behaviors": "behaviours",
    "center": "centre",
    "centers": "centres",
    "organize": "organise",
    "organizes": "organises",
    "organizing": "organising",
    "organized": "organised",
    "organization": "organisation",
    "organizations": "organisations",
    "math": "maths",
    "canceled": "cancelled",
    "canceling": "cancelling",
}


def _match_case(source: str, replacement: str) -> str:
    if source.isupper():
        return replacement.upper()
    if source[:1].isupper():
        return replacement[:1].upper() + replacement[1:]
    return replacement


def _to_uk_english(text: str) -> str:
    if not text:
        return text
    result = text
    for us, uk in UK_WORD_MAP.items():
        pattern = r"\b" + re.escape(us) + r"\b"
        result = re.sub(pattern, lambda m: _match_case(m.group(0), uk), result, flags=re.IGNORECASE)
    return result


def _write_json_atomic(path: str, payload: Dict[str, Any]) -> None:
    tmp = f"{path}.tmp"
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(tmp, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2)
    os.replace(tmp, path)
    try:
        os.chmod(path, 0o600)
    except Exception:
        pass


def _sanitize_payload(payload: Dict[str, Any]) -> Dict[str, Any]:
    if not isinstance(payload, dict):
        return {}
    cleaned = dict(payload)
    sensitive_keys = {
        "api_key",
        "apikey",
        "access_token",
        "auth_token",
        "password",
        "secret",
        "token",
        "llm_api_key",
        "fallback_llm_api_key",
        "gemini_api_key",
        "gemini_access_token",
        "google_api_key",
        "github_token",
    }
    for key in list(cleaned.keys()):
        if key.lower() in sensitive_keys:
            cleaned[key] = "***"
    if "job_secrets" in cleaned:
        secrets = cleaned.get("job_secrets")
        if isinstance(secrets, list):
            masked_list = []
            for entry in secrets:
                if not isinstance(entry, dict):
                    continue
                name = str(entry.get("name") or "").strip()
                if name:
                    masked_list.append({"name": name, "value": "***"})
            cleaned["job_secrets"] = masked_list
        elif isinstance(secrets, dict):
            cleaned["job_secrets"] = {str(k): "***" for k in secrets.keys()}
        else:
            cleaned["job_secrets"] = "***"
    if "requirements_text" in cleaned:
        cleaned["requirements_text"] = "[redacted]"
    return cleaned


VOICE_TEXT_KEYS = ("text", "thought", "todo", "note", "input", "query", "message", "content")


def _extract_bearer_token(auth_header: Optional[str]) -> Optional[str]:
    if not auth_header:
        return None
    parts = auth_header.strip().split()
    if len(parts) == 2 and parts[0].lower() == "bearer":
        return parts[1].strip()
    return auth_header.strip()


def _extract_voice_tokens(payload: Optional[Dict[str, Any]] = None) -> List[str]:
    candidates: List[str] = []
    auth_header = request.headers.get("Authorization") or request.headers.get("authorization")
    token = _extract_bearer_token(auth_header)
    if token:
        candidates.append(token)
    for header in ("X-Voice-Token", "X-Api-Key", "X-Api-Token"):
        value = request.headers.get(header)
        if value:
            candidates.append(value.strip())
    for key in ("token", "voice_token", "api_key", "key"):
        value = request.args.get(key)
        if value:
            candidates.append(str(value).strip())
    for key in ("token", "voice_token", "api_key", "key"):
        value = request.form.get(key)
        if value:
            candidates.append(str(value).strip())
    if isinstance(payload, dict):
        for key in ("token", "auth_token", "access_token"):
            value = payload.get(key)
            if isinstance(value, str) and value.strip():
                candidates.append(value.strip())
        session = payload.get("session") if isinstance(payload.get("session"), dict) else {}
        user = session.get("user") if isinstance(session.get("user"), dict) else {}
        value = user.get("accessToken")
        if isinstance(value, str) and value.strip():
            candidates.append(value.strip())
        context = payload.get("context") if isinstance(payload.get("context"), dict) else {}
        system = context.get("System") if isinstance(context.get("System"), dict) else {}
        system_user = system.get("user") if isinstance(system.get("user"), dict) else {}
        value = system_user.get("accessToken")
        if isinstance(value, str) and value.strip():
            candidates.append(value.strip())
        original = payload.get("originalDetectIntentRequest") if isinstance(payload.get("originalDetectIntentRequest"), dict) else {}
        orig_payload = original.get("payload") if isinstance(original.get("payload"), dict) else {}
        orig_user = orig_payload.get("user") if isinstance(orig_payload.get("user"), dict) else {}
        value = orig_user.get("accessToken")
        if isinstance(value, str) and value.strip():
            candidates.append(value.strip())
    seen = set()
    unique: List[str] = []
    for candidate in candidates:
        if not candidate or candidate in seen:
            continue
        unique.append(candidate)
        seen.add(candidate)
    return unique


def _voice_user_from_request(payload: Optional[Dict[str, Any]] = None) -> Optional[str]:
    for token in _extract_voice_tokens(payload):
        user = None
        if _customers_enabled():
            try:
                identity = _run_external_call(
                    "customers voice token resolve",
                    CUSTOMERS_SERVICE.resolve_voice_token,
                    token,
                    timeout=_service_timeout_seconds(CUSTOMERS_SERVICE, default=max(1.0, CUSTOMERS_TIMEOUT)),
                )
                if identity.get("authenticated"):
                    user = identity.get("user")
            except Exception as exc:
                logger.warning("customers voice token resolve failed: %s", exc)
        if not user:
            user = voice_token_store.verify(token) or VOICE_ENV_TOKEN_MAP.get(token)
        if user:
            user = str(user).strip()
            if not user:
                continue
            if user_store.has_users() and not _user_role(user):
                continue
            return user
    return None


def _validate_voice_user(user: Optional[str]) -> Optional[str]:
    if not user:
        return None
    user = str(user).strip()
    if not user:
        return None
    if user_store.has_users() and not _user_role(user):
        return None
    return user


def _extract_alexa_user_id(payload: Dict[str, Any]) -> Optional[str]:
    session = payload.get("session") if isinstance(payload.get("session"), dict) else {}
    session_user = session.get("user") if isinstance(session.get("user"), dict) else {}
    user_id = session_user.get("userId")
    if isinstance(user_id, str) and user_id.strip():
        return user_id.strip()
    context = payload.get("context") if isinstance(payload.get("context"), dict) else {}
    system = context.get("System") if isinstance(context.get("System"), dict) else {}
    system_user = system.get("user") if isinstance(system.get("user"), dict) else {}
    user_id = system_user.get("userId")
    if isinstance(user_id, str) and user_id.strip():
        return user_id.strip()
    return None


def _extract_google_user_id(payload: Dict[str, Any]) -> Optional[str]:
    original = payload.get("originalDetectIntentRequest") if isinstance(payload.get("originalDetectIntentRequest"), dict) else {}
    orig_payload = original.get("payload") if isinstance(original.get("payload"), dict) else {}
    orig_user = orig_payload.get("user") if isinstance(orig_payload.get("user"), dict) else {}
    user_id = orig_user.get("userId")
    if isinstance(user_id, str) and user_id.strip():
        return user_id.strip()
    user = payload.get("user") if isinstance(payload.get("user"), dict) else {}
    user_id = user.get("userId")
    if isinstance(user_id, str) and user_id.strip():
        return user_id.strip()
    return None


def _map_voice_user(provider: str, external_id: Optional[str]) -> Optional[str]:
    if external_id:
        key = f"{provider}:{external_id}"
        if key in VOICE_USER_MAP:
            return VOICE_USER_MAP.get(key)
        generic_key = f"voice:{external_id}"
        if generic_key in VOICE_USER_MAP:
            return VOICE_USER_MAP.get(generic_key)
    return VOICE_DEFAULT_USER or None


def _voice_user_from_provider(
    provider: str,
    payload: Dict[str, Any],
    *,
    allow_tokens: bool = False,
) -> Optional[str]:
    external_id = None
    if provider == "alexa":
        external_id = _extract_alexa_user_id(payload)
    elif provider == "google":
        external_id = _extract_google_user_id(payload)
    user = _validate_voice_user(_map_voice_user(provider, external_id))
    if user:
        return user
    if allow_tokens:
        return _voice_user_from_request(payload)
    return None


def _stt_learning_context(payload: Optional[Dict[str, Any]] = None) -> str:
    snippets: List[str] = []
    if isinstance(payload, dict):
        for key in ("context", "prompt", "hint", "topic", "query", "message", "text"):
            value = payload.get(key)
            if isinstance(value, str) and value.strip():
                snippets.append(value.strip())
    for key in ("context", "prompt", "hint", "topic", "query", "message"):
        value = request.values.get(key)
        if isinstance(value, str) and value.strip():
            snippets.append(value.strip())
    combined = " ".join(snippets)
    combined = re.sub(r"\s+", " ", combined).strip()
    if len(combined) > 800:
        combined = combined[:800].rstrip()
    return combined


def _stt_prompt_hint(payload: Optional[Dict[str, Any]] = None) -> str:
    if not stt_learning_store:
        return ""
    context = _stt_learning_context(payload)
    try:
        return stt_learning_store.build_prompt_hint(context=context)
    except Exception:
        return ""


def _stt_record_learning(text: Optional[str], source: str) -> None:
    if not text or not stt_learning_store:
        return
    try:
        stt_learning_store.learn_from_text(text, source=source)
    except Exception as exc:
        logger.debug("STT learning update skipped: %s", exc)


STT_ALLOWED_MIME = {
    "audio/webm",
    "audio/ogg",
    "audio/wav",
    "audio/x-wav",
    "audio/mpeg",
    "audio/mp4",
    "audio/aac",
}


def _stt_authorized(payload: Optional[Dict[str, Any]] = None) -> bool:
    if _current_user():
        return True
    if STT_TOKEN:
        for token in _extract_voice_tokens(payload):
            if secrets_lib.compare_digest(token, STT_TOKEN):
                return True
        return False
    if STT_PUBLIC:
        return True
    return False


def _sanitize_lang(value: Optional[str]) -> str:
    if not value:
        return STT_LANG_DEFAULT or "en-GB"
    cleaned = str(value).strip()
    if not cleaned:
        return STT_LANG_DEFAULT or "en-GB"
    if not re.match(r"^[A-Za-z0-9_.-]+$", cleaned):
        return STT_LANG_DEFAULT or "en-GB"
    return cleaned


def _parse_boolish(value: Any) -> Optional[bool]:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if not isinstance(value, str):
        return None
    cleaned = value.strip().lower()
    if not cleaned:
        return None
    if cleaned in {"1", "true", "yes", "on", "office"}:
        return True
    if cleaned in {"0", "false", "no", "off", "chat"}:
        return False
    return None


def _stt_option(payload: Optional[Dict[str, Any]], *keys: str) -> str:
    for key in keys:
        value = request.form.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
        value = request.args.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    if isinstance(payload, dict):
        for key in keys:
            value = payload.get(key)
            if isinstance(value, str) and value.strip():
                return value.strip()
    return ""


def _stt_motion_context(payload: Optional[Dict[str, Any]]) -> Tuple[str, str, Optional[bool]]:
    gesture_raw = _stt_option(
        payload,
        "gesture_mode",
        "gestureMode",
        "motion_style",
        "motionStyle",
        "motion-style",
    )
    avatar_raw = _stt_option(payload, "avatar_mode", "avatarMode")
    office_raw = _stt_option(payload, "office_mode", "officeMode")
    office_flag = _parse_boolish(office_raw)
    if office_flag is None and isinstance(payload, dict):
        office_flag = _parse_boolish(payload.get("office_mode"))
        if office_flag is None:
            office_flag = _parse_boolish(payload.get("officeMode"))
    gesture_mode = sanitize_gesture_mode(
        gesture_raw,
        default_mode=STT_GESTURE_DEFAULT_MODE,
        bsl_enabled=STT_BSL_ENABLED,
    )
    avatar_mode = sanitize_avatar_mode(
        avatar_raw,
        office_mode=office_flag,
        default_mode=STT_GESTURE_DEFAULT_AVATAR_MODE,
    )
    return gesture_mode, avatar_mode, office_flag


def _stt_collaboration_mode(payload: Optional[Dict[str, Any]]) -> bool:
    raw = _stt_option(
        payload,
        "collaboration_mode",
        "collaborationMode",
        "collaboration",
        "multi_speaker",
        "multiSpeaker",
        "multi_speaker_mode",
        "multiSpeakerMode",
    )
    parsed = _parse_boolish(raw)
    if parsed is not None:
        return parsed
    if isinstance(payload, dict):
        for key in (
            "collaboration_mode",
            "collaborationMode",
            "collaboration",
            "multi_speaker",
            "multiSpeaker",
            "multi_speaker_mode",
            "multiSpeakerMode",
        ):
            parsed = _parse_boolish(payload.get(key))
            if parsed is not None:
                return parsed
    return False


def _extract_audio_bytes(payload: Optional[Dict[str, Any]] = None) -> Tuple[Optional[bytes], str, Optional[str]]:
    file = request.files.get("audio")
    if file and file.filename:
        data = file.read()
        mime = (file.mimetype or "").lower()
        if mime and STT_ALLOWED_MIME and mime not in STT_ALLOWED_MIME:
            return None, "", "unsupported_format"
        ext = os.path.splitext(file.filename)[1].lower()
        if not ext:
            ext = ".webm" if "webm" in mime else ".wav" if "wav" in mime else ".ogg" if "ogg" in mime else ".bin"
        return data, ext, None
    if isinstance(payload, dict):
        b64 = payload.get("audio_base64") or payload.get("audio")
        if isinstance(b64, str) and b64.strip():
            try:
                data = base64.b64decode(b64)
            except Exception:
                return None, "", "invalid_audio_base64"
            ext = ".webm"
            return data, ext, None
    raw = request.get_data(cache=True)
    if raw:
        mime = (request.headers.get("Content-Type") or "").split(";")[0].lower()
        ext = ".webm" if "webm" in mime else ".wav" if "wav" in mime else ".ogg" if "ogg" in mime else ".bin"
        return raw, ext, None
    return None, "", "audio_required"


def _write_audio_temp(data: bytes, ext: str) -> str:
    suffix = ext if ext.startswith(".") else f".{ext}"
    handle = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    try:
        handle.write(data)
    finally:
        handle.close()
    return handle.name


def _build_command(command: str, args_template: str, **kwargs: str) -> List[str]:
    args = shlex.split(args_template) if args_template else []
    formatted: List[str] = []
    for arg in args:
        try:
            formatted.append(arg.format(**kwargs))
        except Exception:
            formatted.append(arg)
    return [command] + formatted


def _run_preprocess(input_path: str) -> Tuple[str, Optional[str]]:
    if STT_BACKEND == "server" and not STT_SERVER_PREPROCESS:
        return input_path, None
    if not STT_PREPROCESS_COMMAND:
        return input_path, None
    ext = STT_PREPROCESS_EXT or ".wav"
    if not ext.startswith("."):
        ext = f".{ext}"
    output_path = f"{input_path}{ext}"
    cmd = _build_command(STT_PREPROCESS_COMMAND, STT_PREPROCESS_ARGS, input=input_path, output=output_path)
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=STT_TIMEOUT)
    except Exception:
        return input_path, "stt_preprocess_failed"
    if result.returncode != 0:
        return input_path, "stt_preprocess_failed"
    if not os.path.exists(output_path):
        return input_path, "stt_preprocess_missing_output"
    return output_path, None


def _run_stt(
    audio_path: str,
    lang: str,
    prompt_hint: Optional[str] = None,
    *,
    gesture_mode: Optional[str] = None,
    avatar_mode: Optional[str] = None,
    office_mode: Optional[bool] = None,
    collaboration_mode: Optional[bool] = None,
) -> Tuple[Optional[str], Optional[str], Optional[Dict[str, Any]]]:
    if STT_BACKEND == "server" and STT_SERVER_URL:
        return _run_stt_server(
            audio_path,
            lang,
            prompt_hint=prompt_hint,
            gesture_mode=gesture_mode,
            avatar_mode=avatar_mode,
            office_mode=office_mode,
            collaboration_mode=collaboration_mode,
        )
    transcript, error = _run_stt_command(audio_path, lang, prompt_hint=prompt_hint)
    return transcript, error, None


def _run_stt_command(audio_path: str, lang: str, prompt_hint: Optional[str] = None) -> Tuple[Optional[str], Optional[str]]:
    if not STT_COMMAND:
        return None, "stt_not_configured"
    cmd = _build_command(STT_COMMAND, STT_ARGS, audio=audio_path, lang=lang, prompt=prompt_hint or "")
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=STT_TIMEOUT)
    except Exception:
        return None, "stt_command_failed"
    if result.returncode != 0:
        return None, "stt_command_failed"
    if STT_OUTPUT_MODE == "json":
        output_path = STT_OUTPUT_PATH_TEMPLATE.format(audio=audio_path)
        if not os.path.exists(output_path):
            return None, "stt_output_missing"
        try:
            with open(output_path, "r", encoding="utf-8") as handle:
                data = json.load(handle)
            if isinstance(data, dict):
                text = data.get("text") or data.get("transcript") or data.get("result")
                if isinstance(text, str) and text.strip():
                    return text.strip(), None
        except Exception:
            return None, "stt_output_invalid"
        return None, "stt_output_empty"
    transcript = (result.stdout or "").strip()
    if not transcript:
        transcript = (result.stderr or "").strip()
    if not transcript:
        return None, "stt_output_empty"
    return transcript, None


def _acquire_request_capacity(semaphore: Any, wait_seconds: float, *, owner: str = "") -> bool:
    if hasattr(semaphore, "acquire_for"):
        return bool(semaphore.acquire_for(owner, timeout=max(0.0, float(wait_seconds or 0.0))))
    if wait_seconds > 0:
        return bool(semaphore.acquire(timeout=wait_seconds))
    return bool(semaphore.acquire(blocking=False))


def _release_request_capacity(semaphore: Any, *, owner: str = "") -> None:
    if hasattr(semaphore, "release_for"):
        semaphore.release_for(owner)
        return
    semaphore.release()


def _stt_server_session() -> requests.Session:
    """
    Return a per-thread session for STT server calls.

    `requests.Session` gives us keep-alive connection reuse for lower latency.
    We keep one session per thread to avoid cross-thread mutation issues.
    """
    session = getattr(_STT_SERVER_SESSION_LOCAL, "session", None)
    if isinstance(session, requests.Session):
        return session
    session = requests.Session()
    adapter = requests.adapters.HTTPAdapter(
        pool_connections=STT_SERVER_POOL_CONNECTIONS,
        pool_maxsize=STT_SERVER_POOL_MAXSIZE,
        max_retries=0,
    )
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    _STT_SERVER_SESSION_LOCAL.session = session
    return session


def _stt_server_request_data(
    lang: str,
    prompt_hint: Optional[str] = None,
    *,
    gesture_mode: Optional[str] = None,
    avatar_mode: Optional[str] = None,
    office_mode: Optional[bool] = None,
    collaboration_mode: Optional[bool] = None,
) -> Dict[str, str]:
    """Builds a multi-alias form payload so Rust/Python/frontends remain schema-compatible."""
    data: Dict[str, str] = {"lang": lang}
    if prompt_hint and STT_SERVER_SEND_PROMPT:
        data["prompt"] = prompt_hint
    if gesture_mode:
        gesture = str(gesture_mode).strip()
        data["gesture_mode"] = gesture
        data["gestureMode"] = gesture
        data["motion_style"] = gesture
        data["motionStyle"] = gesture
    if avatar_mode:
        avatar = str(avatar_mode).strip()
        data["avatar_mode"] = avatar
        data["avatarMode"] = avatar
    if office_mode is not None:
        office = "1" if office_mode else "0"
        data["office_mode"] = office
        data["officeMode"] = office
    if collaboration_mode is not None:
        collab = "1" if collaboration_mode else "0"
        data["collaboration_mode"] = collab
        data["collaborationMode"] = collab
        data["multi_speaker"] = collab
        data["multiSpeaker"] = collab
    return data


def _stt_server_mime(ext: str) -> str:
    suffix = ext.strip().lower()
    return {
        ".wav": "audio/wav",
        ".webm": "audio/webm",
        ".ogg": "audio/ogg",
        ".mp3": "audio/mpeg",
        ".m4a": "audio/mp4",
        ".flac": "audio/flac",
    }.get(suffix, "application/octet-stream")


def _stt_server_retry_delay_seconds(attempt: int, resp: Optional[requests.Response]) -> float:
    """Use Retry-After when available, otherwise exponential backoff bounded by env caps."""
    if resp is not None:
        retry_after = (resp.headers.get("Retry-After") or "").strip()
        if retry_after:
            try:
                return min(STT_SERVER_BACKOFF_MAX, max(0.0, float(retry_after)))
            except Exception:
                pass
    delay = STT_SERVER_BACKOFF_BASE * (2 ** max(0, attempt))
    return min(STT_SERVER_BACKOFF_MAX, max(0.0, delay))


def _is_stt_server_retryable_status(code: int) -> bool:
    return code in {408, 409, 425, 429, 500, 502, 503, 504}


def _run_stt_server_bytes(
    audio_bytes: bytes,
    ext: str,
    lang: str,
    prompt_hint: Optional[str] = None,
    *,
    gesture_mode: Optional[str] = None,
    avatar_mode: Optional[str] = None,
    office_mode: Optional[bool] = None,
    collaboration_mode: Optional[bool] = None,
) -> Tuple[Optional[str], Optional[str], Optional[Dict[str, Any]]]:
    """
    Call the STT server directly from in-memory audio bytes.

    This avoids temporary file round-trips when preprocess is not required and
    adds bounded retries for transient transport/server failures.
    """
    if not STT_SERVER_URL:
        return None, "stt_server_not_configured", None
    endpoint = STT_SERVER_URL.rstrip("/") + "/transcribe"
    data = _stt_server_request_data(
        lang,
        prompt_hint=prompt_hint,
        gesture_mode=gesture_mode,
        avatar_mode=avatar_mode,
        office_mode=office_mode,
        collaboration_mode=collaboration_mode,
    )
    suffix = ext if isinstance(ext, str) and ext.startswith(".") else f".{(ext or 'bin').lstrip('.')}"
    files = {"audio": (f"audio{suffix}", audio_bytes, _stt_server_mime(suffix))}
    attempts = max(1, STT_SERVER_RETRIES + 1)
    session = _stt_server_session()
    resp: Optional[requests.Response] = None
    for attempt in range(attempts):
        try:
            resp = _external_http_request(
                "POST",
                endpoint,
                session_obj=session,
                files=files,
                data=data,
                timeout=STT_SERVER_TIMEOUT,
            )
        except requests.RequestException:
            if attempt + 1 >= attempts:
                return None, "stt_server_unreachable", None
            delay = _stt_server_retry_delay_seconds(attempt, None)
            if delay > 0:
                time.sleep(delay)
            continue

        if resp.status_code >= 400 and _is_stt_server_retryable_status(resp.status_code) and attempt + 1 < attempts:
            delay = _stt_server_retry_delay_seconds(attempt, resp)
            if delay > 0:
                time.sleep(delay)
            continue
        break

    if resp is None:
        return None, "stt_server_unreachable", None
    if resp.status_code >= 400:
        try:
            payload = resp.json()
        except Exception:
            payload = {}
        return None, payload.get("error") or "stt_server_error", None
    try:
        payload = resp.json()
    except Exception:
        payload = {}
    text = payload.get("text") or payload.get("transcript")
    if not isinstance(text, str) or not text.strip():
        return None, "stt_output_empty", None

    server_payload = sanitize_nmstt_motion_response(payload)
    if "collaboration_mode" not in server_payload:
        collab_flag = _parse_boolish(payload.get("collaboration_mode"))
        if collab_flag is not None:
            server_payload["collaboration_mode"] = collab_flag

    return text.strip(), None, server_payload or None


def _run_nmstt_gesture_plan(
    text: str,
    *,
    gesture_mode: str,
    avatar_mode: str,
    office_mode: Optional[bool],
) -> Optional[Dict[str, Any]]:
    """Ask nmstt to plan avatar gestures for an already-known transcript."""
    if not STT_SERVER_URL or not STT_GESTURE_NMSTT_FALLBACK:
        return None
    transcript = str(text or "").strip()
    if not transcript:
        return None
    endpoint = STT_SERVER_URL.rstrip("/") + "/gesture-plan"
    request_payload = NmsttGesturePlanRequest(
        text=transcript,
        gesture_mode=str(gesture_mode or "").strip(),
        avatar_mode=str(avatar_mode or "").strip(),
        office_mode=office_mode,
    )
    attempts = max(1, STT_SERVER_RETRIES + 1)
    session = _stt_server_session()
    resp: Optional[requests.Response] = None
    for attempt in range(attempts):
        try:
            resp = session.post(endpoint, json=request_payload.as_dict(), timeout=STT_GESTURE_NMSTT_TIMEOUT)
        except requests.RequestException:
            if attempt + 1 >= attempts:
                return None
            delay = _stt_server_retry_delay_seconds(attempt, None)
            if delay > 0:
                time.sleep(delay)
            continue
        if resp.status_code >= 400 and _is_stt_server_retryable_status(resp.status_code) and attempt + 1 < attempts:
            delay = _stt_server_retry_delay_seconds(attempt, resp)
            if delay > 0:
                time.sleep(delay)
            continue
        break
    if resp is None or resp.status_code >= 400:
        return None
    try:
        data = resp.json()
    except Exception:
        return None
    out = sanitize_nmstt_motion_response(data)
    return out or None


def _run_stt_server(
    audio_path: str,
    lang: str,
    prompt_hint: Optional[str] = None,
    *,
    gesture_mode: Optional[str] = None,
    avatar_mode: Optional[str] = None,
    office_mode: Optional[bool] = None,
    collaboration_mode: Optional[bool] = None,
) -> Tuple[Optional[str], Optional[str], Optional[Dict[str, Any]]]:
    try:
        with open(audio_path, "rb") as handle:
            audio_bytes = handle.read()
    except Exception:
        return None, "invalid_audio", None
    ext = os.path.splitext(audio_path)[1] or ".bin"
    return _run_stt_server_bytes(
        audio_bytes,
        ext,
        lang,
        prompt_hint=prompt_hint,
        gesture_mode=gesture_mode,
        avatar_mode=avatar_mode,
        office_mode=office_mode,
        collaboration_mode=collaboration_mode,
    )


def _extract_text_from_dict(payload: Dict[str, Any]) -> str:
    for key in VOICE_TEXT_KEYS:
        value = payload.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    return ""


def _extract_voice_text(payload: Optional[Dict[str, Any]] = None) -> str:
    if isinstance(payload, dict):
        text = _extract_text_from_dict(payload)
        if text:
            return text
    for key in VOICE_TEXT_KEYS:
        value = request.values.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    raw = request.get_data(cache=True, as_text=True)
    if raw:
        raw = raw.strip()
        if raw and not raw.startswith("{") and not raw.startswith("["):
            return raw
    return ""


def _extract_alexa_text(payload: Dict[str, Any]) -> str:
    request_payload = payload.get("request") if isinstance(payload.get("request"), dict) else {}
    intent = request_payload.get("intent") if isinstance(request_payload.get("intent"), dict) else {}
    slots = intent.get("slots") if isinstance(intent.get("slots"), dict) else {}
    for slot in slots.values():
        if not isinstance(slot, dict):
            continue
        value = slot.get("value")
        if isinstance(value, str) and value.strip():
            return value.strip()
    text = _extract_text_from_dict(intent)
    if text:
        return text
    return _extract_text_from_dict(payload)


def _extract_google_text(payload: Dict[str, Any]) -> str:
    query_result = payload.get("queryResult") if isinstance(payload.get("queryResult"), dict) else {}
    text = query_result.get("queryText")
    if isinstance(text, str) and text.strip():
        return text.strip()
    params = query_result.get("parameters") if isinstance(query_result.get("parameters"), dict) else {}
    text = _extract_text_from_dict(params)
    if text:
        return text
    return _extract_text_from_dict(payload)


def _infer_device_from_user_agent(user_agent: str) -> Optional[str]:
    if not user_agent:
        return None
    ua = user_agent.lower()
    if "carplay" in ua:
        return "carplay"
    if "watch" in ua:
        return "watch"
    if "iphone" in ua:
        return "iphone"
    if "ipad" in ua:
        return "ipad"
    return None


def _extract_voice_device(payload: Optional[Dict[str, Any]] = None) -> Optional[str]:
    if isinstance(payload, dict):
        for key in ("device", "source_device", "client_device", "client"):
            value = payload.get(key)
            if isinstance(value, str) and value.strip():
                return value.strip().lower()
    for key in ("device", "source_device", "client_device", "client"):
        value = request.values.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip().lower()
    inferred = _infer_device_from_user_agent(request.headers.get("User-Agent") or "")
    return inferred


def _alexa_response(text: str, *, end_session: bool = True) -> Response:
    return jsonify(
        {
            "version": "1.0",
            "response": {
                "outputSpeech": {"type": "PlainText", "text": text},
                "shouldEndSession": bool(end_session),
            },
        }
    )


def _google_response(text: str) -> Response:
    return jsonify({"fulfillmentText": text})


_ALEXA_CERT_CACHE: Dict[str, Dict[str, Any]] = {}
_ALEXA_CERT_CACHE_LOCK = threading.Lock()
_GOOGLE_CERT_CACHE: Dict[str, Any] = {"expires_at": 0.0, "keys": {}}
_GOOGLE_CERT_CACHE_LOCK = threading.Lock()


def _load_alexa_cert_cache() -> Dict[str, Any]:
    if not ALEXA_CERT_CACHE_PATH:
        return {}
    try:
        with open(ALEXA_CERT_CACHE_PATH, "r", encoding="utf-8") as handle:
            data = json.load(handle)
        if isinstance(data, dict):
            return data
    except Exception:
        return {}
    return {}


def _write_alexa_cert_cache(cache: Dict[str, Any]) -> None:
    if not ALEXA_CERT_CACHE_PATH:
        return
    _write_json_atomic(ALEXA_CERT_CACHE_PATH, cache)


def _parse_iso8601(value: Optional[str]) -> Optional[dt.datetime]:
    if not value or not isinstance(value, str):
        return None
    cleaned = value.strip()
    if not cleaned:
        return None
    if cleaned.endswith("Z"):
        cleaned = cleaned[:-1] + "+00:00"
    try:
        parsed = dt.datetime.fromisoformat(cleaned)
    except Exception:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=dt.timezone.utc)
    return parsed


def _spki_sha256(cert: Any) -> Optional[str]:
    if not cert or not Encoding or not PublicFormat:
        return None
    try:
        spki = cert.public_key().public_bytes(Encoding.DER, PublicFormat.SubjectPublicKeyInfo)
    except Exception:
        return None
    return hashlib.sha256(spki).hexdigest()


def _alexa_cert_url_valid(url: str) -> bool:
    if not url:
        return False
    try:
        parsed = urlparse(url)
    except Exception:
        return False
    if parsed.scheme.lower() != "https":
        return False
    if parsed.hostname not in {"s3.amazonaws.com"}:
        return False
    if parsed.port not in (None, 443):
        return False
    path = posixpath.normpath(parsed.path or "")
    if not path.startswith("/"):
        path = f"/{path}"
    if not (path == "/echo.api" or path.startswith("/echo.api/")):
        return False
    if parsed.query or parsed.fragment:
        return False
    return True


def _alexa_parse_cert_chain(pem_text: str) -> List[Any]:
    if not x509:
        return []
    matches = re.findall(r"-----BEGIN CERTIFICATE-----.*?-----END CERTIFICATE-----", pem_text, re.S)
    certs: List[Any] = []
    for block in matches:
        try:
            certs.append(x509.load_pem_x509_certificate(block.encode("utf-8")))
        except Exception:
            continue
    return certs


def _alexa_fetch_cert_chain(url: str) -> List[Any]:
    if not url:
        return []
    now = time.time()
    if ALEXA_CERT_CACHE_PATH:
        disk_cache = _load_alexa_cert_cache()
        if isinstance(disk_cache, dict):
            entry = disk_cache.get(url)
            if isinstance(entry, dict) and entry.get("expires_at", 0) > now:
                pem_text = entry.get("pem") or ""
                if isinstance(pem_text, str) and pem_text.strip():
                    return _alexa_parse_cert_chain(pem_text)
    with _ALEXA_CERT_CACHE_LOCK:
        cached = _ALEXA_CERT_CACHE.get(url)
        if cached and cached.get("expires_at", 0) > now:
            return cached.get("certs", [])
    if not VOICE_ALLOW_NETWORK:
        return []
    try:
        resp = _http_request_with_retry("GET", url, timeout=10, retries=1)
    except Exception:
        return []
    if resp.status_code >= 400:
        return []
    certs = _alexa_parse_cert_chain(resp.text or "")
    ttl = ALEXA_CERT_TTL_SEC
    cache_control = resp.headers.get("Cache-Control") or ""
    match = re.search(r"max-age=(\d+)", cache_control)
    if match:
        try:
            ttl = max(60, min(int(match.group(1)), 86400))
        except Exception:
            ttl = ALEXA_CERT_TTL_SEC
    with _ALEXA_CERT_CACHE_LOCK:
        _ALEXA_CERT_CACHE[url] = {"certs": certs, "expires_at": now + float(ttl)}
    if ALEXA_CERT_CACHE_PATH:
        disk_cache = _load_alexa_cert_cache()
        if not isinstance(disk_cache, dict):
            disk_cache = {}
        disk_cache[url] = {"pem": resp.text or "", "expires_at": now + float(ttl)}
        _write_alexa_cert_cache(disk_cache)
    return certs


def _alexa_verify_cert_chain(certs: List[Any]) -> bool:
    if not certs or not x509 or not padding or not hashes:
        return False
    now = dt.datetime.utcnow()
    for cert in certs:
        try:
            if now < cert.not_valid_before or now > cert.not_valid_after:
                return False
        except Exception:
            return False
    leaf = certs[0]
    try:
        san = leaf.extensions.get_extension_for_class(x509.SubjectAlternativeName).value
        dns_names = san.get_values_for_type(x509.DNSName)
        if "echo-api.amazon.com" not in dns_names:
            return False
    except Exception:
        return False
    for idx in range(len(certs) - 1):
        issuer = certs[idx + 1]
        try:
            issuer.public_key().verify(
                certs[idx].signature,
                certs[idx].tbs_certificate_bytes,
                padding.PKCS1v15(),
                certs[idx].signature_hash_algorithm,
            )
        except Exception:
            return False
    root = certs[-1]
    try:
        root.public_key().verify(
            root.signature,
            root.tbs_certificate_bytes,
            padding.PKCS1v15(),
            root.signature_hash_algorithm,
        )
    except Exception:
        return False
    spki_hash = _spki_sha256(root)
    if ALEXA_TRUSTED_ROOT_SPKI and (not spki_hash or spki_hash.lower() not in ALEXA_TRUSTED_ROOT_SPKI):
        return False
    return True


def _alexa_verify_request(payload: Dict[str, Any], body: bytes) -> Tuple[bool, str]:
    if not x509 or not padding or not hashes:
        return False, "crypto_unavailable"
    if not VOICE_ALLOW_NETWORK and not ALEXA_CERT_CACHE_PATH:
        return False, "network_disabled"
    cert_url = request.headers.get("SignatureCertChainUrl") or request.headers.get("signaturecertchainurl")
    signature = request.headers.get("Signature-256") or request.headers.get("Signature") or request.headers.get("signature")
    if not cert_url or not signature:
        return False, "missing_signature_headers"
    if not _alexa_cert_url_valid(cert_url):
        return False, "invalid_cert_url"
    certs = _alexa_fetch_cert_chain(cert_url)
    if not certs or not _alexa_verify_cert_chain(certs):
        return False, "invalid_cert_chain"
    request_payload = payload.get("request") if isinstance(payload.get("request"), dict) else {}
    timestamp = _parse_iso8601(request_payload.get("timestamp") if isinstance(request_payload.get("timestamp"), str) else None)
    if not timestamp:
        return False, "missing_timestamp"
    now = dt.datetime.now(dt.timezone.utc)
    delta = abs((now - timestamp).total_seconds())
    if ALEXA_REQUEST_TTL_SEC > 0 and delta > ALEXA_REQUEST_TTL_SEC:
        return False, "stale_timestamp"
    try:
        signature_bytes = base64.b64decode(signature)
    except Exception:
        return False, "invalid_signature"
    try:
        certs[0].public_key().verify(signature_bytes, body, padding.PKCS1v15(), hashes.SHA256())
    except Exception:
        return False, "signature_mismatch"
    return True, "ok"


def _google_fetch_keys() -> Dict[str, Any]:
    if GOOGLE_CERTS_PATH:
        try:
            with open(GOOGLE_CERTS_PATH, "r", encoding="utf-8") as handle:
                data = json.load(handle)
        except Exception:
            data = {}
        keys: Dict[str, Any] = {}
        if isinstance(data, dict) and "keys" in data and isinstance(data.get("keys"), list):
            for entry in data.get("keys"):
                if not isinstance(entry, dict):
                    continue
                kid = entry.get("kid")
                if not kid:
                    continue
                try:
                    key_obj = jwt.algorithms.RSAAlgorithm.from_jwk(json.dumps(entry)) if jwt else None
                except Exception:
                    key_obj = None
                if key_obj:
                    keys[kid] = key_obj
        elif isinstance(data, dict):
            for kid, pem in data.items():
                if isinstance(kid, str) and isinstance(pem, str):
                    keys[kid] = pem
        return keys
    if not GOOGLE_CERTS_URL or not VOICE_ALLOW_NETWORK:
        return {}
    now = time.time()
    with _GOOGLE_CERT_CACHE_LOCK:
        cached = _GOOGLE_CERT_CACHE
        if cached.get("expires_at", 0) > now and cached.get("keys"):
            return cached["keys"]
    try:
        resp = _http_request_with_retry("GET", GOOGLE_CERTS_URL, timeout=10, retries=1)
    except Exception:
        return {}
    if resp.status_code >= 400:
        return {}
    try:
        data = resp.json()
    except Exception:
        return {}
    keys: Dict[str, Any] = {}
    if isinstance(data, dict) and "keys" in data and isinstance(data.get("keys"), list):
        for entry in data.get("keys"):
            if not isinstance(entry, dict):
                continue
            kid = entry.get("kid")
            if not kid:
                continue
            try:
                key_obj = jwt.algorithms.RSAAlgorithm.from_jwk(json.dumps(entry)) if jwt else None
            except Exception:
                key_obj = None
            if key_obj:
                keys[kid] = key_obj
    elif isinstance(data, dict):
        for kid, pem in data.items():
            if isinstance(kid, str) and isinstance(pem, str):
                keys[kid] = pem
    ttl = 3600
    cache_control = resp.headers.get("Cache-Control") or ""
    match = re.search(r"max-age=(\d+)", cache_control)
    if match:
        try:
            ttl = max(60, min(int(match.group(1)), 86400))
        except Exception:
            ttl = 3600
    with _GOOGLE_CERT_CACHE_LOCK:
        _GOOGLE_CERT_CACHE["keys"] = keys
        _GOOGLE_CERT_CACHE["expires_at"] = now + float(ttl)
    return keys


def _extract_google_jwt() -> Optional[str]:
    header_token = request.headers.get("Google-Assistant-Signature") or request.headers.get("google-assistant-signature")
    if isinstance(header_token, str) and header_token.strip():
        return header_token.strip()
    auth_header = request.headers.get("Authorization") or request.headers.get("authorization")
    token = _extract_bearer_token(auth_header)
    if token:
        return token
    return None


def _google_verify_request() -> Tuple[bool, str]:
    if not jwt:
        return False, "pyjwt_missing"
    if not GOOGLE_AUDIENCES:
        return False, "audience_missing"
    token = _extract_google_jwt()
    if not token:
        return False, "missing_jwt"
    try:
        header = jwt.get_unverified_header(token)
    except Exception:
        header = {}
    kid = header.get("kid") if isinstance(header, dict) else None
    keys = _google_fetch_keys()
    if not keys:
        return False, "no_keys"
    key = keys.get(kid) if kid else None
    if not key and keys:
        key = next(iter(keys.values()))
    if not key:
        return False, "key_not_found"
    try:
        jwt.decode(
            token,
            key=key,
            algorithms=["RS256"],
            audience=GOOGLE_AUDIENCES,
            issuer=GOOGLE_ISSUERS,
        )
    except Exception:
        return False, "jwt_invalid"
    return True, "ok"


def _busy_jobs_snapshot(exclude_user: Optional[str] = None) -> List[Job]:
    busy: List[Job] = []
    for job in manager.list_jobs():
        if job.status in {"queued", "running", "paused"}:
            if exclude_user and job.owner == exclude_user:
                continue
            busy.append(job)
    return busy




def _read_log_tail(path: str, count: int) -> List[Dict[str, Any]]:
    try:
        if count <= 0:
            with open(path, "r", encoding="utf-8", errors="ignore") as handle:
                lines = handle.readlines()
        else:
            lines = deque(maxlen=count)
            with open(path, "r", encoding="utf-8", errors="ignore") as handle:
                for line in handle:
                    lines.append(line)
        return [{"ts": "--", "line": line.rstrip("\n"), "stream": "logfile"} for line in list(lines)]
    except Exception:
        return []


def _smtp_config() -> Optional[Dict[str, Any]]:
    host = (os.getenv("REFINER_SMTP_HOST") or "").strip()
    if not host:
        return None
    port = int(os.getenv("REFINER_SMTP_PORT", "587"))
    user = (os.getenv("REFINER_SMTP_USER") or "").strip()
    password = os.getenv("REFINER_SMTP_PASS") or ""
    sender = (os.getenv("REFINER_SMTP_FROM") or user or "refiner@localhost").strip()
    use_tls = _env_flag("REFINER_SMTP_TLS", True)
    use_ssl = _env_flag("REFINER_SMTP_SSL", False)
    timeout = int(os.getenv("REFINER_SMTP_TIMEOUT", "20"))
    return {
        "host": host,
        "port": port,
        "user": user,
        "password": password,
        "sender": sender,
        "use_tls": use_tls,
        "use_ssl": use_ssl,
        "timeout": timeout,
    }


def _send_email(recipient: str, subject: str, body: str) -> Optional[str]:
    cfg = _smtp_config()
    if not cfg:
        return "SMTP not configured (set REFINER_SMTP_HOST)."
    msg = EmailMessage()
    msg["From"] = cfg["sender"]
    msg["To"] = recipient
    msg["Subject"] = subject
    msg.set_content(body)
    try:
        if cfg["use_ssl"]:
            server = smtplib.SMTP_SSL(cfg["host"], cfg["port"], timeout=cfg["timeout"])
        else:
            server = smtplib.SMTP(cfg["host"], cfg["port"], timeout=cfg["timeout"])
        with server:
            server.ehlo()
            if cfg["use_tls"] and not cfg["use_ssl"]:
                server.starttls()
                server.ehlo()
            if cfg["user"]:
                server.login(cfg["user"], cfg["password"])
            server.send_message(msg)
    except Exception as exc:
        return f"Email send failed: {exc}"
    return None


def _job_notification_body(job: "Job") -> str:
    status = job.status
    project = job.project_name or "Untitled"
    workflow = job.workflow
    finished = job.finished_at or _now_iso()
    lines = [
        f"Refiner job update:",
        f"- Job ID: {job.job_id}",
        f"- Project: {project}",
        f"- Workflow: {workflow}",
        f"- Status: {status}",
        f"- Finished: {finished}",
    ]
    if job.exit_code is not None:
        lines.append(f"- Exit Code: {job.exit_code}")
    return "\\n".join(lines)


def _notify_portal_usage(job: "Job") -> None:
    if not PORTAL_WEBHOOK_URL:
        return
    spent = int(job.token_debited or 0)
    if spent <= 0 and not job.token_shortfall:
        return
    summary = _account_summary("user", job.owner)
    payload = {
        "event": "job_tokens_settled",
        "user": job.owner,
        "job_id": job.job_id,
        "workflow": job.workflow,
        "status": job.status,
        "estimate": job.token_estimate,
        "actual": job.token_actual,
        "debited": job.token_debited,
        "shortfall": job.token_shortfall,
        "balance": summary.get("balance"),
        "ts": _now_iso(),
    }
    headers = {"Content-Type": "application/json"}
    if PORTAL_WEBHOOK_TOKEN:
        headers["Authorization"] = f"Bearer {PORTAL_WEBHOOK_TOKEN}"
    try:
        future = _submit_external_http_request(
            "POST",
            PORTAL_WEBHOOK_URL,
            json_body=payload,
            headers=headers,
            timeout=PORTAL_WEBHOOK_TIMEOUT,
        )
        future.add_done_callback(
            lambda done_future: _log_external_future_exception(
                done_future,
                operation="portal usage webhook notify",
            )
        )
    except Exception:
        pass


manager.reconcile_tokens()


def _estimate_tokens_from_text(text: str) -> int:
    if not text:
        return 0
    return max(1, len(text) // 4)

REQ_LINE_RE = re.compile(r"^\s*(?:[-*+]\s*)?(REQ-\d{3,})\b", re.I)


def _count_req_lines(text: str) -> int:
    if not text:
        return 0
    return sum(1 for line in text.splitlines() if REQ_LINE_RE.match(line))


def _estimate_repo_tokens(root: Optional[str]) -> Dict[str, Any]:
    if not root or not isinstance(root, str):
        return {"tokens": 0, "file_count": 0, "sampled": False}
    root = root.strip()
    if not root or not os.path.isdir(root):
        return {"tokens": 0, "file_count": 0, "sampled": False}
    key = os.path.abspath(root)
    now = time.time()
    with _estimate_repo_cache_lock:
        cached = _estimate_repo_cache.get(key)
        if cached and (now - cached.get("ts", 0)) < ESTIMATE_REPO_TTL_SEC:
            return cached["data"]
    start = time.time()
    total_files = 0
    text_bytes = 0
    sampled = False
    for dirpath, dirs, files in os.walk(key):
        dirs[:] = [d for d in dirs if d not in ESTIMATE_IGNORED_DIRS]
        for filename in files:
            total_files += 1
            if total_files >= ESTIMATE_REPO_MAX_FILES or (time.time() - start) > ESTIMATE_REPO_MAX_SEC:
                sampled = True
                break
            ext = os.path.splitext(filename)[1].lower()
            if ext not in ESTIMATE_TEXT_EXTS:
                continue
            abs_path = os.path.join(dirpath, filename)
            try:
                size = os.path.getsize(abs_path)
            except Exception:
                continue
            if size > ESTIMATE_REPO_MAX_FILE_BYTES:
                continue
            text_bytes += size
        if sampled:
            break
    tokens = int(text_bytes // 4)
    if sampled and tokens:
        tokens = int(tokens * ESTIMATE_REPO_SAMPLE_MULTIPLIER)
    data = {"tokens": tokens, "file_count": total_files, "sampled": sampled}
    with _estimate_repo_cache_lock:
        _estimate_repo_cache[key] = {"ts": now, "data": data}
    return data


def _estimate_job_tokens_raw(payload: Dict[str, Any], *, include_repo: bool = True) -> int:
    workflow = (payload.get("workflow") or "project_solver").strip()
    input_text = ""
    for key in ("requirements_text", "topic_source", "topic_research", "jql", "projects", "space", "extra_args"):
        value = payload.get(key)
        if isinstance(value, str) and value.strip():
            input_text += value.strip() + "\\n"
    req_path = payload.get("requirements_path")
    if isinstance(req_path, str) and req_path.strip() and os.path.exists(req_path):
        input_text += _read_file_limited(req_path, REQUIREMENTS_MAX_BYTES)

    input_tokens = _estimate_tokens_from_text(input_text)
    req_count = _count_req_lines(input_text)
    if req_count:
        input_tokens += req_count * 120

    max_tokens = payload.get("llm_max_tokens")
    try:
        max_tokens = int(max_tokens) if max_tokens is not None else None
    except Exception:
        max_tokens = None

    repo_tokens = 0
    if include_repo and workflow in {"project_solver", "project", "delivery_pipeline"}:
        repo_root = payload.get("project_root") or payload.get("delivery_project_root")
        repo_tokens = _estimate_repo_tokens(repo_root).get("tokens", 0)

    if workflow in {"project_solver", "project"}:
        iterations = payload.get("project_iterations") or 1
        try:
            iterations = int(iterations)
        except Exception:
            iterations = 1
        max_tokens_eff = max_tokens if max_tokens is not None else DEFAULT_LLM_MAX_TOKENS
        calls = 3.2 * max(1, iterations)
        estimate = 900 + input_tokens * 1.8 + max_tokens_eff * calls + repo_tokens * 0.25
        return int(estimate)
    if workflow == "topic_research":
        max_iters = payload.get("max_iterations") or 3
        try:
            max_iters = int(max_iters)
        except Exception:
            max_iters = 3
        max_tokens_eff = max_tokens if max_tokens is not None else 2000
        calls = 2.2 * max(1, max_iters)
        estimate = 800 + input_tokens * 2.2 + max_tokens_eff * calls
        return int(estimate)
    if workflow in {"jira_analysis", "confluence_analysis"}:
        max_tokens_eff = max_tokens if max_tokens is not None else 1600
        calls = 1.6
        estimate = 700 + input_tokens * 1.2 + max_tokens_eff * calls
        return int(estimate)
    if workflow == "delivery_pipeline":
        max_tokens_eff = max_tokens if max_tokens is not None else 2400
        calls = 2.4
        estimate = 1000 + input_tokens * 1.4 + max_tokens_eff * calls + repo_tokens * 0.2
        return int(estimate)
    max_tokens_eff = max_tokens if max_tokens is not None else 1500
    estimate = 600 + input_tokens * 1.4 + max_tokens_eff * 1.2
    return int(estimate)


def _estimate_calibration() -> Dict[str, Dict[str, float]]:
    if "manager" not in globals():
        return {}
    now = time.time()
    job_count = len(manager.jobs)
    cached = _estimate_calibration_cache.get("data") or {}
    if (
        cached
        and (now - _estimate_calibration_cache.get("ts", 0)) < ESTIMATE_CALIBRATION_TTL_SEC
        and _estimate_calibration_cache.get("job_count") == job_count
    ):
        return cached
    ratios: Dict[str, List[float]] = {}
    for job in manager.list_jobs():
        actual = int(job.token_actual or 0)
        if actual <= 0:
            continue
        payload = job.payload if isinstance(job.payload, dict) else {}
        base = _estimate_job_tokens_raw(payload, include_repo=False)
        if base <= 0:
            continue
        ratio = actual / float(base)
        ratio = min(4.0, max(0.35, ratio))
        ratios.setdefault(job.workflow or "unknown", []).append(ratio)
    data: Dict[str, Dict[str, float]] = {}
    for workflow, values in ratios.items():
        if not values:
            continue
        data[workflow] = {"ratio": float(statistics.median(values))}
    _estimate_calibration_cache.update({"ts": now, "data": data, "job_count": job_count})
    return data


def _estimate_job_tokens(payload: Dict[str, Any]) -> int:
    workflow = (payload.get("workflow") or "project_solver").strip()
    estimate = _estimate_job_tokens_raw(payload, include_repo=True)
    calibration = _estimate_calibration()
    if workflow in calibration:
        ratio = calibration[workflow].get("ratio")
        if ratio:
            estimate = int(estimate * ratio)
    return max(300, int(estimate))


def _chargeable_token_estimate(
    owner: Optional[str],
    payload: Optional[Dict[str, Any]],
    estimate: Optional[int],
) -> int:
    total_estimate = max(0, int(estimate or 0))
    if total_estimate <= 0 or not owner or not isinstance(payload, dict):
        return total_estimate
    primary = _resolve_llm_settings(
        user=str(owner),
        provider_hint=payload.get("llm_provider"),
        model_hint=payload.get("llm_model"),
    )
    if payload.get("fallback_llm_provider") or payload.get("fallback_llm_model"):
        fallback = _resolve_llm_settings(
            user=str(owner),
            provider_hint=payload.get("fallback_llm_provider"),
            model_hint=payload.get("fallback_llm_model"),
        )
    else:
        fallback = _fallback_llm_settings(str(owner), primary)
    return total_estimate if (primary.get("chargeable") or fallback.get("chargeable")) else 0


REQ_HEADER_ALIASES = {
    "id": {"id", "req", "req id", "requirement id", "requirement", "requirement_id", "req_id"},
    "title": {"title", "summary", "name", "requirement title"},
    "description": {"description", "details", "detail", "notes", "desc", "statement"},
}


def _normalize_header_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[\s_-]+", " ", text)
    return text


def _detect_req_header(row: List[Any]) -> Optional[Dict[str, int]]:
    mapping: Dict[str, int] = {}
    for idx, cell in enumerate(row):
        key = _normalize_header_cell(cell)
        if not key:
            continue
        for field, aliases in REQ_HEADER_ALIASES.items():
            if key in aliases and field not in mapping:
                mapping[field] = idx
    return mapping or None


def _extract_req_id(text: str) -> Tuple[Optional[str], str]:
    if not text:
        return None, ""
    match = re.search(r"\bREQ-\d{3,}\b", text, re.I)
    if not match:
        return None, text.strip()
    req_id = match.group(0).upper()
    remaining = (text[: match.start()] + text[match.end() :]).strip(" :-–")
    return req_id, remaining or text.strip()


def _coerce_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def _row_to_requirement(row: List[Any], header: Optional[Dict[str, int]]) -> Optional[Dict[str, str]]:
    if not row:
        return None
    values = [_coerce_cell(cell) for cell in row]
    if header:
        req_id = values[header.get("id", -1)] if "id" in header and header.get("id", -1) < len(values) else ""
        title = values[header.get("title", -1)] if "title" in header and header.get("title", -1) < len(values) else ""
        desc = (
            values[header.get("description", -1)]
            if "description" in header and header.get("description", -1) < len(values)
            else ""
        )
    else:
        req_id = values[0] if len(values) > 0 else ""
        title = values[1] if len(values) > 1 else ""
        desc = values[2] if len(values) > 2 else ""
        if len(values) == 2:
            req_id = ""
            title = values[0]
            desc = values[1]
        if len(values) == 1:
            req_id = ""
            title = values[0]
            desc = ""

    req_id = req_id.strip().upper()
    title = title.strip()
    desc = desc.strip()
    if not req_id and title:
        extracted, remaining = _extract_req_id(title)
        if extracted:
            req_id = extracted
            title = remaining
    if not req_id and desc:
        extracted, remaining = _extract_req_id(desc)
        if extracted:
            req_id = extracted
            desc = remaining
    if not title and desc:
        title = desc
        desc = ""
    if not any((req_id, title, desc)):
        return None
    return {"id": req_id, "title": title, "description": desc}


def _rows_to_requirements(rows: List[List[Any]]) -> List[Dict[str, str]]:
    if not rows:
        return []
    header = _detect_req_header(rows[0])
    start_idx = 1 if header else 0
    items: List[Dict[str, str]] = []
    for row in rows[start_idx:]:
        if not row or not any(_coerce_cell(cell) for cell in row):
            continue
        item = _row_to_requirement(row, header)
        if item:
            items.append(item)
    return items


def _parse_csv_rows(data: bytes) -> List[List[str]]:
    try:
        text = data.decode("utf-8-sig")
    except Exception:
        text = data.decode("latin-1", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    return [row for row in reader if row and any(cell.strip() for cell in row)]


def _parse_xlsx_rows(data: bytes) -> List[List[Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError("openpyxl is required to import .xlsx files") from exc
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows: List[List[Any]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def _parse_xls_rows(data: bytes) -> List[List[Any]]:
    try:
        import xlrd  # type: ignore
    except Exception as exc:
        raise RuntimeError("xlrd is required to import .xls files") from exc
    book = xlrd.open_workbook(file_contents=data)
    sheet = book.sheet_by_index(0)
    rows: List[List[Any]] = []
    for idx in range(sheet.nrows):
        rows.append(sheet.row_values(idx))
    return rows


def _parse_ods_rows(data: bytes) -> List[List[Any]]:
    try:
        from odf.opendocument import load  # type: ignore
        from odf.table import Table, TableRow, TableCell  # type: ignore
        from odf.text import P  # type: ignore
    except Exception as exc:
        raise RuntimeError("odfpy is required to import .ods files") from exc
    with tempfile.NamedTemporaryFile(suffix=".ods", delete=False) as handle:
        handle.write(data)
        temp_path = handle.name
    try:
        doc = load(temp_path)
    finally:
        try:
            os.remove(temp_path)
        except Exception:
            pass
    tables = doc.spreadsheet.getElementsByType(Table)
    if not tables:
        return []
    table = tables[0]
    rows: List[List[Any]] = []
    for row in table.getElementsByType(TableRow):
        values: List[Any] = []
        for cell in row.getElementsByType(TableCell):
            repeat = cell.getAttribute("numbercolumnsrepeated")
            try:
                repeat_count = int(repeat) if repeat else 1
            except Exception:
                repeat_count = 1
            text_parts: List[str] = []
            for p in cell.getElementsByType(P):
                if p.firstChild:
                    text_parts.append(str(p.firstChild.data))
            value = " ".join(text_parts).strip()
            if not value:
                string_val = cell.getAttribute("stringvalue")
                if string_val:
                    value = str(string_val)
                else:
                    raw_val = cell.getAttribute("value")
                    if raw_val is not None:
                        value = str(raw_val)
            for _ in range(repeat_count):
                values.append(value)
        rows.append(values)
    return rows


def _export_csv(items: List[Dict[str, str]]) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Title", "Description"])
    for item in items:
        writer.writerow([item.get("id", ""), item.get("title", ""), item.get("description", "")])
    return output.getvalue().encode("utf-8")


def _export_xlsx(items: List[Dict[str, str]]) -> bytes:
    try:
        from openpyxl import Workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError("openpyxl is required to export .xlsx files") from exc
    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements"
    ws.append(["ID", "Title", "Description"])
    for item in items:
        ws.append([item.get("id", ""), item.get("title", ""), item.get("description", "")])
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _export_xls(items: List[Dict[str, str]]) -> bytes:
    try:
        import xlwt  # type: ignore
    except Exception as exc:
        raise RuntimeError("xlwt is required to export .xls files") from exc
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Requirements")
    headers = ["ID", "Title", "Description"]
    for col, name in enumerate(headers):
        ws.write(0, col, name)
    for row_idx, item in enumerate(items, start=1):
        ws.write(row_idx, 0, item.get("id", ""))
        ws.write(row_idx, 1, item.get("title", ""))
        ws.write(row_idx, 2, item.get("description", ""))
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _export_ods(items: List[Dict[str, str]]) -> bytes:
    try:
        from odf.opendocument import OpenDocumentSpreadsheet  # type: ignore
        from odf.table import Table, TableRow, TableCell  # type: ignore
        from odf.text import P  # type: ignore
    except Exception as exc:
        raise RuntimeError("odfpy is required to export .ods files") from exc
    doc = OpenDocumentSpreadsheet()
    table = Table(name="Requirements")
    doc.spreadsheet.addElement(table)
    for row in [["ID", "Title", "Description"]]:
        row_el = TableRow()
        for value in row:
            cell = TableCell(valuetype="string")
            cell.addElement(P(text=str(value)))
            row_el.addElement(cell)
        table.addElement(row_el)
    for item in items:
        row_el = TableRow()
        for value in [item.get("id", ""), item.get("title", ""), item.get("description", "")]:
            cell = TableCell(valuetype="string")
            cell.addElement(P(text=str(value)))
            row_el.addElement(cell)
        table.addElement(row_el)
    with tempfile.NamedTemporaryFile(suffix=".ods", delete=False) as handle:
        temp_path = handle.name
    output_path = temp_path
    try:
        doc.save(temp_path)
        if not os.path.exists(output_path) and os.path.exists(f"{temp_path}.ods"):
            output_path = f"{temp_path}.ods"
        with open(output_path, "rb") as handle:
            return handle.read()
    finally:
        for path in {temp_path, output_path}:
            try:
                os.remove(path)
            except Exception:
                pass


def _normalize_requirements_items(items: Any) -> List[Dict[str, str]]:
    normalized: List[Dict[str, str]] = []
    if not isinstance(items, list):
        return normalized
    for item in items:
        if not isinstance(item, dict):
            continue
        req_id = str(item.get("id") or "").strip().upper()
        title = str(item.get("title") or "").strip()
        desc = str(item.get("description") or "").strip()
        if not any((req_id, title, desc)):
            continue
        normalized.append({"id": req_id, "title": title, "description": desc})
    return normalized


REFUND_STATUSES = {
    "requested",
    "awaiting approval",
    "approved",
    "settled",
    "rejected",
    "partial-refund",
}


def _find_refund_request(job: "Job", request_id: str) -> Optional[Dict[str, Any]]:
    if not request_id:
        return None
    for refund in job.refunds:
        if refund.get("id") == request_id:
            return refund
    return None


def _latest_refund_request(job: "Job") -> Optional[Dict[str, Any]]:
    if not job.refunds:
        return None
    return max(job.refunds, key=lambda entry: _timestamp_sort_key(entry.get("requested_at")))


def _refund_job_snapshot(job: "Job") -> Dict[str, Any]:
    tokens = job.metrics.get("token_usage") if isinstance(job.metrics, dict) else {}
    return {
        "job_id": job.job_id,
        "project_name": job.project_name,
        "workflow": job.workflow,
        "status": job.status,
        "started_at": job.started_at,
        "finished_at": job.finished_at,
        "token_estimate": job.token_estimate,
        "token_actual": job.token_actual,
        "token_debited": job.token_debited,
        "token_shortfall": job.token_shortfall,
        "token_usage_total": tokens.get("total") if isinstance(tokens, dict) else None,
    }


def _refund_max_amount(job: "Job") -> int:
    actual = int(job.token_debited or 0)
    if actual <= 0:
        tokens = job.metrics.get("token_usage") if isinstance(job.metrics, dict) else {}
        actual = int(tokens.get("chargeable_total") or 0) if isinstance(tokens, dict) else 0
    return max(0, actual)


def _safe_amount(value: Any) -> Optional[int]:
    try:
        amount = int(float(value))
    except Exception:
        return None
    if amount <= 0:
        return None
    return amount


def _store_refund_files(job_id: str, request_id: str, files: List[Any]) -> List[Dict[str, str]]:
    stored: List[Dict[str, str]] = []
    if not files:
        return stored
    job_dir = os.path.join(JOB_ROOT, job_id)
    refund_dir = os.path.join(job_dir, "refunds", request_id)
    os.makedirs(refund_dir, exist_ok=True)
    for idx, file in enumerate(files, start=1):
        filename = secure_filename(file.filename or "")
        if not filename:
            continue
        ext = os.path.splitext(filename)[1].lower()
        if ext not in REFUND_ALLOWED_EXTS:
            continue
        file.stream.seek(0, os.SEEK_END)
        size = file.stream.tell()
        file.stream.seek(0)
        if size > REFUND_SCREENSHOT_MAX_BYTES:
            continue
        if idx > REFUND_MAX_FILES:
            break
        base, ext = os.path.splitext(filename)
        safe_name = filename
        if os.path.exists(os.path.join(refund_dir, safe_name)):
            safe_name = f"{base}_{idx}{ext}"
        abs_path = os.path.join(refund_dir, safe_name)
        file.save(abs_path)
        stored.append(
            {
                "filename": safe_name,
                "path": os.path.join("refunds", request_id, safe_name),
                "uploaded_at": _now_iso(),
            }
        )
    return stored


def _extract_json_payload(text: str) -> Optional[Any]:
    if not text:
        return None
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```[a-zA-Z]*", "", cleaned).strip()
        cleaned = re.sub(r"```$", "", cleaned).strip()
    try:
        return json.loads(cleaned)
    except Exception:
        pass
    start = None
    end = None
    for i, ch in enumerate(cleaned):
        if ch in "{[":
            start = i
            break
    for i in range(len(cleaned) - 1, -1, -1):
        if cleaned[i] in "]}":
            end = i + 1
            break
    if start is not None and end is not None and end > start:
        try:
            return json.loads(cleaned[start:end])
        except Exception:
            return None
    return None


def _current_user() -> Optional[str]:
    if not has_request_context():
        return None
    if getattr(g, "auth_user_resolved", False):
        return getattr(g, "auth_user", None)
    session_user = session.get("user")
    if session_user:
        identity = _normalize_auth_identity(_local_identity_payload(str(session_user).strip(), include_directory=True))
        _set_request_identity(identity, resolved=True)
        g.auth_profile = identity
        g.auth_profile_resolved = True
        return str(session_user).strip()
    auth_header, cookie_header = _current_request_auth_headers()
    token = _extract_bearer_token(auth_header)
    if _customers_enabled() and (auth_header or cookie_header):
        _customers_profile_prefetch(
            authorization=auth_header,
            cookie_header=cookie_header,
        )
        try:
            payload = _run_external_call(
                "customers session lookup",
                CUSTOMERS_SERVICE.session_from_headers,
                authorization=auth_header,
                cookie_header=cookie_header,
                timeout=min(
                    _service_timeout_seconds(CUSTOMERS_SERVICE, default=max(1.0, CUSTOMERS_TIMEOUT)),
                    CUSTOMERS_SESSION_LOOKUP_WAIT_SEC,
                ),
            )
        except Exception as exc:
            logger.warning("customers session lookup failed: %s", exc)
            payload = {}
        identity = _normalize_auth_identity(payload)
        if not (identity and identity.get("authenticated") and identity.get("user")):
            prefetched_payload = _customers_profile_from_prefetch(wait_timeout=CUSTOMERS_PROFILE_PREFETCH_WAIT_SEC)
            prefetched_identity = _normalize_auth_identity(prefetched_payload or {})
            if prefetched_identity and prefetched_identity.get("authenticated") and prefetched_identity.get("user"):
                identity = prefetched_identity
        if identity and identity.get("authenticated") and identity.get("user"):
            _set_request_identity(identity, resolved=True)
            return str(identity.get("user") or "").strip() or None
    if token and CENTRAL_STORE is not None:
        try:
            identity = CENTRAL_STORE.access_tokens.verify(token)
        except Exception as exc:
            logger.warning("central access token verification failed: %s", exc)
            identity = None
        normalized_identity = _normalize_auth_identity(identity)
        if normalized_identity and normalized_identity.get("user"):
            _set_request_identity(normalized_identity, resolved=True)
            return str(normalized_identity.get("user") or "").strip() or None
    _set_request_identity(None, resolved=True)
    return None


def _issue_access_token_payload(user: str, *, source: str) -> Dict[str, Any]:
    if CENTRAL_STORE is None or not user:
        return {}
    try:
        issued = CENTRAL_STORE.access_tokens.issue(
            user,
            label=source,
            meta={"source": source},
        )
    except Exception as exc:
        logger.warning("central access token issue failed for %s: %s", user, exc)
        return {}
    payload = {
        "access_token": issued.get("token"),
    }
    expires_at = issued.get("expires_at")
    if expires_at:
        payload["access_expires_at"] = expires_at
    return payload


def _touch_user_activity(user: Optional[str]) -> None:
    if not user:
        return
    with _user_activity_lock:
        _user_activity[user] = time.time()


def _is_user_active(user: Optional[str], window_seconds: Optional[int] = None) -> bool:
    if not user:
        return False
    window = window_seconds if window_seconds is not None else ACTIVE_WINDOW_SEC
    with _user_activity_lock:
        last_seen = _user_activity.get(user)
    if last_seen is None:
        return False
    return (time.time() - last_seen) <= window


def _clear_user_activity(user: Optional[str]) -> None:
    if not user:
        return
    with _user_activity_lock:
        _user_activity.pop(user, None)


def _active_users_snapshot(window_seconds: Optional[int] = None) -> List[Dict[str, object]]:
    window = window_seconds if window_seconds is not None else ACTIVE_WINDOW_SEC
    now = time.time()
    with _user_activity_lock:
        items = list(_user_activity.items())
    active: List[Dict[str, object]] = []
    for user, last_seen in items:
        age = now - last_seen
        if age <= window:
            active.append({"user": user, "last_seen_sec": int(age)})
    active.sort(key=lambda entry: entry.get("last_seen_sec", 0))
    return active


def _max_timestamp(left: Optional[str], right: Optional[str]) -> Optional[str]:
    if not left:
        return right
    if not right:
        return left
    return left if _timestamp_sort_key(left) >= _timestamp_sort_key(right) else right


def _user_token_snapshot(user: str) -> Dict[str, Any]:
    summary = _account_summary("user", user)
    reserved = manager.reserved_tokens(owner=user, source="user")
    in_use = manager.in_use_tokens(owner=user, source="user")
    balance = int(summary.get("balance") or 0)
    paid_balance = int(summary.get("paid_balance") or balance)
    free_balance = int(summary.get("free_balance") or 0)
    spent_total = int(summary.get("spent_total") or 0)
    cashout_total = int(summary.get("cashout_total") or 0)
    shortfall_total = int(summary.get("shortfall_total") or 0)
    free_grant_total = int(summary.get("free_grant_total") or 0)
    available = max(0, balance - reserved)
    capacity = int(summary.get("last_topup_tokens") or balance or 0)
    display_capacity = max(1, capacity, balance)
    low_threshold = int(round(capacity * 0.2)) if capacity else 0
    status = "low" if capacity and balance <= low_threshold else "ok"
    return {
        "balance": balance,
        "tokens": balance,
        "paid_balance": paid_balance,
        "free_balance": free_balance,
        "available": available,
        "reserved": reserved,
        "in_use": in_use,
        "last_topup_tokens": int(summary.get("last_topup_tokens") or 0),
        "capacity": capacity,
        "display_capacity": display_capacity,
        "low_threshold": low_threshold,
        "status": status,
        "last_topup_at": summary.get("last_topup_at"),
        "updated_at": summary.get("updated_at"),
        "spent_total": spent_total,
        "cashout_total": cashout_total,
        "shortfall_total": shortfall_total,
        "free_grant_total": free_grant_total,
        "identity": summary.get("identity") if isinstance(summary.get("identity"), dict) else None,
    }


def _team_token_snapshot(team_id: str) -> Dict[str, Any]:
    summary = _account_summary("team", team_id)
    reserved = manager.reserved_tokens(team_id=team_id, source="team")
    in_use = manager.in_use_tokens(team_id=team_id, source="team")
    balance = int(summary.get("balance") or 0)
    paid_balance = int(summary.get("paid_balance") or balance)
    free_balance = int(summary.get("free_balance") or 0)
    spent_total = int(summary.get("spent_total") or 0)
    cashout_total = int(summary.get("cashout_total") or 0)
    shortfall_total = int(summary.get("shortfall_total") or 0)
    free_grant_total = int(summary.get("free_grant_total") or 0)
    available = max(0, balance - reserved)
    capacity = int(summary.get("last_topup_tokens") or balance or 0)
    display_capacity = max(1, capacity, balance)
    low_threshold = int(round(capacity * 0.2)) if capacity else 0
    status = "low" if capacity and balance <= low_threshold else "ok"
    return {
        "balance": balance,
        "tokens": balance,
        "paid_balance": paid_balance,
        "free_balance": free_balance,
        "available": available,
        "reserved": reserved,
        "in_use": in_use,
        "last_topup_tokens": int(summary.get("last_topup_tokens") or 0),
        "capacity": capacity,
        "display_capacity": display_capacity,
        "low_threshold": low_threshold,
        "status": status,
        "last_topup_at": summary.get("last_topup_at"),
        "updated_at": summary.get("updated_at"),
        "spent_total": spent_total,
        "cashout_total": cashout_total,
        "shortfall_total": shortfall_total,
        "free_grant_total": free_grant_total,
    }


def _token_snapshot(user: str, team_id: Optional[str] = None) -> Dict[str, Any]:
    user_snapshot = _user_token_snapshot(user)
    if not team_id:
        return {
            **user_snapshot,
            "btc_rate": TOKEN_BTC_RATE,
            "scope": "personal",
        }
    team_snapshot = _team_token_snapshot(team_id)
    team_name = None
    team = access_store.get_team(team_id)
    if team:
        team_name = team.get("name")
    balance = user_snapshot["balance"] + team_snapshot["balance"]
    paid_balance = user_snapshot["paid_balance"] + team_snapshot["paid_balance"]
    free_balance = user_snapshot["free_balance"] + team_snapshot["free_balance"]
    reserved = user_snapshot["reserved"] + team_snapshot["reserved"]
    in_use = user_snapshot["in_use"] + team_snapshot["in_use"]
    available = max(0, balance - reserved)
    capacity = user_snapshot["capacity"] + team_snapshot["capacity"]
    display_capacity = max(1, capacity, balance)
    low_threshold = int(round(capacity * 0.2)) if capacity else 0
    status = "low" if capacity and balance <= low_threshold else "ok"
    return {
        "balance": balance,
        "tokens": balance,
        "paid_balance": paid_balance,
        "free_balance": free_balance,
        "available": available,
        "reserved": reserved,
        "in_use": in_use,
        "capacity": capacity,
        "display_capacity": display_capacity,
        "low_threshold": low_threshold,
        "status": status,
        "btc_rate": TOKEN_BTC_RATE,
        "last_topup_at": _max_timestamp(user_snapshot.get("last_topup_at"), team_snapshot.get("last_topup_at")),
        "updated_at": _max_timestamp(user_snapshot.get("updated_at"), team_snapshot.get("updated_at")),
        "user_balance": user_snapshot.get("balance"),
        "user_paid_balance": user_snapshot.get("paid_balance"),
        "user_free_balance": user_snapshot.get("free_balance"),
        "user_available": user_snapshot.get("available"),
        "user_reserved": user_snapshot.get("reserved"),
        "user_in_use": user_snapshot.get("in_use"),
        "user_capacity": user_snapshot.get("capacity"),
        "user_last_topup_tokens": user_snapshot.get("last_topup_tokens"),
        "team_balance": team_snapshot.get("balance"),
        "team_paid_balance": team_snapshot.get("paid_balance"),
        "team_free_balance": team_snapshot.get("free_balance"),
        "team_available": team_snapshot.get("available"),
        "team_reserved": team_snapshot.get("reserved"),
        "team_in_use": team_snapshot.get("in_use"),
        "team_capacity": team_snapshot.get("capacity"),
        "team_last_topup_tokens": team_snapshot.get("last_topup_tokens"),
        "team_id": team_id,
        "team_name": team_name,
        "scope": "team",
    }


def _can_access_team_tokens(user: Optional[str], team_id: Optional[str]) -> bool:
    if not user or not team_id:
        return False
    if _is_admin_user(user):
        return True
    if _team_role_for_user(user, team_id):
        return True
    for project in access_store.list_projects():
        if project.get("team_id") != team_id:
            continue
        caps = access_store.project_capabilities(user, project.get("id"))
        if caps.get("read"):
            return True
    return False


def _is_team_leader(user: Optional[str], team_id: Optional[str]) -> bool:
    if not user or not team_id:
        return False
    if _is_admin_user(user):
        return True
    return _team_role_for_user(user, team_id) == "leader"


def _is_secure_request() -> bool:
    if request.is_secure:
        return True
    forwarded_proto = request.headers.get("X-Forwarded-Proto", "")
    if forwarded_proto:
        proto = forwarded_proto.split(",")[0].strip().lower()
        if proto == "https":
            return True
    return False


def _enforce_https() -> Optional[Response]:
    if not ENFORCE_HTTPS:
        return None
    if _is_secure_request():
        return None
    if request.method == "GET":
        https_url = request.url.replace("http://", "https://", 1)
        return redirect(https_url, code=308)
    return jsonify({"error": "https_required"}), 403


def _allowed_origin() -> Optional[str]:
    origin = request.headers.get("Origin")
    if not origin:
        return None
    normalized = origin.rstrip("/")
    if normalized in CORS_ORIGINS:
        return normalized
    return None


def _origin_matches_host(origin: str) -> bool:
    if not origin:
        return False
    host_url = request.host_url.rstrip("/")
    if origin.rstrip("/") == host_url:
        return True
    if SITE_BASE and origin.rstrip("/") == SITE_BASE.rstrip("/"):
        return True
    return False


def _check_origin() -> Optional[Response]:
    if not CSRF_ORIGIN_CHECK:
        return None
    if request.method in {"GET", "HEAD", "OPTIONS"}:
        return None
    origin = request.headers.get("Origin")
    if origin:
        if _allowed_origin() or _origin_matches_host(origin):
            return None
        return jsonify({"error": "origin_not_allowed"}), 403
    referer = request.headers.get("Referer")
    if referer:
        if _origin_matches_host(referer):
            return None
        return jsonify({"error": "referer_not_allowed"}), 403
    return None


def _add_vary(response: Response, value: str) -> None:
    current = response.headers.get("Vary")
    if not current:
        response.headers["Vary"] = value
        return
    existing = [item.strip() for item in current.split(",") if item.strip()]
    if value not in existing:
        response.headers["Vary"] = ", ".join(existing + [value])


def _apply_cors(response: Response) -> Response:
    origin = _allowed_origin()
    if not origin:
        return response
    response.headers["Access-Control-Allow-Origin"] = origin
    response.headers["Access-Control-Allow-Credentials"] = "true"
    response.headers["Access-Control-Allow-Headers"] = ", ".join(CORS_ALLOW_HEADERS)
    response.headers["Access-Control-Allow-Methods"] = ", ".join(CORS_ALLOW_METHODS)
    response.headers["Access-Control-Max-Age"] = str(CORS_MAX_AGE)
    _add_vary(response, "Origin")
    _add_vary(response, "Access-Control-Request-Method")
    _add_vary(response, "Access-Control-Request-Headers")
    if ENFORCE_HTTPS and _is_secure_request():
        response.headers.setdefault("Strict-Transport-Security", "max-age=63072000; includeSubDomains")
    return response


def _metrics_path_label() -> str:
    if request.url_rule and request.url_rule.rule:
        return request.url_rule.rule
    path = request.path or ""
    if path.startswith("/static/"):
        return "/static/*"
    if path.startswith("/public/"):
        return "/public/*"
    return path or "unknown"


def _require_login() -> Optional[Response]:
    path = request.path or ""
    if (
        path.startswith("/static/")
        or path.startswith("/public/")
        or path.startswith("/favicon")
        or path.startswith("/billing/assets/")
    ):
        return None
    if path == METRICS_PATH:
        return None
    # Allow public access to API documentation and health endpoints
    if path.startswith("/api/docs") or path in {"/health", "/api/version"}:
        return None
    if path in {
        "/login",
        "/register",
        "/auth/external-login",
        "/sso",
        "/oidc/login",
        "/oidc/callback",
        "/setup",
        "/api/auth/config",
        "/api/login",
        "/api/login/mfa/totp",
        "/api/logout",
        "/api/oidc/exchange",
        "/api/register",
        "/api/session",
        "/api/setup",
        "/api/passkeys/authenticate/options",
        "/api/passkeys/authenticate/verify",
        "/api/profile/mfa/totp/start",
        "/api/profile/mfa/totp/verify",
        "/api/profile/mfa/totp/disable",
        "/api/profile/passkeys/register/options",
        "/api/profile/passkeys/register/verify",
    } or path.startswith("/api/health"):
        return None
    if path.startswith("/api/profile/passkeys/"):
        return None
    if path.startswith("/api/voice/") and not path.startswith("/api/voice/tokens"):
        return None
    is_api = path.startswith("/api/")
    if not user_store.has_users():
        if AUTH_MODE in {"oidc", "mixed"}:
            return None
        if is_api:
            return jsonify({"error": "unauthorized"}), 401
        if path != "/setup":
            return redirect(url_for("setup"))
        return None
    if not _current_user():
        if is_api:
            return jsonify({"error": "unauthorized"}), 401
        if AUTH_MODE == "oidc":
            return redirect(url_for("oidc_login"))
        return redirect(url_for("login"))
    identity = _current_auth_profile() or _current_auth_identity()
    if not _identity_has_service_access(identity, "refiner", SERVICE_ACCESS_USE):
        if is_api:
            return (
                jsonify(
                    {
                        "error": "forbidden",
                        "service": "refiner",
                        "required_access": SERVICE_ACCESS_USE,
                    }
                ),
                403,
            )
        return make_response("Forbidden", 403)
    return None


@app.before_request
def _before_request() -> Optional[Response]:
    if METRICS_ENABLED:
        g.metrics_start = time.time()
        INFLIGHT.inc()
    https_block = _enforce_https()
    if https_block:
        return https_block
    if request.method == "OPTIONS":
        return Response(status=204)
    origin_block = _check_origin()
    if origin_block:
        return origin_block
    auth_block = _require_login()
    if auth_block:
        return auth_block
    _touch_user_activity(_current_user())
    return None


@app.after_request
def _after_request(response: Response) -> Response:
    if METRICS_ENABLED:
        try:
            path_label = _metrics_path_label()
            elapsed = time.time() - getattr(g, "metrics_start", time.time())
            REQUEST_LATENCY.labels(request.method, path_label).observe(elapsed)
            REQUEST_COUNT.labels(request.method, path_label, response.status_code).inc()
        finally:
            INFLIGHT.dec()
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "DENY")
    response.headers.setdefault("Referrer-Policy", "same-origin")
    if CSP_POLICY:
        response.headers.setdefault("Content-Security-Policy", CSP_POLICY)
    if request.path.startswith("/api/"):
        response.headers.setdefault("Cache-Control", "no-store")
    return _apply_cors(response)


def index() -> str:
    """Render or serve the index route."""
    user = _current_user()
    return render_template(
        "index.html",
        current_user=user,
        user_role=_user_role(user) if user else None,
        api_base=API_BASE,
        site_base=SITE_BASE,
    )


def playground() -> str:
    """Render or serve the playground route."""
    user = _current_user()
    return render_template(
        "playground.html",
        current_user=user,
        user_role=_user_role(user) if user else None,
        api_base=API_BASE,
        site_base=SITE_BASE,
    )


def admin_dashboard() -> Response:
    """Render the admin dashboard page."""
    user = _current_user()
    role = _user_role(user) if user else None
    if not _is_admin_user(user):
        return Response("forbidden", status=403)
    return render_template(
        "admin.html",
        current_user=user,
        user_role=role,
        api_base=API_BASE,
        site_base=SITE_BASE,
        active_window_sec=ACTIVE_WINDOW_SEC,
    )


def public_asset(filename: str) -> Response:
    """Serve files from the public web assets directory."""
    return send_from_directory(PUBLIC_DIR, filename)


def favicon() -> Response:
    """Render or serve the favicon route."""
    return send_from_directory(PUBLIC_DIR, "favicon.ico")


def metrics() -> Response:
    """Render or serve the metrics route."""
    if not METRICS_ENABLED:
        return jsonify({"error": "metrics_disabled"}), 404

    try:
        WORKER_COUNT.set(len(manager.workers))
        JOB_QUEUE_DEPTH.set(manager.queue.qsize())
        JOB_ACTION_QUEUE_DEPTH.set(job_action_manager.queue_depth())
        JOB_ACTION_INFLIGHT.set(job_action_manager.inflight())
        UPTIME.set(time.time() - APP_START_TIME)

        status_counts: Dict[str, int] = {}
        with manager.lock:
            jobs_snapshot = list(manager.jobs.values())
        for job in jobs_snapshot:
            status_counts[job.status] = status_counts.get(job.status, 0) + 1

        for status in ("queued", "running", "paused", "stopped", "completed", "failed"):
            JOBS_BY_STATUS.labels(status=status).set(status_counts.get(status, 0))
    except Exception:
        pass

    payload = generate_latest()
    return Response(payload, mimetype=CONTENT_TYPE_LATEST)


def login() -> Response:
    """Render or serve the login route."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    next_path = _safe_next_path(request.args.get("next") or session.get("login_next"))
    if AUTH_MODE == "oidc":
        session["login_next"] = next_path
        return redirect(url_for("oidc_login", next=next_path))
    if not user_store.has_users() and not OIDC_ENABLED:
        session["login_next"] = next_path
        return redirect(url_for("setup"))
    error = None
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = (request.form.get("password") or "").strip()
        if _login_throttled(username):
            _audit_event("login", actor=username, status="throttled")
            error = "Too many attempts. Please try again later."
        elif user_store.verify(username, password):
            session["user"] = username
            _record_identity_event(
                username,
                role=user_store.get_role(username),
                email=user_store.get_email(username),
                provider="local",
                meta={"source": "login_form"},
            )
            _record_login_event(username, auth_mode="local", provider="local", meta={"source": "login_form"})
            _record_login_attempt(username, ok=True)
            _audit_event("login", actor=username, status="success")
            session.pop("login_next", None)
            return redirect(next_path)
        else:
            _record_login_attempt(username, ok=False)
            _audit_event("login", actor=username, status="failed")
            error = "Invalid username or password."
    return render_template(
        "login.html",
        error=error,
        api_base=API_BASE,
        site_base=SITE_BASE,
        oidc_enabled=OIDC_ENABLED,
        oidc_label=OIDC_BUTTON_LABEL,
        local_enabled=user_store.has_users(),
        next_path=next_path,
        self_registration_enabled=bool(SELF_REGISTRATION_ENABLED and user_store.has_users()),
        password_min_length=PASSWORD_MIN_LEN,
        passkeys_supported=False,
    )


def register() -> Response:
    """Render or serve the self-registration route."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    next_path = _safe_next_path(request.args.get("next") or session.get("login_next"))
    if AUTH_MODE == "oidc":
        return redirect(url_for("login", next=next_path))
    if not SELF_REGISTRATION_ENABLED:
        return redirect(url_for("login", next=next_path))
    if not user_store.has_users():
        return redirect(url_for("setup", next=next_path))
    error = None
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = (request.form.get("password") or "").strip()
        confirm = (request.form.get("confirm") or "").strip()
        email = (request.form.get("email") or "").strip()
        if not email:
            error = "Email is required."
        elif not username or not password:
            error = "Username and password are required."
        elif not USERNAME_RE.match(username):
            error = "Username must be 3-32 chars (letters, numbers, underscore, dash)."
        elif not EMAIL_RE.match(email):
            error = "Enter a valid email address."
        elif len(password) < PASSWORD_MIN_LEN:
            error = f"Password must be at least {PASSWORD_MIN_LEN} characters."
        elif confirm and password != confirm:
            error = "Passwords do not match."
        elif user_store.get_user(username):
            error = "That username is already in use."
        else:
            user_store.create_user(username, password, role="user", email=email)
            session["user"] = username
            session.pop("login_next", None)
            _record_identity_event(
                username,
                role=user_store.get_role(username),
                email=user_store.get_email(username),
                provider="register",
                meta={"source": "register_form"},
            )
            _record_login_event(username, auth_mode="register", provider="local", meta={"source": "register_form"})
            _audit_event("register", actor=username, status="success")
            return redirect(next_path)
    return render_template(
        "register.html",
        error=error,
        api_base=API_BASE,
        site_base=SITE_BASE,
        next_path=next_path,
        password_min_length=PASSWORD_MIN_LEN,
        passkeys_supported=False,
    )


def _parse_kv_params(raw: str) -> Dict[str, str]:
    params: Dict[str, str] = {}
    if not raw:
        return params
    for chunk in raw.split(","):
        chunk = chunk.strip()
        if not chunk or "=" not in chunk:
            continue
        key, value = chunk.split("=", 1)
        key = key.strip()
        value = value.strip()
        if key:
            params[key] = value
    return params


def oidc_login() -> Response:
    """Handle the oidc login route."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    if not OIDC_ENABLED:
        return jsonify({"error": "oidc_not_enabled"}), 404
    next_path = _safe_next_path(request.args.get("next") or session.get("login_next"))
    session["login_next"] = next_path
    config = _oidc_discovery()
    if not config:
        return jsonify({"error": "oidc_config_missing"}), 500
    auth_endpoint = config.get("authorization_endpoint")
    if not auth_endpoint:
        return jsonify({"error": "oidc_authorization_missing"}), 500
    state = secrets_lib.token_urlsafe(32)
    nonce = secrets_lib.token_urlsafe(16)
    code_verifier = secrets_lib.token_urlsafe(64)
    code_challenge = base64.urlsafe_b64encode(hashlib.sha256(code_verifier.encode("utf-8")).digest()).decode("utf-8").rstrip("=")
    session["oidc_state"] = state
    session["oidc_nonce"] = nonce
    session["oidc_code_verifier"] = code_verifier
    session["oidc_started_at"] = _now_iso()
    params = {
        "response_type": "code",
        "client_id": OIDC_CLIENT_ID,
        "redirect_uri": _oidc_redirect_uri(),
        "scope": OIDC_SCOPE,
        "state": state,
        "nonce": nonce,
        "code_challenge": code_challenge,
        "code_challenge_method": "S256",
    }
    extra_params = _parse_kv_params(os.getenv("REFINER_OIDC_EXTRA_PARAMS", ""))
    params.update(extra_params)
    return redirect(auth_endpoint + "?" + urlencode(params), code=302)


def oidc_callback() -> Response:
    """Handle the oidc callback route."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    if not OIDC_ENABLED:
        return jsonify({"error": "oidc_not_enabled"}), 404
    error = (request.args.get("error") or "").strip()
    if error:
        _audit_event("oidc_callback", actor=None, status="failed", details={"error": error})
        return redirect(url_for("login"))
    code = (request.args.get("code") or "").strip()
    state = (request.args.get("state") or "").strip()
    if not code or not state:
        _audit_event("oidc_callback", actor=None, status="failed", details={"error": "missing_code_or_state"})
        return redirect(url_for("login"))
    if state != session.get("oidc_state"):
        _audit_event("oidc_callback", actor=None, status="failed", details={"error": "state_mismatch"})
        return redirect(url_for("login"))
    config = _oidc_discovery()
    if not config:
        return jsonify({"error": "oidc_config_missing"}), 500
    token_endpoint = config.get("token_endpoint")
    if not token_endpoint:
        return jsonify({"error": "oidc_token_endpoint_missing"}), 500
    code_verifier = session.get("oidc_code_verifier")
    token_payload = {
        "grant_type": "authorization_code",
        "code": code,
        "redirect_uri": _oidc_redirect_uri(),
        "client_id": OIDC_CLIENT_ID,
    }
    if code_verifier:
        token_payload["code_verifier"] = code_verifier
    auth = None
    if OIDC_CLIENT_SECRET:
        if OIDC_CLIENT_AUTH == "post":
            token_payload["client_secret"] = OIDC_CLIENT_SECRET
        else:
            auth = (OIDC_CLIENT_ID, OIDC_CLIENT_SECRET)
    token_resp = _http_request_with_retry(
        "POST",
        token_endpoint,
        data=token_payload,
        auth=auth,
        timeout=15,
        retries=1,
    )
    if token_resp.status_code >= 400:
        _audit_event("oidc_callback", actor=None, status="failed", details={"error": "token_exchange_failed"})
        return redirect(url_for("login"))
    token_data = token_resp.json()
    id_token = token_data.get("id_token")
    access_token = token_data.get("access_token")
    if not id_token:
        _audit_event("oidc_callback", actor=None, status="failed", details={"error": "id_token_missing"})
        return redirect(url_for("login"))
    try:
        claims = _verify_jwt(id_token, nonce=session.get("oidc_nonce"))
    except Exception as exc:
        _audit_event("oidc_callback", actor=None, status="failed", details={"error": str(exc)})
        return redirect(url_for("login"))
    claims = _oidc_maybe_enrich_claims(claims, access_token, config=config)
    username = _oidc_username_from_claims(claims)
    if not username:
        _audit_event("oidc_callback", actor=None, status="failed", details={"error": "username_missing"})
        return redirect(url_for("login"))
    email = claims.get(OIDC_EMAIL_CLAIM) if isinstance(claims.get(OIDC_EMAIL_CLAIM), str) else None
    role = _oidc_role_from_claims(claims)
    subject = claims.get("sub") if isinstance(claims.get("sub"), str) else None
    user_store.upsert_external_user(username, role=role, email=email, provider="oidc", subject=subject)
    session["user"] = username
    _record_identity_event(
        username,
        role=role,
        email=email,
        provider="oidc",
        subject=subject,
        meta={"source": "oidc_callback"},
    )
    _record_login_event(
        username,
        auth_mode="oidc",
        provider="oidc",
        session_id=subject,
        meta={"source": "oidc_callback"},
    )
    for key in ("oidc_state", "oidc_nonce", "oidc_code_verifier", "oidc_started_at"):
        session.pop(key, None)
    _audit_event("oidc_login", actor=username, status="success", details={"role": role})
    next_path = _safe_next_path(session.pop("login_next", None))
    return redirect(next_path or url_for("index"))


def api_oidc_exchange() -> Response:
    """API endpoint for oidc exchange."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    if not OIDC_ENABLED or not OIDC_EXCHANGE_ENABLED:
        return jsonify({"error": "oidc_not_enabled"}), 404
    payload = request.get_json(force=True, silent=True)
    if not isinstance(payload, dict):
        payload = {}
    code = (payload.get("code") or "").strip()
    code_verifier = (payload.get("code_verifier") or "").strip()
    client_id = (payload.get("client_id") or "").strip()
    redirect_uri = (payload.get("redirect_uri") or "").strip()
    id_token = (payload.get("id_token") or "").strip()
    access_token = (payload.get("access_token") or "").strip()

    if redirect_uri and not _oidc_is_redirect_allowed(redirect_uri):
        _audit_event("oidc_exchange", actor=None, status="failed", details={"error": "redirect_uri_not_allowed"})
        return jsonify({"error": "redirect_uri_not_allowed"}), 400

    if code:
        if client_id and client_id != OIDC_CLIENT_ID:
            _audit_event("oidc_exchange", actor=None, status="failed", details={"error": "client_id_mismatch"})
            return jsonify({"error": "client_id_mismatch"}), 400
        config = _oidc_discovery()
        if not config:
            return jsonify({"error": "oidc_config_missing"}), 500
        token_endpoint = config.get("token_endpoint")
        if not token_endpoint:
            return jsonify({"error": "oidc_token_endpoint_missing"}), 500
        token_payload = {
            "grant_type": "authorization_code",
            "code": code,
            "redirect_uri": redirect_uri or _oidc_redirect_uri(),
            "client_id": OIDC_CLIENT_ID,
        }
        if code_verifier:
            token_payload["code_verifier"] = code_verifier
        auth = None
        if OIDC_CLIENT_SECRET:
            if OIDC_CLIENT_AUTH == "post":
                token_payload["client_secret"] = OIDC_CLIENT_SECRET
            else:
                auth = (OIDC_CLIENT_ID, OIDC_CLIENT_SECRET)
        token_resp = _http_request_with_retry(
            "POST",
            token_endpoint,
            data=token_payload,
            auth=auth,
            timeout=15,
            retries=1,
        )
        if token_resp.status_code >= 400:
            _audit_event("oidc_exchange", actor=None, status="failed", details={"error": "token_exchange_failed"})
            return jsonify({"error": "token_exchange_failed"}), 401
        try:
            token_data = token_resp.json()
        except Exception:
            token_data = {}
        id_token = (token_data.get("id_token") or "").strip()
        access_token = (token_data.get("access_token") or "").strip()

    if not id_token:
        _audit_event("oidc_exchange", actor=None, status="failed", details={"error": "id_token_missing"})
        return jsonify({"error": "id_token_required"}), 400
    try:
        claims = _verify_jwt(id_token, nonce=None)
    except Exception as exc:
        _audit_event("oidc_exchange", actor=None, status="failed", details={"error": str(exc)})
        return jsonify({"error": "invalid_id_token"}), 401
    claims = _oidc_maybe_enrich_claims(claims, access_token, config=_oidc_discovery())
    username = _oidc_username_from_claims(claims)
    if not username:
        _audit_event("oidc_exchange", actor=None, status="failed", details={"error": "username_missing"})
        return jsonify({"error": "username_missing"}), 400
    email = claims.get(OIDC_EMAIL_CLAIM) if isinstance(claims.get(OIDC_EMAIL_CLAIM), str) else None
    role = _oidc_role_from_claims(claims)
    subject = claims.get("sub") if isinstance(claims.get("sub"), str) else None
    user_store.upsert_external_user(username, role=role, email=email, provider="oidc", subject=subject)
    session["user"] = username
    _record_identity_event(
        username,
        role=role,
        email=email,
        provider="oidc",
        subject=subject,
        meta={"source": "api_oidc_exchange"},
    )
    _record_login_event(
        username,
        auth_mode="oidc",
        provider="oidc",
        session_id=subject,
        meta={"source": "api_oidc_exchange"},
    )
    sso_token = _issue_sso_token(username)
    access_token_payload = _issue_access_token_payload(username, source="api_oidc_exchange")
    _audit_event("oidc_exchange", actor=username, status="success", details={"role": role})
    return jsonify(
        {
            "status": "ok",
            "user": username,
            "role": role,
            "sso_token": sso_token,
            "sso_expires_in": SSO_TTL_SECONDS,
            **access_token_payload,
        }
    )

def sso_login() -> Response:
    """Handle the sso login route."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    token = (request.args.get("token") or "").strip()
    next_path = _safe_next_path(request.args.get("next"))
    user = _consume_sso_token(token)
    if not user:
        _audit_event("sso_login", actor=None, status="failed")
        return redirect(url_for("login"))
    session["user"] = user
    _record_identity_event(
        user,
        role=_user_role(user),
        email=user_store.get_email(user),
        provider="sso",
        meta={"source": "sso_login"},
    )
    _record_login_event(user, auth_mode="sso", provider="sso", meta={"source": "sso_login"})
    _audit_event("sso_login", actor=user, status="success")
    return redirect(next_path)


def external_login() -> Response:
    """Bridge Refiner login back to sibling protected services."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    target = _safe_external_redirect(request.args.get("rd"))
    user = _current_user()
    if user:
        _touch_user_activity(user)
        return redirect(target)
    next_path = url_for("external_login", rd=target)
    session["login_next"] = next_path
    if AUTH_MODE == "oidc":
        return redirect(url_for("oidc_login", next=next_path))
    return redirect(url_for("login", next=next_path))


def logout() -> Response:
    """Render or serve the logout route."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    user = _current_user()
    _clear_user_activity(user)
    session.pop("login_next", None)
    session.pop("user", None)
    _audit_event("logout", actor=user, status="success")
    return redirect(url_for("login"))


def api_login() -> Response:
    """API endpoint for login."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    if not user_store.has_users():
        return jsonify({"error": "setup_required"}), 400
    if AUTH_MODE == "oidc":
        return jsonify({"error": "oidc_required"}), 403
    payload = request.get_json(force=True, silent=True)
    if not isinstance(payload, dict):
        payload = {}
    username = (payload.get("username") or "").strip()
    password = (payload.get("password") or "").strip()
    if not username or not password:
        return jsonify({"error": "username_and_password_required"}), 400
    if _login_throttled(username):
        _audit_event("api_login", actor=username, status="throttled")
        return jsonify({"error": "too_many_attempts"}), 429
    if not user_store.verify(username, password):
        _record_login_attempt(username, ok=False)
        _audit_event("api_login", actor=username, status="failed")
        return jsonify({"error": "invalid_credentials"}), 401
    session["user"] = username
    session.pop("login_next", None)
    _record_identity_event(
        username,
        role=user_store.get_role(username),
        email=user_store.get_email(username),
        provider="local",
        meta={"source": "api_login"},
    )
    _record_login_event(username, auth_mode="local", provider="local", meta={"source": "api_login"})
    _record_login_attempt(username, ok=True)
    _audit_event("api_login", actor=username, status="success")
    sso_token = _issue_sso_token(username)
    access_token_payload = _issue_access_token_payload(username, source="api_login")
    return jsonify(
        {
            "status": "ok",
            **_local_identity_payload(username, include_directory=False),
            "sso_token": sso_token,
            "sso_expires_in": SSO_TTL_SECONDS,
            **access_token_payload,
        }
    )


def api_login_mfa_totp() -> Response:
    """API endpoint for TOTP login completion."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    return jsonify({"error": "not_available"}), 404


def api_setup() -> Response:
    """API endpoint for setup."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    if AUTH_MODE == "oidc":
        return jsonify({"error": "oidc_required"}), 403
    if user_store.has_users():
        return jsonify({"error": "setup_not_allowed"}), 409
    payload = request.get_json(force=True, silent=True)
    if not isinstance(payload, dict):
        payload = {}
    username = (payload.get("username") or "").strip()
    password = (payload.get("password") or "").strip()
    confirm = (payload.get("confirm") or "").strip()
    email = (payload.get("email") or "").strip()
    if not username or not password:
        return jsonify({"error": "username_and_password_required"}), 400
    if not USERNAME_RE.match(username):
        return jsonify(
            {
                "error": "invalid_username",
                "details": "Username must be 3-32 chars (letters, numbers, underscore, dash).",
            }
        ), 400
    if len(password) < PASSWORD_MIN_LEN:
        return jsonify(
            {"error": "password_too_short", "details": f"Password must be at least {PASSWORD_MIN_LEN} characters."}
        ), 400
    if confirm and password != confirm:
        return jsonify({"error": "password_mismatch", "details": "Passwords do not match."}), 400
    if email and not EMAIL_RE.match(email):
        return jsonify({"error": "invalid_email", "details": "Enter a valid email address."}), 400
    user_store.create_user(username, password, role="admin", email=email or None)
    session["user"] = username
    session.pop("login_next", None)
    _record_identity_event(
        username,
        role=user_store.get_role(username),
        email=user_store.get_email(username),
        provider="setup",
        meta={"source": "api_setup"},
    )
    _record_login_event(username, auth_mode="setup", provider="local", meta={"source": "api_setup"})
    _audit_event("setup", actor=username, status="success")
    sso_token = _issue_sso_token(username)
    access_token_payload = _issue_access_token_payload(username, source="api_setup")
    return (
        jsonify(
            {
                "status": "ok",
                **_local_identity_payload(username, include_directory=False),
                "sso_token": sso_token,
                "sso_expires_in": SSO_TTL_SECONDS,
                **access_token_payload,
            }
        ),
        201,
    )


def api_register() -> Response:
    """API endpoint for self-service user registration."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    if AUTH_MODE == "oidc":
        return jsonify({"error": "oidc_required"}), 403
    if not SELF_REGISTRATION_ENABLED:
        return jsonify({"error": "registration_not_allowed"}), 403
    if not user_store.has_users():
        return jsonify({"error": "setup_required"}), 400
    payload = request.get_json(force=True, silent=True)
    if not isinstance(payload, dict):
        payload = {}
    username = (payload.get("username") or "").strip()
    password = (payload.get("password") or "").strip()
    confirm = (payload.get("confirm") or "").strip()
    email = (payload.get("email") or "").strip()
    if not email:
        return jsonify({"error": "email_required", "details": "Email is required."}), 400
    if not username or not password:
        return jsonify({"error": "username_and_password_required"}), 400
    if not USERNAME_RE.match(username):
        return jsonify(
            {
                "error": "invalid_username",
                "details": "Username must be 3-32 chars (letters, numbers, underscore, dash).",
            }
        ), 400
    if not EMAIL_RE.match(email):
        return jsonify({"error": "invalid_email", "details": "Enter a valid email address."}), 400
    if len(password) < PASSWORD_MIN_LEN:
        return jsonify(
            {"error": "password_too_short", "details": f"Password must be at least {PASSWORD_MIN_LEN} characters."}
        ), 400
    if confirm and password != confirm:
        return jsonify({"error": "password_mismatch", "details": "Passwords do not match."}), 400
    if user_store.get_user(username):
        return jsonify({"error": "user_exists", "details": "That username is already in use."}), 409
    user_store.create_user(username, password, role="user", email=email)
    session["user"] = username
    session.pop("login_next", None)
    _record_identity_event(
        username,
        role=user_store.get_role(username),
        email=user_store.get_email(username),
        provider="register",
        meta={"source": "api_register"},
    )
    _record_login_event(username, auth_mode="register", provider="local", meta={"source": "api_register"})
    _audit_event("api_register", actor=username, status="success")
    sso_token = _issue_sso_token(username)
    access_token_payload = _issue_access_token_payload(username, source="api_register")
    return (
        jsonify(
            {
                "status": "ok",
                **_local_identity_payload(username, include_directory=False),
                "sso_token": sso_token,
                "sso_expires_in": SSO_TTL_SECONDS,
                **access_token_payload,
            }
        ),
        201,
    )


def api_auth_config() -> Response:
    """Public auth capability metadata for SPA consumers."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    return jsonify(_refiner_auth_config_payload())


def api_sso_issue() -> Response:
    """API endpoint for sso issue."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    sso_token = _issue_sso_token(user)
    return jsonify(
        {
            "status": "ok",
            "token": sso_token,
            "expires_in": SSO_TTL_SECONDS,
            **_local_identity_payload(user, include_directory=False),
        }
    )


def api_logout() -> Response:
    """API endpoint for logout."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    user = _current_user()
    _clear_user_activity(user)
    session.pop("login_next", None)
    session.pop("user", None)
    _audit_event("api_logout", actor=user, status="success")
    return jsonify({"status": "ok"})


def api_session() -> Response:
    """API endpoint for session."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    user = _current_user()
    if not user:
        return jsonify({"authenticated": False, "user": None}), 200
    return jsonify(_local_identity_payload(user, include_directory=False))


def api_authz_nginx() -> Response:
    """Nginx auth_request-compatible session probe."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    user = _current_user()
    if not user:
        response = make_response("", 401)
        response.headers["Cache-Control"] = "no-store"
        return response
    role = getattr(g, "auth_user_role", None) or _user_role(user)
    response = make_response("", 204)
    response.headers["Cache-Control"] = "no-store"
    response.headers["X-Auth-Request-User"] = user
    if role:
        response.headers["X-Auth-Request-Role"] = str(role)
    active_team = getattr(g, "auth_active_team", None)
    if isinstance(active_team, dict) and active_team.get("team_id"):
        response.headers["X-Auth-Request-Team"] = str(active_team.get("team_id") or "")
    return response


def api_profile() -> Response:
    """API endpoint for profile."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if request.method == "GET":
        return jsonify(_local_identity_payload(user, include_directory=True))
    payload = request.get_json(force=True, silent=True) or {}
    email = None
    if "email" in payload:
        email = str(payload.get("email") or "").strip()
        if email and not EMAIL_RE.match(email):
            return jsonify({"error": "invalid_email", "details": "Enter a valid email address."}), 400
    if "settings" in payload:
        try:
            _update_user_settings(user, payload.get("settings"))
        except SettingsValidationError as exc:
            return jsonify({"error": "invalid_settings", "details": exc.issues}), 400
        except KeyError:
            return jsonify({"error": "user_not_found"}), 404
    if "email" in payload:
        user_store.set_email(user, email or None)
    _record_identity_event(
        user,
        role=_user_role(user),
        email=user_store.get_email(user),
        provider="profile",
        meta={"source": "api_profile", "settings_updated": "settings" in payload},
    )
    return jsonify({"status": "ok", **_local_identity_payload(user, include_directory=True)})


def api_profile_password() -> Response:
    """API endpoint for profile password changes."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    return jsonify({"error": "not_available"}), 404


def api_profile_mfa_totp_start() -> Response:
    """API endpoint for starting TOTP enrolment."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    return jsonify({"error": "not_available"}), 404


def api_profile_mfa_totp_verify() -> Response:
    """API endpoint for verifying TOTP enrolment."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    return jsonify({"error": "not_available"}), 404


def api_profile_mfa_totp_disable() -> Response:
    """API endpoint for disabling TOTP."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    return jsonify({"error": "not_available"}), 404


def api_profile_passkeys_register_options() -> Response:
    """API endpoint for starting passkey registration."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    return jsonify({"error": "not_available"}), 404


def api_profile_passkeys_register_verify() -> Response:
    """API endpoint for verifying passkey registration."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    return jsonify({"error": "not_available"}), 404


def api_profile_passkey_delete(credential_id: str) -> Response:
    """API endpoint for deleting a passkey."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    return jsonify({"error": "not_available"}), 404


def api_passkeys_authenticate_options() -> Response:
    """API endpoint for starting passkey authentication."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    return jsonify({"error": "not_available"}), 404


def api_passkeys_authenticate_verify() -> Response:
    """API endpoint for completing passkey authentication."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    return jsonify({"error": "not_available"}), 404


def api_users() -> Response:
    """API endpoint for centralized user management."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    return jsonify({"error": "not_available"}), 404


def api_user_password(username: str) -> Response:
    """API endpoint for centralized user password management."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    return jsonify({"error": "not_available"}), 404


def api_team_invite(team_id: str) -> Response:
    """API endpoint for centralized team invitations."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    return jsonify({"error": "not_available"}), 404


def api_team_invitation_accept(membership_id: str) -> Response:
    """API endpoint for accepting centralized team invitations."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    return jsonify({"error": "not_available"}), 404


def api_team_invitation_reject(membership_id: str) -> Response:
    """API endpoint for rejecting centralized team invitations."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    return jsonify({"error": "not_available"}), 404


def api_team_leave(team_id: str) -> Response:
    """API endpoint for leaving a centralized team."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    return jsonify({"error": "not_available"}), 404


def _todo_meta_from_payload(payload: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """Normalize optional session/job/project linkage metadata for inbox items."""
    meta = payload.get("meta") if isinstance(payload.get("meta"), dict) else {}
    meta = dict(meta or {})
    for key in (
        "session_id",
        "session_ids",
        "room_id",
        "room_ids",
        "job_id",
        "job_ids",
        "project_id",
        "project_ids",
        "team_id",
        "team_ids",
    ):
        value = payload.get(key)
        if value in (None, "", []):
            continue
        meta[key] = value
    return meta or None


def api_todos() -> Response:
    """API endpoint for todos."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if request.method == "GET":
        status_raw = (request.args.get("status") or "").strip()
        statuses = [item.strip().lower() for item in status_raw.split(",") if item.strip()] if status_raw else None
        limit_raw = request.args.get("limit")
        try:
            limit_val = int(limit_raw) if limit_raw else 50
        except Exception:
            limit_val = 50
        limit_val = max(0, min(limit_val, 200))
        query = (request.args.get("query") or request.args.get("q") or "").strip() or None
        ready_only = str(request.args.get("ready") or "").strip().lower() in {"1", "true", "yes", "y"}
        include_routes = str(request.args.get("include_route") or "").strip().lower() in {"1", "true", "yes", "y"}
        items = todo_store.list_items(
            user,
            statuses=statuses,
            limit=None,
            query=query,
            ready_only=ready_only,
            include_routes=include_routes,
        )
        defer_raw = request.args.get("defer")
        if defer_raw is not None:
            want_defer = str(defer_raw).strip().lower() in {"1", "true", "yes", "y"}
            items = [item for item in items if bool(item.get("defer_until_idle")) == want_defer]
        if limit_val >= 0:
            items = items[:limit_val]
        return jsonify({"items": items})

    payload = request.get_json(force=True, silent=True) or {}
    text = str(payload.get("text") or payload.get("thought") or payload.get("todo") or "").strip()
    if not text:
        return jsonify({"error": "text_required"}), 400
    source = str(payload.get("source") or "manual").strip().lower() or "manual"
    device = str(payload.get("device") or "").strip() or None
    defer_until_idle = payload.get("defer_until_idle")
    if defer_until_idle is None:
        defer_until_idle = False
    available_after = _normalise_timestamp(payload.get("available_after")) if payload.get("available_after") else None
    meta = _todo_meta_from_payload(payload)
    item = todo_store.add_item(
        user,
        text,
        source=source,
        device=device,
        meta=meta,
        defer_until_idle=bool(defer_until_idle),
        available_after=available_after,
    )
    _audit_event("todo_create", actor=user, status="success", details={"todo_id": item.get("id"), "source": source})
    return jsonify({"status": "ok", "todo": item}), 201


def api_todo_next() -> Response:
    """API endpoint for todo next."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    idle_only = str(request.args.get("idle") or "").strip().lower() in {"1", "true", "yes", "y"}
    claim_item = str(request.args.get("claim") or "").strip().lower() in {"1", "true", "yes", "y"}
    if idle_only:
        busy_jobs = _busy_jobs_snapshot(exclude_user=user)
        if busy_jobs:
            return jsonify({"status": "busy", "busy_jobs": len(busy_jobs)}), 409
    item = (
        todo_store.claim_next_item(user, idle_only=idle_only)
        if claim_item
        else todo_store.peek_next_item(user, idle_only=idle_only)
    )
    route = item.get("route") if isinstance(item, dict) else None
    return jsonify({"todo": item, "route": route})


def api_todo_detail(todo_id: str) -> Response:
    """API endpoint for todo detail."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if request.method == "DELETE":
        deleted = todo_store.delete_item(user, todo_id)
        if not deleted:
            return jsonify({"error": "not_found"}), 404
        _audit_event("todo_delete", actor=user, status="success", details={"todo_id": todo_id})
        return jsonify({"status": "deleted", "todo_id": todo_id})

    payload = request.get_json(force=True, silent=True) or {}
    updates: Dict[str, Any] = {}
    if "text" in payload or "thought" in payload or "todo" in payload:
        text = str(payload.get("text") or payload.get("thought") or payload.get("todo") or "").strip()
        if not text:
            return jsonify({"error": "text_required"}), 400
        updates["text"] = text
    if "status" in payload:
        status = str(payload.get("status") or "").strip().lower()
        if status not in {"todo", "done", "archived"}:
            return jsonify({"error": "invalid_status"}), 400
        updates["status"] = status
    if "notes" in payload:
        updates["notes"] = payload.get("notes")
    if "priority" in payload:
        updates["priority"] = payload.get("priority")
    if "tags" in payload:
        tags = payload.get("tags")
        if isinstance(tags, str):
            tags = [item.strip() for item in tags.split(",") if item.strip()]
        updates["tags"] = tags
    if "defer_until_idle" in payload:
        updates["defer_until_idle"] = bool(payload.get("defer_until_idle"))
    if "available_after" in payload:
        value = payload.get("available_after")
        updates["available_after"] = _normalise_timestamp(value) if value else None
    updated = todo_store.update_item(user, todo_id, updates)
    if not updated:
        return jsonify({"error": "not_found"}), 404
    _audit_event("todo_update", actor=user, status="success", details={"todo_id": todo_id})
    return jsonify({"status": "ok", "todo": updated})


def api_todo_route(todo_id: str) -> Response:
    """Return the recommended Refiner workflow for a captured thought."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    item = todo_store.get_item(user, todo_id)
    if not item:
        return jsonify({"error": "not_found"}), 404
    route = build_route_suggestion(item)
    return jsonify({"status": "ok", "todo": item, "route": route})


def _coerce_schedule_run_at(payload: Dict[str, Any], item: Optional[Dict[str, Any]] = None) -> Optional[str]:
    run_at_raw = payload.get("run_at") or payload.get("scheduled_for") or payload.get("available_after")
    if run_at_raw:
        normalized = _normalise_timestamp(str(run_at_raw))
        return normalized if _parse_timestamp(normalized) else None
    delay_sec = max(0.0, _safe_float(payload.get("delay_sec"), 0.0))
    if delay_sec > 0:
        run_at = dt.datetime.now(UK_TZ) + dt.timedelta(seconds=delay_sec)
        return run_at.strftime(UK_DATETIME_FORMAT)
    if isinstance(item, dict) and item.get("available_after"):
        normalized = _normalise_timestamp(item.get("available_after"))
        return normalized if _parse_timestamp(normalized) else None
    return _now_iso()


def api_todo_schedule(todo_id: str) -> Response:
    """Create or list deferred execution schedules for a TODO item."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    item = todo_store.get_item(user, todo_id)
    if not item:
        return jsonify({"error": "not_found"}), 404
    if request.method == "GET":
        try:
            limit = int(request.args.get("limit") or 20)
        except Exception:
            limit = 20
        limit = max(1, min(limit, 100))
        schedules = schedule_store.list_items(user=user, todo_id=todo_id, limit=limit)
        return jsonify({"schedules": schedules, "todo": item})

    payload = request.get_json(force=True, silent=True) or {}
    run_at = _coerce_schedule_run_at(payload, item)
    if not run_at:
        return jsonify({"error": "invalid_run_at", "details": "Provide run_at, scheduled_for, available_after, or delay_sec."}), 400
    route_override = payload.get("route_override")
    if not isinstance(route_override, dict):
        route_override = payload.get("route") if isinstance(payload.get("route"), dict) else None
    schedule = schedule_store.create(
        user=user,
        todo_id=todo_id,
        run_at=run_at,
        route_override=route_override if isinstance(route_override, dict) else None,
    )
    todo_store.append_link(user, todo_id, "schedules", schedule.get("id") or "")
    _audit_event("todo_schedule_create", actor=user, status="success", details={"schedule_id": schedule.get("id"), "todo_id": todo_id})
    return jsonify({"status": "scheduled", "schedule": schedule, "todo": todo_store.get_item(user, todo_id)}), 201


def api_schedules() -> Response:
    """List the current user's TODO schedules."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    status_raw = str(request.args.get("status") or "").strip()
    statuses = [item.strip().lower() for item in status_raw.split(",") if item.strip()] if status_raw else None
    todo_id = str(request.args.get("todo_id") or "").strip() or None
    try:
        limit = int(request.args.get("limit") or 50)
    except Exception:
        limit = 50
    limit = max(1, min(limit, 200))
    schedules = schedule_store.list_items(user=user, todo_id=todo_id, statuses=statuses, limit=limit)
    return jsonify({"schedules": schedules, "scheduler": todo_scheduler.status()})


def api_schedule_detail(schedule_id: str) -> Response:
    """Return details for one deferred TODO schedule."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    schedule = schedule_store.get_item(schedule_id, user=user)
    if not schedule:
        return jsonify({"error": "not_found"}), 404
    return jsonify({"schedule": schedule})


def api_schedule_cancel(schedule_id: str) -> Response:
    """Cancel a deferred TODO schedule."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    schedule = schedule_store.get_item(schedule_id, user=user)
    if not schedule:
        return jsonify({"error": "not_found"}), 404
    status = str(schedule.get("status") or "").strip().lower()
    if status in {"completed", "failed"}:
        return jsonify({"error": "schedule_not_cancellable"}), 409
    subtask_id = str(schedule.get("subtask_id") or "").strip()
    if subtask_id:
        subtask_manager.cancel(subtask_id, owner=user)
    updated = schedule_store.cancel(schedule_id, reason="cancelled_by_user")
    todo_id = str(schedule.get("todo_id") or "").strip()
    todo = todo_store.get_item(user, todo_id)
    if todo and str(todo.get("status") or "todo").strip().lower() == "todo":
        todo_store.release_execution(user, todo_id, execution_state="ready", error="cancelled_by_user")
    _audit_event("todo_schedule_cancel", actor=user, status="success", details={"schedule_id": schedule_id, "todo_id": todo_id})
    return jsonify({"schedule": updated or schedule_store.get_item(schedule_id, user=user)})


def api_subtasks() -> Response:
    """Create or list generic background subtasks."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if request.method == "GET":
        try:
            limit = int(request.args.get("limit") or 50)
        except Exception:
            limit = 50
        limit = max(1, min(limit, 200))
        include_results = _is_truthy(request.args.get("include_results"))
        scope_type = str(request.args.get("scope_type") or "").strip().lower()
        scope_id = str(request.args.get("scope_id") or "").strip()
        if scope_type and scope_id:
            tasks = subtask_manager.list_for_scope(scope_type, scope_id, limit=limit, include_results=include_results)
            if not _is_admin_user(user):
                tasks = [task for task in tasks if task.get("owner") == user]
        else:
            tasks = subtask_manager.list_for_owner(user, limit=limit, include_results=include_results)
        return jsonify({"tasks": tasks, "queue_depth": subtask_manager.queue_depth(), "inflight": subtask_manager.inflight()})

    payload = request.get_json(force=True, silent=True) or {}
    action = str(payload.get("action") or "").strip()
    if action not in {"todo_execute", "assistant_requirements", "playground_plan", "execution_plan", "submit_job"}:
        return jsonify({"error": "invalid_action"}), 400
    subtask_payload = payload.get("payload") if isinstance(payload.get("payload"), dict) else None
    if subtask_payload is None:
        subtask_payload = {
            key: value
            for key, value in payload.items()
            if key not in {"action", "scope_type", "scope_id", "timeout_sec"}
        }
    scope_type = str(payload.get("scope_type") or "user").strip().lower() or "user"
    scope_id = str(payload.get("scope_id") or "").strip() or None
    timeout_sec = max(1.0, min(_safe_float(payload.get("timeout_sec"), SUBTASK_TIMEOUT_SEC), 600.0))
    try:
        task = subtask_manager.submit(
            owner=user,
            action=action,
            payload=subtask_payload,
            scope_type=scope_type,
            scope_id=scope_id,
            timeout_sec=timeout_sec,
        )
    except queue.Full:
        return jsonify({"error": "subtask_capacity_unavailable"}), 503
    _audit_event("subtask_create", actor=user, status="success", details={"task_id": task.task_id, "action": action, "scope_type": scope_type, "scope_id": scope_id})
    return jsonify({"task": task.to_dict(include_result=False)}), 202


def api_subtask_detail(task_id: str) -> Response:
    """Return one generic background subtask."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    task = subtask_manager.get_task(task_id)
    if not task or (task.owner != user and not _is_admin_user(user)):
        return jsonify({"error": "task_not_found"}), 404
    include_result = _is_truthy(request.args.get("include_result")) or _is_admin_user(user)
    return jsonify({"task": task.to_dict(include_result=include_result)})


def api_subtask_cancel(task_id: str) -> Response:
    """Cancel one generic background subtask."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    task = subtask_manager.get_task(task_id)
    if not task or (task.owner != user and not _is_admin_user(user)):
        return jsonify({"error": "task_not_found"}), 404
    if task.is_terminal() and task.status != "cancelled":
        return jsonify({"error": "task_not_cancellable"}), 409
    if not subtask_manager.cancel(task_id, owner=task.owner):
        return jsonify({"error": "task_not_cancellable"}), 409
    _audit_event("subtask_cancel", actor=user, status="success", details={"task_id": task_id, "action": task.action})
    refreshed = subtask_manager.get_task(task_id)
    return jsonify({"task": refreshed.to_dict(include_result=True) if refreshed else {"task_id": task_id, "status": "cancelled"}})


def api_voice_tokens() -> Response:
    """API endpoint for voice tokens."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if request.method == "GET":
        target = (request.args.get("user") or "").strip() or None
        if target and not _is_admin_user(user):
            return jsonify({"error": "forbidden"}), 403
        if not target and not _is_admin_user(user):
            target = user
        tokens = voice_token_store.list_tokens(target)
        return jsonify({"tokens": tokens})

    payload = request.get_json(force=True, silent=True) or {}
    target = str(payload.get("user") or user).strip()
    if target != user and not _is_admin_user(user):
        return jsonify({"error": "forbidden"}), 403
    if user_store.has_users() and not user_store.get_role(target):
        return jsonify({"error": "user_not_found"}), 404
    label = payload.get("label")
    issued = voice_token_store.issue(target, label=label if isinstance(label, str) else None)
    _audit_event(
        "voice_token_issue",
        actor=user,
        status="success",
        details={"token_id": issued.get("id"), "target": target},
    )
    return (
        jsonify(
            {
                "status": "ok",
                "token": issued.get("token"),
                "id": issued.get("id"),
                "user": issued.get("user"),
                "label": issued.get("label"),
                "created_at": issued.get("created_at"),
            }
        ),
        201,
    )


def api_voice_token_delete(token_id: str) -> Response:
    """API endpoint for voice token delete."""
    if _customers_enabled():
        return _proxy_service_request(
            CUSTOMERS_API_BASE,
            CUSTOMERS_TIMEOUT,
            path=f"/api/voice/tokens/{quote(token_id)}",
        )
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if not _is_admin_user(user):
        allowed_ids = {entry.get("id") for entry in voice_token_store.list_tokens(user)}
        if token_id not in allowed_ids:
            return jsonify({"error": "forbidden"}), 403
    revoked = voice_token_store.revoke(token_id)
    if not revoked:
        return jsonify({"error": "not_found"}), 404
    _audit_event("voice_token_revoke", actor=user, status="success", details={"token_id": token_id})
    return jsonify({"status": "revoked", "id": token_id})


def api_voice_capture() -> Response:
    """API endpoint for voice capture."""
    payload = request.get_json(force=False, silent=True) or {}
    user = _voice_user_from_request(payload)
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    text = _extract_voice_text(payload)
    if not text:
        return jsonify({"error": "text_required"}), 400
    source = str(payload.get("source") or request.args.get("source") or "voice").strip().lower() or "voice"
    device = _extract_voice_device(payload)
    meta = {}
    locale = payload.get("locale")
    if isinstance(locale, str) and locale.strip():
        meta["locale"] = locale.strip()
    item = todo_store.add_item(
        user,
        text,
        source=source,
        device=device,
        meta=meta or None,
        defer_until_idle=True,
    )
    _audit_event(
        "voice_capture",
        actor=user,
        status="success",
        details={"todo_id": item.get("id"), "source": source, "device": device},
    )
    return jsonify({"status": "ok", "todo": item})


def api_voice_siri() -> Response:
    """API endpoint for voice siri."""
    payload = request.get_json(force=False, silent=True) or {}
    user = _voice_user_from_request(payload)
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    text = _extract_voice_text(payload)
    if not text:
        return jsonify({"error": "text_required"}), 400
    device = _extract_voice_device(payload)
    meta = {}
    shortcut = payload.get("shortcut") or request.values.get("shortcut")
    if isinstance(shortcut, str) and shortcut.strip():
        meta["shortcut"] = shortcut.strip()
    item = todo_store.add_item(
        user,
        text,
        source="siri",
        device=device,
        meta=meta or None,
        defer_until_idle=True,
    )
    _audit_event(
        "voice_capture",
        actor=user,
        status="success",
        details={"todo_id": item.get("id"), "source": "siri", "device": device},
    )
    message = f"Captured: {item.get('text')}"
    wants_plain = "text/plain" in (request.headers.get("Accept") or "")
    wants_plain = wants_plain or (request.args.get("format") or "").strip().lower() in {"text", "plain"}
    if wants_plain:
        return Response(message, mimetype="text/plain")
    return jsonify({"status": "ok", "message": message, "todo": item})


def api_voice_alexa() -> Response:
    """API endpoint for voice alexa."""
    payload = request.get_json(force=False, silent=True) or {}
    if VOICE_VERIFY_ALEXA:
        ok, reason = _alexa_verify_request(payload, request.get_data(cache=True))
        if not ok:
            logger.warning("Alexa verification failed: %s", reason)
            return _alexa_response("Request validation failed.", end_session=True), 401
        user = _voice_user_from_provider("alexa", payload, allow_tokens=VOICE_ALLOW_TOKENS_WITH_SIGNATURE)
    else:
        user = _voice_user_from_request(payload) if VOICE_ALLOW_TOKENS else None
    if not user:
        return _alexa_response("Please link your Refiner account to use this skill.", end_session=True), 401
    text = _extract_alexa_text(payload)
    if not text:
        return _alexa_response("Sorry, I didn't catch that. What should I capture?", end_session=False)
    request_payload = payload.get("request") if isinstance(payload.get("request"), dict) else {}
    intent = request_payload.get("intent") if isinstance(request_payload.get("intent"), dict) else {}
    meta = {
        "intent": intent.get("name"),
        "locale": request_payload.get("locale"),
    }
    item = todo_store.add_item(
        user,
        text,
        source="alexa",
        device="alexa",
        meta=meta,
        defer_until_idle=True,
    )
    _audit_event(
        "voice_capture",
        actor=user,
        status="success",
        details={"todo_id": item.get("id"), "source": "alexa"},
    )
    return _alexa_response("Captured. I'll save that for later.", end_session=True)


def api_voice_google() -> Response:
    """API endpoint for voice google."""
    payload = request.get_json(force=False, silent=True) or {}
    if VOICE_VERIFY_GOOGLE:
        ok, reason = _google_verify_request()
        if not ok:
            logger.warning("Google verification failed: %s", reason)
            return _google_response("Request validation failed."), 401
        user = _voice_user_from_provider("google", payload, allow_tokens=VOICE_ALLOW_TOKENS_WITH_SIGNATURE)
    else:
        user = _voice_user_from_request(payload) if VOICE_ALLOW_TOKENS else None
    if not user:
        return _google_response("Please link your Refiner account to use this action."), 401
    text = _extract_google_text(payload)
    if not text:
        return _google_response("Sorry, I didn't catch that. What should I capture?")
    query_result = payload.get("queryResult") if isinstance(payload.get("queryResult"), dict) else {}
    intent = query_result.get("intent") if isinstance(query_result.get("intent"), dict) else {}
    meta = {
        "intent": intent.get("displayName") or intent.get("name"),
        "language": query_result.get("languageCode"),
    }
    item = todo_store.add_item(
        user,
        text,
        source="google",
        device="google_home",
        meta=meta,
        defer_until_idle=True,
    )
    _audit_event(
        "voice_capture",
        actor=user,
        status="success",
        details={"todo_id": item.get("id"), "source": "google"},
    )
    return _google_response("Captured. I'll save that for later.")


def api_voice_stt() -> Response:
    """API endpoint for voice stt."""
    request_started = time.perf_counter()
    payload = request.get_json(force=False, silent=True) or {}
    if not _stt_authorized(payload):
        return jsonify({"error": "unauthorized"}), 401
    voice_user = _voice_user_from_request(payload)
    actor = _current_user() or voice_user or "voice"
    gesture_mode, avatar_mode, office_flag = _stt_motion_context(payload)
    collaboration_mode = _stt_collaboration_mode(payload)
    data, ext, error = _extract_audio_bytes(payload)
    if error:
        return jsonify({"error": error}), 400
    if data is None:
        return jsonify({"error": "audio_required"}), 400
    if STT_MAX_BYTES and len(data) > STT_MAX_BYTES:
        return jsonify({"error": "audio_too_large"}), 413
    lang = _sanitize_lang(request.form.get("lang") or request.args.get("lang") or payload.get("lang"))
    prompt_hint = _stt_prompt_hint(payload)
    capacity_acquired = _acquire_request_capacity(_STT_REQUEST_CAPACITY, STT_CAPACITY_WAIT_SEC, owner=actor)
    if not capacity_acquired:
        return jsonify({"error": "stt_capacity_unavailable"}), 503
    audio_path = None
    processed_path = None
    server_motion_payload: Optional[Dict[str, Any]] = None
    direct_server_mode = (
        STT_BACKEND == "server"
        and bool(STT_SERVER_URL)
        and not (STT_SERVER_PREPROCESS and STT_PREPROCESS_COMMAND)
    )
    timings_ms: Dict[str, int] = {"preprocess": 0, "stt": 0, "planner": 0}
    stt_transport = "server_direct" if direct_server_mode else "command"
    if STT_BACKEND == "server" and STT_SERVER_URL and not direct_server_mode:
        stt_transport = "server_file"
    try:
        if direct_server_mode:
            stt_started = time.perf_counter()
            transcript, stt_error, server_motion_payload = _run_stt_server_bytes(
                data,
                ext,
                lang,
                prompt_hint=prompt_hint,
                gesture_mode=gesture_mode,
                avatar_mode=avatar_mode,
                office_mode=office_flag,
                collaboration_mode=collaboration_mode,
            )
            timings_ms["stt"] = int((time.perf_counter() - stt_started) * 1000)
        else:
            preprocess_started = time.perf_counter()
            audio_path = _write_audio_temp(data, ext)
            processed_path, preprocess_error = _run_preprocess(audio_path)
            timings_ms["preprocess"] = int((time.perf_counter() - preprocess_started) * 1000)
            if preprocess_error:
                return jsonify({"error": preprocess_error}), 500
            stt_started = time.perf_counter()
            transcript, stt_error, server_motion_payload = _run_stt(
                processed_path,
                lang,
                prompt_hint=prompt_hint,
                gesture_mode=gesture_mode,
                avatar_mode=avatar_mode,
                office_mode=office_flag,
                collaboration_mode=collaboration_mode,
            )
            timings_ms["stt"] = int((time.perf_counter() - stt_started) * 1000)
        if stt_error or not transcript:
            # Passive-listening mode: treat no-speech / invalid-audio outcomes as non-fatal.
            benign_errors = {"stt_output_empty", "invalid_audio", "unsupported_format"}
            error_code = (stt_error or "").strip().lower()
            if error_code in benign_errors:
                return jsonify({"status": "ok", "text": "", "lang": lang, "reason": error_code}), 200
            return jsonify({"error": stt_error or "stt_failed"}), 500
    finally:
        for path in {audio_path, processed_path}:
            if path and os.path.exists(path):
                try:
                    os.remove(path)
                except Exception:
                    pass
        _release_request_capacity(_STT_REQUEST_CAPACITY, owner=actor)
    _stt_record_learning(transcript, source="voice_stt")
    response_payload: Dict[str, Any] = {
        "status": "ok",
        "text": transcript,
        "lang": lang,
        "gesture_mode": gesture_mode,
        "avatar_mode": avatar_mode,
        "collaboration_mode": bool(collaboration_mode),
    }
    planner_used = "disabled" if not STT_GESTURE_ENABLED else "none"
    has_server_clip = (
        isinstance(server_motion_payload, dict)
        and isinstance(server_motion_payload.get("avatar_motion"), (dict, list))
    )
    if STT_GESTURE_ENABLED and has_server_clip and STT_GESTURE_PREFER_SERVER:
        response_payload.update(server_motion_payload or {})
        response_payload["gesture_mode"] = sanitize_gesture_mode(
            response_payload.get("gesture_mode"),
            default_mode=gesture_mode,
            bsl_enabled=STT_BSL_ENABLED,
        )
        response_payload["avatar_mode"] = sanitize_avatar_mode(
            response_payload.get("avatar_mode"),
            office_mode=office_flag,
            default_mode=avatar_mode,
        )
        planner_used = "nmstt_server"

    if STT_GESTURE_ENABLED and planner_used != "nmstt_server":
        planner_started = time.perf_counter()
        try:
            nmstt_motion_payload = None
            if STT_GESTURE_NMSTT_FALLBACK and STT_BACKEND == "server" and STT_SERVER_URL:
                nmstt_motion_payload = _run_nmstt_gesture_plan(
                    transcript,
                    gesture_mode=gesture_mode,
                    avatar_mode=avatar_mode,
                    office_mode=office_flag,
                )
            if isinstance(nmstt_motion_payload, dict):
                response_payload.update(nmstt_motion_payload)
                response_payload["gesture_mode"] = sanitize_gesture_mode(
                    response_payload.get("gesture_mode"),
                    default_mode=gesture_mode,
                    bsl_enabled=STT_BSL_ENABLED,
                )
                response_payload["avatar_mode"] = sanitize_avatar_mode(
                    response_payload.get("avatar_mode"),
                    office_mode=office_flag,
                    default_mode=avatar_mode,
                )
                planner_used = "nmstt_gesture_plan"
            else:
                motion_payload = plan_stt_avatar_motion(
                    transcript,
                    gesture_mode=gesture_mode,
                    avatar_mode=avatar_mode,
                    bsl_enabled=STT_BSL_ENABLED,
                )
                if isinstance(motion_payload, dict):
                    response_payload.update(motion_payload)
                    planner_used = "python_fallback"
        except Exception as exc:
            logger.debug("STT gesture planning skipped: %s", exc)
            if planner_used == "none":
                planner_used = "error"
        finally:
            timings_ms["planner"] = int((time.perf_counter() - planner_started) * 1000)
    timings_ms["total"] = int((time.perf_counter() - request_started) * 1000)
    _audit_event(
        "voice_stt",
        actor=actor,
        status="success",
        details={
            "bytes": len(data),
            "lang": lang,
            "gesture_mode": response_payload.get("gesture_mode"),
            "avatar_mode": response_payload.get("avatar_mode"),
            "collaboration_mode": response_payload.get("collaboration_mode"),
            "gesture_planner": planner_used,
            "transport": stt_transport,
            "timings_ms": timings_ms,
        },
    )
    return jsonify(response_payload)


def _parse_user_list_payload(value: Any) -> List[str]:
    if value is None:
        return []
    if isinstance(value, str):
        return [item.strip() for item in value.split(",") if item.strip()]
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    return []


def _diff_list(old: Optional[List[str]], new: Optional[List[str]]) -> Dict[str, List[str]]:
    old_set = {str(item).strip() for item in (old or []) if str(item).strip()}
    new_set = {str(item).strip() for item in (new or []) if str(item).strip()}
    added = sorted(new_set - old_set)
    removed = sorted(old_set - new_set)
    return {"added": added, "removed": removed}


def _diff_permissions(old: Optional[Dict[str, Any]], new: Optional[Dict[str, Any]]) -> Dict[str, Dict[str, List[str]]]:
    old_perm = AccessStore._normalise_permissions(old)
    new_perm = AccessStore._normalise_permissions(new)
    changes: Dict[str, Dict[str, List[str]]] = {}
    for key in ("read", "write", "grant"):
        diff = _diff_list(old_perm.get(key), new_perm.get(key))
        if diff["added"] or diff["removed"]:
            changes[key] = diff
    return changes


def api_projects() -> Response:
    """API endpoint for projects."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if request.method == "GET":
        include_viewers = request.args.get("include_viewers", "1").strip().lower() in {"1", "true", "yes"}
        if _is_admin_user(user):
            projects = []
            for project in access_store.list_projects():
                entry = dict(project)
                entry["role"] = "admin"
                entry["capabilities"] = {"read": True, "write": True, "grant": True}
                team = _team_display_record(project.get("team_id")) if project.get("team_id") else None
                if team:
                    entry["team_name"] = team.get("name")
                projects.append(entry)
        else:
            projects = access_store.projects_for_user(user, include_viewers=include_viewers)
        return jsonify({"projects": projects})
    payload = request.get_json(force=True, silent=True) or {}
    name = str(payload.get("name") or "").strip()
    if not name:
        return jsonify({"error": "invalid_name"}), 400
    team_id = payload.get("team_id") or None
    leaders = _parse_user_list_payload(payload.get("leaders"))
    contributors = _parse_user_list_payload(payload.get("contributors"))
    viewers = _parse_user_list_payload(payload.get("viewers"))
    permissions = payload.get("permissions")
    if permissions is None:
        permissions = {
            "read": payload.get("permissions_read"),
            "write": payload.get("permissions_write"),
            "grant": payload.get("permissions_grant"),
        }
        if not any(permissions.values()):
            permissions = None
    if not _is_admin_user(user):
        if not team_id:
            return jsonify({"error": "team_required", "details": "Team ID required for non-admin creation."}), 400
        if not access_store.can_create_project(user, team_id):
            return jsonify({"error": "forbidden", "details": "Team grant required to create projects."}), 403
        if user not in leaders:
            leaders.append(user)
    try:
        project = access_store.create_project(
            name,
            team_id=team_id,
            leaders=leaders,
            contributors=contributors,
            viewers=viewers,
            permissions=permissions,
        )
    except ValueError as exc:
        return jsonify({"error": "invalid_team", "details": str(exc)}), 400
    _audit_event(
        "project_create",
        actor=user,
        status="success",
        details={
            "project_id": project.get("id"),
            "team_id": project.get("team_id"),
            "leaders": project.get("leaders"),
            "contributors": project.get("contributors"),
            "viewers": project.get("viewers"),
            "permissions": project.get("permissions"),
        },
    )
    return jsonify(project), 201


def api_project_detail(project_id: str) -> Response:
    """API endpoint for project detail."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if request.method == "DELETE":
        if not _is_admin_user(user):
            return jsonify({"error": "forbidden"}), 403
        if not access_store.delete_project(project_id):
            return jsonify({"error": "delete_failed"}), 409
        _audit_event("project_delete", actor=user, status="success", details={"project_id": project_id})
        return jsonify({"status": "deleted", "project_id": project_id})
    if not (_is_admin_user(user) or access_store.can_manage_project(user, project_id)):
        return jsonify({"error": "forbidden"}), 403
    before = dict(access_store.get_project(project_id) or {})
    payload = request.get_json(force=True, silent=True) or {}
    name = payload.get("name")
    team_id = payload.get("team_id")
    leaders = payload.get("leaders")
    contributors = payload.get("contributors")
    viewers = payload.get("viewers")
    permissions = payload.get("permissions")
    if permissions is None:
        permissions = {
            "read": payload.get("permissions_read"),
            "write": payload.get("permissions_write"),
            "grant": payload.get("permissions_grant"),
        }
        if not any(permissions.values()):
            permissions = None
    try:
        project = access_store.update_project(
            project_id,
            name=name,
            team_id=team_id,
            leaders=_parse_user_list_payload(leaders) if leaders is not None else None,
            contributors=_parse_user_list_payload(contributors) if contributors is not None else None,
            viewers=_parse_user_list_payload(viewers) if viewers is not None else None,
            permissions=permissions,
        )
    except ValueError as exc:
        return jsonify({"error": "invalid_team", "details": str(exc)}), 400
    if not project:
        return jsonify({"error": "not_found"}), 404
    _audit_event("project_update", actor=user, status="success", details={"project_id": project_id})
    if before:
        changes: Dict[str, Any] = {}
        leader_diff = _diff_list(before.get("leaders"), project.get("leaders"))
        if leader_diff["added"] or leader_diff["removed"]:
            changes["leaders"] = leader_diff
        contributor_diff = _diff_list(before.get("contributors"), project.get("contributors"))
        if contributor_diff["added"] or contributor_diff["removed"]:
            changes["contributors"] = contributor_diff
        viewer_diff = _diff_list(before.get("viewers"), project.get("viewers"))
        if viewer_diff["added"] or viewer_diff["removed"]:
            changes["viewers"] = viewer_diff
        perm_diff = _diff_permissions(before.get("permissions"), project.get("permissions"))
        if perm_diff:
            changes["permissions"] = perm_diff
        if changes:
            _audit_event(
                "project_permissions_change",
                actor=user,
                status="success",
                details={"project_id": project_id, "changes": changes},
            )
    return jsonify(project)


def api_teams() -> Response:
    """API endpoint for teams."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if request.method == "GET":
        if not _is_admin_user(user):
            return jsonify({"error": "forbidden"}), 403
        return jsonify({"teams": access_store.list_teams()})
    if not _is_admin_user(user):
        return jsonify({"error": "forbidden"}), 403
    payload = request.get_json(force=True, silent=True) or {}
    name = str(payload.get("name") or "").strip()
    if not name:
        return jsonify({"error": "invalid_name"}), 400
    parent_id = payload.get("parent_id") or None
    leaders = _parse_user_list_payload(payload.get("leaders"))
    members = _parse_user_list_payload(payload.get("members"))
    team = access_store.create_team(name, parent_id=parent_id, leaders=leaders, members=members)
    _audit_event(
        "team_create",
        actor=user,
        status="success",
        details={
            "team_id": team.get("id"),
            "parent_id": team.get("parent_id"),
            "leaders": team.get("leaders"),
            "members": team.get("members"),
        },
    )
    return jsonify(team), 201


def api_team_detail(team_id: str) -> Response:
    """API endpoint for team detail."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if not _is_admin_user(user):
        return jsonify({"error": "forbidden"}), 403
    if request.method == "DELETE":
        if not access_store.delete_team(team_id):
            return jsonify({"error": "delete_failed"}), 409
        _audit_event("team_delete", actor=user, status="success", details={"team_id": team_id})
        return jsonify({"status": "deleted", "team_id": team_id})
    before = dict(access_store.get_team(team_id) or {})
    payload = request.get_json(force=True, silent=True) or {}
    name = payload.get("name")
    parent_id = payload.get("parent_id")
    leaders = payload.get("leaders")
    members = payload.get("members")
    team = access_store.update_team(
        team_id,
        name=name,
        parent_id=parent_id,
        leaders=_parse_user_list_payload(leaders) if leaders is not None else None,
        members=_parse_user_list_payload(members) if members is not None else None,
    )
    if not team:
        return jsonify({"error": "not_found"}), 404
    _audit_event("team_update", actor=user, status="success", details={"team_id": team_id})
    if before:
        changes: Dict[str, Any] = {}
        leader_diff = _diff_list(before.get("leaders"), team.get("leaders"))
        if leader_diff["added"] or leader_diff["removed"]:
            changes["leaders"] = leader_diff
        member_diff = _diff_list(before.get("members"), team.get("members"))
        if member_diff["added"] or member_diff["removed"]:
            changes["members"] = member_diff
        if changes:
            _audit_event(
                "team_membership_change",
                actor=user,
                status="success",
                details={"team_id": team_id, "changes": changes},
            )
    return jsonify(team)


def api_team_tokens(team_id: str) -> Response:
    """API endpoint for team tokens."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    team = access_store.get_team(team_id)
    if not team:
        return jsonify({"error": "team_not_found"}), 404
    if request.method == "GET":
        if not _can_access_team_tokens(user, team_id):
            return jsonify({"error": "forbidden"}), 403
        snapshot = _team_token_snapshot(team_id)
        return jsonify({"team_id": team_id, "team_name": team.get("name"), **snapshot})
    if not _is_admin_user(user):
        return jsonify({"error": "forbidden"}), 403
    payload = request.get_json(force=True, silent=True) or {}
    action = (payload.get("action") or "grant").strip().lower()
    tokens_raw = payload.get("token_amount")
    tokens = 0
    if tokens_raw not in (None, ""):
        try:
            tokens = int(float(tokens_raw))
        except Exception:
            tokens = 0
    if action in {"add", "grant"}:
        if tokens <= 0:
            return jsonify({"error": "invalid_amount", "details": "Token amount must be positive."}), 400
        meta = {
            "tokens": tokens,
            "team_id": team_id,
            "team_name": team.get("name"),
            "source": payload.get("source") or "admin",
            "granted_by": user,
        }
        entry_type = "topup" if action == "add" else "grant"
        _record_token_event(
            "team",
            team_id,
            entry_type,
            tokens,
            meta,
            request_id=_chain_request_id("team-token", team_id, entry_type, uuid.uuid4().hex),
        )
        _audit_event(
            "team_tokens_topup" if action == "add" else "team_tokens_grant",
            actor=user,
            status="success",
            details={"team_id": team_id, "amount": tokens},
        )
        snapshot = _team_token_snapshot(team_id)
        return jsonify({"message": "Team tokens updated.", "team_id": team_id, **snapshot})
    if action == "sync":
        target = payload.get("balance")
        if target is None:
            return jsonify({"error": "balance_required"}), 400
        try:
            target_balance = int(float(target))
        except Exception:
            return jsonify({"error": "invalid_balance"}), 400
        if target_balance < 0:
            return jsonify({"error": "invalid_balance"}), 400
        capacity = payload.get("capacity") or payload.get("last_topup_tokens")
        try:
            capacity_val = int(float(capacity)) if capacity not in (None, "") else None
        except Exception:
            capacity_val = None
        target_paid = payload.get("paid_balance")
        target_free = payload.get("free_balance")
        try:
            target_paid_val = int(float(target_paid)) if target_paid not in (None, "") else None
        except Exception:
            target_paid_val = None
        try:
            target_free_val = int(float(target_free)) if target_free not in (None, "") else None
        except Exception:
            target_free_val = None
        snapshot = _team_token_snapshot(team_id)
        delta = target_balance - snapshot["balance"]
        status = "matched" if delta == 0 else "adjusted"
        _record_token_event(
            "team",
            team_id,
            "sync",
            delta,
            {
                "target_balance": target_balance,
                "capacity": capacity_val,
                "source": payload.get("source") or "admin",
                "sync_user": payload.get("user") or user,
                "sync_role": payload.get("role"),
                "target_paid_balance": target_paid_val,
                "target_free_balance": target_free_val,
            },
            request_id=None,
        )
        _audit_event("team_tokens_sync", actor=user, status=status, details={"team_id": team_id, "delta": delta})
        snapshot = _team_token_snapshot(team_id)
        return jsonify({"message": "Sync complete.", "status": status, "team_id": team_id, **snapshot})
    return jsonify({"error": "invalid_action"}), 400


def api_access_tree() -> Response:
    """API endpoint for access tree."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if _customers_enabled():
        return jsonify({"tree": _customer_access_tree()})
    if _is_admin_user(user):
        tree = access_store.tree_all()
    else:
        tree = access_store.tree_for_user(user)
    return jsonify({"tree": tree})


def _sse_event(event: str, entry: Dict[str, Any]) -> str:
    payload = json.dumps(entry)
    return f"event: {event}\ndata: {payload}\n\n"


def api_sessions() -> Response:
    """API endpoint for sessions."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    payload = request.get_json(force=True, silent=True) or {}
    job_id = str(payload.get("job_id") or "").strip()
    room_id = str(payload.get("room_id") or "").strip() or None
    if not job_id:
        return jsonify({"error": "job_id_required"}), 400
    job = manager.get_job(job_id)
    if not job or not _can_view_job(user, job):
        return jsonify({"error": "job_not_found"}), 404
    project_id = _job_project_id(job)
    role = _job_role_for_user(user, job)
    session_obj = session_store.get_or_create(job_id, project_id, user, role, room_id=room_id)
    snapshot = session_obj.snapshot()
    snapshot["project_role"] = role
    return jsonify(snapshot)


def api_session_detail(session_id: str) -> Response:
    """API endpoint for session detail."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    session_obj = session_store.get(session_id)
    if not session_obj:
        return jsonify({"error": "not_found"}), 404
    job = manager.get_job(session_obj.job_id)
    if not job or not _can_view_job(user, job):
        return jsonify({"error": "forbidden"}), 403
    snapshot = session_obj.snapshot()
    snapshot["project_role"] = _job_role_for_user(user, job)
    history = session_history.load(session_obj.room_id)
    if history and isinstance(history.get("events"), list):
        snapshot["history_count"] = len(history.get("events"))
    return jsonify(snapshot)


def api_session_leave(session_id: str) -> Response:
    """API endpoint for session leave."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    session_obj = session_store.get(session_id)
    if not session_obj:
        return jsonify({"error": "not_found"}), 404
    session_store.leave(session_id, user)
    return jsonify({"status": "ok"})


def api_session_stream(session_id: str) -> Response:
    """API endpoint for session stream."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    session_obj = session_store.get(session_id)
    if not session_obj:
        return jsonify({"error": "not_found"}), 404
    job = manager.get_job(session_obj.job_id)
    if not job or not _can_view_job(user, job):
        return jsonify({"error": "forbidden"}), 403
    role = _job_role_for_user(user, job)
    session_obj.join(user, role)

    def generate():
        q = session_obj.add_listener()
        last_status = None
        last_progress = None
        last_status_event = None
        try:
            while True:
                try:
                    entry = q.get(timeout=1.0)
                except queue.Empty:
                    # periodic job state updates
                    status = job.status
                    progress = job.progress
                    if status != last_status or progress != last_progress:
                        last_status = status
                        last_progress = progress
                        if status != last_status_event:
                            last_status_event = status
                            session_obj._record_event(
                                "job_status",
                                user=None,
                                detail={"status": status, "progress": progress},
                            )
                        payload = {
                            "job_id": job.job_id,
                            "status": status,
                            "progress": progress,
                            "updated_at": job.updated_at,
                        }
                        yield _sse_event("job", payload)
                    yield ": keep-alive\n\n"
                    continue
                event_type = entry.get("event") or "presence"
                payload = entry.get("payload") or {}
                yield _sse_event(event_type, payload)
        finally:
            session_obj.remove_listener(q)
            session_store.leave(session_id, user)

    return Response(generate(), mimetype="text/event-stream", headers={"Cache-Control": "no-cache"})


def api_session_history(session_id: str) -> Response:
    """API endpoint for session history."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    session_obj = session_store.get(session_id)
    job = manager.get_job(session_obj.job_id) if session_obj else None
    if session_obj and job:
        if not _can_view_job(user, job):
            return jsonify({"error": "forbidden"}), 403
    else:
        history = session_history.load(session_id)
        if history and history.get("job_id"):
            job = manager.get_job(history.get("job_id"))
            if job and not _can_view_job(user, job):
                return jsonify({"error": "forbidden"}), 403
    history = session_history.load(session_id)
    if not history:
        return jsonify({"history": [], "room_id": session_id})
    events = history.get("events") if isinstance(history.get("events"), list) else []
    return jsonify({"room_id": session_id, "history": events})


def api_sessions_history() -> Response:
    """API endpoint for sessions history."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if not _is_admin_user(user):
        return jsonify({"error": "forbidden"}), 403
    try:
        limit = int(request.args.get("limit") or 50)
    except Exception:
        limit = 50
    rooms = session_history.list_rooms(limit=limit, tail=5)
    return jsonify({"rooms": rooms})


def setup() -> Response:
    """Render or serve the setup route."""
    if _customers_enabled():
        return _proxy_service_request(CUSTOMERS_API_BASE, CUSTOMERS_TIMEOUT)
    next_path = _safe_next_path(request.args.get("next") or session.get("login_next"))
    if AUTH_MODE == "oidc":
        session["login_next"] = next_path
        return redirect(url_for("oidc_login", next=next_path))
    if user_store.has_users():
        return redirect(url_for("login", next=next_path))
    error = None
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = (request.form.get("password") or "").strip()
        confirm = (request.form.get("confirm") or "").strip()
        email = (request.form.get("email") or "").strip()
        if not USERNAME_RE.match(username):
            error = "Username must be 3-32 chars (letters, numbers, underscore, dash)."
        elif email and not EMAIL_RE.match(email):
            error = "Please enter a valid email address."
        elif len(password) < PASSWORD_MIN_LEN:
            error = f"Password must be at least {PASSWORD_MIN_LEN} characters."
        elif password != confirm:
            error = "Passwords do not match."
        else:
            user_store.create_user(username, password, role="admin", email=email or None)
            session["user"] = username
            _record_identity_event(
                username,
                role=user_store.get_role(username),
                email=user_store.get_email(username),
                provider="setup",
                meta={"source": "setup_form"},
            )
            _record_login_event(username, auth_mode="setup", provider="local", meta={"source": "setup_form"})
            _audit_event("setup", actor=username, status="success")
            session.pop("login_next", None)
            return redirect(next_path)
    return render_template(
        "setup.html",
        error=error,
        api_base=API_BASE,
        site_base=SITE_BASE,
        next_path=next_path,
        password_min_length=PASSWORD_MIN_LEN,
        passkeys_supported=False,
    )


def health() -> Response:
    """Render or serve the health route."""
    version = get_public_version_info()
    learning = None
    autoscaler_status = _continuum_autoscaler_status()
    job_queue = _job_queue_snapshot(manager, top_limit=3, include_owner_lists=False)
    ai_status = _ai_orchestration_status(probe_engines=False, candidate_limit=5)
    if stt_learning_store:
        try:
            learning = stt_learning_store.stats()
        except Exception:
            learning = {"error": "unavailable"}
    return jsonify(
        {
            "status": "ok",
            "version": version["version"],
            "jobs": len(manager.jobs),
            "workers": len(manager.workers),
            "job_queue": job_queue,
            "job_actions": job_action_manager.snapshot(),
            "subtasks": {
                "workers": len(subtask_manager.workers),
                "queue_depth": subtask_manager.queue_depth(),
                "inflight": subtask_manager.inflight(),
                "queue_capacity": SUBTASK_MAX_QUEUE,
            },
            "todo_scheduler": todo_scheduler.status(),
            "continuum_autoscaler": autoscaler_status,
            "workers_summary": autoscaler_status.get("workers") if isinstance(autoscaler_status, dict) else {},
            "sso": _sso_store_health(),
            "stt_learning": learning,
            "ai_orchestration": ai_status,
            "service_integrations": {
                "nmchain": {"enabled": _chain_enabled(), "base_url": NMCHAIN.base_url if _chain_enabled() else None},
                "customers": {
                    "enabled": _customers_enabled(),
                    "base_url": CUSTOMERS_SERVICE.base_url if _customers_enabled() else None,
                },
                "billing": {
                    "enabled": _billing_enabled(),
                    "base_url": BILLING_SERVICE.base_url if _billing_enabled() else None,
                },
            },
        }
    )


def api_version() -> Response:
    """Return the running application version payload."""
    return jsonify(get_public_version_info())


def capabilities_report() -> Response:
    """Return a capabilities snapshot for UI/API consumers."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    refresh = str(request.args.get("refresh") or "").strip().lower()
    force_refresh = refresh in {"1", "true", "yes", "y"}
    report = get_capabilities(force_refresh=force_refresh)
    return jsonify(report)


def admin_stats() -> Response:
    """Return aggregated admin statistics."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if not _is_admin_user(user):
        return jsonify({"error": "forbidden"}), 403
    total_users = user_store.count_users()
    jobs_snapshot = manager.list_jobs()
    job_queue = _job_queue_snapshot(manager, top_limit=3, include_owner_lists=True)
    jobs_by_status: Dict[str, int] = {}
    for job in jobs_snapshot:
        status = job.status or "unknown"
        jobs_by_status[status] = jobs_by_status.get(status, 0) + 1
    active_users = _active_users_snapshot()
    retention_status = _llm_telemetry_retention_status()
    ai_status = _ai_orchestration_status(probe_engines=True, candidate_limit=20)
    llm_request_telemetry: Dict[str, Any] = {
        "enabled": False,
        "retention": retention_status,
    }
    telemetry_store = _llm_request_telemetry_store()
    if telemetry_store is not None:
        try:
            llm_request_telemetry = telemetry_store.summary(hours=72, limit=12)
            llm_request_telemetry["retention"] = retention_status
        except Exception as exc:
            logger.warning("llm request telemetry summary failed: %s", exc)
            llm_request_telemetry = {
                "enabled": True,
                "degraded": True,
                "error": "telemetry_summary_unavailable",
                "retention": retention_status,
            }
    return jsonify(
        {
            "active_users": active_users,
            "active_users_count": len(active_users),
            "active_window_sec": ACTIVE_WINDOW_SEC,
            "total_users": total_users,
            "workers": len(manager.workers),
            "job_queue": job_queue,
            "job_action_workers": len(job_action_manager.workers),
            "job_action_queue_depth": job_action_manager.queue_depth(),
            "job_action_inflight": job_action_manager.inflight(),
            "job_action_queue_capacity": job_action_manager.max_queue,
            "job_action_max_outstanding_per_owner": job_action_manager.max_outstanding_per_owner,
            "job_action_max_inflight_per_owner": job_action_manager.max_inflight_per_owner,
            "job_action_active_owner_count": job_action_manager.active_owner_count(),
            "subtask_workers": len(subtask_manager.workers),
            "subtask_queue_depth": subtask_manager.queue_depth(),
            "subtask_inflight": subtask_manager.inflight(),
            "subtask_queue_capacity": SUBTASK_MAX_QUEUE,
            "todo_scheduler": todo_scheduler.status(),
            "jobs_total": len(jobs_snapshot),
            "jobs_running": jobs_by_status.get("running", 0),
            "jobs_queued": jobs_by_status.get("queued", 0),
            "jobs_failed": jobs_by_status.get("failed", 0),
            "jobs_completed": jobs_by_status.get("completed", 0),
            "jobs_paused": jobs_by_status.get("paused", 0),
            "jobs_by_status": jobs_by_status,
            "llm_request_telemetry": llm_request_telemetry,
            "ai_orchestration": ai_status,
            "continuum_autoscaler": _continuum_autoscaler_status(),
            "uptime_sec": int(time.time() - APP_START_TIME),
        }
    )


def admin_llm_telemetry() -> Response:
    """Return filtered LLM request telemetry for admin drill-down views."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if not _is_admin_user(user):
        return jsonify({"error": "forbidden"}), 403

    scope_raw = str(request.args.get("scope") or "").strip().lower()
    if scope_raw not in {"", "user", "team"}:
        return jsonify({"error": "invalid_scope"}), 400
    hours = max(1, min(_safe_int(request.args.get("hours"), 72), 24 * 365))
    limit = max(1, min(_safe_int(request.args.get("limit"), 20), 100))
    include_subjects = _is_truthy(request.args.get("include_subjects"))
    subject_limit = max(1, min(_safe_int(request.args.get("subject_limit"), 12), 100))
    filters = {
        "scope": scope_raw or None,
        "subject": str(request.args.get("subject") or "").strip() or None,
        "provider": str(request.args.get("provider") or "").strip() or None,
        "model": str(request.args.get("model") or "").strip() or None,
        "category": str(request.args.get("category") or "").strip() or None,
    }
    retention_status = _llm_telemetry_retention_status()
    telemetry_store = _llm_request_telemetry_store()
    if telemetry_store is None:
        return jsonify(
            {
                "enabled": False,
                "window_hours": hours,
                "retention": retention_status,
                "filters": filters,
            }
        )
    try:
        payload = telemetry_store.summary(
            scope=filters["scope"],
            subject=filters["subject"],
            hours=hours,
            limit=limit,
            provider=filters["provider"],
            model=filters["model"],
            category=filters["category"],
            include_subjects=include_subjects,
            subject_limit=subject_limit,
        )
    except Exception as exc:
        logger.warning("admin llm telemetry summary failed: %s", exc)
        return jsonify(
            {
                "enabled": True,
                "degraded": True,
                "error": "telemetry_summary_unavailable",
                "message": "Failed to load LLM telemetry summary.",
                "retention": retention_status,
                "filters": filters,
            }
        )

    payload["retention"] = retention_status
    payload["filters"] = filters
    payload["limit"] = limit
    payload["include_subjects"] = include_subjects
    payload["subject_limit"] = subject_limit
    return jsonify(payload)


def admin_ai_orchestration() -> Response:
    """Return AI orchestration registry status for admin drill-down views."""

    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if not _is_admin_user(user):
        return jsonify({"error": "forbidden"}), 403

    probe_engines = _is_truthy(request.args.get("probe_engines"))
    limit = max(1, min(_safe_int(request.args.get("limit"), 20), 100))
    payload = _ai_orchestration_status(probe_engines=probe_engines, candidate_limit=limit)
    payload["probe_engines"] = probe_engines
    payload["limit"] = limit
    payload["fetched_at"] = dt.datetime.now(dt.timezone.utc).isoformat()
    return jsonify(payload)


def workers_telemetry() -> Response:
    """Return worker capacity and autoscaler telemetry for the Control Room."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    refresh = str(request.args.get("refresh") or "").strip().lower() in {"1", "true", "yes", "y"}
    include_cluster = str(request.args.get("include_cluster") or "").strip().lower() in {"1", "true", "yes", "y"}
    try:
        limit = int(request.args.get("limit") or 180)
    except Exception:
        limit = 180
    limit = max(10, min(limit, CONTINUUM_AUTOSCALE_HISTORY_MAX))
    try:
        payload = _workers_telemetry_payload(limit=limit, refresh=refresh, include_cluster=include_cluster)
    except Exception as exc:
        logger.warning("Workers telemetry endpoint degraded: %s", exc)
        payload = {
            "ok": False,
            "degraded": True,
            "message": _friendly_continuum_error(str(exc)),
            "warnings": [_friendly_continuum_error(str(exc))],
            "autoscaler": _continuum_autoscaler_status(),
            "summary": {},
            "job_queue": {},
            "history": [],
        }
    return jsonify(payload)


def api_audit() -> Response:
    """API endpoint for audit."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if not _is_admin_user(user):
        return jsonify({"error": "forbidden"}), 403
    limit = request.args.get("limit")
    try:
        limit_val = int(limit) if limit else 60
    except Exception:
        limit_val = 60
    limit_val = max(1, min(limit_val, 200))
    actions_raw = request.args.get("actions") or ""
    actions = [item.strip() for item in actions_raw.split(",") if item.strip()]
    entries = _read_audit_entries(limit_val, actions=actions or None)
    return jsonify({"entries": entries})


def job_estimate() -> Response:
    """Job endpoint for estimate."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    payload = request.get_json(force=True, silent=True) or {}
    project_id = payload.get("project_id") or payload.get("project")
    team_id = payload.get("team_id") if isinstance(payload, dict) else None
    if project_id:
        project = access_store.get_project(project_id)
        if not project:
            return jsonify({"error": "project_not_found"}), 404
        if not (_is_admin_user(user) or access_store.can_submit_project(user, project_id)):
            return jsonify({"error": "forbidden", "details": "Project access denied."}), 403
        if not team_id:
            team_id = project.get("team_id")
            if team_id:
                payload["team_id"] = team_id
    elif team_id and not _can_access_team_tokens(user, team_id):
        return jsonify({"error": "forbidden", "details": "Team access denied."}), 403
    if isinstance(payload, dict) and str(payload.get("token_scope") or "").lower() == "personal":
        team_id = None
    estimate = _estimate_job_tokens(payload)
    chargeable_estimate = _chargeable_token_estimate(user, payload, estimate)
    snapshot = _token_snapshot(user, team_id)
    return jsonify({"estimate": estimate, "chargeable_estimate": chargeable_estimate, **snapshot})


def import_requirements() -> Response:
    """Handle the import requirements route."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "file_required"}), 400
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in {".csv", ".xls", ".xlsx", ".ods"}:
        return jsonify({"error": "unsupported_format"}), 400
    data = file.read()
    if len(data) > REQUIREMENTS_IMPORT_MAX_BYTES:
        return jsonify({"error": "file_too_large"}), 413
    try:
        if ext == ".csv":
            rows = _parse_csv_rows(data)
        elif ext == ".xlsx":
            rows = _parse_xlsx_rows(data)
        elif ext == ".xls":
            rows = _parse_xls_rows(data)
        else:
            rows = _parse_ods_rows(data)
    except Exception as exc:
        return jsonify({"error": "import_failed", "details": str(exc)}), 400
    items = _rows_to_requirements(rows)
    return jsonify({"items": items, "count": len(items)})


def export_requirements() -> Response:
    """Handle the export requirements route."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    payload = request.get_json(force=True, silent=True) or {}
    fmt = str(payload.get("format") or "csv").lower()
    if fmt not in {"csv", "xls", "xlsx", "ods"}:
        return jsonify({"error": "unsupported_format"}), 400
    items = _normalize_requirements_items(payload.get("items"))
    try:
        if fmt == "csv":
            data = _export_csv(items)
            mime = "text/csv"
        elif fmt == "xlsx":
            data = _export_xlsx(items)
            mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif fmt == "xls":
            data = _export_xls(items)
            mime = "application/vnd.ms-excel"
        else:
            data = _export_ods(items)
            mime = "application/vnd.oasis.opendocument.spreadsheet"
    except Exception as exc:
        return jsonify({"error": "export_failed", "details": str(exc)}), 400
    filename = f"requirements_register.{fmt}"
    response = Response(data, mimetype=mime)
    response.headers["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def mcp_servers() -> Response:
    """MCP endpoint for servers."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if not _is_admin_user(user):
        return jsonify({"error": "forbidden"}), 403
    if request.method == "GET":
        servers = [_mcp_masked_server(user, server) for server in mcp_store.list_servers(user)]
        _audit_event("mcp_list", actor=user, status="success", details={"count": len(servers)})
        return jsonify({"servers": servers})
    payload = request.get_json(force=True, silent=True) or {}
    name = str(payload.get("name") or "").strip()
    current = mcp_store.get_server(user, name)
    if not name or (not current and not str(payload.get("base_url") or "").strip()):
        return jsonify({"error": "name_and_base_url_required"}), 400
    try:
        config = _prepare_mcp_server_config(user, payload, current=current)
    except ValueError as exc:
        return jsonify({"error": "invalid_mcp_server", "details": str(exc)}), 400
    if hasattr(user_store, "ensure_user"):
        try:
            user_store.ensure_user(user, role=_user_role(user) or "user", email=user_store.get_email(user))
        except Exception:
            pass
    mcp_store.save_server(user, config)
    saved = mcp_store.get_server(user, config.name) or config
    _audit_event(
        "mcp_save",
        actor=user,
        status="success",
        details={"server": config.name, "host": urlparse(config.base_url).hostname or ""},
    )
    return jsonify({"server": _mcp_masked_server(user, saved)})


def mcp_server_delete(name: str) -> Response:
    """MCP endpoint for server delete."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if not _is_admin_user(user):
        return jsonify({"error": "forbidden"}), 403
    server = mcp_store.get_server(user, name)
    if not server:
        return jsonify({"error": "not_found"}), 404
    deleted = mcp_store.delete_server(user, name)
    if not deleted:
        return jsonify({"error": "not_found"}), 404
    _mcp_cleanup_managed_secrets(user, server)
    _audit_event(
        "mcp_delete",
        actor=user,
        status="success",
        details={"server": name, "host": urlparse(server.base_url).hostname or ""},
    )
    return jsonify({"status": "deleted", "name": name})


def mcp_server_tools(name: str) -> Response:
    """MCP endpoint for server tools."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if not _is_admin_user(user):
        return jsonify({"error": "forbidden"}), 403
    try:
        tools = _mcp_execute(
            user,
            name,
            "list_tools",
            lambda client: client.list_tools(),
            runtime_from_result=lambda result: {"tool_count": len(_mcp_result_items(result, "tools"))},
        )
    except KeyError:
        return jsonify({"error": "not_found"}), 404
    except Exception as exc:
        return jsonify({"error": "mcp_request_failed", "details": str(exc)}), 400
    return jsonify({"tools": tools})


def mcp_server_call(name: str) -> Response:
    """MCP endpoint for server call."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if not _is_admin_user(user):
        return jsonify({"error": "forbidden"}), 403
    payload = request.get_json(force=True, silent=True) or {}
    tool = str(payload.get("tool") or "").strip()
    if not tool:
        return jsonify({"error": "tool_required"}), 400
    arguments = payload.get("arguments")
    if arguments is not None and not isinstance(arguments, dict):
        return jsonify({"error": "invalid_arguments"}), 400
    try:
        result = _mcp_execute(
            user,
            name,
            "call",
            lambda client: client.call_tool(tool, arguments or {}),
            audit_details={"tool": tool},
            runtime_from_result=lambda _result: {"last_tool": tool},
        )
    except KeyError:
        return jsonify({"error": "not_found"}), 404
    except Exception as exc:
        return jsonify({"error": "mcp_request_failed", "details": str(exc)}), 400
    return jsonify({"result": result})


def mcp_server_resources(name: str) -> Response:
    """MCP endpoint for server resources."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if not _is_admin_user(user):
        return jsonify({"error": "forbidden"}), 403
    try:
        resources = _mcp_execute(
            user,
            name,
            "list_resources",
            lambda client: client.list_resources(),
            runtime_from_result=lambda result: {"resource_count": len(_mcp_result_items(result, "resources"))},
        )
    except KeyError:
        return jsonify({"error": "not_found"}), 404
    except Exception as exc:
        return jsonify({"error": "mcp_request_failed", "details": str(exc)}), 400
    return jsonify({"resources": resources})


def mcp_server_resource(name: str) -> Response:
    """MCP endpoint for server resource."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if not _is_admin_user(user):
        return jsonify({"error": "forbidden"}), 403
    payload = request.get_json(force=True, silent=True) or {}
    uri = str(payload.get("uri") or "").strip()
    if not uri:
        return jsonify({"error": "uri_required"}), 400
    try:
        resource = _mcp_execute(
            user,
            name,
            "read_resource",
            lambda client: client.read_resource(uri),
            audit_details={"uri": uri[:256]},
            runtime_from_result=lambda _result: {"last_resource_uri": uri[:512]},
        )
    except KeyError:
        return jsonify({"error": "not_found"}), 404
    except Exception as exc:
        return jsonify({"error": "mcp_request_failed", "details": str(exc)}), 400
    return jsonify({"resource": resource})


def request_refund(job_id: str) -> Response:
    """Handle the request refund route."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    job = manager.get_job(job_id, owner=user)
    if not job:
        return jsonify({"error": "job not found"}), 404
    latest = _latest_refund_request(job)
    if latest and latest.get("status") not in {"rejected", "settled", "partial-refund"}:
        return jsonify({"error": "refund_already_open"}), 409
    amount = _safe_amount(request.form.get("amount"))
    reason = (request.form.get("reason") or "").strip()
    details = (request.form.get("details") or "").strip()
    if amount is None:
        return jsonify({"error": "invalid_amount"}), 400
    if not reason:
        return jsonify({"error": "reason_required"}), 400
    max_refund = _refund_max_amount(job)
    if max_refund and amount > max_refund:
        return jsonify({"error": "amount_exceeds_max", "max_refund": max_refund}), 400
    files = request.files.getlist("screenshots")
    if not files:
        return jsonify({"error": "screenshots_required"}), 400
    if len(files) > REFUND_MAX_FILES:
        return jsonify({"error": "too_many_files"}), 400
    request_id = uuid.uuid4().hex
    stored_files = _store_refund_files(job.job_id, request_id, files)
    if not stored_files:
        return jsonify({"error": "invalid_screenshots"}), 400
    refund = {
        "id": request_id,
        "status": "requested",
        "requested_amount": amount,
        "requested_at": _now_iso(),
        "requested_by": user,
        "reason": reason,
        "details": details,
        "screenshots": stored_files,
        "job_snapshot": _refund_job_snapshot(job),
        "history": [
            {
                "status": "requested",
                "at": _now_iso(),
                "by": user,
                "note": reason,
            }
        ],
    }
    job.refunds.append(refund)
    job.updated_at = _now_iso()
    job.persist(force=True)
    return jsonify({"refund": refund})


def list_refunds() -> Response:
    """Handle the list refunds route."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if not _is_admin_user(user):
        return jsonify({"error": "forbidden"}), 403
    requests: List[Dict[str, Any]] = []
    for job in manager.list_jobs():
        for refund in job.refunds:
            requests.append(
                {
                    "job_id": job.job_id,
                    "project_name": job.project_name,
                    "owner": job.owner,
                    "workflow": job.workflow,
                    "job_status": job.status,
                    "tokens": {
                        "estimate": job.token_estimate,
                        "actual": job.token_actual,
                        "debited": job.token_debited,
                        "shortfall": job.token_shortfall,
                    },
                    "refund": refund,
                }
            )
    requests.sort(
        key=lambda item: _timestamp_sort_key(item.get("refund", {}).get("requested_at")),
        reverse=True,
    )
    return jsonify({"requests": requests})


def screen_refund(job_id: str, request_id: str) -> Response:
    """Handle the screen refund route."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if not _is_admin_user(user):
        return jsonify({"error": "forbidden"}), 403
    job = manager.get_job(job_id)
    if not job:
        return jsonify({"error": "job not found"}), 404
    refund = _find_refund_request(job, request_id)
    if not refund:
        return jsonify({"error": "refund not found"}), 404
    settings = _resolve_llm_settings(user=user)
    try:
        provider = _build_request_llm_provider(
            user,
            settings,
            workflow="admin_refund_screening",
            role="reviewer",
        )
    except Exception as exc:
        return jsonify({"error": "llm_unavailable", "details": str(exc)}), 400
    system = (
        "You are a refund screening assistant. Review the request and recommend a decision. "
        "Return JSON with keys: decision (approve/reject/partial), suggested_amount (integer), "
        "confidence (0-1), rationale."
    )
    job_snapshot = refund.get("job_snapshot") or _refund_job_snapshot(job)
    screenshots = refund.get("screenshots") or []
    prompt = (
        "Refund request:\n"
        f"- Requested amount: {refund.get('requested_amount')}\n"
        f"- Reason: {refund.get('reason')}\n"
        f"- Details: {refund.get('details')}\n"
        f"- Screenshots: {len(screenshots)}\n"
        "\nJob snapshot:\n"
        f"- Workflow: {job_snapshot.get('workflow')}\n"
        f"- Status: {job_snapshot.get('status')}\n"
        f"- Token estimate: {job_snapshot.get('token_estimate')}\n"
        f"- Token actual: {job_snapshot.get('token_actual')}\n"
        f"- Token debited: {job_snapshot.get('token_debited')}\n"
        f"- Token shortfall: {job_snapshot.get('token_shortfall')}\n"
    )
    try:
        response = provider.predict(messages=[{"role": "user", "content": prompt}], system=system)
    except Exception as exc:
        return jsonify({"error": "llm_request_failed", "details": str(exc)}), 400
    suggestion = _extract_json_payload(response.text) or {}
    if not isinstance(suggestion, dict):
        suggestion = {"raw": response.text}
    suggested_amount = _safe_amount(suggestion.get("suggested_amount")) if isinstance(suggestion, dict) else None
    if suggested_amount is not None:
        suggestion["suggested_amount"] = suggested_amount
    suggestion.update(
        {
            "provider": response.provider or settings.get("provider"),
            "model": response.model or settings.get("model"),
            "screened_at": _now_iso(),
        }
    )
    refund["llm_screening"] = suggestion
    job.persist(force=True)
    return jsonify({"suggestion": suggestion})


def decide_refund(job_id: str, request_id: str) -> Response:
    """Handle the decide refund route."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if not _is_admin_user(user):
        return jsonify({"error": "forbidden"}), 403
    job = manager.get_job(job_id)
    if not job:
        return jsonify({"error": "job not found"}), 404
    refund = _find_refund_request(job, request_id)
    if not refund:
        return jsonify({"error": "refund not found"}), 404
    payload = request.get_json(force=True, silent=True) or {}
    status = str(payload.get("status") or "").strip().lower()
    amount = _safe_amount(payload.get("amount")) if payload.get("amount") is not None else None
    note = (payload.get("note") or "").strip()
    if status not in REFUND_STATUSES:
        return jsonify({"error": "invalid_status"}), 400
    if status in {"approved", "partial-refund", "settled"} and amount is None:
        return jsonify({"error": "amount_required"}), 400
    max_refund = _refund_max_amount(job)
    if amount is not None and amount > max_refund:
        return jsonify({"error": "amount_exceeds_max", "max_refund": max_refund}), 400
    refund.setdefault("history", []).append(
        {
            "status": status,
            "at": _now_iso(),
            "by": user,
            "note": note,
            "amount": amount,
        }
    )
    refund["status"] = status
    if amount is not None:
        refund["approved_amount"] = amount
    refund["admin_decision"] = {
        "status": status,
        "amount": amount,
        "note": note,
        "admin": user,
        "decided_at": _now_iso(),
    }
    if status in {"settled", "partial-refund"}:
        if refund.get("settled_at"):
            return jsonify({"error": "already_settled"}), 409
        entry = _record_token_event(
            "user",
            job.owner,
            "refund",
            int(amount or 0),
            {
                "job_id": job.job_id,
                "refund_id": request_id,
                "decision": status,
                "requested_amount": refund.get("requested_amount"),
                "approved_amount": amount,
                "admin": user,
                "note": note,
            },
            request_id=_chain_request_id("refund", request_id),
        )
        refund["settled_at"] = _now_iso()
        refund["ledger_entry"] = entry
    job.updated_at = _now_iso()
    job.persist(force=True)
    return jsonify({"refund": refund, "max_refund": max_refund})


def refund_file(job_id: str, request_id: str, filename: str) -> Response:
    """Handle the refund file route."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    job = manager.get_job(job_id)
    if not job:
        return jsonify({"error": "job not found"}), 404
    if not _is_admin_user(user) and job.owner != user:
        return jsonify({"error": "forbidden"}), 403
    refund = _find_refund_request(job, request_id)
    if not refund:
        return jsonify({"error": "refund not found"}), 404
    allowed = {entry.get("filename") for entry in refund.get("screenshots", []) if entry.get("filename")}
    if filename not in allowed:
        return jsonify({"error": "file not found"}), 404
    refund_dir = os.path.join(JOB_ROOT, job_id, "refunds", request_id)
    return send_from_directory(refund_dir, filename)


def jobs() -> Response:
    """Jobs collection endpoint (submit/list)."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if request.method == "POST":
        payload = request.get_json(force=True, silent=True) or {}
        payload["owner"] = user
        _apply_user_job_defaults(payload, user)
        project_id = payload.get("project_id") or payload.get("project")
        team_id = payload.get("team_id") if isinstance(payload, dict) else None
        if project_id:
            project = access_store.get_project(project_id)
            if not project:
                return jsonify({"error": "project_not_found"}), 404
            if not (_is_admin_user(user) or access_store.can_submit_project(user, project_id)):
                return jsonify({"error": "forbidden", "details": "Project access denied."}), 403
            if not payload.get("team_id"):
                payload["team_id"] = project.get("team_id")
                team_id = payload.get("team_id")
        elif team_id and not _can_access_team_tokens(user, team_id):
            return jsonify({"error": "forbidden", "details": "Team access denied."}), 403
        if isinstance(payload, dict) and str(payload.get("token_scope") or "").lower() == "personal":
            payload.pop("team_id", None)
            team_id = None
        estimate = _estimate_job_tokens(payload)
        chargeable_estimate = _chargeable_token_estimate(user, payload, estimate)
        snapshot = _token_snapshot(user, team_id)
        if chargeable_estimate > snapshot["available"]:
            return (
                jsonify(
                    {
                        "error": "insufficient_tokens",
                        "details": "Insufficient tokens to submit this job.",
                        "estimate": estimate,
                        "chargeable_estimate": chargeable_estimate,
                        "available": snapshot["available"],
                        "balance": snapshot["balance"],
                    }
                ),
                402,
            )
        job = manager.submit_job(payload, owner=user)
        manager._reserve_tokens(job, estimate)
        return jsonify(_augment_job_dict_for_user(job.to_dict(), user, job))
    status = request.args.get("status")
    scope = (request.args.get("scope") or "team").strip().lower()
    project_filter = (request.args.get("project_id") or request.args.get("project") or "").strip()
    jobs_source = manager.list_jobs(status=status)
    visible_jobs: List[Job] = []
    if _is_admin_user(user) and scope == "all":
        visible_jobs = list(jobs_source)
    elif scope == "personal":
        visible_jobs = [job for job in jobs_source if job.owner == user]
    else:
        visible_jobs = [job for job in jobs_source if _can_view_job(user, job)]
    if project_filter:
        visible_jobs = [job for job in visible_jobs if _job_project_id(job) == project_filter]
    jobs_list = [_augment_job_dict_for_user(job.to_dict(), user, job) for job in visible_jobs]
    return jsonify({"jobs": jobs_list})


def job_detail(job_id: str) -> Response:
    """Job endpoint for detail."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    job = manager.get_job(job_id)
    if not job or not _can_view_job(user, job):
        return jsonify({"error": "job not found"}), 404
    is_admin = _is_admin_user(user)
    if request.method == "DELETE":
        payload = request.get_json(force=True, silent=True) or {}
        stop = bool(payload.get("stop")) or request.args.get("stop") in {"1", "true", "yes"}
        if not _can_manage_job(user, job):
            return jsonify({"error": "forbidden"}), 403
        if job.status in {"queued", "running", "paused"} and not stop:
            return jsonify({"error": "job_active", "details": "Stop the job before deleting."}), 409
        deleted = manager.delete_job(job_id, owner=None, stop_if_active=stop)
        if not deleted:
            return jsonify({"error": "delete_failed"}), 409
        return jsonify({"status": "deleted", "job_id": job_id})
    data = job.to_dict(include_logs=True, log_tail=DEFAULT_TAIL)
    data["logs"] = _redact_log_entries(data.get("logs", []), is_admin)
    if isinstance(job.output_paths, dict):
        data.update(_job_output_insights_from_output(job.output_paths.get("primary")))
    return jsonify(_augment_job_dict_for_user(data, user, job))


def _workspace_capabilities() -> Dict[str, Any]:
    return {
        "continuum": _continuum_enabled(),
        "continuum_ready": _continuum_ready(),
        "ide_template": bool(CONTINUUM_IDE_URL_TEMPLATE),
        "preview_template": bool(CONTINUUM_PREVIEW_URL_TEMPLATE),
        "ide_file_template": bool(CONTINUUM_IDE_FILE_URL_TEMPLATE),
    }


def _workspace_response_payload(
    job: Job,
    *,
    status: Optional[str] = None,
    error: Optional[str] = None,
    task: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    payload: Dict[str, Any] = {
        "workspace": job.workspace_env if isinstance(job.workspace_env, dict) else {},
        "capabilities": _workspace_capabilities(),
        "status": status,
        "error": error,
    }
    if task is not None:
        payload["task"] = task
    return payload


def _workspace_task_links(job_id: str, task_id: str) -> Dict[str, str]:
    return {
        "status_url": url_for("job_task_detail", job_id=job_id, task_id=task_id),
        "cancel_url": url_for("job_task_cancel", job_id=job_id, task_id=task_id),
    }


def _workspace_action_refresh(job: Job, *, timeout_sec: Optional[float] = None) -> Dict[str, Any]:
    return _workspace_action_refresh_impl(
        job,
        config=_continuum_client_config(),
        now_iso=lambda: _now_iso(),
        request=lambda method, path, **kwargs: _continuum_request(method, path, **kwargs),
        format_workspace_template=_format_workspace_template,
        error_factory=lambda code, details, status_code: JobActionExecutionError(
            code,
            details,
            status_code=status_code,
        ),
        timeout_sec=timeout_sec,
    )


def _workspace_action_create(
    job: Job,
    payload: Dict[str, Any],
    *,
    timeout_sec: Optional[float] = None,
) -> Dict[str, Any]:
    return _workspace_action_create_impl(
        job,
        payload,
        config=_continuum_client_config(),
        now_iso=lambda: _now_iso(),
        is_truthy=_is_truthy,
        request=lambda method, path, **kwargs: _continuum_request(method, path, **kwargs),
        format_workspace_template=_format_workspace_template,
        error_factory=lambda code, details, status_code: JobActionExecutionError(
            code,
            details,
            status_code=status_code,
        ),
        timeout_sec=timeout_sec,
    )


def _invoke_internal_post_json(
    *,
    user: str,
    path: str,
    handler,
    payload: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Execute an existing POST handler under a synthetic internal request context."""
    request_payload = payload if isinstance(payload, dict) else {}
    with app.test_request_context(path, method="POST", json=request_payload):
        session["user"] = user
        response = app.make_response(handler())
        status_code = int(response.status_code or 200)
        data = response.get_json(silent=True)
    if 200 <= status_code < 300:
        return data if isinstance(data, dict) else {"data": data}
    if isinstance(data, dict):
        error_code = str(data.get("error") or "route_execution_failed").strip() or "route_execution_failed"
        details = str(data.get("details") or data.get("message") or error_code).strip() or error_code
    else:
        error_code = "route_execution_failed"
        details = f"Internal route {path} returned HTTP {status_code}."
    raise SubtaskExecutionError(error_code, details, status_code=status_code)


def _execute_todo_route(
    *,
    owner: str,
    todo_id: str,
    schedule_id: Optional[str] = None,
    route_override: Optional[Dict[str, Any]] = None,
    todo_store_override: Optional[Any] = None,
) -> Dict[str, Any]:
    """Execute the routed action for a TODO item and persist the outcome back to the inbox."""
    todo_id = str(todo_id or "").strip()
    if not todo_id:
        raise SubtaskExecutionError("todo_id_required", "A todo_id is required.", status_code=400)
    execution_store = todo_store_override if todo_store_override is not None else todo_store
    item = execution_store.get_item(owner, todo_id)
    if not item:
        raise SubtaskExecutionError("todo_not_found", "TODO item no longer exists.", status_code=404)
    if str(item.get("status") or "todo").strip().lower() != "todo":
        raise SubtaskExecutionError("todo_not_pending", "TODO item is no longer pending.", status_code=409)
    current_state = str(item.get("execution_state") or "ready").strip().lower()
    if current_state not in {"claimed", "processing"}:
        claimed = execution_store.claim_item(owner, todo_id)
        if not claimed:
            raise SubtaskExecutionError("todo_not_ready", "TODO item is not ready for execution.", status_code=409)
    processing = execution_store.begin_processing(owner, todo_id)
    if not processing:
        raise SubtaskExecutionError("todo_not_ready", "TODO item could not enter processing state.", status_code=409)
    item = processing
    try:
        route = route_override if isinstance(route_override, dict) and route_override else build_route_suggestion(item)
        target = str(route.get("target") or "").strip().lower()
        workflow = str(route.get("workflow") or "").strip()
        route_payload = route.get("payload") if isinstance(route.get("payload"), dict) else {}
        if target == "job":
            response = _invoke_internal_post_json(
                user=owner,
                path="/api/jobs",
                handler=jobs,
                payload=route_payload,
            )
            job_id = str(response.get("job_id") or "").strip()
            result = {
                "todo_id": todo_id,
                "schedule_id": schedule_id,
                "route": route,
                "response": response,
            }
            links: Dict[str, Any] = {}
            if schedule_id:
                links["schedules"] = schedule_id
            if job_id:
                result["job_id"] = job_id
                links["jobs"] = job_id
            execution_store.complete_execution(owner, todo_id, result, status="done", links=links)
            return result
        if target != "assistant":
            raise SubtaskExecutionError("unsupported_todo_route", f"Unsupported TODO target: {target or '--'}", status_code=400)
        if workflow == "assistant_requirements":
            response = _invoke_internal_post_json(
                user=owner,
                path="/api/assistant/requirements",
                handler=assistant_requirements,
                payload=route_payload,
            )
        elif workflow == "playground_plan":
            response = _invoke_internal_post_json(
                user=owner,
                path="/api/playground/plan",
                handler=playground_plan,
                payload=route_payload,
            )
        elif workflow == "execution_plan":
            response = _invoke_internal_post_json(
                user=owner,
                path="/api/execution/plan",
                handler=execution_plan,
                payload=route_payload,
            )
        else:
            raise SubtaskExecutionError("unsupported_todo_route", f"Unsupported TODO assistant workflow: {workflow or '--'}", status_code=400)
        result = {
            "todo_id": todo_id,
            "schedule_id": schedule_id,
            "route": route,
            "response": response,
        }
        links: Dict[str, Any] = {}
        if schedule_id:
            links["schedules"] = schedule_id
        response_job_id = str(response.get("job_id") or "").strip() if isinstance(response, dict) else ""
        if response_job_id:
            result["job_id"] = response_job_id
            links["jobs"] = response_job_id
        execution_store.complete_execution(owner, todo_id, result, status="done", links=links)
        return result
    except SubtaskExecutionError as exc:
        execution_store.fail_execution(
            owner,
            todo_id,
            exc.code,
            result={"error": exc.code, "details": exc.details, "schedule_id": schedule_id},
        )
        raise
    except Exception as exc:
        execution_store.fail_execution(
            owner,
            todo_id,
            "todo_execution_failed",
            result={"error": "todo_execution_failed", "details": str(exc), "schedule_id": schedule_id},
        )
        raise


def _mark_subtask_cancelled_side_effects(task: "Subtask") -> None:
    """Best-effort cleanup for queued subtasks that are cancelled before execution."""

    if not task or task.action != "rag_collection_build":
        return
    payload = task.payload if isinstance(task.payload, dict) else {}
    name = str(payload.get("name") or "").strip()
    version_id = str(payload.get("_rag_version_id") or payload.get("version_id") or "").strip()
    if not name or not version_id:
        return
    metadata_store = getattr(CENTRAL_STORE, "rag_metadata", None) if CENTRAL_STORE is not None else None
    if metadata_store is None:
        return
    try:
        metadata_store.fail_collection_build(
            task.owner,
            name,
            version_id=version_id,
            status="cancelled",
            metadata={"error_code": "cancelled", "error_detail": "Cancelled before execution started."},
        )
    except Exception as exc:  # pragma: no cover - best effort only
        logger.debug("Cancelled RAG subtask metadata update skipped for %s/%s: %s", task.owner, name, exc)


def _execute_subtask(task: Subtask) -> Dict[str, Any]:
    """Dispatch generic subtasks to existing API handlers or TODO route execution helpers."""
    if task.cancel_requested:
        raise SubtaskExecutionError("cancelled", "Subtask was cancelled before execution.", status_code=409)
    payload = task.payload if isinstance(task.payload, dict) else {}
    if task.action == "todo_execute":
        return _execute_todo_route(
            owner=task.owner,
            todo_id=str(payload.get("todo_id") or "").strip(),
            schedule_id=str(payload.get("schedule_id") or "").strip() or None,
            route_override=payload.get("route_override") if isinstance(payload.get("route_override"), dict) else None,
            todo_store_override=payload.get("_todo_store"),
        )
    if task.action == "assistant_requirements":
        response = _invoke_internal_post_json(
            user=task.owner,
            path="/api/assistant/requirements",
            handler=assistant_requirements,
            payload=payload,
        )
        return {"action": task.action, "response": response}
    if task.action == "playground_plan":
        response = _invoke_internal_post_json(
            user=task.owner,
            path="/api/playground/plan",
            handler=playground_plan,
            payload=payload,
        )
        return {"action": task.action, "response": response}
    if task.action == "execution_plan":
        response = _invoke_internal_post_json(
            user=task.owner,
            path="/api/execution/plan",
            handler=execution_plan,
            payload=payload,
        )
        return {"action": task.action, "response": response}
    if task.action == "submit_job":
        response = _invoke_internal_post_json(
            user=task.owner,
            path="/api/jobs",
            handler=jobs,
            payload=payload,
        )
        result = {"action": task.action, "response": response}
        job_id = str(response.get("job_id") or "").strip()
        if job_id:
            result["job_id"] = job_id
        return result
    if task.action == "rag_collection_build":
        try:
            result = assistant_service.rag_collection_build(
                _assistant_pipeline_dependencies(),
                user=task.owner,
                payload=payload,
            )
        except ServiceError as exc:
            raise SubtaskExecutionError(exc.code, exc.details or exc.code, status_code=exc.status_code) from exc
        response = result.payload if isinstance(result.payload, dict) else {"data": result.payload}
        return {"action": task.action, "response": response, "version_id": response.get("version_id")}
    raise SubtaskExecutionError("invalid_action", f"Unsupported subtask action: {task.action}", status_code=400)


def _execute_job_action_task(task: JobActionTask) -> Dict[str, Any]:
    """Dispatch background action tasks to the relevant job/workspace handlers."""
    job = manager.get_job(task.job_id)
    if not job:
        raise JobActionExecutionError("job_not_found", "Job no longer exists.", status_code=404)
    if task.action == "workspace_refresh":
        return _workspace_action_refresh(job, timeout_sec=task.timeout_sec)
    if task.action == "workspace_create":
        payload = task.payload if isinstance(task.payload, dict) else {}
        return _workspace_action_create(job, payload, timeout_sec=task.timeout_sec)
    raise JobActionExecutionError("invalid_action", f"Unsupported task action: {task.action}", status_code=400)


def _enqueue_workspace_task(
    job: Job,
    *,
    owner: str,
    action: str,
    payload: Optional[Dict[str, Any]] = None,
    timeout_sec: float = JOB_ACTION_TIMEOUT_SEC,
) -> JobActionTask:
    try:
        return job_action_manager.submit(
            job_id=job.job_id,
            owner=owner,
            action=action,
            payload=payload or {},
            timeout_sec=timeout_sec,
        )
    except queue.Full as exc:
        raise JobActionExecutionError(
            "job_action_capacity_unavailable",
            "Background action queue is full. Retry later.",
            status_code=503,
        ) from exc


def job_workspace(job_id: str) -> Response:
    """Job endpoint for workspace."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    job = manager.get_job(job_id)
    if not job or not _can_view_job(user, job):
        return jsonify({"error": "job not found"}), 404

    if request.method == "GET":
        return jsonify(_workspace_response_payload(job))
    if not _can_manage_job(user, job):
        return jsonify({"error": "forbidden"}), 403

    payload = request.get_json(force=True, silent=True) or {}
    action = str(payload.get("action") or "create").strip().lower()
    blocking = _is_truthy(payload.get("blocking")) or _is_truthy(payload.get("wait"))
    timeout_sec = max(1.0, min(_safe_float(payload.get("timeout_sec"), JOB_ACTION_TIMEOUT_SEC), 120.0))
    now = _now_iso()

    if action == "clear":
        job.workspace_env = {}
        job.persist(force=True)
        job.append_log("Workspace cleared by user.")
        return jsonify(_workspace_response_payload(job, status="cleared"))

    if action == "attach":
        ide_url = str(payload.get("ide_url") or "").strip()
        preview_url = str(payload.get("preview_url") or "").strip()
        if not ide_url and not preview_url:
            return jsonify({"error": "missing_urls", "details": "Provide an IDE URL or preview URL."}), 400
        workspace_env = job.workspace_env if isinstance(job.workspace_env, dict) else {}
        workspace_env.update(
            {
                "provider": workspace_env.get("provider") or "manual",
                "status": workspace_env.get("status") or "ready",
                "ide_url": ide_url or workspace_env.get("ide_url") or "",
                "preview_url": preview_url or workspace_env.get("preview_url") or "",
                "updated_at": now,
                "requested_at": workspace_env.get("requested_at") or now,
                "details": workspace_env.get("details") or "Attached URLs",
            }
        )
        job.workspace_env = workspace_env
        job.persist(force=True)
        job.append_log("Workspace URLs attached by user.")
        return jsonify(_workspace_response_payload(job, status="attached"))

    if action == "refresh":
        workspace_env = job.workspace_env if isinstance(job.workspace_env, dict) else {}
        if workspace_env.get("provider") != "continuum" or not workspace_env.get("vm_id"):
            result = _workspace_action_refresh(job, timeout_sec=timeout_sec)
            return jsonify(_workspace_response_payload(job, status=result.get("status")))
        if not _continuum_enabled():
            return jsonify({"error": "continuum_unavailable", "details": "Continuum API base is not configured."}), 400
        if blocking:
            try:
                result = _workspace_action_refresh(job, timeout_sec=timeout_sec)
            except JobActionExecutionError as exc:
                return jsonify(exc.to_payload()), exc.status_code
            return jsonify(_workspace_response_payload(job, status=result.get("status")))
        try:
            task = _enqueue_workspace_task(
                job,
                owner=user,
                action="workspace_refresh",
                payload={},
                timeout_sec=timeout_sec,
            )
        except JobActionExecutionError as exc:
            return jsonify(exc.to_payload()), exc.status_code
        task_view = task.to_dict(include_result=False)
        task_view.update(_workspace_task_links(job.job_id, task.task_id))
        return jsonify(_workspace_response_payload(job, status="queued", task=task_view)), 202

    if action != "create":
        return jsonify({"error": "invalid_action"}), 400
    if not _continuum_ready():
        missing = []
        if not CONTINUUM_API_BASE:
            missing.append("CONTINUUM_API_BASE")
        if not CONTINUUM_VM_PUBLIC_KEY_ID:
            missing.append("CONTINUUM_VM_PUBLIC_KEY_ID")
        detail = "Continuum not configured." if not missing else f"Missing: {', '.join(missing)}"
        return jsonify({"error": "continuum_unavailable", "details": detail}), 400
    workspace_env = job.workspace_env if isinstance(job.workspace_env, dict) else {}
    force_create = _is_truthy(payload.get("force"))
    if workspace_env.get("provider") == "continuum" and workspace_env.get("vm_id") and not force_create:
        return jsonify(_workspace_response_payload(job, status="exists"))
    if blocking:
        try:
            result = _workspace_action_create(job, payload, timeout_sec=timeout_sec)
        except JobActionExecutionError as exc:
            return jsonify(exc.to_payload()), exc.status_code
        return jsonify(_workspace_response_payload(job, status=result.get("status")))
    task_payload = {
        "force": force_create,
        "name": payload.get("name"),
        "init_script": payload.get("init_script"),
        "sku": payload.get("sku"),
        "region": payload.get("region"),
        "os_image": payload.get("os_image"),
        "public_key_id": payload.get("public_key_id"),
    }
    try:
        task = _enqueue_workspace_task(
            job,
            owner=user,
            action="workspace_create",
            payload=task_payload,
            timeout_sec=timeout_sec,
        )
    except JobActionExecutionError as exc:
        return jsonify(exc.to_payload()), exc.status_code
    task_view = task.to_dict(include_result=False)
    task_view.update(_workspace_task_links(job.job_id, task.task_id))
    return jsonify(_workspace_response_payload(job, status="queued", task=task_view)), 202


def job_tasks(job_id: str) -> Response:
    """Job endpoint for tasks."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    job = manager.get_job(job_id)
    if not job or not _can_view_job(user, job):
        return jsonify({"error": "job not found"}), 404
    limit = _safe_int(request.args.get("limit"), 20)
    include_results = _is_truthy(request.args.get("include_results")) or _can_manage_job(user, job)
    tasks = job_action_manager.list_for_job(job_id, limit=limit, include_results=include_results)
    for task in tasks:
        task.update(_workspace_task_links(job_id, task["task_id"]))
    return jsonify(
        {
            "tasks": tasks,
            "queue_depth": job_action_manager.queue_depth(),
            "inflight": job_action_manager.inflight(),
            "queue_capacity": job_action_manager.max_queue,
            "max_outstanding_per_owner": job_action_manager.max_outstanding_per_owner,
            "max_inflight_per_owner": job_action_manager.max_inflight_per_owner,
            "active_owner_count": job_action_manager.active_owner_count(),
        }
    )


def job_task_detail(job_id: str, task_id: str) -> Response:
    """Job endpoint for task detail."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    job = manager.get_job(job_id)
    if not job or not _can_view_job(user, job):
        return jsonify({"error": "job not found"}), 404
    task = job_action_manager.get_task(task_id)
    if not task or task.job_id != job_id:
        return jsonify({"error": "task_not_found"}), 404
    include_result = _is_truthy(request.args.get("include_result")) or _can_manage_job(user, job)
    payload = task.to_dict(include_result=include_result)
    payload.update(_workspace_task_links(job_id, task_id))
    return jsonify({"task": payload})


def job_task_cancel(job_id: str, task_id: str) -> Response:
    """Job endpoint for task cancel."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    job = manager.get_job(job_id)
    if not job or not _can_view_job(user, job):
        return jsonify({"error": "job not found"}), 404
    if not _can_manage_job(user, job):
        return jsonify({"error": "forbidden"}), 403
    task = job_action_manager.get_task(task_id)
    if not task or task.job_id != job_id:
        return jsonify({"error": "task_not_found"}), 404
    if task.is_terminal() and task.status != "cancelled":
        return jsonify({"error": "task_not_cancellable"}), 409
    if not job_action_manager.cancel(task_id, job_id=job_id):
        return jsonify({"error": "task_not_cancellable"}), 409
    task = job_action_manager.get_task(task_id)
    payload = task.to_dict(include_result=True) if task else {"task_id": task_id, "status": "cancelled"}
    payload.update(_workspace_task_links(job_id, task_id))
    return jsonify({"task": payload})


def job_workspace_open(job_id: str) -> Response:
    """Job endpoint for workspace open."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    job = manager.get_job(job_id)
    if not job or not _can_view_job(user, job):
        return jsonify({"error": "job not found"}), 404
    payload = request.get_json(force=True, silent=True) or {}
    root_id = str(payload.get("root") or "").strip()
    rel_path = payload.get("path")
    if not root_id:
        return jsonify({"error": "root_required"}), 400
    root_entry = _editor_root_by_id(job, root_id)
    if not root_entry:
        return jsonify({"error": "invalid_root"}), 400
    try:
        rel = _editor_normalize_rel(rel_path)
        if not rel:
            return jsonify({"error": "path_required"}), 400
    except ValueError:
        return jsonify({"error": "invalid_path"}), 400
    url, opened_file = _workspace_file_open_url(job, root_entry, rel)
    if not url:
        return jsonify({"error": "workspace_unavailable", "details": "IDE URL not configured."}), 400
    return jsonify({"url": url, "opened_file": opened_file, "path": rel})


def job_editor_roots(job_id: str) -> Response:
    """Job endpoint for editor roots."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    job = manager.get_job(job_id)
    if not job or not _can_view_job(user, job):
        return jsonify({"error": "job not found"}), 404
    if not _editor_allowed(job):
        return jsonify({"roots": [], "enabled": False, "workflow": job.workflow}), 200
    roots = []
    for entry in _editor_root_candidates(job):
        root_path = entry.get("path")
        if not root_path:
            continue
        default_path = _editor_default_path(job, root_path)
        roots.append(
            {
                "id": entry.get("id"),
                "label": entry.get("label"),
                "default_path": default_path,
            }
        )
    return jsonify({"roots": roots, "enabled": True, "workflow": job.workflow})


def job_editor_list(job_id: str) -> Response:
    """Job endpoint for editor list."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    job = manager.get_job(job_id)
    if not job or not _can_view_job(user, job):
        return jsonify({"error": "job not found"}), 404
    if not _editor_allowed(job):
        return jsonify({"error": "editor_disabled"}), 400
    root_id = str(request.args.get("root") or "").strip()
    rel_path = request.args.get("path")
    root_entry = _editor_root_by_id(job, root_id)
    if not root_entry:
        return jsonify({"error": "invalid_root"}), 400
    try:
        rel = _editor_normalize_rel(rel_path)
        entries, truncated = _editor_list_dir(root_entry["path"], rel)
    except FileNotFoundError:
        return jsonify({"error": "not_found"}), 404
    except ValueError:
        return jsonify({"error": "invalid_path"}), 400
    return jsonify({"root": root_id, "path": rel, "entries": entries, "truncated": truncated})


def job_editor_file(job_id: str) -> Response:
    """Job endpoint for editor file."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    job = manager.get_job(job_id)
    if not job or not _can_view_job(user, job):
        return jsonify({"error": "job not found"}), 404
    if not _editor_allowed(job):
        return jsonify({"error": "editor_disabled"}), 400
    if request.method == "GET":
        root_id = str(request.args.get("root") or "").strip()
        rel_path = request.args.get("path")
    else:
        payload = request.get_json(force=True, silent=True) or {}
        root_id = str(payload.get("root") or "").strip()
        rel_path = payload.get("path")
    root_entry = _editor_root_by_id(job, root_id)
    if not root_entry:
        return jsonify({"error": "invalid_root"}), 400
    try:
        rel = _editor_normalize_rel(rel_path)
        if not rel:
            return jsonify({"error": "path_required"}), 400
    except ValueError:
        return jsonify({"error": "invalid_path"}), 400

    if request.method == "GET":
        try:
            data = _editor_read_file(root_entry["path"], rel)
        except FileNotFoundError:
            return jsonify({"error": "not_found"}), 404
        except IsADirectoryError:
            return jsonify({"error": "is_directory"}), 400
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        return jsonify({"root": root_id, "path": rel, **data})

    if not _can_manage_job(user, job):
        return jsonify({"error": "forbidden"}), 403
    content = payload.get("content")
    if not isinstance(content, str):
        return jsonify({"error": "content_required"}), 400
    try:
        data = _editor_write_file(root_entry["path"], rel, content)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    job.append_log(f"Editor saved {rel}.")
    job.persist(force=True)
    return jsonify({"root": root_id, "path": rel, **data})


def job_editor_ops(job_id: str) -> Response:
    """Job endpoint for editor ops."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    job = manager.get_job(job_id)
    if not job or not _can_view_job(user, job):
        return jsonify({"error": "job not found"}), 404
    if not _editor_allowed(job):
        return jsonify({"error": "editor_disabled"}), 400
    if not _can_manage_job(user, job):
        return jsonify({"error": "forbidden"}), 403
    payload = request.get_json(force=True, silent=True) or {}
    action = str(payload.get("action") or "").strip().lower()
    root_id = str(payload.get("root") or "").strip()
    root_entry = _editor_root_by_id(job, root_id)
    if not root_entry:
        return jsonify({"error": "invalid_root"}), 400

    def normalize(value: Optional[str]) -> str:
        return _editor_normalize_rel(value)

    if action == "create":
        rel_path = payload.get("path")
        content = payload.get("content") if isinstance(payload.get("content"), str) else ""
        try:
            rel = normalize(rel_path)
            if not rel:
                return jsonify({"error": "path_required"}), 400
            data = _editor_create_file(root_entry["path"], rel, content)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        job.append_log(f"Editor created {rel}.")
        job.persist(force=True)
        return jsonify({"root": root_id, "path": rel, **data})

    if action == "delete":
        rel_path = payload.get("path")
        force = bool(payload.get("force"))
        try:
            rel = normalize(rel_path)
            if not rel:
                return jsonify({"error": "path_required"}), 400
            data = _editor_delete_path(root_entry["path"], rel, force=force)
        except FileNotFoundError:
            return jsonify({"error": "not_found"}), 404
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        job.append_log(f"Editor deleted {rel}.")
        job.persist(force=True)
        return jsonify({"root": root_id, "path": rel, **data})

    if action == "mkdir":
        rel_path = payload.get("path")
        try:
            rel = normalize(rel_path)
            if not rel:
                return jsonify({"error": "path_required"}), 400
            data = _editor_create_dir(root_entry["path"], rel)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        job.append_log(f"Editor created folder {rel}.")
        job.persist(force=True)
        return jsonify({"root": root_id, "path": rel, **data})

    if action == "rename" or action == "move":
        rel_path = payload.get("path")
        new_path = payload.get("new_path")
        try:
            rel = normalize(rel_path)
            new_rel = normalize(new_path)
            if not rel or not new_rel:
                return jsonify({"error": "path_required"}), 400
            data = _editor_rename_path(root_entry["path"], rel, new_rel)
        except FileNotFoundError:
            return jsonify({"error": "not_found"}), 404
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        job.append_log(f"Editor renamed {rel} -> {new_rel}.")
        job.persist(force=True)
        return jsonify({"root": root_id, "path": new_rel, **data})

    return jsonify({"error": "invalid_action"}), 400


def job_requirements_progress(job_id: str) -> Response:
    """Job endpoint for requirements progress."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    job = manager.get_job(job_id)
    if not job or not _can_view_job(user, job):
        return jsonify({"error": "job not found"}), 404

    payload = {
        "total": 0,
        "completed": 0,
        "in_progress": 0,
        "remaining": 0,
        "status": "pending",
        "source": "none",
        "message": "Awaiting requirements output.",
        "updated_at": None,
    }
    req_path = None
    if isinstance(job.payload, dict):
        req_path = job.payload.get("requirements_path")
    include_global = bool(job.payload.get("include_global_requirements")) if isinstance(job.payload, dict) else False
    solution_path = None
    if isinstance(job.output_paths, dict):
        solution_path = job.output_paths.get("primary")
    if not solution_path or not os.path.exists(solution_path):
        fallback = _requirements_progress_from_requirements_file(req_path)
        if fallback:
            payload.update(fallback)
            payload["status"] = "partial"
            payload["message"] = "Using requirements file totals."
            if include_global:
                global_count = _global_requirements_count()
                if global_count:
                    payload["total"] = _safe_int(payload.get("total")) + global_count
                    payload["remaining"] = _safe_int(payload.get("remaining")) + global_count
                    source = payload.get("source") or "requirements_file"
                    if "global" not in str(source):
                        payload["source"] = f"{source}+global"
                    payload["message"] = f"{payload['message']} Includes global requirements."
        return jsonify(payload)

    solution_mtime = _mtime_dt(solution_path)
    payload["updated_at"] = _iso_from_mtime(solution_path)
    job_active = job.status in {"queued", "running", "paused"}
    job_marker = _parse_timestamp(job.started_at) or _parse_timestamp(job.updated_at)
    if job_active and solution_mtime and job_marker and solution_mtime < job_marker:
        fallback = _requirements_progress_from_requirements_file(req_path)
        if fallback:
            payload.update(fallback)
            payload["status"] = "running"
            payload["message"] = "Job is running; using requirements totals until new output is written."
            source = payload.get("source") or "requirements_file"
            if "stale" not in str(source):
                payload["source"] = f"{source}+stale"
        else:
            payload["status"] = "running"
            payload["message"] = "Job is running; requirements progress will update when output is written."
        return jsonify(payload)

    solution = _read_json_file(solution_path)
    if not isinstance(solution, dict):
        payload["status"] = "unreadable"
        payload["message"] = "Unable to read requirements output."
        fallback = _requirements_progress_from_requirements_file(req_path)
        if fallback:
            payload.update(fallback)
            payload["status"] = "partial"
            payload["message"] = "Using requirements file totals."
        return jsonify(payload)

    progress = _requirements_progress_from_solution(solution)
    payload.update(progress)
    if payload["total"] > 0:
        payload["message"] = ""
        payload["status"] = "ready"
        if payload["source"] == "register":
            payload["status"] = "partial"
            payload["message"] = "Traceability not yet generated; using register totals."
    return jsonify(payload)


def job_requirements_summary(job_id: str) -> Response:
    """Job endpoint for requirements summary."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    job = manager.get_job(job_id)
    if not job or not _can_view_job(user, job):
        return jsonify({"error": "job not found"}), 404
    is_admin = _is_admin_user(user)

    payload = {
        "summary": "",
        "items": [],
        "total": 0,
        "source": "none",
        "updated_at": None,
        "message": "No requirements summary available.",
        "redacted": False,
    }

    req_path = None
    if isinstance(job.payload, dict):
        req_path = job.payload.get("requirements_path")
    include_global = bool(job.payload.get("include_global_requirements")) if isinstance(job.payload, dict) else False

    solution_path = None
    if isinstance(job.output_paths, dict):
        solution_path = job.output_paths.get("primary")

    if solution_path and os.path.exists(solution_path):
        solution = _read_json_file(solution_path)
        payload["updated_at"] = _iso_from_mtime(solution_path)
        if isinstance(solution, dict):
            summary = _requirements_summary_from_register(solution)
            if summary:
                summary = _redact_global_requirements_summary(summary, is_admin)
                payload.update(summary)
                if summary.get("redacted"):
                    payload["redacted"] = True
                payload["message"] = ""
                if req_path and os.path.exists(req_path):
                    text = _read_file_limited(req_path, 80_000)
                    if text:
                        lead = _requirements_summary_from_text(text).get("summary") or ""
                        if lead:
                            payload["summary"] = lead
                return jsonify(payload)
        payload["message"] = "Requirements summary unavailable in output."

    if req_path and os.path.exists(req_path):
        text = _read_file_limited(req_path, 250_000)
        if text:
            summary = _requirements_summary_from_text(text)
            if include_global:
                summary = _append_global_requirements_summary(summary, redact=not is_admin)
            summary = _redact_global_requirements_summary(summary, is_admin)
            payload.update(summary)
            if summary.get("redacted"):
                payload["redacted"] = True
            payload["updated_at"] = _iso_from_mtime(req_path)
            payload["message"] = "" if summary.get("items") else payload["message"]
            return jsonify(payload)

    return jsonify(payload)


def job_logs(job_id: str) -> Response:
    """Job endpoint for logs."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    job = manager.get_job(job_id)
    if not job or not _can_view_job(user, job):
        return jsonify({"error": "job not found"}), 404
    is_admin = _is_admin_user(user)
    tail = request.args.get("tail")
    try:
        tail_count = int(tail) if tail else DEFAULT_TAIL
    except Exception:
        tail_count = DEFAULT_TAIL
    logs = job.get_log_tail(tail_count)
    return jsonify({"logs": _redact_log_entries(logs, is_admin)})


def job_logs_stream(job_id: str) -> Response:
    """Job endpoint for logs stream."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    job = manager.get_job(job_id)
    if not job or not _can_view_job(user, job):
        return jsonify({"error": "job not found"}), 404
    is_admin = _is_admin_user(user)

    def generate():
        q = job.add_listener()
        try:
            while True:
                try:
                    entry = q.get(timeout=1.0)
                except queue.Empty:
                    yield ": keep-alive\n\n"
                    continue
                if not is_admin:
                    redacted_entries = _redact_log_entries([entry], is_admin)
                    if redacted_entries:
                        entry = redacted_entries[0]
                yield _sse(entry)
        finally:
            job.remove_listener(q)

    return Response(generate(), mimetype="text/event-stream", headers={"Cache-Control": "no-cache"})


def job_actions(job_id: str) -> Response:
    """Job endpoint for actions."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    job = manager.get_job(job_id)
    if not job or not _can_view_job(user, job):
        return jsonify({"error": "job not found"}), 404
    if not _can_manage_job(user, job):
        return jsonify({"error": "forbidden"}), 403
    payload = request.get_json(force=True, silent=True) or {}
    action = payload.get("action")
    success = False
    if action == "pause":
        success = manager.pause_job(job_id)
    elif action == "resume":
        success = manager.resume_job(job_id)
    elif action == "stop":
        success = manager.stop_job(job_id)
    elif action == "restart":
        estimate = job.token_estimate or _estimate_job_tokens(job.payload)
        chargeable_estimate = _chargeable_token_estimate(job.owner, job.payload, estimate)
        team_id = _job_team_id(job)
        snapshot = _token_snapshot(job.owner, team_id)
        if chargeable_estimate > snapshot["available"] and job.token_reserved <= 0:
            return (
                jsonify(
                    {
                        "error": "insufficient_tokens",
                        "details": "Insufficient tokens to restart this job.",
                        "estimate": estimate,
                        "chargeable_estimate": chargeable_estimate,
                        "available": snapshot["available"],
                    }
                ),
                402,
            )
        if job.token_reserved <= 0:
            manager._reserve_tokens(job, estimate)
        success = manager.restart_job(job_id)
    else:
        return jsonify({"error": "unknown action"}), 400
    if not success:
        return jsonify({"error": "action failed"}), 409
    return jsonify(_augment_job_dict_for_user(job.to_dict(), user, job))


def job_transfer(job_id: str) -> Response:
    """Job endpoint for transfer."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    job = manager.get_job(job_id)
    if not job:
        return jsonify({"error": "job not found"}), 404
    payload = request.get_json(force=True, silent=True) or {}
    action = str(payload.get("action") or "").strip().lower()
    if not action:
        return jsonify({"error": "action_required"}), 400
    transfer = job.transfer_request if isinstance(job.transfer_request, dict) else None
    if action == "request":
        if job.owner != user and not _is_admin_user(user):
            return jsonify({"error": "forbidden"}), 403
        if transfer and transfer.get("status") == "pending":
            return jsonify({"error": "transfer_pending"}), 409
        team_id = str(payload.get("team_id") or "").strip()
        if not team_id:
            return jsonify({"error": "team_id_required"}), 400
        team = access_store.get_team(team_id)
        if not team:
            return jsonify({"error": "team_not_found"}), 404
        if not _can_access_team_tokens(user, team_id):
            return jsonify({"error": "forbidden"}), 403
        job.transfer_request = {
            "team_id": team_id,
            "team_name": team.get("name"),
            "requested_by": user,
            "requested_at": _now_iso(),
            "status": "pending",
        }
        job.append_log(f"Transfer requested to team {team.get('name') or team_id}.")
        job.persist(force=True)
        _audit_event(
            "job_team_invite",
            actor=user,
            status="success",
            details={"job_id": job.job_id, "team_id": team_id},
        )
        return jsonify(_augment_job_dict_for_user(job.to_dict(), user, job))
    if action in {"cancel", "withdraw"}:
        if job.owner != user and not _is_admin_user(user):
            return jsonify({"error": "forbidden"}), 403
        if not transfer or transfer.get("status") != "pending":
            return jsonify({"error": "transfer_not_pending"}), 409
        transfer.update({"status": "cancelled", "decided_by": user, "decided_at": _now_iso()})
        job.transfer_request = transfer
        job.append_log("Transfer request cancelled.")
        job.persist(force=True)
        _audit_event(
            "job_team_invite_cancelled",
            actor=user,
            status="success",
            details={"job_id": job.job_id, "team_id": transfer.get("team_id")},
        )
        return jsonify(_augment_job_dict_for_user(job.to_dict(), user, job))
    if action in {"accept", "decline"}:
        if not transfer or transfer.get("status") != "pending":
            return jsonify({"error": "transfer_not_pending"}), 409
        team_id = transfer.get("team_id")
        if not (_is_admin_user(user) or _is_team_leader(user, team_id)):
            return jsonify({"error": "forbidden"}), 403
        if action == "decline":
            transfer.update({"status": "declined", "decided_by": user, "decided_at": _now_iso()})
            job.transfer_request = transfer
            job.append_log("Transfer request declined.")
            job.persist(force=True)
            _audit_event(
                "job_team_invite_declined",
                actor=user,
                status="success",
                details={"job_id": job.job_id, "team_id": team_id},
            )
            return jsonify(_augment_job_dict_for_user(job.to_dict(), user, job))
        if job.status in {"running"}:
            return jsonify({"error": "job_active", "details": "Pause or stop the job before transferring."}), 409
        team = access_store.get_team(team_id) if team_id else None
        if not team:
            return jsonify({"error": "team_not_found"}), 404
        project_id = str(payload.get("project_id") or "").strip()
        if project_id:
            project = access_store.get_project(project_id)
            if not project or project.get("team_id") != team_id:
                return jsonify({"error": "project_invalid"}), 400
            if not (_is_admin_user(user) or access_store.can_manage_project(user, project_id)):
                return jsonify({"error": "forbidden"}), 403
        estimate = job.token_estimate or _estimate_job_tokens(job.payload)
        chargeable_estimate = _chargeable_token_estimate(job.owner, job.payload, estimate)
        snapshot = _token_snapshot(job.owner, team_id)
        if chargeable_estimate > snapshot["available"]:
            return (
                jsonify(
                    {
                        "error": "insufficient_tokens",
                        "details": "Insufficient tokens to transfer this job.",
                        "estimate": estimate,
                        "chargeable_estimate": chargeable_estimate,
                        "available": snapshot["available"],
                    }
                ),
                402,
            )
        if job.token_reserved:
            manager._release_tokens(job, reason="transfer")
        if not isinstance(job.payload, dict):
            job.payload = {}
        job.payload.pop("token_scope", None)
        job.payload["team_id"] = team_id
        if project_id:
            job.payload["project_id"] = project_id
            job.payload.pop("project", None)
            job.project_name = project.get("name") if project else job.project_name
        transfer.update(
            {
                "status": "accepted",
                "decided_by": user,
                "decided_at": _now_iso(),
                "team_name": team.get("name") if team else transfer.get("team_name"),
            }
        )
        job.transfer_request = transfer
        job.append_log(f"Transfer accepted to team {team.get('name') or team_id}.")
        if estimate > 0:
            manager._reserve_tokens(job, estimate)
        job.persist(force=True)
        _audit_event(
            "job_team_invite_accepted",
            actor=user,
            status="success",
            details={"job_id": job.job_id, "team_id": team_id, "project_id": project_id or None},
        )
        return jsonify(_augment_job_dict_for_user(job.to_dict(), user, job))
    if action in {"assign_user", "demote"}:
        team_id = _job_team_id(job)
        if not team_id:
            return jsonify({"error": "not_team_job"}), 400
        if not (_is_admin_user(user) or _is_team_leader(user, team_id)):
            return jsonify({"error": "forbidden"}), 403
        target_user = str(payload.get("target_user") or "").strip()
        if not target_user:
            return jsonify({"error": "target_user_required"}), 400
        if not access_store._team_role(target_user, team_id):
            return jsonify({"error": "target_not_in_team"}), 400
        if job.status in {"running"}:
            return jsonify({"error": "job_active", "details": "Pause or stop the job before reassigning."}), 409
        estimate = job.token_estimate or _estimate_job_tokens(job.payload)
        chargeable_estimate = _chargeable_token_estimate(target_user, job.payload, estimate)
        snapshot = _token_snapshot(target_user)
        if chargeable_estimate > snapshot["available"]:
            return (
                jsonify(
                    {
                        "error": "insufficient_tokens",
                        "details": "Target user does not have enough personal tokens.",
                        "estimate": estimate,
                        "chargeable_estimate": chargeable_estimate,
                        "available": snapshot["available"],
                    }
                ),
                402,
            )
        if job.token_reserved:
            manager._release_tokens(job, reason="assign_user")
        if not isinstance(job.payload, dict):
            job.payload = {}
        job.payload.pop("team_id", None)
        job.payload.pop("project_id", None)
        job.payload.pop("project", None)
        job.payload["owner"] = target_user
        job.owner = target_user
        job.transfer_request = None
        job.append_log(f"Job assigned to {target_user}.")
        if estimate > 0:
            manager._reserve_tokens(job, estimate)
        job.persist(force=True)
        _audit_event(
            "job_assign_user",
            actor=user,
            status="success",
            details={"job_id": job.job_id, "target": target_user, "previous_team": team_id},
        )
        return jsonify(_augment_job_dict_for_user(job.to_dict(), user, job))
    return jsonify({"error": "invalid_action"}), 400


def job_archive(job_id: str) -> Response:
    """Job endpoint for archive."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    job = manager.get_job(job_id)
    if not job or not _can_view_job(user, job):
        return jsonify({"error": "job not found"}), 404
    if not _can_manage_job(user, job):
        return jsonify({"error": "forbidden"}), 403
    payload = request.get_json(force=True, silent=True) or {}
    archived = bool(payload.get("archived", True))
    stop = bool(payload.get("stop"))
    if archived and job.status in {"queued", "running", "paused"} and not stop:
        return jsonify({"error": "job_active", "details": "Stop the job before archiving."}), 409
    if archived and job.status in {"queued", "running", "paused"}:
        manager.stop_job(job_id)
    if not manager.set_archived(job_id, archived):
        return jsonify({"error": "archive_failed"}), 409
    return jsonify(_augment_job_dict_for_user(job.to_dict(), user, job))


def jobs_bulk_delete() -> Response:
    """Handle the jobs bulk delete route."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    payload = request.get_json(force=True, silent=True) or {}
    scope = (payload.get("scope") or "queue").strip().lower()
    if scope not in {"queue", "archive"}:
        return jsonify({"error": "invalid_scope"}), 400
    stop = bool(payload.get("stop"))
    target_archived = scope == "archive"
    jobs_source = manager.list_jobs()
    jobs_list = [
        job
        for job in jobs_source
        if bool(job.archived) == target_archived and _can_manage_job(user, job)
    ]
    active_jobs = [job for job in jobs_list if job.status in {"queued", "running", "paused"}]
    if active_jobs and not stop:
        return jsonify({"error": "job_active", "details": "Stop active jobs before deleting."}), 409
    deleted: List[str] = []
    for job in jobs_list:
        if manager.delete_job(job.job_id, owner=None, stop_if_active=stop):
            deleted.append(job.job_id)
    return jsonify({"deleted": deleted, "count": len(deleted)})


def tokens() -> Response:
    """Token-balance management endpoint."""
    if _billing_enabled():
        project_id = (request.args.get("project_id") or request.args.get("project") or "").strip()
        team_id = (request.args.get("team_id") or "").strip()
        if not project_id and not team_id:
            return _proxy_service_request(BILLING_API_BASE, BILLING_TIMEOUT)
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    if request.method == "GET":
        project_id = (request.args.get("project_id") or request.args.get("project") or "").strip()
        team_id = (request.args.get("team_id") or "").strip() or None
        if project_id:
            project = access_store.get_project(project_id)
            if not project:
                return jsonify({"error": "project_not_found"}), 404
            if not (_is_admin_user(user) or access_store.can_view_project(user, project_id)):
                return jsonify({"error": "forbidden", "details": "Project access denied."}), 403
            if not team_id:
                team_id = project.get("team_id")
        if team_id and not _can_access_team_tokens(user, team_id):
            return jsonify({"error": "forbidden", "details": "Team access denied."}), 403
        snapshot = _token_snapshot(user, team_id or None)
        return jsonify(snapshot)

    payload = request.get_json(force=True, silent=True) or {}
    action = (payload.get("action") or "review").strip().lower()
    username = (payload.get("username") or user).strip()
    if action != "grant" and username and username != user:
        return jsonify({"error": "invalid_user", "details": "Username mismatch."}), 403

    snapshot = _token_snapshot(user)
    if action == "review":
        return jsonify(snapshot)

    if action in {"add", "cashout", "grant"}:
        password = (payload.get("password") or "").strip()
        if not password or not user_store.verify(user, password):
            return jsonify({"error": "invalid_credentials", "details": "Password verification failed."}), 401

    if action == "add":
        tokens_raw = payload.get("token_amount")
        btc_raw = payload.get("btc_amount") or payload.get("btc_value")
        tokens = 0
        btc_value = None
        if tokens_raw not in (None, ""):
            try:
                tokens = int(float(tokens_raw))
            except Exception:
                tokens = 0
        elif btc_raw not in (None, ""):
            try:
                btc_value = float(btc_raw)
                tokens = int(round(btc_value / TOKEN_BTC_RATE))
            except Exception:
                tokens = 0
        if tokens <= 0:
            return jsonify({"error": "invalid_amount", "details": "Token amount must be positive."}), 400
        meta = {
            "tokens": tokens,
            "btc_amount": btc_value,
            "btc_rate": TOKEN_BTC_RATE,
            "btc_txid": payload.get("btc_txid"),
            "btc_address": payload.get("btc_address"),
            "source": payload.get("source") or "portal",
            "currency": payload.get("settlement_currency") or "GBP",
            "payment_provider": payload.get("payment_provider"),
            "payment_channel": payload.get("payment_channel"),
            "payment_method": payload.get("payment_method"),
            "checkout_flow": payload.get("checkout_flow"),
            "payment_id": payload.get("payment_id") or payload.get("checkout_id"),
        }
        payment_request_id = str(meta.get("payment_id") or meta.get("btc_txid") or "").strip() or _chain_request_id(
            "portal-topup", user, uuid.uuid4().hex
        )
        _capture_payment_event(user, tokens, meta, request_id=payment_request_id)
        _audit_event("tokens_topup", actor=user, status="success", details={"amount": tokens})
        snapshot = _token_snapshot(user)
        return jsonify({"message": "Tokens added.", **snapshot})

    if action == "cashout":
        tokens_raw = payload.get("token_amount")
        try:
            tokens = int(float(tokens_raw))
        except Exception:
            tokens = 0
        if tokens <= 0:
            return jsonify({"error": "invalid_amount", "details": "Token amount must be positive."}), 400
        if tokens > snapshot.get("paid_balance", snapshot["balance"]):
            return jsonify({"error": "insufficient_tokens", "details": "Not enough tokens to cash out."}), 409
        meta = {
            "tokens": tokens,
            "btc_address": payload.get("btc_address"),
            "source": payload.get("source") or "portal",
        }
        _record_token_event(
            "user",
            user,
            "cashout",
            -tokens,
            meta,
            request_id=str(payload.get("payout_reference") or payload.get("settlement_reference") or "").strip()
            or _chain_request_id("cashout", user, uuid.uuid4().hex),
        )
        _audit_event("tokens_cashout", actor=user, status="success", details={"amount": tokens})
        snapshot = _token_snapshot(user)
        return jsonify({"message": "Cashout recorded.", **snapshot})

    if action == "grant":
        if not _is_admin_user(user):
            return jsonify({"error": "forbidden", "details": "Admin role required."}), 403
        target_user = (payload.get("target_user") or payload.get("recipient") or payload.get("username") or "").strip()
        if not target_user:
            return jsonify({"error": "target_required", "details": "Target user is required."}), 400
        tokens_raw = payload.get("token_amount")
        try:
            tokens = int(float(tokens_raw))
        except Exception:
            tokens = 0
        if tokens <= 0:
            return jsonify({"error": "invalid_amount", "details": "Token amount must be positive."}), 400
        meta = {
            "tokens": tokens,
            "granted_by": user,
            "note": payload.get("note") or payload.get("reason"),
            "source": payload.get("source") or "admin",
        }
        _record_token_event(
            "user",
            target_user,
            "grant",
            tokens,
            meta,
            request_id=_chain_request_id("grant", target_user, uuid.uuid4().hex),
        )
        _audit_event(
            "tokens_grant",
            actor=user,
            status="success",
            details={"amount": tokens, "target": target_user},
        )
        snapshot = _token_snapshot(target_user)
        return jsonify({"message": "Free tokens granted.", "target": target_user, **snapshot})

    if action == "debit":
        tokens_raw = payload.get("token_amount")
        try:
            tokens = int(float(tokens_raw))
        except Exception:
            tokens = 0
        if tokens <= 0:
            return jsonify({"error": "invalid_amount", "details": "Token amount must be positive."}), 400
        meta = {
            "tokens": tokens,
            "source": payload.get("source") or "api",
            "note": payload.get("note") or payload.get("reason"),
            "operation": payload.get("operation"),
            "workspace_id": payload.get("workspace_id"),
        }
        request_id = str(payload.get("request_id") or payload.get("operation_id") or "").strip() or _chain_request_id(
            "debit",
            user,
            uuid.uuid4().hex,
        )
        entry = _record_token_event(
            "user",
            user,
            "debit",
            -tokens,
            meta,
            request_id=request_id,
        )
        shortfall = int(entry.get("shortfall") or 0)
        used_total = abs(int(entry.get("delta") or 0))
        if shortfall > 0 or used_total < tokens:
            _audit_event(
                "tokens_debit",
                actor=user,
                status="failed",
                details={"requested": tokens, "used": used_total, "shortfall": max(shortfall, tokens - used_total)},
            )
            snapshot = _token_snapshot(user)
            return (
                jsonify(
                    {
                        "error": "insufficient_tokens",
                        "details": "Not enough tokens to debit.",
                        "requested": tokens,
                        "used": used_total,
                        "shortfall": max(shortfall, tokens - used_total),
                        **snapshot,
                    }
                ),
                409,
            )
        _audit_event("tokens_debit", actor=user, status="success", details={"amount": tokens})
        snapshot = _token_snapshot(user)
        return jsonify({"message": "Tokens debited.", "entry": entry, **snapshot})

    if action == "refund":
        tokens_raw = payload.get("token_amount")
        try:
            tokens = int(float(tokens_raw))
        except Exception:
            tokens = 0
        if tokens <= 0:
            return jsonify({"error": "invalid_amount", "details": "Token amount must be positive."}), 400
        meta = {
            "tokens": tokens,
            "source": payload.get("source") or "api",
            "note": payload.get("note") or payload.get("reason"),
            "operation": payload.get("operation"),
            "workspace_id": payload.get("workspace_id"),
        }
        request_id = str(payload.get("request_id") or payload.get("operation_id") or "").strip() or _chain_request_id(
            "refund",
            user,
            uuid.uuid4().hex,
        )
        entry = _record_token_event(
            "user",
            user,
            "refund",
            tokens,
            meta,
            request_id=request_id,
        )
        _audit_event("tokens_refund", actor=user, status="success", details={"amount": tokens})
        snapshot = _token_snapshot(user)
        return jsonify({"message": "Tokens refunded.", "entry": entry, **snapshot})

    if action == "sync":
        target = payload.get("balance")
        if target is None:
            return jsonify({"error": "balance_required"}), 400
        try:
            target_balance = int(float(target))
        except Exception:
            return jsonify({"error": "invalid_balance"}), 400
        if target_balance < 0:
            return jsonify({"error": "invalid_balance"}), 400
        capacity = payload.get("capacity") or payload.get("last_topup_tokens")
        try:
            capacity_val = int(float(capacity)) if capacity not in (None, "") else None
        except Exception:
            capacity_val = None
        target_paid = payload.get("paid_balance")
        target_free = payload.get("free_balance")
        try:
            target_paid_val = int(float(target_paid)) if target_paid not in (None, "") else None
        except Exception:
            target_paid_val = None
        try:
            target_free_val = int(float(target_free)) if target_free not in (None, "") else None
        except Exception:
            target_free_val = None
        delta = target_balance - snapshot["balance"]
        status = "matched" if delta == 0 else "adjusted"
        _record_token_event(
            "user",
            user,
            "sync",
            delta,
            {
                "target_balance": target_balance,
                "capacity": capacity_val,
                "source": payload.get("source") or "portal",
                "sync_user": payload.get("user") or user,
                "sync_role": payload.get("role"),
                "target_paid_balance": target_paid_val,
                "target_free_balance": target_free_val,
            },
            request_id=None,
        )
        _audit_event("tokens_sync", actor=user, status=status, details={"delta": delta})
        snapshot = _token_snapshot(user)
        return jsonify({"message": "Sync complete.", "status": status, **snapshot})

    return jsonify({"error": "invalid_action"}), 400


def tokens_ledger() -> Response:
    """Return token ledger history entries."""
    if _billing_enabled():
        return _proxy_service_request(BILLING_API_BASE, BILLING_TIMEOUT)
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    target = (request.args.get("user") or "").strip()
    if target:
        if not _is_admin_user(user):
            return jsonify({"error": "forbidden"}), 403
        user = target
    limit = request.args.get("limit")
    try:
        limit_val = int(limit) if limit else 50
    except Exception:
        limit_val = 50
    entries = _account_entries("user", user, limit=limit_val)
    return jsonify({"entries": entries})


def billing_dashboard() -> Response:
    """Proxy the customer billing dashboard through the public Refiner origin."""
    if not _billing_enabled():
        return make_response("Billing dashboard unavailable.", 503)
    return _proxy_service_request(BILLING_API_BASE, BILLING_TIMEOUT)


def billing_dashboard_admin() -> Response:
    """Proxy the admin billing dashboard through the public Refiner origin."""
    if not _billing_enabled():
        return make_response("Billing dashboard unavailable.", 503)
    return _proxy_service_request(BILLING_API_BASE, BILLING_TIMEOUT)


def billing_dashboard_asset(filename: str) -> Response:
    """Proxy Billing dashboard assets so the browser stays on the public API host."""
    _ = filename
    if not _billing_enabled():
        return make_response("Billing dashboard unavailable.", 503)
    return _proxy_service_request(BILLING_API_BASE, BILLING_TIMEOUT)


def billing_dashboard_customer() -> Response:
    """Proxy customer dashboard JSON so the website uses a single backend origin."""
    if not _billing_enabled():
        return jsonify({"error": "billing_unavailable"}), 503
    return _proxy_service_request(BILLING_API_BASE, BILLING_TIMEOUT)


def billing_dashboard_admin_data() -> Response:
    """Proxy admin dashboard JSON so billing telemetry stays under the public API host."""
    if not _billing_enabled():
        return jsonify({"error": "billing_unavailable"}), 503
    return _proxy_service_request(BILLING_API_BASE, BILLING_TIMEOUT)


def secrets() -> Response:
    """Handle the secrets route."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    store = _get_secret_store(user)
    if request.method == "GET":
        return jsonify({"secrets": store.list_masked()})
    payload = request.get_json(force=True, silent=True) or {}
    name = str(payload.get("name") or "").strip()
    value = str(payload.get("value") or "").strip()
    if not name or not value:
        return jsonify({"error": "name and value are required"}), 400
    if not SECRET_NAME_RE.match(name):
        return jsonify({"error": "invalid secret name"}), 400
    store.set(name, value)
    _audit_event("secret_set", actor=user, status="success", details={"name": name})
    return jsonify({"name": name, "masked": SecretStore._mask(value), "updated_at": _now_iso()})


def delete_secret(name: str) -> Response:
    """Handle the delete secret route."""
    if not SECRET_NAME_RE.match(name):
        return jsonify({"error": "invalid secret name"}), 400
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    store = _get_secret_store(user)
    deleted = store.delete(name)
    if not deleted:
        return jsonify({"error": "secret not found"}), 404
    _audit_event("secret_delete", actor=user, status="success", details={"name": name})
    return jsonify({"status": "deleted"})


def github_tree() -> Response:
    """Handle the github tree route."""
    user = _current_user()
    if not user:
        return jsonify({"error": "unauthorized"}), 401
    payload = request.get_json(force=True, silent=True) or {}
    repo_input = str(payload.get("repo_url") or payload.get("repo") or "").strip()
    branch = str(payload.get("branch") or "").strip()
    if not repo_input:
        return jsonify({"error": "repo is required"}), 400
    owner, repo = JobManager._parse_repo_input(repo_input)
    if not owner or not repo:
        return jsonify({"error": "invalid repo input"}), 400
    token = _get_github_api_token(user)
    headers = {"Accept": "application/vnd.github+json"}
    if token:
        headers["Authorization"] = f"token {token}"

    repo_url = f"https://api.github.com/repos/{owner}/{repo}"
    try:
        repo_resp = _http_request_with_retry("GET", repo_url, headers=headers, timeout=20)
    except requests.RequestException as exc:
        return jsonify({"error": "repo_lookup_failed", "details": str(exc)}), 502
    if repo_resp.status_code != 200:
        return jsonify({"error": "repo lookup failed", "details": repo_resp.text}), 400
    repo_data = repo_resp.json()
    default_branch = repo_data.get("default_branch") or "main"
    branch = branch or default_branch

    tree_url = f"https://api.github.com/repos/{owner}/{repo}/git/trees/{branch}"
    try:
        tree_resp = _http_request_with_retry(
            "GET",
            tree_url,
            headers=headers,
            params={"recursive": "1"},
            timeout=30,
        )
    except requests.RequestException as exc:
        return jsonify({"error": "tree_lookup_failed", "details": str(exc)}), 502
    if tree_resp.status_code != 200:
        return jsonify({"error": "tree lookup failed", "details": tree_resp.text}), 400
    tree_data = tree_resp.json()
    items = tree_data.get("tree") or []
    max_entries = int(payload.get("max_entries") or 4000)
    trimmed = items[:max_entries]
    response_items = [
        {
            "path": item.get("path"),
            "type": item.get("type"),
            "size": item.get("size"),
        }
        for item in trimmed
        if item.get("path") and item.get("type") in {"blob", "tree"}
    ]
    return jsonify({"items": response_items, "branch": branch, "owner": owner, "repo": repo})


REQ_DRAFT_LINE_RE = re.compile(r"^\s*(?:[-*+]\s*)?(REQ-\d{3,})\s*(?:[:\-–]\s*)?(.*)$", re.I)
REQ_DRAFT_HEADING_RE = re.compile(r"^\s*#{1,6}\s+(.*)$")
REQ_DRAFT_BULLET_RE = re.compile(r"^\s*(?:[-*+]|\d+\.)\s+(.*)$")


def _has_req_draft_lines(text: str) -> bool:
    for line in text.splitlines():
        if REQ_DRAFT_LINE_RE.match(line):
            return True
    return False


def _strip_req_register(text: str) -> str:
    lines = text.splitlines()
    output: List[str] = []
    skipping = False
    for line in lines:
        heading = REQ_DRAFT_HEADING_RE.match(line)
        if heading:
            title = heading.group(1).strip().lower()
            if title.startswith("requirements register"):
                skipping = True
                continue
            if skipping:
                skipping = False
        if not skipping:
            output.append(line)
    return "\n".join(output).strip()


def _collect_req_sections(text: str) -> Dict[str, List[str]]:
    sections: Dict[str, List[str]] = {}
    current = None
    for line in text.splitlines():
        heading = REQ_DRAFT_HEADING_RE.match(line)
        if heading:
            current = heading.group(1).strip().lower()
            sections.setdefault(current, [])
            continue
        if current is not None:
            sections[current].append(line)
    return sections


def _extract_bullets(lines: List[str]) -> List[str]:
    items: List[str] = []
    for line in lines:
        match = REQ_DRAFT_BULLET_RE.match(line)
        if match:
            item = match.group(1).strip()
            if item:
                items.append(item)
    return items


def _extract_requirement_candidates(text: str) -> List[str]:
    sections = _collect_req_sections(text)
    preferred_sections = [
        key
        for key in sections
        if "functional requirements" in key or "non-functional requirements" in key
    ]
    candidates: List[str] = []
    for key in preferred_sections:
        candidates.extend(_extract_bullets(sections.get(key, [])))
    if not candidates:
        candidates = _extract_bullets(text.splitlines())
    if not candidates:
        content = ""
        for key in preferred_sections:
            content += " ".join(sections.get(key, [])) + " "
        if not content.strip():
            content = text
        content = re.sub(r"\s+", " ", content).strip()
        if content:
            for sentence in re.split(r"(?<=[.!?])\s+", content):
                cleaned = sentence.strip()
                if cleaned:
                    candidates.append(cleaned)
                if len(candidates) >= 8:
                    break
    seen = set()
    unique: List[str] = []
    for item in candidates:
        key = item.lower()
        if key in seen:
            continue
        seen.add(key)
        unique.append(item)
        if len(unique) >= 12:
            break
    return unique


def _build_requirements_register(items: List[str]) -> str:
    lines = ["## Requirements Register"]
    for idx, item in enumerate(items, start=1):
        lines.append(f"- REQ-{idx:03d}: {item}")
    return "\n".join(lines)


def _ensure_req_register_in_draft(text: str) -> str:
    if _has_req_draft_lines(text):
        return text
    candidates = _extract_requirement_candidates(text)
    if not candidates:
        candidates = ["Review draft and extract requirements from the sections above."]
    register = _build_requirements_register(candidates)
    stripped = _strip_req_register(text)
    if not stripped:
        return register
    return f"{stripped}\n\n{register}"


def _is_simple_greeting(text: str) -> bool:
    if not isinstance(text, str):
        return False
    cleaned = re.sub(r"[^a-z0-9\s]", " ", text.lower())
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    if not cleaned:
        return False
    greeting_set = {
        "hi",
        "hello",
        "hey",
        "hi there",
        "hello there",
        "hey there",
        "good morning",
        "good afternoon",
        "good evening",
    }
    return cleaned in greeting_set


def _is_marketing_assistant_request(payload: Dict[str, Any], requirements_text: str) -> bool:
    assistant_profile = str(payload.get("assistant_profile") or "").strip().lower()
    if assistant_profile in {"marketing", "neuralmimicry_marketing", "nm_marketing"}:
        return True
    lowered_context = (requirements_text or "").strip().lower()
    if not lowered_context:
        return False
    if "neuralmimicry marketing assistant" in lowered_context:
        return True
    if "questions about neuralmimicry" in lowered_context and "products" in lowered_context and "services" in lowered_context:
        return True
    return False


def _assistant_memory_store(user: str) -> SolverEpisodeStore:
    """Return the per-user episodic memory store for assistant-style routes."""

    ensure_dir_permissions(ASSISTANT_MEMORY_ROOT, mode=0o700)
    try:
        max_entries = max(20, min(int(os.getenv("REFINER_ASSISTANT_MEMORY_MAX_ENTRIES", "240")), 1000))
    except Exception:
        max_entries = 240
    try:
        compact_every = max(1, min(int(os.getenv("REFINER_ASSISTANT_MEMORY_COMPACT_EVERY", "20")), 200))
    except Exception:
        compact_every = 20
    user_key = hash_identifier(user or "") or "anonymous"
    return SolverEpisodeStore(
        os.path.join(ASSISTANT_MEMORY_ROOT, f"{user_key}.jsonl"),
        max_entries=max_entries,
        compact_every=compact_every,
    )


def _central_assistant_episode_store() -> Optional[Any]:
    """Return the shared assistant episode store when central metadata is enabled."""

    if CENTRAL_STORE is None:
        return None
    return getattr(CENTRAL_STORE, "assistant_episodes", None)


def _assistant_memory_scope(route: str, *, mode: str = "", profile: str = "") -> str:
    """Compatibility wrapper for the extracted assistant memory scope helper."""

    return _assistant_memory_scope_impl(route, mode=mode, profile=profile)


def _assistant_memory_query_text(
    *,
    prompt: str = "",
    requirements_text: str = "",
    messages: Optional[List[Dict[str, Any]]] = None,
    extra_parts: Optional[List[str]] = None,
) -> str:
    """Compatibility wrapper for the extracted assistant memory query helper."""

    return _assistant_memory_query_text_impl(
        prompt=prompt,
        requirements_text=requirements_text,
        messages=messages,
        extra_parts=extra_parts,
    )


def _assistant_memory_matches(
    user: str,
    *,
    scope: str,
    query_text: str,
    limit: int,
) -> List[SolverEpisode]:
    """Compatibility wrapper for the extracted assistant memory lookup helper."""

    return _assistant_memory_matches_impl(
        user,
        scope=scope,
        query_text=query_text,
        limit=limit,
        episode_store_for_user=_assistant_memory_store,
        get_central_store=_central_assistant_episode_store,
        logger=logger,
    )


def _assistant_memory_entry_line(entry: SolverEpisode) -> str:
    """Compatibility wrapper for the extracted assistant memory render helper."""

    return _assistant_memory_entry_line_impl(entry)


def _assistant_memory_prompt_block(
    user: str,
    *,
    scope: str,
    query_text: str,
    header: str,
    limit: int = 3,
    max_chars: int = 1400,
) -> str:
    """Compatibility wrapper for the extracted assistant prompt memory helper."""

    return _assistant_memory_prompt_block_impl(
        user,
        scope=scope,
        query_text=query_text,
        header=header,
        episode_store_for_user=_assistant_memory_store,
        get_central_store=_central_assistant_episode_store,
        logger=logger,
        limit=limit,
        max_chars=max_chars,
    )


def _assistant_memory_reference_payload(
    user: str,
    *,
    scope: str,
    query_text: str,
    limit: int = 3,
) -> List[Dict[str, Any]]:
    """Compatibility wrapper for the extracted assistant memory reference helper."""

    return _assistant_memory_reference_payload_impl(
        user,
        scope=scope,
        query_text=query_text,
        episode_store_for_user=_assistant_memory_store,
        get_central_store=_central_assistant_episode_store,
        logger=logger,
        limit=limit,
    )


def _assistant_memory_requirement_ids(text: str, limit: int = 12) -> List[str]:
    """Compatibility wrapper for the extracted assistant requirement-id helper."""

    return _assistant_memory_requirement_ids_impl(text, limit=limit)


def _assistant_memory_summary(text: str, *, fallback: str = "", max_chars: int = 260) -> str:
    """Compatibility wrapper for the extracted assistant summary helper."""

    return _assistant_memory_summary_impl(text, fallback=fallback, max_chars=max_chars)


def _record_assistant_memory(
    user: str,
    *,
    scope: str,
    prompt_text: str,
    requirements_text: str,
    reply_text: str,
    project_name: str = "",
    steps: Optional[List[str]] = None,
    extra_notes: Optional[List[str]] = None,
    metadata: Optional[Dict[str, Any]] = None,
) -> None:
    """Compatibility wrapper for the extracted assistant write helper."""

    _record_assistant_memory_impl(
        user,
        scope=scope,
        prompt_text=prompt_text,
        requirements_text=requirements_text,
        reply_text=reply_text,
        episode_store_for_user=_assistant_memory_store,
        get_central_store=_central_assistant_episode_store,
        logger=logger,
        project_name=project_name,
        steps=steps,
        extra_notes=extra_notes,
        metadata=metadata,
    )


def _should_use_assistant_ask_memory(
    prompt: str,
    *,
    requirements_text: str,
    messages: Optional[List[Dict[str, Any]]],
    is_marketing_assistant: bool,
) -> bool:
    """Compatibility wrapper for the extracted assistant memory policy helper."""

    return _should_use_assistant_ask_memory_impl(
        prompt,
        requirements_text=requirements_text,
        messages=messages,
        is_marketing_assistant=is_marketing_assistant,
    )


def _assistant_reply_payload(
    reply_text: str,
    provider: str,
    model: str,
    payload: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    gesture_mode, avatar_mode, office_flag = _stt_motion_context(payload)
    response_payload: Dict[str, Any] = {
        "reply": reply_text,
        "provider": provider,
        "model": model,
        "gesture_mode": gesture_mode,
        "avatar_mode": avatar_mode,
    }
    if not STT_GESTURE_ENABLED:
        return response_payload
    if STT_GESTURE_NMSTT_FALLBACK and STT_BACKEND == "server" and STT_SERVER_URL:
        try:
            nmstt_payload = _run_nmstt_gesture_plan(
                reply_text,
                gesture_mode=gesture_mode,
                avatar_mode=avatar_mode,
                office_mode=office_flag,
            )
            if isinstance(nmstt_payload, dict):
                response_payload.update(nmstt_payload)
                response_payload["gesture_mode"] = sanitize_gesture_mode(
                    response_payload.get("gesture_mode"),
                    default_mode=gesture_mode,
                    bsl_enabled=STT_BSL_ENABLED,
                )
                response_payload["avatar_mode"] = sanitize_avatar_mode(
                    response_payload.get("avatar_mode"),
                    office_mode=office_flag,
                    default_mode=avatar_mode,
                )
                return response_payload
        except Exception as exc:
            logger.debug("Assistant nmstt gesture planning skipped: %s", exc)
    try:
        motion_payload = plan_stt_avatar_motion(
            reply_text,
            gesture_mode=gesture_mode,
            avatar_mode=avatar_mode,
            bsl_enabled=STT_BSL_ENABLED,
        )
        if isinstance(motion_payload, dict):
            response_payload.update(motion_payload)
            response_payload["gesture_mode"] = sanitize_gesture_mode(
                response_payload.get("gesture_mode"),
                default_mode=gesture_mode,
                bsl_enabled=STT_BSL_ENABLED,
            )
            response_payload["avatar_mode"] = sanitize_avatar_mode(
                response_payload.get("avatar_mode"),
                office_mode=office_flag,
                default_mode=avatar_mode,
            )
    except Exception as exc:
        logger.debug("Assistant gesture planning skipped: %s", exc)
    return response_payload


def _assistant_pipeline_rag_config() -> Dict[str, Any]:
    return {
        "max_docs": RAG_MAX_DOCS,
        "max_doc_bytes": RAG_MAX_DOC_BYTES,
        "default_chunk_size": RAG_DEFAULT_CHUNK_SIZE,
        "default_chunk_overlap": RAG_DEFAULT_CHUNK_OVERLAP,
        "default_max_chunks": RAG_DEFAULT_MAX_CHUNKS,
        "async_index_builds": RAG_ASYNC_INDEX_BUILDS,
        "build_timeout_sec": RAG_BUILD_TIMEOUT_SEC,
    }


def _assistant_pipeline_runtime_config() -> Dict[str, Any]:
    return {
        "request_capacity": _ASSISTANT_REQUEST_CAPACITY,
        "capacity_wait_sec": ASSISTANT_CAPACITY_WAIT_SEC,
    }


def _assistant_pipeline_security_config() -> Dict[str, Any]:
    return {
        "policy_enabled": ASSISTANT_SECURITY_POLICY_ENABLED,
        "strict_message_roles": ASSISTANT_SECURITY_STRICT_MESSAGE_ROLES,
        "block_prompt_leak_requests": ASSISTANT_SECURITY_BLOCK_PROMPT_LEAK,
        "admin_only_mcp": ASSISTANT_MCP_ADMIN_ONLY,
        "block_unsafe_tool_requests": ASSISTANT_BLOCK_UNSAFE_TOOL_REQUESTS,
        "validate_rag_source_urls": ASSISTANT_SECURITY_VALIDATE_RAG_SOURCE_URLS,
        "redact_output_pii": ASSISTANT_OUTPUT_REDACT_PII,
        "validate_output_shapes": ASSISTANT_OUTPUT_VALIDATE_SHAPES,
    }


def _assistant_pipeline_routing_config() -> Dict[str, Any]:
    return {
        "enabled": ASSISTANT_INTENT_ROUTING_ENABLED,
        "skill_hint_limit": ASSISTANT_ROUTING_SKILL_HINT_LIMIT,
        "capability_hint_max_items": ASSISTANT_ROUTING_CAPABILITY_MAX_ITEMS,
    }


def _assistant_pipeline_cache_config() -> Dict[str, Any]:
    return {
        "enabled": ASSISTANT_SEMANTIC_CACHE_ENABLED,
        "ttl_hours": ASSISTANT_SEMANTIC_CACHE_TTL_HOURS,
        "min_similarity": ASSISTANT_SEMANTIC_CACHE_MIN_SIMILARITY,
        "max_candidates": ASSISTANT_SEMANTIC_CACHE_MAX_CANDIDATES,
    }


def _assistant_pipeline_retrieval_config() -> Dict[str, Any]:
    return {
        "enabled": ASSISTANT_HYBRID_RETRIEVAL_ENABLED,
        "sparse_weight": ASSISTANT_HYBRID_RETRIEVAL_SPARSE_WEIGHT,
        "dense_weight": ASSISTANT_HYBRID_RETRIEVAL_DENSE_WEIGHT,
        "candidate_multiplier": ASSISTANT_HYBRID_RETRIEVAL_CANDIDATE_MULTIPLIER,
        "min_dense_score": ASSISTANT_HYBRID_RETRIEVAL_MIN_DENSE_SCORE,
        "coverage_enabled": ASSISTANT_RETRIEVAL_COVERAGE_ENABLED,
        "min_query_term_coverage": ASSISTANT_RETRIEVAL_MIN_QUERY_TERM_COVERAGE,
        "min_match_count": ASSISTANT_RETRIEVAL_MIN_MATCH_COUNT,
        "min_context_chars": ASSISTANT_RETRIEVAL_MIN_CONTEXT_CHARS,
        "retry_enabled": ASSISTANT_RETRIEVAL_RETRY_ENABLED,
        "max_retry_queries": ASSISTANT_RETRIEVAL_MAX_RETRY_QUERIES,
        "min_clause_terms": ASSISTANT_RETRIEVAL_MIN_CLAUSE_TERMS,
        "rerank_enabled": ASSISTANT_RETRIEVAL_RERANK_ENABLED,
        "rerank_max_phrase_terms": ASSISTANT_RETRIEVAL_RERANK_MAX_PHRASE_TERMS,
        "refuse_on_insufficient": ASSISTANT_RETRIEVAL_REFUSE_ON_INSUFFICIENT,
    }


def _assistant_pipeline_playground_config() -> Dict[str, Any]:
    return {
        "project_min_iterations": PLAYGROUND_PROJECT_MIN_ITERATIONS,
        "project_max_iterations": PLAYGROUND_PROJECT_MAX_ITERATIONS,
        "project_max_steps": PLAYGROUND_PROJECT_MAX_STEPS,
        "llm_max_tokens": PLAYGROUND_LLM_MAX_TOKENS,
    }


def _assistant_pipeline_dependencies() -> AssistantPipelineDependencies:
    """Return dynamic dependency accessors for the extracted assistant pipeline."""

    return AssistantPipelineDependencies(
        current_user=lambda: _current_user(),
        logger=logger,
        json_dumps=lambda value, **kwargs: json.dumps(value, **kwargs),
        new_trace_id=lambda: uuid.uuid4().hex,
        get_rag_store=lambda: rag_store,
        get_rag_metadata_store=lambda: getattr(CENTRAL_STORE, "rag_metadata", None) if CENTRAL_STORE is not None else None,
        get_assistant_conversation_store=lambda: (
            getattr(CENTRAL_STORE, "assistant_conversations", None) if CENTRAL_STORE is not None else None
        ),
        get_assistant_trace_store=lambda: getattr(CENTRAL_STORE, "assistant_traces", None) if CENTRAL_STORE is not None else None,
        get_assistant_cache_store=lambda: (
            getattr(CENTRAL_STORE, "assistant_semantic_cache", None) if CENTRAL_STORE is not None else None
        ),
        get_stt_learning_store=lambda: stt_learning_store,
        get_rag_config=lambda: _assistant_pipeline_rag_config(),
        get_assistant_runtime_config=lambda: _assistant_pipeline_runtime_config(),
        get_assistant_security_config=lambda: _assistant_pipeline_security_config(),
        get_assistant_routing_config=lambda: _assistant_pipeline_routing_config(),
        get_assistant_cache_config=lambda: _assistant_pipeline_cache_config(),
        get_assistant_retrieval_config=lambda: _assistant_pipeline_retrieval_config(),
        get_playground_config=lambda: _assistant_pipeline_playground_config(),
        safe_int=lambda value, default=0: _safe_int(value, default),
        coerce_rag_sources=lambda payload: _coerce_rag_sources(payload),
        build_rag_documents=lambda sources, **kwargs: _build_rag_documents(sources, **kwargs),
        build_rag_index=lambda **kwargs: RagIndex.build(**kwargs),
        render_rag_context=lambda matches: _render_rag_context(matches),
        serialize_rag_match=lambda match: _serialize_rag_match(match),
        capability_summary=lambda **kwargs: capability_summary(**kwargs),
        select_skills=lambda prompt, limit=4: select_skills(prompt, limit=limit),
        format_skill_brief=lambda skills: format_skill_brief(skills),
        is_admin_user=lambda user: _is_admin_user(user),
        mcp_execute=lambda *args, **kwargs: _mcp_execute(*args, **kwargs),
        execute_atlassian_action=lambda user, action_request: _assistant_execute_atlassian_action(user, action_request),
        resolve_llm_settings=lambda **kwargs: _resolve_llm_settings(**kwargs),
        build_request_llm_provider=lambda user, settings, workflow, role: _build_request_llm_provider(
            user,
            settings,
            workflow=workflow,
            role=role,
        ),
        acquire_request_capacity=lambda semaphore, wait_seconds, owner="": _acquire_request_capacity(
            semaphore,
            wait_seconds,
            owner=owner,
        ),
        guardrail_scan=lambda text: _guardrail_scan(text),
        stt_motion_context=lambda payload=None: _stt_motion_context(payload),
        is_marketing_assistant_request=lambda payload, requirements_text: _is_marketing_assistant_request(payload, requirements_text),
        is_simple_greeting=lambda text: _is_simple_greeting(text),
        assistant_memory_scope=lambda route, **kwargs: _assistant_memory_scope(route, **kwargs),
        assistant_memory_query_text=lambda **kwargs: _assistant_memory_query_text(**kwargs),
        assistant_memory_prompt_block=lambda user, **kwargs: _assistant_memory_prompt_block(user, **kwargs),
        assistant_memory_reference_payload=lambda user, **kwargs: _assistant_memory_reference_payload(user, **kwargs),
        should_use_assistant_ask_memory=lambda prompt, **kwargs: _should_use_assistant_ask_memory(prompt, **kwargs),
        record_assistant_memory=lambda user, **kwargs: _record_assistant_memory(user, **kwargs),
        assistant_reply_payload=lambda reply_text, provider, model, payload=None: _assistant_reply_payload(
            reply_text,
            provider,
            model,
            payload=payload,
        ),
        ensure_req_register_in_draft=lambda text: _ensure_req_register_in_draft(text),
        stt_record_learning=lambda text, **kwargs: _stt_record_learning(text, **kwargs),
        extract_json_payload=lambda text: _extract_json_payload(text),
        to_uk_english=lambda text: _to_uk_english(text),
        global_requirements_titles=lambda: _global_requirements_titles(),
        global_requirements_count=lambda: _global_requirements_count(),
        estimate_job_tokens=lambda payload: _estimate_job_tokens(payload),
        opencode_available_for_playground=lambda: _opencode_available_for_playground(),
        submit_subtask=lambda **kwargs: subtask_manager.submit(**kwargs),
        record_runtime_telemetry=lambda owner, event: _record_llm_request_telemetry(
            owner,
            event,
            team_id=_current_team_telemetry_subject(),
        ),
    )


_EXTRACTED_ASSISTANT_HANDLERS = build_assistant_handlers(_assistant_pipeline_dependencies())
_EXTRACTED_RAG_HANDLERS = build_rag_handlers(_assistant_pipeline_dependencies())
_EXTRACTED_ASSISTANT_ADMIN_HANDLERS = build_assistant_admin_handlers(
    current_user=lambda: _current_user(),
    is_admin_user=lambda user: _is_admin_user(user),
    get_assistant_conversation_store=lambda: (
        getattr(CENTRAL_STORE, "assistant_conversations", None) if CENTRAL_STORE is not None else None
    ),
    get_assistant_trace_store=lambda: getattr(CENTRAL_STORE, "assistant_traces", None) if CENTRAL_STORE is not None else None,
    safe_int=lambda value, default=0: _safe_int(value, default),
)


def rag_indexes() -> Response:
    """RAG endpoint for indexes."""

    return _EXTRACTED_RAG_HANDLERS["rag_indexes"]()


def rag_index_create() -> Response:
    """RAG endpoint for index create."""

    return _EXTRACTED_RAG_HANDLERS["rag_index_create"]()


def rag_index_delete(name: str) -> Response:
    """RAG endpoint for index delete."""

    return _EXTRACTED_RAG_HANDLERS["rag_index_delete"](name)


def rag_query() -> Response:
    """RAG endpoint for query."""

    return _EXTRACTED_RAG_HANDLERS["rag_query"]()


def assistant_rag_mcp() -> Response:
    """Handle the assistant RAG + MCP route via the extracted assistant pipeline."""

    return _EXTRACTED_ASSISTANT_HANDLERS["assistant_rag_mcp"]()


def assistant_requirements() -> Response:
    """Handle the assistant requirements route via the extracted assistant pipeline."""

    return _EXTRACTED_ASSISTANT_HANDLERS["assistant_requirements"]()


def assistant_form_fill() -> Response:
    """Handle the assistant form-fill route via the extracted assistant pipeline."""

    return _EXTRACTED_ASSISTANT_HANDLERS["assistant_form_fill"]()


def playground_plan() -> Response:
    """Handle the playground plan route via the extracted assistant pipeline."""

    return _EXTRACTED_ASSISTANT_HANDLERS["playground_plan"]()


def execution_plan() -> Response:
    """Handle the execution plan route via the extracted assistant pipeline."""

    return _EXTRACTED_ASSISTANT_HANDLERS["execution_plan"]()


def _sse(entry: Dict[str, Any]) -> str:
    payload = json.dumps(entry)
    return f"data: {payload}\n\n"


if hasattr(app, "add_url_rule"):
    register_admin_routes(
        app,
        metrics_path=METRICS_PATH,
        index=index,
        playground=playground,
        admin_dashboard=admin_dashboard,
        public_asset=public_asset,
        favicon=favicon,
        metrics=metrics,
        setup=setup,
        health=health,
        api_version=api_version,
        capabilities_report=capabilities_report,
        admin_stats=admin_stats,
        admin_llm_telemetry=admin_llm_telemetry,
        admin_ai_orchestration=admin_ai_orchestration,
        assistant_admin_conversations=_EXTRACTED_ASSISTANT_ADMIN_HANDLERS["assistant_admin_conversations"],
        assistant_admin_conversation_detail=_EXTRACTED_ASSISTANT_ADMIN_HANDLERS["assistant_admin_conversation_detail"],
        assistant_admin_traces=_EXTRACTED_ASSISTANT_ADMIN_HANDLERS["assistant_admin_traces"],
        assistant_admin_trace_detail=_EXTRACTED_ASSISTANT_ADMIN_HANDLERS["assistant_admin_trace_detail"],
        workers_telemetry=workers_telemetry,
        api_audit=api_audit,
    )
    register_auth_routes(
        app,
        login=login,
        register=register,
        external_login=external_login,
        oidc_login=oidc_login,
        oidc_callback=oidc_callback,
        api_oidc_exchange=api_oidc_exchange,
        sso_login=sso_login,
        logout=logout,
        api_login=api_login,
        api_login_mfa_totp=api_login_mfa_totp,
        api_setup=api_setup,
        api_register=api_register,
        api_auth_config=api_auth_config,
        api_sso_issue=api_sso_issue,
        api_logout=api_logout,
        api_session=api_session,
        api_authz_nginx=api_authz_nginx,
        api_profile=api_profile,
        api_profile_password=api_profile_password,
        api_profile_mfa_totp_start=api_profile_mfa_totp_start,
        api_profile_mfa_totp_verify=api_profile_mfa_totp_verify,
        api_profile_mfa_totp_disable=api_profile_mfa_totp_disable,
        api_profile_passkeys_register_options=api_profile_passkeys_register_options,
        api_profile_passkeys_register_verify=api_profile_passkeys_register_verify,
        api_profile_passkey_delete=api_profile_passkey_delete,
        api_passkeys_authenticate_options=api_passkeys_authenticate_options,
        api_passkeys_authenticate_verify=api_passkeys_authenticate_verify,
        api_users=api_users,
        api_user_password=api_user_password,
    )
    register_voice_routes(
        app,
        api_voice_tokens=api_voice_tokens,
        api_voice_token_delete=api_voice_token_delete,
        api_voice_capture=api_voice_capture,
        api_voice_siri=api_voice_siri,
        api_voice_alexa=api_voice_alexa,
        api_voice_google=api_voice_google,
        api_voice_stt=api_voice_stt,
    )
    register_assistant_routes(
        app,
        assistant_rag_mcp=assistant_rag_mcp,
        assistant_requirements=assistant_requirements,
        assistant_form_fill=assistant_form_fill,
        playground_plan=playground_plan,
        execution_plan=execution_plan,
    )
    register_rag_routes(
        app,
        rag_indexes=rag_indexes,
        rag_index_create=rag_index_create,
        rag_index_delete=rag_index_delete,
        rag_query=rag_query,
    )
    register_jobs_routes(
        app,
        {
            "api_todos": api_todos,
            "api_todo_next": api_todo_next,
            "api_todo_route": api_todo_route,
            "api_todo_detail": api_todo_detail,
            "api_todo_schedule": api_todo_schedule,
            "api_schedules": api_schedules,
            "api_schedule_detail": api_schedule_detail,
            "api_schedule_cancel": api_schedule_cancel,
            "api_subtasks": api_subtasks,
            "api_subtask_detail": api_subtask_detail,
            "api_subtask_cancel": api_subtask_cancel,
            "api_projects": api_projects,
            "api_project_detail": api_project_detail,
            "api_teams": api_teams,
            "api_team_detail": api_team_detail,
            "api_team_invite": api_team_invite,
            "api_team_invitation_accept": api_team_invitation_accept,
            "api_team_invitation_reject": api_team_invitation_reject,
            "api_team_leave": api_team_leave,
            "api_team_tokens": api_team_tokens,
            "api_access_tree": api_access_tree,
            "api_sessions": api_sessions,
            "api_session_detail": api_session_detail,
            "api_session_leave": api_session_leave,
            "api_session_stream": api_session_stream,
            "api_session_history": api_session_history,
            "api_sessions_history": api_sessions_history,
            "job_estimate": job_estimate,
            "import_requirements": import_requirements,
            "export_requirements": export_requirements,
            "mcp_servers": mcp_servers,
            "mcp_server_delete": mcp_server_delete,
            "mcp_server_tools": mcp_server_tools,
            "mcp_server_call": mcp_server_call,
            "mcp_server_resources": mcp_server_resources,
            "mcp_server_resource": mcp_server_resource,
            "request_refund": request_refund,
            "list_refunds": list_refunds,
            "screen_refund": screen_refund,
            "decide_refund": decide_refund,
            "refund_file": refund_file,
            "jobs": jobs,
            "job_detail": job_detail,
            "job_workspace": job_workspace,
            "job_workspace_open": job_workspace_open,
            "job_tasks": job_tasks,
            "job_task_detail": job_task_detail,
            "job_task_cancel": job_task_cancel,
            "job_editor_roots": job_editor_roots,
            "job_editor_list": job_editor_list,
            "job_editor_file": job_editor_file,
            "job_editor_ops": job_editor_ops,
            "job_requirements_progress": job_requirements_progress,
            "job_requirements_summary": job_requirements_summary,
            "job_logs": job_logs,
            "job_logs_stream": job_logs_stream,
            "job_actions": job_actions,
            "job_transfer": job_transfer,
            "job_archive": job_archive,
            "jobs_bulk_delete": jobs_bulk_delete,
            "billing_dashboard": billing_dashboard,
            "billing_dashboard_admin": billing_dashboard_admin,
            "billing_dashboard_asset": billing_dashboard_asset,
            "billing_dashboard_customer": billing_dashboard_customer,
            "billing_dashboard_admin_data": billing_dashboard_admin_data,
            "tokens": tokens,
            "tokens_ledger": tokens_ledger,
            "secrets": secrets,
            "delete_secret": delete_secret,
            "github_tree": github_tree,
        },
    )
    if TODO_SCHEDULER_ENABLED:
        todo_scheduler.start()
    if LLM_TELEMETRY_RETENTION_HOURS > 0:
        llm_telemetry_janitor.start()
    ai_model_inventory_monitor.start()


def main() -> int:
    """Run the Refiner backend Flask application."""
    _rehydrate_job_manager_from_disk()

    # Register API documentation and health endpoints
    try:
        from refiner.api_docs import add_api_documentation_support
        add_api_documentation_support(
            app,
            stt_server_url=STT_SERVER_URL,
            redis_enabled=lambda: _env_flag("REFINER_ENABLE_REDIS", False),
            continuum_enabled=_continuum_enabled,
        )
        logger.info("API documentation enabled at /api/docs")
    except Exception as exc:
        logger.warning(f"Failed to register API documentation: {exc}")

    host = os.getenv("REFINER_HOST", "127.0.0.1")
    port = int(os.getenv("REFINER_PORT", "5001"))
    debug = os.getenv("REFINER_DEBUG", "0") in {"1", "true", "yes"}
    app.run(host=host, port=port, debug=debug, threaded=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
